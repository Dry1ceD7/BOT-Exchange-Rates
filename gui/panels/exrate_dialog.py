#!/usr/bin/env python3
"""
gui/panels/exrate_dialog.py
---------------------------------------------------------------------------
ExRate Sheet creation dialog — standalone TopLevel window.
Extracted from gui/app.py to reduce God Object line count.
---------------------------------------------------------------------------
"""

import asyncio
import calendar
import contextlib
import logging
import os
import shutil
import tempfile
import threading
from datetime import date
from pathlib import Path
from tkinter import filedialog

import customtkinter as ctk
import httpx

from core.constants import humanize_save_error
from core.i18n import tr
from gui.theme import get_theme

logger = logging.getLogger(__name__)

# Remembered across dialog invocations so the save picker reopens where the
# user last saved instead of defaulting to an arbitrary directory each time.
_LAST_SAVE_DIR: str | None = None

# ── Constants ────────────────────────────────────────────────────────────
EXRATE_CURRENCIES = [
    "USD", "EUR", "GBP", "JPY", "CNY", "HKD", "SGD", "AUD", "CHF",
]

EXRATE_RATE_TYPES = {
    "Buying TT":    "buying_transfer",
    "Buying Sight": "buying_sight",
    "Selling":      "selling",
    "Mid Rate":     "mid_rate",
}


class _ExRateCancelled(Exception):
    """Raised inside the progress callback to abort an in-flight fetch (#2).

    The standalone updater calls ``progress_cb`` at every step — before each
    currency fetch and the holiday fetch — and does not swallow its exceptions,
    so raising from the callback cooperatively unwinds the worker's asyncio
    loop between network calls (the same between-step granularity the batch
    engine uses for its stop_event). This lets the dialog offer a Cancel
    affordance without the worker busy-waiting on a 4GB PC.
    """


def show_exrate_dialog(app) -> None:
    """Show the ExRate creation options dialog. Calls back into *app* for
    status/progress updates.

    Args:
        app: The BOTExrateApp instance (parent window).
    """
    t = get_theme()

    dialog = ctk.CTkToplevel(app)
    dialog.title(tr("exrate.window_title"))
    dialog.geometry("440x680")
    dialog.resizable(False, False)
    dialog.configure(fg_color=t["card_bg"])
    dialog.transient(app)
    dialog.grab_set()

    dialog.update_idletasks()
    sx = (dialog.winfo_screenwidth() - 440) // 2
    sy = (dialog.winfo_screenheight() - 680) // 2
    dialog.geometry(f"440x680+{sx}+{sy}")

    ctk.CTkLabel(
        dialog, text=tr("exrate.heading"),
        font=ctk.CTkFont(size=18, weight="bold"),
        text_color=t["text_primary"],
    ).pack(pady=(16, 12))

    # ── Currencies ────────────────────────────────────────────────────
    ctk.CTkLabel(
        dialog, text=tr("exrate.section_currencies"),
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color=t["text_secondary"],
    ).pack(anchor="w", padx=24, pady=(0, 4))

    cur_frame = ctk.CTkFrame(
        dialog, fg_color=t["section_bg"], corner_radius=8,
    )
    cur_frame.pack(fill="x", padx=24, pady=(0, 12))

    cur_vars = {}
    DEFAULTS_ON = {"USD", "EUR"}
    row_frame = None
    for i, ccy in enumerate(EXRATE_CURRENCIES):
        if i % 3 == 0:
            row_frame = ctk.CTkFrame(cur_frame, fg_color="transparent")
            row_frame.pack(fill="x", padx=8, pady=2)
        var = ctk.BooleanVar(value=ccy in DEFAULTS_ON)
        cur_vars[ccy] = var
        ctk.CTkCheckBox(
            row_frame, text=ccy, variable=var,
            font=ctk.CTkFont(size=13),
            text_color=t["text_primary"],
            fg_color=t["accent_indigo"], hover_color=t["accent_indigo_hover"],
            width=120,
        ).pack(side="left", padx=4, pady=2)

    # ── Rate Types ────────────────────────────────────────────────────
    ctk.CTkLabel(
        dialog, text=tr("exrate.section_rate_types"),
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color=t["text_secondary"],
    ).pack(anchor="w", padx=24, pady=(0, 4))

    rate_frame = ctk.CTkFrame(
        dialog, fg_color=t["section_bg"], corner_radius=8,
    )
    rate_frame.pack(fill="x", padx=24, pady=(0, 16))

    rate_vars = {}
    RATE_DEFAULTS = {"Buying TT", "Selling"}
    row_frame2 = None
    for i, (label, _) in enumerate(EXRATE_RATE_TYPES.items()):
        if i % 2 == 0:
            row_frame2 = ctk.CTkFrame(rate_frame, fg_color="transparent")
            row_frame2.pack(fill="x", padx=8, pady=2)
        var = ctk.BooleanVar(value=label in RATE_DEFAULTS)
        rate_vars[label] = var
        ctk.CTkCheckBox(
            row_frame2, text=label, variable=var,
            font=ctk.CTkFont(size=13),
            text_color=t["text_primary"],
            fg_color=t["accent_indigo"], hover_color=t["accent_indigo_hover"],
            width=180,
        ).pack(side="left", padx=4, pady=2)

    # ── Date Range ────────────────────────────────────────────────────
    ctk.CTkLabel(
        dialog, text=tr("exrate.section_date_range"),
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color=t["text_secondary"],
    ).pack(anchor="w", padx=24, pady=(0, 4))

    date_range_frame = ctk.CTkFrame(
        dialog, fg_color=t["section_bg"], corner_radius=8,
    )
    date_range_frame.pack(fill="x", padx=24, pady=(0, 16))

    date_mode_var = ctk.StringVar(value="auto")
    today = date.today()

    auto_label = ctk.CTkLabel(
        date_range_frame,
        text=f"  Auto: {today.year}-01-01 → {today.strftime('%Y-%m-%d')}",
        font=ctk.CTkFont(size=12),
        text_color=t["text_secondary"],
    )

    ctk.CTkSwitch(
        date_range_frame,
        text=tr("exrate.manual_toggle"),
        variable=date_mode_var,
        onvalue="manual", offvalue="auto",
        font=ctk.CTkFont(size=13),
        text_color=t["text_primary"],
        progress_color=t["accent_indigo"],
        command=lambda: _toggle_date_mode(),
    ).pack(anchor="w", padx=12, pady=(8, 0))

    auto_label.pack(anchor="w", padx=12, pady=(2, 8))

    # Manual date inputs (initially hidden)
    manual_frame = ctk.CTkFrame(date_range_frame, fg_color="transparent")
    years = [str(y) for y in range(today.year - 5, today.year + 2)]
    months = [f"{m:02d}" for m in range(1, 13)]

    def _refresh_day_options(year_box, month_box, day_box):
        """Constrain a day combobox to the days valid for its month/year (#1).

        Uses calendar.monthrange so February shows 28/29 and 30-day months
        show 30 — no more offering 31 everywhere and failing only at Create
        time. The currently selected day is clamped down if it now exceeds the
        month length (e.g. 31 → 30 when switching to April, or → 28/29 for Feb).
        """
        try:
            yr = int(year_box.get())
            mo = int(month_box.get())
        except (ValueError, TypeError):
            return
        if not 1 <= mo <= 12:
            return
        _, last_day = calendar.monthrange(yr, mo)
        day_box.configure(values=[f"{d:02d}" for d in range(1, last_day + 1)])
        with contextlib.suppress(ValueError, TypeError):
            if int(day_box.get()) > last_day:
                day_box.set(f"{last_day:02d}")

    # Start date row
    start_row = ctk.CTkFrame(manual_frame, fg_color="transparent")
    start_row.pack(fill="x", padx=8, pady=2)
    ctk.CTkLabel(start_row, text=tr("exrate.label_start"),
                 font=ctk.CTkFont(size=12),
                 text_color=t["text_secondary"], width=40).pack(side="left")
    start_year = ctk.CTkComboBox(
        start_row, values=years, width=80,
        font=ctk.CTkFont(size=12),
        fg_color=t["combo_bg"], border_color=t["combo_border"],
        text_color=t["text_primary"],
        command=lambda _v: _refresh_day_options(start_year, start_month, start_day),
    )
    start_year.set(str(today.year))
    start_year.pack(side="left", padx=2)
    start_month = ctk.CTkComboBox(
        start_row, values=months, width=60,
        font=ctk.CTkFont(size=12),
        fg_color=t["combo_bg"], border_color=t["combo_border"],
        text_color=t["text_primary"],
        command=lambda _v: _refresh_day_options(start_year, start_month, start_day),
    )
    start_month.set("01")
    start_month.pack(side="left", padx=2)
    start_day = ctk.CTkComboBox(
        start_row, values=[f"{d:02d}" for d in range(1, 32)], width=60,
        font=ctk.CTkFont(size=12),
        fg_color=t["combo_bg"], border_color=t["combo_border"],
        text_color=t["text_primary"],
    )
    start_day.set("01")
    start_day.pack(side="left", padx=2)

    # End date row
    end_row = ctk.CTkFrame(manual_frame, fg_color="transparent")
    end_row.pack(fill="x", padx=8, pady=2)
    ctk.CTkLabel(end_row, text=tr("exrate.label_end"),
                 font=ctk.CTkFont(size=12),
                 text_color=t["text_secondary"], width=40).pack(side="left")
    end_year = ctk.CTkComboBox(
        end_row, values=years, width=80,
        font=ctk.CTkFont(size=12),
        fg_color=t["combo_bg"], border_color=t["combo_border"],
        text_color=t["text_primary"],
        command=lambda _v: _refresh_day_options(end_year, end_month, end_day),
    )
    end_year.set(str(today.year))
    end_year.pack(side="left", padx=2)
    end_month = ctk.CTkComboBox(
        end_row, values=months, width=60,
        font=ctk.CTkFont(size=12),
        fg_color=t["combo_bg"], border_color=t["combo_border"],
        text_color=t["text_primary"],
        command=lambda _v: _refresh_day_options(end_year, end_month, end_day),
    )
    end_month.set(f"{today.month:02d}")
    end_month.pack(side="left", padx=2)
    end_day = ctk.CTkComboBox(
        end_row, values=[f"{d:02d}" for d in range(1, 32)], width=60,
        font=ctk.CTkFont(size=12),
        fg_color=t["combo_bg"], border_color=t["combo_border"],
        text_color=t["text_primary"],
    )
    end_day.set(f"{today.day:02d}")
    end_day.pack(side="left", padx=2)

    # Constrain both day lists to the initial month/year on build so the very
    # first manual selection already reflects the real month length (#1).
    _refresh_day_options(start_year, start_month, start_day)
    _refresh_day_options(end_year, end_month, end_day)

    def _toggle_date_mode():
        if date_mode_var.get() == "manual":
            auto_label.pack_forget()
            manual_frame.pack(fill="x", pady=(0, 8))
        else:
            manual_frame.pack_forget()
            auto_label.pack(anchor="w", padx=12, pady=(2, 8))

    # ── Validation error label (reused, not stacked) ────────────────────
    _err_label = ctk.CTkLabel(
        dialog, text="",
        text_color=t["error_text"],
        font=ctk.CTkFont(size=12),
    )
    _err_label.pack(pady=(0, 4))

    # ── Create Button ─────────────────────────────────────────────────
    def _on_create():
        _err_label.configure(text="")  # clear previous error

        # Refuse to start while a batch owns the shared status/progress UI.
        # The ExRate Sheet button is disabled during a batch, but the dialog
        # may have been opened first; without this guard a batch started after
        # the dialog opened would have its progress bar/status hijacked by a
        # second concurrent engine run (#4).
        if getattr(app, "_batch_running", False):
            _err_label.configure(text=tr("exrate.err_batch_running"))
            return

        currencies = [c for c, v in cur_vars.items() if v.get()]
        rate_types = {
            lbl: EXRATE_RATE_TYPES[lbl]
            for lbl, v in rate_vars.items() if v.get()
        }
        if not currencies:
            _err_label.configure(text=tr("exrate.err_no_currency"))
            return
        if not rate_types:
            _err_label.configure(text=tr("exrate.err_no_rate_type"))
            return

        # Get date range
        if date_mode_var.get() == "manual":
            try:
                s_date = date(int(start_year.get()), int(start_month.get()),
                              int(start_day.get()))
                e_date = date(int(end_year.get()), int(end_month.get()),
                              int(end_day.get()))
            except ValueError:
                _err_label.configure(text=tr("exrate.err_invalid_date"))
                return
            date_range = (s_date, e_date)
        else:
            date_range = None  # auto = current year

        dialog.destroy()
        _create_exrate_file(app, currencies, rate_types, date_range=date_range)

    create_btn = ctk.CTkButton(
        dialog, text=tr("exrate.btn_create"),
        fg_color=t["accent_indigo"], hover_color=t["accent_indigo_hover"],
        font=ctk.CTkFont(size=14, weight="bold"),
        corner_radius=10, height=44,
        command=_on_create,
    )
    create_btn.pack(padx=24, fill="x", pady=(0, 12))

    # ── Keyboard handling — match every other modal in the app (#3) ─────
    # Escape cancels, Return confirms. focus_set gives the dialog keyboard
    # focus so the bindings fire without a prior mouse click.
    dialog.bind("<Escape>", lambda _e: dialog.destroy())
    dialog.bind("<Return>", lambda _e: _on_create())
    dialog.focus_set()
    create_btn.focus_set()


def _build_exrate_summary(currencies, rate_types, date_range) -> str:
    """Build a one-line summary of what an ExRate creation will write (#1).

    Reports the populated date span (day count), the currencies, and the rate
    types requested, so a file full of blanks no longer looks identical to a
    full success. The day count is the inclusive calendar span the standalone
    writer iterates (it writes one row per day in the range).

    Args:
        currencies: List of currency codes.
        rate_types: Dict of {label: api_key} (only the labels are surfaced).
        date_range: Optional (start_date, end_date) tuple; None means auto
            (current year, Jan 1 → today).

    Returns:
        A human-readable summary string (no leading/trailing markers).
    """
    if date_range:
        s_date, e_date = date_range
    else:
        today = date.today()
        s_date, e_date = date(today.year, 1, 1), today
    days = (e_date - s_date).days + 1 if e_date >= s_date else 0
    ccy_list = ", ".join(currencies)
    rate_list = ", ".join(rate_types.keys())
    return (
        f"{days} day{'s' if days != 1 else ''} "
        f"({s_date:%Y-%m-%d} → {e_date:%Y-%m-%d}) · {ccy_list} · {rate_list}"
    )


def _verify_exrate_dest(dest: str) -> None:
    """Light structural read-back of the moved destination file (F50/F201).

    The engine path already proved every written cell against the in-memory
    expected values BEFORE its atomic save (core/workbook_io.atomic_save's
    verify hook), so this only re-checks the file that actually landed at
    ``dest`` after shutil.move: it must reopen as a workbook, carry an
    "ExRate" sheet, and have a populated row 2. One cheap read-only pass.

    Raises:
        ValueError: On any structural failure (including an unreadable
            file), so the worker's existing error path surfaces it.
    """
    import gc

    from openpyxl import load_workbook

    wb = None
    try:
        try:
            wb = load_workbook(dest, read_only=True, data_only=False)
            if "ExRate" not in wb.sheetnames:
                raise ValueError("no ExRate sheet present")
            ws = wb["ExRate"]
            row2 = next(
                ws.iter_rows(min_row=2, max_row=2, values_only=True), None,
            )
            if row2 is None or all(v in (None, "") for v in row2):
                raise ValueError("ExRate sheet has no data in row 2")
        except Exception as exc:
            raise ValueError(
                "Post-write verification failed for "
                f"{Path(dest).name}: {exc}"
            ) from exc
    finally:
        if wb is not None:
            with contextlib.suppress(OSError):
                wb.close()
        del wb
        gc.collect()


def _create_exrate_file(app, currencies, rate_types, date_range=None):
    """Create a new standalone ExRate file — fully independent, pulls from BOT API.

    Uses a callback interface to communicate with the parent window.
    The `app` parameter is only used for:
      - app.after() (thread-safe scheduling)
      - app.event_bus (for LedgerEngine progress)
    All UI updates go through callbacks registered on `app` via the
    _get_ui_callbacks() helper.

    Args:
        app: The BOTExrateApp instance for scheduling + event bus.
        currencies: List of currency codes.
        rate_types: Dict of {label: api_key}.
        date_range: Optional (start_date, end_date) tuple.
    """
    global _LAST_SAVE_DIR
    t = get_theme()

    # Belt-and-suspenders: never let an ExRate run start while a batch owns the
    # shared progress/status widgets (#4). _on_create already guards, but the
    # save picker below can sit open long enough for a batch to start.
    if getattr(app, "_batch_running", False):
        if hasattr(app, "lbl_status"):
            app.lbl_status.configure(
                text=tr("exrate.err_batch_running"),
                text_color=t["warning"],
            )
        return

    dest = filedialog.asksaveasfilename(
        title="Save ExRate File",
        initialdir=_LAST_SAVE_DIR or None,
        initialfile="ExRate.xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        defaultextension=".xlsx",
        confirmoverwrite=True,
    )
    if not dest:
        return
    # Remember the chosen directory so the next save reopens there (#2).
    _LAST_SAVE_DIR = str(Path(dest).parent) or _LAST_SAVE_DIR

    summary = _build_exrate_summary(currencies, rate_types, date_range)

    # ── Callback interface — decoupled from widget internals ─────────
    def _set_export_button(state: str):
        if hasattr(app, "btn_export_exrate"):
            app.btn_export_exrate.configure(state=state)

    def _set_status(text: str, color: str):
        if hasattr(app, "lbl_status"):
            app.lbl_status.configure(text=text, text_color=color)

    def _set_progress(mode: str = "determinate", value: float = 0, running: bool = False):
        if hasattr(app, "progressbar"):
            if mode:
                app.progressbar.configure(mode=mode)
            if running:
                app.progressbar.start()
            else:
                app.progressbar.stop()
                app.progressbar.set(value)

    def _on_complete(dest_path: str):
        if hasattr(app, "last_processed_path"):
            app.last_processed_path = dest_path
        if hasattr(app, "btn_reveal"):
            app.btn_reveal.pack(pady=(12, 14))

    # ── Busy/Cancel affordance (#2) ──────────────────────────────────
    # The dialog is already gone by the time the fetch runs; without this the
    # only sign of life is the indeterminate bar on the main card. A multi-
    # currency span fetch can take many seconds on a 4GB PC with no way out.
    # We surface a transient Cancel button on the main card (app.card) that
    # sets cancel_event; the worker's status callback checks it between network
    # steps and unwinds cooperatively.
    cancel_event = threading.Event()
    cancel_btn: ctk.CTkButton | None = None

    def _show_cancel_button():
        nonlocal cancel_btn
        card = getattr(app, "card", None)
        if card is None or cancel_btn is not None:
            return
        with contextlib.suppress(Exception):
            cancel_btn = ctk.CTkButton(
                card, text=tr("exrate.btn_cancel"),
                height=32, width=160,
                fg_color=t["warning"], hover_color=t["warning_hover"],
                font=ctk.CTkFont(size=12, weight="bold"),
                corner_radius=8,
                command=lambda: (cancel_event.set(), _set_status(
                    tr("exrate.cancelling"), t["warning"])),
            )
            cancel_btn.pack(pady=(6, 0))

    def _destroy_cancel_button():
        nonlocal cancel_btn
        if cancel_btn is not None:
            with contextlib.suppress(Exception):
                cancel_btn.destroy()
            cancel_btn = None

    # ── Start: disable button + indeterminate progress ───────────────
    _set_export_button("disabled")
    _set_status(tr("exrate.creating"), t["text_secondary"])
    _set_progress("indeterminate", running=True)
    _show_cancel_button()
    app.update_idletasks()

    def _status_cb(msg: str):
        # Cooperative cancellation: the standalone updater calls this at every
        # step and propagates the exception, so raising here aborts the fetch
        # between currency/holiday network calls (#2).
        if cancel_event.is_set():
            raise _ExRateCancelled
        # app destroyed during ExRate generation
        with contextlib.suppress(RuntimeError):
            app.after(0, _set_status, msg, t["text_secondary"])

    def _done(success: bool, message: str, *, color: str | None = None):
        """Main-thread callback to restore UI state."""
        _destroy_cancel_button()
        _set_progress("determinate", value=1.0 if success else 0.0)
        if success:
            _set_status(message, t["success"])
            _on_complete(dest)
        else:
            _set_status(message, color or t["error_text"])
        _set_export_button("normal")

    def _fail_message(exc: BaseException) -> str:
        """Translate a worker exception into actionable guidance (#5).

        A locked/open destination becomes the shared "close it in Excel"
        message via humanize_save_error; everything else keeps a short
        plain-language reason instead of leaking a raw errno/traceback string.
        """
        humanized = humanize_save_error(Path(dest).name, exc)
        if humanized:
            return humanized
        if isinstance(exc, (httpx.RequestError, httpx.HTTPStatusError)):
            return "Could not reach the BOT server — check your connection."
        return f"Failed: {exc}"

    event_bus = getattr(app, "event_bus", None)

    def _worker():
        from openpyxl import Workbook

        from core.api_client import CLIENT_TIMEOUT, BOTClient
        from core.backup_manager import BackupError
        from core.engine import LedgerEngine

        # Build into a temp file in the destination directory and only move it
        # into place once the fetch+write fully succeeds (#2). The previous
        # flow saved a BLANK workbook straight over `dest` before fetching, so
        # an API/network failure left the user's chosen file destroyed. Writing
        # to a sibling temp keeps the original intact until success; on failure
        # the temp is discarded and `dest` is never touched.
        tmp_path: str | None = None
        # Captured from the engine inside _run so the success branch can back
        # up an existing dest with the engine's OWN BackupManager (F10) — no
        # second manager instance, same data/backups/ directory and naming.
        backup_mgr = None
        try:
            dest_dir = Path(dest).parent
            fd, tmp_path = tempfile.mkstemp(
                suffix=".xlsx", prefix=".exrate_tmp_", dir=dest_dir
            )
            os.close(fd)
            wb = Workbook()
            ws = wb.active
            ws.title = "ExRate"
            wb.save(tmp_path)
            wb.close()

            async def _run():
                nonlocal backup_mgr
                async with httpx.AsyncClient(timeout=CLIENT_TIMEOUT) as client:
                    api = BOTClient(client)
                    engine = LedgerEngine(api, event_bus=event_bus)
                    backup_mgr = getattr(engine, "backup", None)
                    return await engine.update_exrate_standalone(
                        tmp_path,
                        progress_cb=_status_cb,
                        currencies=currencies,
                        rate_types=rate_types,
                        date_range=date_range,
                    )

            loop = asyncio.new_event_loop()
            try:
                loop.run_until_complete(_run())
                # Backup-first (F10): with confirmoverwrite the user may pick
                # an EXISTING workbook as dest. The engine only backed up the
                # blank TEMP (its run target), so without this step the real
                # file would be replaced with NO backup — making Revert
                # impossible. Back dest up BEFORE the move; if the backup
                # fails, BackupError aborts the move below and dest is left
                # intact (same fail-safe rule as the batch pipeline). The
                # useless .exrate_tmp_* backup of the blank temp remains —
                # suppressing it needs a flag on the updater's run() signature
                # (owned elsewhere), so its cleanup is deferred to a later wave.
                if backup_mgr is not None and Path(dest).exists():
                    backup_mgr.create_backup(dest)
                # Success: atomically move the fully-written temp over dest.
                shutil.move(tmp_path, dest)
                tmp_path = None
                # Structural read-back of what actually landed at dest — a
                # botched move/unreadable file must fail loudly, not report
                # success (the per-cell verification already ran inside the
                # engine's atomic save against the temp).
                _verify_exrate_dest(dest)
                with contextlib.suppress(RuntimeError):
                    app.after(0, _done, True,
                              f"✓ ExRate created: {Path(dest).name} — {summary}")
            except _ExRateCancelled:
                # User pressed Cancel — the temp is discarded in the outer
                # finally and `dest` was never touched, so no data is lost (#2).
                logger.info("ExRate standalone cancelled by user")
                with contextlib.suppress(RuntimeError):
                    app.after(0, lambda: _done(
                        False, tr("exrate.cancelled"), color=t["warning"]))
            except (httpx.RequestError, httpx.HTTPStatusError,
                    OSError, ValueError, BackupError) as e:
                logger.error("ExRate standalone failed: %s", e)
                with contextlib.suppress(RuntimeError):
                    app.after(0, _done, False, _fail_message(e))
            finally:
                loop.close()
        except (OSError, ValueError) as e:
            logger.error("ExRate file creation failed: %s", e)
            with contextlib.suppress(RuntimeError):
                app.after(0, _done, False, _fail_message(e))
        finally:
            # Discard the temp file on any failure path so we never leave a
            # blank/partial sibling behind next to the user's real files.
            if tmp_path:
                with contextlib.suppress(OSError):
                    Path(tmp_path).unlink()

    threading.Thread(target=_worker, daemon=True, name="ExRateWorker").start()

