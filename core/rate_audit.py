#!/usr/bin/env python3
"""
core/rate_audit.py
---------------------------------------------------------------------------
BOT Exchange Rate Processor — ExRate Sheet Rate Auditor
---------------------------------------------------------------------------
Verifies the hard USD/EUR rate values already written into an existing
workbook's "ExRate" master sheet against the authoritative Bank of Thailand
values, and proposes / applies exact 4-decimal Decimal corrections.

This module is the PURE-LOGIC layer: the comparison and the cell rewrite. It
holds no network, no backup, and no file I/O orchestration — those live in the
caller (the GUI handler / a standalone runner), which fetches BOT rates, backs
the file up first, and saves. Keeping the comparison pure makes the financial
logic unit-testable without a workbook on disk or a live API.

CRITICAL INVARIANT — weekend / holiday rows are NEVER touched. BOT publishes
no rate on those days, and the ExRate sheet intentionally keeps their rate
cells blank (the v3.2.8 behavior restored in core/exrate_sheet.py). Filling
them would re-introduce the carry-forward bug. The scanner skips any row whose
Date is a Saturday/Sunday or a BOT holiday.
"""

import contextlib
import gc
import logging
import os
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from core.audit_logger import AuditLogger
from core.constants import MIN_DISK_SPACE_MB
from core.constants import parse_date as _parse_date

# ExRate fixed rate columns + Date header come from the single layout source
# (core/exrate_sheet.py). Re-exported here so existing importers keep working.
from core.exrate_sheet import EXRATE_DATE_HEADER, EXRATE_RATE_COLUMNS
from core.logic import safe_to_decimal
from core.workbook_io import atomic_save, build_cell_verifier, ensure_disk_space

logger = logging.getLogger(__name__)

DATA_START_ROW = 2

# Refusal message for a sheet whose A-E headers do not match the standard
# layout. Surfaced verbatim in the report and the run() status/error.
LAYOUT_ERROR_MSG = (
    "Non-standard ExRate layout — audit supports only the standard USD/EUR sheet"
)


def validate_exrate_layout(ws) -> str | None:
    """Return :data:`LAYOUT_ERROR_MSG` when row 1 isn't the standard layout.

    The scanner hard-codes the B-E column meanings (EXRATE_RATE_COLUMNS), but
    the app's ExRate dialog can also produce CUSTOM-layout sheets named
    "ExRate" (e.g. "Date | GBP Buying TT | GBP Selling | Holidays/Weekend").
    Auditing one of those would overwrite foreign-currency cells with USD/EUR
    values, so columns A-E must carry the canonical core/exrate_sheet.py
    header labels (whitespace/case tolerant). Extra columns beyond E — a
    standard sheet with appended extra currencies — are valid and ignored.
    """
    expected = [(1, EXRATE_DATE_HEADER)]
    expected += [(col, label) for col, label, _ccy, _rt in EXRATE_RATE_COLUMNS]
    for col, label in expected:
        actual = ws.cell(row=1, column=col).value
        if (
            not isinstance(actual, str)
            or actual.strip().casefold() != label.casefold()
        ):
            return LAYOUT_ERROR_MSG
    return None


def rate_key(currency: str, rate_type: str) -> str:
    """Lookup key for the bot_rates map: ``"USD:buying_transfer"`` etc."""
    return f"{currency}:{rate_type}"


def _stored_decimal(raw) -> Decimal | None:
    """The cell's literal stored payload as an UNquantized Decimal.

    Unlike safe_to_decimal this does NOT quantize — it exposes the exact
    representation sitting in the cell, so the scanner can tell a canonical
    4dp value apart from a >4dp float-noise encoding of the same value
    (e.g. 32.50009999 for BOT 32.5001). None when the payload isn't numeric.
    """
    try:
        return Decimal(str(raw))
    except (InvalidOperation, TypeError, ValueError):
        return None


@dataclass(slots=True)
class RateChange:
    """A single proposed/applied correction to one ExRate rate cell."""

    row: int
    col: int
    cell: str  # e.g. "B5", for display
    rate_date: date
    column_label: str
    currency: str
    rate_type: str
    old_value: Decimal | None
    new_value: Decimal
    reason: str


@dataclass
class RateAuditReport:
    """Outcome of auditing one workbook's ExRate sheet against BOT."""

    file: str = ""
    sheet: str = "ExRate"
    scanned_rows: int = 0
    compared_cells: int = 0
    changes: list[RateChange] = field(default_factory=list)
    # Trading-day cells that held a value but BOT had no rate to compare
    # against (e.g. a date beyond the available BOT history). Left untouched.
    unverifiable: int = 0
    # Set (LAYOUT_ERROR_MSG) when row 1 isn't the standard A-E layout — the
    # scan aborted with zero corrections and nothing may be written.
    layout_error: str | None = None
    backup_path: str | None = None
    applied: bool = False

    @property
    def change_count(self) -> int:
        return len(self.changes)


def scan_exrate_corrections(
    ws,
    bot_rates: dict[str, dict[date, Decimal]],
    holidays_set: set[date] | None,
    *,
    data_start_row: int = DATA_START_ROW,
) -> RateAuditReport:
    """Compare each TRADING-DAY ExRate rate cell to the authoritative BOT value.

    Args:
        ws: The openpyxl "ExRate" worksheet.
        bot_rates: ``{f"{ccy}:{rate_type}": {date: Decimal}}`` — the BOT values
            keyed by currency + api field, as fetched by the caller.
        holidays_set: BOT holidays (any iterable of ``date``). Rows on these
            dates — and on weekends — are skipped entirely.
        data_start_row: First data row (row 1 is the header).

    Returns:
        A :class:`RateAuditReport` listing proposed corrections. Nothing is
        written; the caller decides whether to :func:`apply_corrections`.

    A cell is flagged when, on a trading day, the stored 4dp value differs from
    BOT's published value — including a blank trading-day cell that BOT *does*
    publish (recorded as a "missing rate filled" correction). A cell whose
    QUANTIZED value matches BOT but whose stored payload carries >4dp float
    noise is also rewritten to the canonical 4dp value (flagged "normalized
    precision"). A blank weekend/holiday cell is correct by design and is
    never filled.

    GUARD — the row-1 headers of columns A-E must match the standard
    core/exrate_sheet.py layout (see :func:`validate_exrate_layout`). On a
    custom layout the scan aborts immediately: the report carries
    ``layout_error`` and ZERO corrections, so nothing can be applied.
    """
    report = RateAuditReport(sheet=ws.title)
    report.layout_error = validate_exrate_layout(ws)
    if report.layout_error:
        return report
    holidays = set(holidays_set or ())
    max_row = ws.max_row or 0

    for row in range(data_start_row, max_row + 1):
        rate_date = _parse_date(ws.cell(row=row, column=1).value)
        if rate_date is None:
            continue
        report.scanned_rows += 1

        # INVARIANT: weekend/holiday rows carry no BOT rate — never touch them.
        if rate_date.weekday() >= 5 or rate_date in holidays:
            continue

        for col, label, ccy, rate_type in EXRATE_RATE_COLUMNS:
            bot_val = bot_rates.get(rate_key(ccy, rate_type), {}).get(rate_date)
            raw = ws.cell(row=row, column=col).value
            old_val = safe_to_decimal(raw)
            # A non-empty cell we cannot parse (e.g. a stray formula string) is
            # neither a blank to fill nor a value we can verify — leave it.
            unparseable = old_val is None and raw not in (None, "")

            if bot_val is None:
                # No authoritative value to compare against — cannot verify.
                if raw not in (None, ""):
                    report.unverifiable += 1
                continue

            # Quantize the BOT value to 4dp through the SAME helper as every
            # other rate write: the cache path can hand back an unquantized
            # Decimal(str(float)), and a >4dp value must never reach a cell
            # (the "exact 4dp Decimal" invariant). safe_to_decimal accepts a
            # Decimal and re-quantizes it.
            bot_q = safe_to_decimal(bot_val)
            if bot_q is None:
                continue

            if unparseable:
                report.unverifiable += 1
                continue

            report.compared_cells += 1
            if old_val == bot_q:
                stored = _stored_decimal(raw)
                if stored is None or stored == bot_q:
                    continue  # already correct and canonically stored
                # F128: the cell holds a >4dp representation of the RIGHT
                # value (legacy float noise) — quantized it matches BOT, but
                # the stored payload does not. Rewrite the canonical 4dp
                # Decimal so the sheet is exact; flagged distinctly so the
                # report/CSV shows it as normalization, not a rate change.
                report.changes.append(
                    RateChange(
                        row=row,
                        col=col,
                        cell=f"{get_column_letter(col)}{row}",
                        rate_date=rate_date,
                        column_label=label,
                        currency=ccy,
                        rate_type=rate_type,
                        old_value=stored,
                        new_value=bot_q,
                        reason=(
                            f"normalized precision: stored {stored} "
                            f"rewritten as canonical 4dp {bot_q}"
                        ),
                    )
                )
                continue

            reason = (
                "missing rate filled from BOT"
                if old_val is None
                else f"value {old_val} != BOT {rate_type} {bot_q}"
            )
            report.changes.append(
                RateChange(
                    row=row,
                    col=col,
                    cell=f"{get_column_letter(col)}{row}",
                    rate_date=rate_date,
                    column_label=label,
                    currency=ccy,
                    rate_type=rate_type,
                    old_value=old_val,
                    new_value=bot_q,
                    reason=reason,
                )
            )
    return report


def apply_corrections(ws, report: RateAuditReport) -> RateAuditReport:
    """Write each proposed change's ``new_value`` into the sheet as a 4dp Decimal.

    Mutates ``ws`` in place and marks the report ``applied``. The caller owns
    backup-before-write and the workbook save/close/gc lifecycle.
    """
    for ch in report.changes:
        cell = ws.cell(row=ch.row, column=ch.col)
        cell.value = ch.new_value
        cell.number_format = "0.0000"
    report.applied = True
    return report


def write_audit_csv(report: RateAuditReport) -> str:
    """Write the report's changes to a timestamped CSV via AuditLogger.

    Returns the CSV path. Each change is one row (Rate Audit as the source).
    Always writes a file even with zero changes, so there is an auditable
    record that the verification ran.
    """
    audit = AuditLogger()
    try:
        fname = Path(report.file).name if report.file else "ExRate"
        for ch in report.changes:
            audit.log_row_change(
                filename=fname,
                sheet=report.sheet,
                row=ch.row,
                cell_date=ch.rate_date.strftime("%Y-%m-%d"),
                currency=f"{ch.currency} {ch.rate_type}",
                original_value="" if ch.old_value is None else str(ch.old_value),
                new_value=str(ch.new_value),
                rate_source="Rate Audit (BOT re-verify)",
            )
        return audit.finalize()
    except Exception:
        # Close the handle on the error path (finalize is idempotent), then
        # re-raise — the caller treats a failed CSV write as non-fatal.
        with contextlib.suppress(Exception):
            audit.finalize()
        raise


class StandaloneRateAuditor:
    """Audit an existing workbook's ExRate sheet against BOT, optionally
    applying 4dp corrections (the file is backed up first).

    Live-binding contract (mirrors StandaloneExRateUpdater): stores ONLY the
    live engine and dereferences engine attributes (backup, cache,
    _preload_api_data, _parse_date, _check_memory_guardrail) AT CALL TIME, so a
    test that reassigns them post-construction is honored.
    """

    def __init__(self, engine) -> None:
        self._engine = engine

    async def run(
        self, filepath: str, *, apply: bool = True, status_cb=None,
    ) -> RateAuditReport:
        """Verify (and optionally correct) the ExRate sheet in ``filepath``.

        With ``apply=False`` this is a dry run: it returns the proposed changes
        without backing up or writing anything. With ``apply=True`` (default)
        it backs the file up, rewrites only the differing trading-day cells,
        and saves atomically. Weekend/holiday rows are never touched.
        """

        def _status(msg: str) -> None:
            if status_cb is not None:
                with contextlib.suppress(Exception):
                    status_cb(msg)

        # noqa: PTH100 — keep os.path.abspath (no symlink resolution) so the
        # in-place save target string matches the rest of the write pipeline.
        filepath = os.path.abspath(filepath)  # noqa: PTH100
        self._engine._check_memory_guardrail(filepath)

        # Pass 1: read the ExRate dates read-only, then release the handle
        # BEFORE the (slow) network fetch — never hold a file lock over I/O.
        _status("Reading ExRate sheet...")
        existing_dates = self._read_exrate_dates(filepath)
        if existing_dates is None:
            raise ValueError("No ExRate sheet found in the selected file.")
        if not existing_dates:
            return RateAuditReport(file=filepath)

        # Authoritative BOT values for the sheet's span (cache-first, same
        # path the standard ExRate build uses).
        _status("Fetching BOT rates for verification...")
        target_year = min(existing_dates).year
        start_str = f"{target_year - 1}-12-20"
        # Audit only verifies dates already in the sheet, so bound the BOT
        # fetch to the sheet's own span. extend_to_today=False keeps the
        # preload from injecting bot_today() (the standard build does, to
        # extend to today): for an old archived ledger that would pull years
        # of unrelated rates on the 4GB target.
        all_target_dates = set(existing_dates)
        (
            logic_engine, usd_selling, eur_selling,
            usd_buying, eur_buying, _usd_data, _eur_data,
        ) = await self._engine._preload_api_data(
            all_target_dates, start_str, extend_to_today=False,
        )
        bot_rates = {
            rate_key("USD", "buying_transfer"): usd_buying,
            rate_key("USD", "selling"): usd_selling,
            rate_key("EUR", "buying_transfer"): eur_buying,
            rate_key("EUR", "selling"): eur_selling,
        }
        holidays_set = set(logic_engine.holidays)

        # Pass 2: load read-write, compare, optionally back up + apply + save.
        # Macro-enabled workbooks need keep_vba or the save silently strips
        # their VBA project (same rule as the ledger write pipeline).
        _status("Comparing against BOT...")
        keep_vba = Path(filepath).suffix.lower() in (".xlsm", ".xltm")
        wb = openpyxl.load_workbook(filepath, keep_vba=keep_vba)
        try:
            if "ExRate" not in wb.sheetnames:
                raise ValueError("No ExRate sheet found in the selected file.")
            ws = wb["ExRate"]
            report = scan_exrate_corrections(ws, bot_rates, holidays_set)
            report.file = filepath

            # Custom-layout refusal (F9): the scan produced ZERO corrections
            # by contract — abort before any backup/apply/save so the file is
            # untouched, and surface the reason as a hard error.
            if report.layout_error:
                _status(report.layout_error)
                raise ValueError(report.layout_error)

            if report.changes and apply:
                # Back up the on-disk original BEFORE overwriting (enables
                # Revert via the existing restore_latest). atomic_save writes a
                # temp file + os.replace, so the original stays intact on disk
                # until the swap — the backup captures the pre-correction file.
                report.backup_path = self._engine.backup.create_backup(filepath)
                _status("Backup created")
                apply_corrections(ws, report)
                ensure_disk_space(Path(filepath).parent, MIN_DISK_SPACE_MB)
                # Layer-1 hard gate (F201): reopen the saved TEMP and prove
                # each corrected cell round-trips as exactly the intended 4dp
                # Decimal BEFORE it replaces the original. A mismatch aborts
                # the swap — the user's file stays byte-for-byte untouched.
                expected_cells: dict[int, dict[int, object]] = {}
                for ch in report.changes:
                    expected_cells.setdefault(ch.row, {})[ch.col] = ch.new_value
                atomic_save(
                    wb, filepath,
                    verify=build_cell_verifier({ws.title: expected_cells}),
                )
                _status(f"OK: Applied {report.change_count} correction(s)")
            elif report.changes:
                _status(f"{report.change_count} difference(s) found (preview)")
            else:
                _status("All rates already match BOT")
            return report
        finally:
            with contextlib.suppress(OSError):
                wb.close()
            del wb
            gc.collect()

    def _read_exrate_dates(self, filepath: str) -> set[date] | None:
        """Return the set of dates in the ExRate sheet (read-only).

        Returns None when the workbook has no ExRate sheet (caller raises),
        an empty set when the sheet exists but has no parseable dates.
        """
        wb = None
        try:
            with Path(filepath).open("rb") as f:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                if "ExRate" not in wb.sheetnames:
                    return None
                ws = wb["ExRate"]
                dates: set[date] = set()
                for row in ws.iter_rows(
                    min_row=DATA_START_ROW, min_col=1, max_col=1,
                    values_only=True,
                ):
                    parsed = self._engine._parse_date(row[0])
                    if parsed:
                        dates.add(parsed)
                return dates
        finally:
            if wb is not None:
                with contextlib.suppress(OSError):
                    wb.close()
