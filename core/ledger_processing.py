#!/usr/bin/env python3
"""
core/ledger_processing.py
---------------------------------------------------------------------------
BOT Exchange Rate Processor — Ledger Processing Helpers
---------------------------------------------------------------------------
Near-pure helpers extracted from core/engine.py to keep the orchestrator
slim. These functions take all of their dependencies as explicit parameters
(no engine ``self`` state), so they are independently testable.

  - run_anomaly_check    -> anomaly orchestration over loaded rate dicts
  - prescan_target_dates -> read-only workbook scan for target dates
"""

import contextlib
import gc
import logging
import os
import threading
from collections.abc import Callable
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl

from core.constants import (
    LEDGER_HOME_CURRENCY,
    LEDGER_SUPPORTED_CURRENCIES,
    is_skip_sheet,
    parse_date,
)
from core.excel_io import find_header_row

logger = logging.getLogger(__name__)

# ── Prescan memoization (opt-in via use_cache=True) ──────────────────────
# The Smart-Date pass (core.prescan.prescan_oldest_date) opens + scans every
# queued file that process_ledger immediately rescans — a provably duplicate
# full read-only workbook open per file. Results are memoized keyed on file
# identity (abspath, st_mtime_ns, st_size) plus the header labels, so the
# second scan is skipped when the file is byte-identical. Featherweight: each
# entry is two small frozensets (dates + currency codes), bounded FIFO.
# The cache is OPT-IN (use_cache=True) and only the two production callers
# that share the canonical parse_date semantics use it; callers injecting a
# custom parse_date_fn must not opt in (results are keyed on file identity,
# not on the parser).
_PRESCAN_CACHE: dict[tuple, tuple[frozenset, frozenset]] = {}
_PRESCAN_CACHE_LOCK = threading.Lock()
_PRESCAN_CACHE_MAX = 128


def run_anomaly_check(
    anomaly_guard,
    emit_fn: Callable[[str, str], None],
    usd_buying: dict[date, Decimal],
    usd_selling: dict[date, Decimal],
    eur_buying: dict[date, Decimal],
    eur_selling: dict[date, Decimal],
    extra_currency_rates: dict[str, dict[date, Decimal]] | None = None,
    extra_rate_type: str = "buying_transfer",
    anomalous_out: set[tuple[str, date]] | None = None,
) -> int:
    """Run anomaly detection across all loaded rates (v3.1.0).

    Alert-only by contract: this function reports via ``emit_fn`` and the
    return count — it has no channel to veto, rewrite, or skip a rate.

    Args:
        anomaly_guard: An object exposing ``check_rates_bulk(rates_bundle)``
            that returns a list of anomaly records.
        emit_fn: Callback ``emit(msg, etype)`` for status events.
        usd_buying: USD buying-transfer rates keyed by date.
        usd_selling: USD selling rates keyed by date.
        eur_buying: EUR buying-transfer rates keyed by date.
        eur_selling: EUR selling rates keyed by date.
        extra_currency_rates: Optional extra (non-USD/EUR) ledger currencies
            as ``{ccy: {date: Decimal}}`` (F42). Each series joins the bundle
            under ``"{CCY}_{extra_rate_type}"`` so GBP/CNY/etc. jumps are
            flagged exactly like the four fixed USD/EUR series.
        extra_rate_type: Rate-type label for the extra series (the engine's
            snapshotted rate type — the extra fetch carries only that type).
        anomalous_out: Optional set that receives one ``(currency, date)``
            tuple per anomaly, so callers can thread the flagged cells into
            the audit trail (F25) without changing the return contract.

    Returns:
        The number of anomalies found.
    """
    rates_bundle = {
        "USD_buying_transfer": usd_buying,
        "USD_selling": usd_selling,
        "EUR_buying_transfer": eur_buying,
        "EUR_selling": eur_selling,
    }
    for ccy, series in (extra_currency_rates or {}).items():
        rates_bundle[f"{ccy}_{extra_rate_type}"] = series
    anomalies = anomaly_guard.check_rates_bulk(rates_bundle)
    if anomalous_out is not None:
        anomalous_out.update((a.currency, a.check_date) for a in anomalies)
    for a in anomalies:
        emit_fn(
            f"WARNING: ANOMALY: {a.currency} {a.rate_type} on "
            f"{a.check_date.strftime('%d %b %Y')}: "
            f"{a.pct_change:.2f}% change "
            f"({a.prev_value} -> {a.new_value})",
            "warning",
        )
    if anomalies:
        logger.warning(
            "Anomaly guard: %d suspicious rate(s) detected",
            len(anomalies),
        )
    return len(anomalies)


def prescan_target_dates(
    filepath: str,
    target_cols: dict[str, str],
    parse_date_fn: Callable[[object], date | None] = parse_date,
    emit_fn: Callable[[str], None] | None = None,
) -> set[date]:
    """Scan a workbook in read-only mode to extract all target dates.

    Thin backward-compatible wrapper around
    :func:`prescan_target_dates_and_currencies` that returns only the date set
    (existing callers expect a ``set[date]``).

    Args:
        filepath: Path to the .xlsx/.xlsm workbook.
        target_cols: Column-name mapping; ``target_cols["source_date"]`` is
            the header label of the date column.
        parse_date_fn: Cell-value -> date parser (defaults to shared parser).
        emit_fn: Optional status callback ``emit(msg)``.

    Returns:
        Set of all parsed dates found in the source date column.
    """
    dates, _currencies = prescan_target_dates_and_currencies(
        filepath, target_cols, parse_date_fn=parse_date_fn, emit_fn=emit_fn,
    )
    return dates


def prescan_target_dates_and_currencies(
    filepath: str,
    target_cols: dict[str, str],
    parse_date_fn: Callable[[object], date | None] = parse_date,
    emit_fn: Callable[[str], None] | None = None,
    use_cache: bool = False,
) -> tuple[set[date], set[str]]:
    """Scan a workbook once for both target dates AND distinct currency codes.

    Opens the workbook in read-only mode, scans all non-skipped sheets for the
    source date + currency columns, and returns ``(dates, currencies)``. The
    currency set powers the multi-currency ledger path: the engine fetches a
    rate column per detected currency rather than silently leaving non-USD/EUR
    rows blank. The workbook is closed + garbage-collected after scanning.

    Args:
        filepath: Path to the .xlsx/.xlsm workbook.
        target_cols: Column-name mapping; ``target_cols["source_date"]`` and
            ``target_cols["currency"]`` are the header labels scanned.
        parse_date_fn: Cell-value -> date parser (defaults to shared parser).
        emit_fn: Optional status callback ``emit(msg)``.
        use_cache: When True, memoize the scan keyed on the file's
            (abspath, mtime_ns, size) + header labels and serve repeat scans
            of a byte-identical file from the memo — this is what lets the
            Smart-Date prescan and process_ledger share ONE open per file.
            Only opt in when ``parse_date_fn`` is behaviorally the shared
            ``core.constants.parse_date`` (the cache key does not include
            the parser). Defensive copies are returned on every hit.

    Returns:
        ``(set of parsed dates, set of upper-cased currency codes)``.
    """
    all_target_dates: set[date] = set()
    all_currencies: set[str] = set()
    source_label = target_cols["source_date"]
    currency_label = target_cols.get("currency")

    # ── Memo lookup (file identity + labels) ──────────────────────────
    cache_key: tuple | None = None
    if use_cache:
        try:
            st = Path(filepath).stat()
            # noqa rationale: os.path.abspath (not Path.resolve) — abspath
            # normalizes WITHOUT resolving symlinks, matching the exact
            # normalization process_ledger applies to its save target.
            cache_key = (
                os.path.abspath(filepath), st.st_mtime_ns, st.st_size,  # noqa: PTH100
                source_label, currency_label, target_cols.get("out_rate"),
            )
        except OSError:
            cache_key = None  # unstat-able → fall through to a real scan
        if cache_key is not None:
            with _PRESCAN_CACHE_LOCK:
                hit = _PRESCAN_CACHE.get(cache_key)
            if hit is not None:
                if emit_fn is not None:
                    emit_fn("Reusing pre-scanned dates (file unchanged)")
                return set(hit[0]), set(hit[1])

    if emit_fn is not None:
        emit_fn("Scanning dates from workbook")

    wb_scan = None
    try:
        wb_scan = openpyxl.load_workbook(
            filepath, read_only=True, data_only=True,
        )
        for sheet_name in wb_scan.sheetnames:
            if is_skip_sheet(sheet_name):
                continue
            ws = wb_scan[sheet_name]
            # Header location + duplicate resolution are owned by
            # core.excel_io.find_header_row (shared with the ledger write
            # path's scan_sheet_headers and core.prescan). The out_rate
            # label is mapped purely as the duplicate-'Date' resolution
            # anchor: the date window fetched here MUST be the same
            # export-entry-date column the written formulas look up, or
            # the master sheet misses exactly the dates the formulas need.
            header_row_idx, col_indices = find_header_row(
                ws,
                (
                    ("source", source_label),
                    ("currency", currency_label),
                    ("out_rate", target_cols.get("out_rate")),
                ),
                sheet_name=sheet_name,
                resolve_left_of={"source": "out_rate"},
            )

            if header_row_idx is None or "source" not in col_indices:
                continue

            src0 = col_indices["source"]
            cur0 = col_indices.get("currency")
            # ONE forward pass via iter_rows. ws.cell() random access on a
            # READ-ONLY worksheet re-parses the sheet XML from row 1 on
            # every call (openpyxl ReadOnlyWorksheet._get_cell) — the old
            # per-row cell loop was O(n^2) and dominated the whole batch
            # (measured: 193s for one 2000-row sheet vs 0.1s with
            # iter_rows, identical dates + currencies).
            for row in ws.iter_rows(
                min_row=header_row_idx + 1, values_only=True,
            ):
                raw_date = row[src0] if src0 < len(row) else None
                parsed_date = parse_date_fn(raw_date)
                if parsed_date:
                    all_target_dates.add(parsed_date)
                if cur0 is not None:
                    cur_val = row[cur0] if cur0 < len(row) else None
                    if cur_val is not None:
                        code = str(cur_val).strip().upper()
                        if code:
                            all_currencies.add(code)
    except (ValueError, TypeError, KeyError,
            openpyxl.utils.exceptions.InvalidFileException):
        raise
    finally:
        if wb_scan is not None:
            with contextlib.suppress(OSError):
                wb_scan.close()
            del wb_scan
            wb_scan = None
        gc.collect()

    # ── Memo store (successful scans only) ────────────────────────────
    if cache_key is not None:
        with _PRESCAN_CACHE_LOCK:
            if len(_PRESCAN_CACHE) >= _PRESCAN_CACHE_MAX:
                # FIFO eviction: drop the oldest entry (dict preserves
                # insertion order) so a long-running process stays bounded.
                _PRESCAN_CACHE.pop(next(iter(_PRESCAN_CACHE)))
            _PRESCAN_CACHE[cache_key] = (
                frozenset(all_target_dates), frozenset(all_currencies),
            )

    return all_target_dates, all_currencies


def classify_currencies(
    currencies: set[str],
) -> tuple[list[str], list[str]]:
    """Split scanned ledger currencies into (extra-supported, unsupported).

    The home currency (THB) and the two fixed master-sheet currencies
    (USD/EUR) are filtered out — they are always handled by the core IFS
    branches. The first list holds supported NON-USD/EUR codes the engine
    should fetch a dynamic column for (sorted for deterministic column order);
    the second holds codes the tool cannot fill, which must be surfaced as a
    warning rather than left silently blank.

    Args:
        currencies: Distinct upper-cased currency codes found in the ledger.

    Returns:
        ``(extra_supported_sorted, unsupported_sorted)``.
    """
    extra: set[str] = set()
    unsupported: set[str] = set()
    for code in currencies:
        if code in (LEDGER_HOME_CURRENCY, "USD", "EUR"):
            continue
        if code in LEDGER_SUPPORTED_CURRENCIES:
            extra.add(code)
        else:
            unsupported.add(code)
    return sorted(extra), sorted(unsupported)
