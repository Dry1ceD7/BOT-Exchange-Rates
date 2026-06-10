#!/usr/bin/env python3
"""
core/exrate_updater.py
---------------------------------------------------------------------------
BOT Exchange Rate Processor — Workbook write pipeline + standalone updater
---------------------------------------------------------------------------
Two extracted units that previously lived inline in core/engine.py:

  - WorkbookWriter          → the process_ledger openpyxl write pipeline.
  - StandaloneExRateUpdater → the update_exrate_standalone method body.

LIVE-BINDING CONTRACT (mandatory): both classes store ONLY the live engine
object (``self._engine = engine``) and dereference every engine attribute
(backup, cache, api, _preload_api_data, _parse_date, _check_memory_guardrail,
_emit, target_cols) AT CALL TIME inside write()/run(). Construction must never
snapshot bound methods or attributes, so post-construction reassignment of
``engine.backup`` / ``engine._preload_api_data`` (done in tests) is honored.
"""

import contextlib
import gc
import logging
import os
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

from core.audit_logger import AuditCollector, AuditRecord
from core.constants import BACKUP_MAX_AGE_DAYS, MIN_DISK_SPACE_MB, bot_today
from core.excel_io import (
    build_exrate_index,
    inject_xlookup_formulas,
    scan_sheet_headers,
    write_custom_exrate_data,
)
from core.exrate_sheet import update_master_exrate_sheet
from core.logic import (
    build_holiday_lookup,
    compute_year_start_date,
    safe_to_decimal,
)
from core.workbook_io import atomic_save as _atomic_save
from core.workbook_io import build_cell_verifier, ensure_disk_space

logger = logging.getLogger(__name__)


def _snapshot_cells(
    ws, cols, start_row: int = 2,
) -> dict[int, dict[int, object]]:
    """Snapshot the in-memory values of ``cols`` (1-based) for every data row.

    The result feeds ``core.workbook_io.build_cell_verifier``: after the
    save, the TEMP file is reopened and each snapshotted cell must round-trip
    to exactly this value (Decimal-string equality for numbers, blank for
    None) BEFORE the temp replaces the original — the Layer-1 exactness gate.
    """
    rows: dict[int, dict[int, object]] = {}
    for row_idx in range(start_row, (ws.max_row or 0) + 1):
        rows[row_idx] = {
            col: ws.cell(row=row_idx, column=col).value for col in cols
        }
    return rows


def _collect_ledger_expectations(
    wb, sheet_maps: dict, exrate_col_map: dict[str, str] | None,
) -> dict[str, dict[int, dict[int, object]]]:
    """Expected post-save cell values for everything THIS RUN wrote.

    - ExRate sheet: every rate cell (fixed B-E plus appended extra-currency
      columns). update_master_exrate_sheet rebuilds ALL data rows in one
      shot (delete_rows + rewrite), so the whole range is this run's output;
      blank weekend/holiday cells are asserted blank.
    - Monthly tabs: every EX Rate cell now holding a formula string — the
      IFS formula injected (or confirmed identical) this run must round-trip
      verbatim. ExRate is in SKIP_SHEET_NAMES, so sheet_maps never collides
      with the ExRate entry above.
    """
    expected: dict[str, dict[int, dict[int, object]]] = {}
    if "ExRate" in wb.sheetnames:
        rate_cols = [2, 3, 4, 5] + [
            column_index_from_string(letter)
            for letter in (exrate_col_map or {}).values()
        ]
        expected["ExRate"] = _snapshot_cells(wb["ExRate"], rate_cols)
    for sheet_name, mapping in sheet_maps.items():
        out_idx = mapping["columns"].get("out_rate")
        if out_idx is None:
            continue
        out_col = out_idx + 1
        ws = wb[sheet_name]
        rows = expected.setdefault(sheet_name, {})
        for row_idx in range(mapping["header_row"] + 1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=out_col).value
            if isinstance(value, str) and value.startswith("="):
                rows.setdefault(row_idx, {})[out_col] = value
    return expected


def _is_macro_workbook(filepath) -> bool:
    """True for macro-enabled containers (.xlsm / .xltm).

    Every load_workbook in this module must pass ``keep_vba=`` this value:
    without it openpyxl silently drops xl/vbaProject.bin on save, and
    atomic_save then replaces the original — destroying the user's macros.
    """
    return str(filepath).lower().endswith((".xlsm", ".xltm"))


def _close_vba_archive(wb) -> None:
    """Deterministically close the keep_vba ZipFile (wb.close() skips it).

    openpyxl stores the preserved VBA container as an append-mode in-memory
    ZipFile on ``wb.vba_archive``; left to GC teardown its ``__del__`` can
    fire after the backing buffer is cleared and emit "Exception ignored"
    noise. Closing it alongside wb keeps the handle-release discipline exact.
    """
    vba = getattr(wb, "vba_archive", None)
    if vba is not None:
        with contextlib.suppress(OSError, ValueError):
            vba.close()


class WorkbookWriter:
    """openpyxl write pipeline for process_ledger (extracted verbatim).

    Reads the live engine (target_cols, _parse_date, _emit) at call time.
    """

    def __init__(self, engine) -> None:
        self._engine = engine

    async def write(
        self,
        filepath: str,
        dry_run: bool,
        usd_buying,
        usd_selling,
        eur_buying,
        eur_selling,
        master_holidays_set,
        holidays_names,
        computed_start,
        rate_type: str | None = None,
        extra_currency_rates: dict | None = None,
        unsupported_currencies: list[str] | None = None,
        audit: AuditCollector | None = None,
    ) -> str:
        # keep_vba preserves xl/vbaProject.bin for .xlsm/.xltm inputs (the
        # in-place overwrite would otherwise strip the user's macros).
        wb = openpyxl.load_workbook(
            filepath, keep_vba=_is_macro_workbook(filepath)
        )

        # rate_type is snapshotted by process_ledger at the start of this
        # file's run and threaded through here, so a Settings "Save" mid-batch
        # cannot change the rate basis for an in-flight file. Only fall back to
        # SettingsManager when a caller did not pass one (backward compat).
        if rate_type is None:
            from core.config_manager import SettingsManager
            rate_type = SettingsManager().load().get(
                "rate_type", "buying_transfer"
            )
        extra_currency_rates = extra_currency_rates or {}
        unsupported_currencies = unsupported_currencies or []

        # try/finally guarantees the OS file handle is released and gc runs on
        # BOTH the success and error exits (the standalone paths do the same).
        # The previous except-and-reraise structure left the trailing
        # gc.collect() unreachable on the error path.
        try:
            # Scan monthly tabs for header/column mappings
            sheet_maps = scan_sheet_headers(wb, self._engine.target_cols)

            # ── STEP 1: Build ExRate master sheet ────────────────────────
            # Returns {ccy: column_letter} for the appended extra currencies so
            # their columns can be wired into the ledger XLOOKUP formula.
            exrate_col_map = update_master_exrate_sheet(
                wb, usd_buying, usd_selling, eur_buying, eur_selling,
                list(master_holidays_set), holidays_names, computed_start,
                extra_currency_rates=extra_currency_rates,
            )

            # ── STEP 2: Build in-memory ExRate lookup index ──────────────
            # Pass the extra-currency column map so the index also carries the
            # appended GBP/JPY/etc. columns for the unfilled-row check below.
            exrate_index = build_exrate_index(wb, exrate_col_map)

            # ── STEP 2b: Snapshot original EX Rate cells for the audit trail ─
            # inject_xlookup_formulas overwrites the EX Rate column with the IFS
            # formula, so the pre-edit value must be captured FIRST. Cheap: a
            # dict of {(sheet, row): original_value} only built when auditing.
            originals = (
                self._snapshot_out_rate_cells(wb, sheet_maps)
                if audit is not None else {}
            )

            # ── STEP 3: Inject XLOOKUP formulas into monthly tabs ────────
            exrate_last_row = 2
            if "ExRate" in wb.sheetnames:
                ws_ex = wb["ExRate"]
                exrate_last_row = max(ws_ex.max_row or 2, 2)

            inject_xlookup_formulas(
                wb, sheet_maps, exrate_last_row,
                parse_date_fn=self._engine._parse_date,
                emit_fn=self._engine._emit,
                dry_run=dry_run,
                rate_type=rate_type,
                exrate_col_map=exrate_col_map,
            )

            # ── STEP 4: Surface rows the formula cannot fill ─────────────
            # Blank EX Rate cells are otherwise silent. Count both
            # unavailable-rate rows (date beyond the rollback / no API data)
            # and unsupported-currency rows, and emit per-file warnings so the
            # accountant is told instead of filing empty rates.
            self._warn_unfilled_rows(
                wb, sheet_maps, exrate_index, exrate_col_map,
                rate_type, unsupported_currencies, Path(filepath).name,
            )

            # ── STEP 5: Record per-cell changes into the audit trail ─────
            # Every EX Rate cell that resolved to a rate is logged with its
            # before/after value, currency, date, and holiday-rollback flag so
            # the auditor-facing CSV is no longer a hollow header (#1).
            if audit is not None:
                self._collect_audit_records(
                    wb, sheet_maps, exrate_index, exrate_col_map, rate_type,
                    originals, Path(filepath).name, master_holidays_set, audit,
                )

            # ── Save ─────────────────────────────────────────────────────
            if dry_run:
                self._engine._emit(
                    "[SIM] File NOT saved (dry run) "
                    f"— {Path(filepath).name}"
                )
            else:
                # ERR-03: Check disk space before saving
                ensure_disk_space(Path(filepath).parent, MIN_DISK_SPACE_MB)
                # Layer-1 hard gate (F50/F63/F72): reopen the saved TEMP and
                # prove every cell this run wrote round-trips exactly (rates
                # as 4dp Decimals, formulas verbatim) BEFORE it replaces the
                # original. A mismatch aborts the swap — file untouched.
                expected = _collect_ledger_expectations(
                    wb, sheet_maps, exrate_col_map
                )
                _atomic_save(
                    wb, filepath, verify=build_cell_verifier(expected)
                )
                logger.info(
                    "Overwritten in-place: %s",
                    Path(filepath).name,
                )
        finally:
            # Release the file handle + reclaim memory on EVERY exit.
            _close_vba_archive(wb)
            try:
                wb.close()
            except OSError:
                logger.debug("Failed to close workbook (ledger write path)")
            del wb  # release file handle immediately
            gc.collect()

        self._engine._emit("File saved and memory cleaned", etype="success")
        return filepath

    def _warn_unfilled_rows(
        self,
        wb,
        sheet_maps: dict,
        exrate_index: dict,
        exrate_col_map: dict[str, str],
        rate_type: str,
        unsupported_currencies: list[str],
        fname: str,
    ) -> None:
        """Emit per-file warnings for ledger rows that resolve to a blank rate.

        Mirrors the ledger IFS formula's resolution so the count matches what
        Excel will actually leave blank:
          - THB always resolves (literal 1),
          - USD/EUR/extra currencies resolve when the row's date is present in
            the ExRate index with a non-empty value for the chosen rate column,
          - any other currency is unsupported and never resolves.

        Counts are reported but the file still saves: a visible warning beats a
        hard failure on what may be a single stray row in a large ledger.
        """
        # USD/EUR fixed columns vary by rate type (B/D buying, C/E selling).
        if rate_type == "selling":
            fixed = {"USD": "usd_selling", "EUR": "eur_selling"}
        else:
            fixed = {"USD": "usd_buying", "EUR": "eur_buying"}

        no_rate = 0
        unsupported_seen: dict[str, int] = {}
        unsupported_set = set(unsupported_currencies)

        for sheet_name, mapping in sheet_maps.items():
            cols = mapping["columns"]
            cur_idx = cols.get("currency")
            src_idx = cols["source"] + 1
            if cur_idx is None:
                continue
            cur_col = cur_idx + 1
            ws = wb[sheet_name]
            for row_idx in range(mapping["header_row"] + 1, ws.max_row + 1):
                cur_val = ws.cell(row=row_idx, column=cur_col).value
                if cur_val is None:
                    continue
                ccy = str(cur_val).strip().upper()
                if not ccy:
                    continue
                if ccy == "THB":
                    continue
                row_date = self._engine._parse_date(
                    ws.cell(row=row_idx, column=src_idx).value
                )
                if ccy in unsupported_set:
                    unsupported_seen[ccy] = unsupported_seen.get(ccy, 0) + 1
                    continue
                if row_date is None:
                    continue
                if not self._rate_available(
                    ccy, row_date, exrate_index, exrate_col_map, fixed,
                ):
                    no_rate += 1

        if no_rate:
            self._engine._emit(
                f"{fname}: {no_rate} row(s) had no rate available "
                "(date beyond the 10-day rollback) — EX Rate left blank",
                etype="warning",
            )
        for ccy, count in sorted(unsupported_seen.items()):
            self._engine._emit(
                f"{fname}: {count} row(s) with unsupported currency {ccy} "
                "left blank",
                etype="warning",
            )

    def _snapshot_out_rate_cells(
        self, wb, sheet_maps: dict,
    ) -> dict[tuple[str, int], object]:
        """Capture each monthly tab's EX Rate cell value BEFORE injection.

        Returns ``{(sheet_name, row_idx): original_value}``. Only the EX Rate
        (out_rate) column is snapshotted, because inject_xlookup_formulas
        overwrites exactly those cells; everything else is left untouched.
        """
        originals: dict[tuple[str, int], object] = {}
        for sheet_name, mapping in sheet_maps.items():
            cols = mapping["columns"]
            out_idx = cols.get("out_rate")
            if out_idx is None:
                continue
            out_col = out_idx + 1
            ws = wb[sheet_name]
            for row_idx in range(mapping["header_row"] + 1, ws.max_row + 1):
                originals[(sheet_name, row_idx)] = ws.cell(
                    row=row_idx, column=out_col
                ).value
        return originals

    def _collect_audit_records(
        self,
        wb,
        sheet_maps: dict,
        exrate_index: dict,
        exrate_col_map: dict[str, str],
        rate_type: str,
        originals: dict[tuple[str, int], object],
        fname: str,
        holidays_set,
        audit: "AuditCollector",
    ) -> None:
        """Append one ``AuditRecord`` per EX Rate cell that resolved to a rate.

        Mirrors the IFS formula's resolution (THB→1, USD/EUR→fixed columns,
        extra currencies→exrate_col_map) so ``new_value`` matches what Excel
        will display. ``original_value`` is the pre-injection cell snapshot.
        Rows that cannot resolve (no rate / unsupported currency) are skipped —
        those are surfaced separately by ``_warn_unfilled_rows`` and would only
        clutter the per-cell change trail with empty "after" values.
        """
        if rate_type == "selling":
            fixed = {"USD": "usd_selling", "EUR": "eur_selling"}
        else:
            fixed = {"USD": "usd_buying", "EUR": "eur_buying"}
        holiday_dates = set(holidays_set or ())

        for sheet_name, mapping in sheet_maps.items():
            cols = mapping["columns"]
            cur_idx = cols.get("currency")
            out_idx = cols.get("out_rate")
            src_idx = cols["source"] + 1
            if cur_idx is None or out_idx is None:
                continue
            cur_col = cur_idx + 1
            ws = wb[sheet_name]
            for row_idx in range(mapping["header_row"] + 1, ws.max_row + 1):
                cur_val = ws.cell(row=row_idx, column=cur_col).value
                if cur_val is None:
                    continue
                ccy = str(cur_val).strip().upper()
                if not ccy:
                    continue
                row_date = self._engine._parse_date(
                    ws.cell(row=row_idx, column=src_idx).value
                )
                resolved = self._resolve_rate_value(
                    ccy, row_date, exrate_index, exrate_col_map, fixed,
                )
                if resolved is None:
                    continue
                audit.add(AuditRecord(
                    filename=fname,
                    sheet=sheet_name,
                    row=row_idx,
                    cell_date=row_date.strftime("%Y-%m-%d") if row_date else "",
                    currency=ccy,
                    original_value=self._fmt_value(
                        originals.get((sheet_name, row_idx))
                    ),
                    new_value=self._fmt_value(resolved),
                    rate_source="Cache/API",
                    holiday_rollback=(
                        row_date is not None and row_date in holiday_dates
                    ),
                ))

    @staticmethod
    def _resolve_rate_value(
        ccy: str,
        row_date,
        exrate_index: dict,
        exrate_col_map: dict[str, str],
        fixed: dict[str, str],
    ):
        """Return the resolved rate for (ccy, date), or None if unresolved.

        THB always resolves to 1 (the formula's literal). USD/EUR read the
        rate-type column; extra currencies read their appended column.
        """
        if ccy == "THB":
            return 1
        if row_date is None:
            return None
        row = exrate_index.get(row_date)
        if row is None:
            return None
        if ccy in fixed:
            return row.get(fixed[ccy])
        if ccy in exrate_col_map:
            return row.get(f"extra:{ccy}")
        return None

    @staticmethod
    def _fmt_value(value) -> str:
        """Render a cell value for the audit CSV (blank cells → empty string)."""
        if value is None:
            return ""
        return str(value)

    @staticmethod
    def _rate_available(
        ccy: str,
        row_date,
        exrate_index: dict,
        exrate_col_map: dict[str, str],
        fixed: dict[str, str],
    ) -> bool:
        """True if the ExRate sheet has a non-empty rate for (ccy, date)."""
        row = exrate_index.get(row_date)
        if row is None:
            return False
        if ccy in fixed:
            return row.get(fixed[ccy]) is not None
        if ccy in exrate_col_map:
            return row.get(f"extra:{ccy}") is not None
        return False


class StandaloneExRateUpdater:
    """Standalone ExRate update path (extracted verbatim).

    Reads the live engine (backup, cache, api, _preload_api_data, _parse_date,
    _check_memory_guardrail, _emit) at call time.
    """

    def __init__(self, engine) -> None:
        self._engine = engine

    async def run(
        self,
        filepath: str,
        progress_cb,
        currencies,
        rate_types,
        date_range,
    ) -> str:
        # ── Defaults ──────────────────────────────────────────────────
        if not currencies:
            currencies = ["USD", "EUR"]
        if not rate_types:
            rate_types = {
                "Buying TT": "buying_transfer",
                "Selling": "selling",
            }

        # If standard USD+EUR with Buying TT + Selling, use the original
        # function for backward compatibility with ledger processing.
        is_standard = (
            set(currencies) == {"USD", "EUR"}
            and set(rate_types.values()) == {"buying_transfer", "selling"}
        )

        def _status(msg: str):
            if progress_cb:
                progress_cb(msg)
            self._engine._emit(msg)

        # ── Fail-safe preconditions (mirror process_ledger) ───────────
        # Every in-place overwrite path must enforce the size guardrail and
        # create a pre-edit backup BEFORE load_workbook. If the file does not
        # yet exist this is a creation path — skip backup, but the loaded
        # save below would fail anyway, so a missing file is still an error.
        # noqa: PTH100 — keep os.path.abspath (no symlink resolution) so the
        # in-place save target string stays identical to the legacy behavior.
        filepath = os.path.abspath(filepath)  # noqa: PTH100
        self._engine._check_memory_guardrail(filepath)
        _status("Size check passed")
        if Path(filepath).exists():
            self._engine.backup.create_backup(filepath)
            _status("Backup created")
            # Prune stale backups on the standalone path too — the engine only
            # runs the 7-day cleanup inside process_batch, so without this the
            # standalone updater would never reclaim old backup disk space.
            self._engine.backup.cleanup_old_backups(
                max_age_days=BACKUP_MAX_AGE_DAYS
            )

        _status("Opening ExRate file...")
        # keep_vba preserves xl/vbaProject.bin for .xlsm/.xltm inputs (the
        # in-place overwrite would otherwise strip the user's macros).
        wb = openpyxl.load_workbook(
            filepath, keep_vba=_is_macro_workbook(filepath)
        )

        if "ExRate" not in wb.sheetnames:
            _close_vba_archive(wb)
            wb.close()
            del wb
            gc.collect()
            raise ValueError("No ExRate sheet found in the selected file.")

        # ── Standard path (backward compatible) ───────────────────────
        if is_standard:
            from core.exrate_sheet import update_master_exrate_sheet

            # try/finally guarantees the OS file handle is released and gc
            # runs even if API fetch / write / save raises (mirrors
            # process_ledger). Otherwise an exception would leak the handle.
            try:
                ws_ex = wb["ExRate"]
                existing_dates = set()
                for row_idx in range(2, (ws_ex.max_row or 1) + 1):
                    cell_val = ws_ex.cell(row=row_idx, column=1).value
                    parsed = self._engine._parse_date(cell_val)
                    if parsed:
                        existing_dates.add(parsed)

                target_year = min(existing_dates).year if existing_dates else bot_today().year

                # Override with manual date_range if provided
                if date_range:
                    dr_start, dr_end = date_range
                    target_year = dr_start.year
                    all_target_dates = {dr_start, dr_end, bot_today()}
                    start_date_str = f"{dr_start.year - 1}-12-20"
                else:
                    dr_start = dr_end = None
                    start_date_str = f"{target_year - 1}-12-20"
                    all_target_dates = existing_dates | {bot_today()}

                _status("Fetching exchange rates from BOT API...")
                (
                    logic_engine, usd_selling, eur_selling,
                    usd_buying, eur_buying, _usd_data, _eur_data,
                ) = await self._engine._preload_api_data(
                    all_target_dates, start_date_str
                )

                computed_start = compute_year_start_date(
                    target_year, logic_engine.holidays
                )

                master_holidays_set, holidays_names = build_holiday_lookup(
                    self._engine.cache, all_target_dates, computed_start, logic_engine,
                )

                _status("Writing exchange rate data...")
                # Manual range → honor the user's exact (dr_start, dr_end).
                # No range → prior-year-December computed_start, end defaults
                # to today() inside update_master_exrate_sheet.
                sheet_start = dr_start if dr_start is not None else computed_start
                update_master_exrate_sheet(
                    wb, usd_buying, usd_selling, eur_buying, eur_selling,
                    sorted(master_holidays_set), holidays_names,
                    sheet_start, end_date=dr_end,
                )
                # ERR-03: Check disk space before saving
                ensure_disk_space(Path(filepath).parent, MIN_DISK_SPACE_MB)
                # Layer-1 hard gate: every rebuilt ExRate rate cell (B-E)
                # must round-trip from the TEMP file before it replaces the
                # original — a mismatch leaves the user's file untouched.
                expected = {
                    "ExRate": _snapshot_cells(wb["ExRate"], (2, 3, 4, 5))
                }
                _atomic_save(
                    wb, filepath, verify=build_cell_verifier(expected)
                )
                _status(f"✓ ExRate updated: {Path(filepath).name}")
                return filepath
            finally:
                _close_vba_archive(wb)
                try:
                    wb.close()
                except OSError:
                    logger.debug("Failed to close workbook (standard path)")
                del wb
                gc.collect()

        # ── Custom path (any currencies / rate types) ─────────────────
        # try/finally guarantees handle release + gc on any error below.
        try:
            _status(f"Fetching rates for {', '.join(currencies)}...")

            # Fetch from API for each currency
            if date_range:
                start_dt, end_dt = date_range
            else:
                target_year = bot_today().year
                start_dt = date(target_year - 1, 12, 20)
                end_dt = bot_today()

            # rate_data[currency][api_field][date] = value
            rate_data: dict[str, dict[str, dict[date, float]]] = {}

            for ccy in currencies:
                _status(f"Fetching {ccy} rates...")
                raw_results = await self._engine.api.get_exchange_rates(
                    start_dt, end_dt, ccy,
                )
                rate_data[ccy] = {}
                for _label, api_field in rate_types.items():
                    rate_data[ccy][api_field] = {}

                for rec in raw_results:
                    try:
                        rec_date = datetime.strptime(
                            rec.period, "%Y-%m-%d"
                        ).date()
                    except (ValueError, TypeError):
                        continue
                    for _label, api_field in rate_types.items():
                        val = getattr(rec, api_field, None)
                        if val is not None:
                            # Mathematical Truth: quantize to 4dp Decimal, same
                            # discipline as the standard USD/EUR path — never
                            # write the raw API float.
                            rate_data[ccy][api_field][rec_date] = safe_to_decimal(val)

            # Fetch holidays
            _status("Fetching holidays...")
            all_target_dates = {bot_today()}
            (
                logic_engine, _, _, _, _, _, _,
            ) = await self._engine._preload_api_data(all_target_dates, str(start_dt))

            holidays_set = set(logic_engine.holidays)
            holidays_names_map: dict[date, str] = {}
            for year in {start_dt.year, end_dt.year}:
                for h_str, h_name in self._engine.cache.get_holidays(year):
                    with contextlib.suppress(ValueError, TypeError):
                        holidays_names_map[
                            datetime.strptime(h_str, "%Y-%m-%d").date()
                        ] = h_name

            # Build column headers: Date + (CCY RateType)... + Holidays
            headers = ["Date"]
            col_specs = []  # (currency, api_field) per data column
            for ccy in currencies:
                for label, api_field in rate_types.items():
                    headers.append(f"{ccy} {label}")
                    col_specs.append((ccy, api_field))
            headers.append("Holidays/Weekend")

            # Build date range
            all_dates = []
            d = start_dt
            while d <= end_dt:
                all_dates.append(d)
                d += timedelta(days=1)
            all_dates.sort()

            # ── Write to sheet ────────────────────────────────────────
            ws = wb["ExRate"]
            _status("Writing custom ExRate data...")

            write_custom_exrate_data(
                ws, rate_data, col_specs, headers,
                all_dates, holidays_set, holidays_names_map,
            )

            # ERR-03: Check disk space before saving
            ensure_disk_space(Path(filepath).parent, MIN_DISK_SPACE_MB)
            # Layer-1 hard gate: every custom-layout rate cell written above
            # (one column per (ccy, rate_type) spec) must round-trip from
            # the TEMP file before it replaces the original.
            expected = {
                "ExRate": _snapshot_cells(
                    ws, tuple(range(2, 2 + len(col_specs)))
                )
            }
            _atomic_save(wb, filepath, verify=build_cell_verifier(expected))
            _status(f"✓ ExRate created: {Path(filepath).name}")
            return filepath
        finally:
            _close_vba_archive(wb)
            try:
                wb.close()
            except OSError:
                logger.debug("Failed to close workbook (custom path)")
            del wb
            gc.collect()
