#!/usr/bin/env python3
"""
core/excel_io.py
---------------------------------------------------------------------------
BOT Exchange Rate Processor — Excel I/O Operations
---------------------------------------------------------------------------
Extracted from engine.py (C-01 decomposition) to enforce Single
Responsibility and keep engine.py under the 200 LOC SFFB guideline.

Contains:
  - zero_touch_write: Write cell value without touching formatting
  - build_exrate_index: Build in-memory lookup from ExRate sheet
  - scan_sheet_headers: Scan monthly tabs for column mappings
  - inject_xlookup_formulas: Write XLOOKUP formulas into monthly tabs
  - write_custom_exrate_data: Write multi-currency ExRate data
"""

import logging
import re
from collections.abc import Callable
from datetime import date, datetime

from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from core.constants import PREFORMAT_BUFFER_ROWS, is_skip_sheet
from core.exrate_sheet import (
    EXRATE_RATE_COLUMNS,
    exrate_fixed_letters,
    exrate_index_key,
)

logger = logging.getLogger(__name__)

# Validation patterns for values interpolated into Excel formula strings.
# A malformed currency code or column letter could otherwise corrupt the
# entire IFS formula for a row, so we validate before interpolating.
_CCY_RE = re.compile(r"^[A-Z]{2,4}$")
_COL_RE = re.compile(r"^[A-Z]{1,3}$")


def _is_blank_value(value) -> bool:
    """True for cell values that render as blank (None or whitespace-only)."""
    return value is None or (isinstance(value, str) and not value.strip())


def zero_touch_write(ws, row: int, col: int, value) -> None:
    """
    Write a value to a monthly-tab cell WITHOUT touching formatting.

    Zero-Touch Protocol: ONLY writes cell.value.
    NEVER reads, copies, or re-applies font/fill/border/alignment.

    In openpyxl, assigning cell.value does NOT alter the cell's
    existing styles. Touching style attributes (even via .copy())
    creates new style objects that can differ from the originals.

    Silently skips MergedCell instances (read-only).
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def build_exrate_index(
    wb,
    exrate_col_map: dict[str, str] | None = None,
) -> dict[date, dict]:
    """
    Build an in-memory ExRate lookup index from the ExRate sheet.

    Reads dates from column A and rate values from columns B-E. When
    ``exrate_col_map`` ({ccy: column_letter}) is supplied, the value of each
    extra-currency column is also indexed under the key ``f"extra:{ccy}"`` so
    callers can tell whether a multi-currency row will resolve or stay blank.

    Returns a dict mapping date → {usd_buying, usd_selling, eur_buying,
    eur_selling[, "extra:<CCY>" ...]}.
    """
    exrate_index: dict[date, dict] = {}
    if "ExRate" not in wb.sheetnames:
        return exrate_index

    extra_cols = {
        ccy: column_index_from_string(col)
        for ccy, col in (exrate_col_map or {}).items()
    }

    ws_exrate = wb["ExRate"]
    for row_idx in range(2, (ws_exrate.max_row or 1) + 1):
        cell_val = ws_exrate.cell(row=row_idx, column=1).value
        row_date = None
        if isinstance(cell_val, datetime):
            row_date = cell_val.date()
        elif isinstance(cell_val, date):
            row_date = cell_val
        if row_date:
            # Fixed B-E columns from the single layout source
            # (core/exrate_sheet.py EXRATE_RATE_COLUMNS).
            entry = {
                exrate_index_key(ccy, rt): ws_exrate.cell(
                    row=row_idx, column=col
                ).value
                for col, _label, ccy, rt in EXRATE_RATE_COLUMNS
            }
            for ccy, col_idx in extra_cols.items():
                entry[f"extra:{ccy}"] = ws_exrate.cell(
                    row=row_idx, column=col_idx
                ).value
            exrate_index[row_date] = entry

    return exrate_index


def find_header_row(
    ws,
    labels: tuple[tuple[str, str | None], ...],
    *,
    scan_depth: int = 10,
    sheet_name: str = "?",
    warn_duplicates: bool = True,
    resolve_left_of: dict[str, str] | None = None,
) -> tuple[int | None, dict[str, int]]:
    """Locate a sheet's header row and map labelled columns (single owner).

    Scans the first ``scan_depth`` rows for the ANCHOR label — ``labels[0]``
    — and, on the first row containing it, maps every ``(key, label)`` pair
    to its 0-based column index. Duplicate labelled columns resolve
    deterministically to the FIRST occurrence (never last-wins, which
    silently depends on column order), EXCEPT keys listed in
    ``resolve_left_of``: ``{key: ref_key}`` resolves a duplicated ``key``
    to the occurrence nearest LEFT of ``ref_key``'s column. The real
    production ledgers carry TWO ``Date`` columns per sheet — the invoice
    date (column B) and the export-entry date immediately left of
    ``EX Rate`` — and the legacy formula contract resolves rates by the
    export-entry date, so the source-date column must bind to the
    occurrence adjacent to ``EX Rate``, not the first one. With
    ``warn_duplicates`` the collision and the applied resolution are
    logged so the operator can see which column won. ``None`` labels
    are skipped.

    Canonical implementation behind :func:`scan_sheet_headers` (ledger write
    path), ``core.ledger_processing.prescan_target_dates_and_currencies``,
    ``core.prescan._scan_xlsx``, and the standalone-ExRate routing PROBE
    (``core.workbook_io.is_standalone_exrate_workbook``) — the probe reuses
    this primitive so its month-tab recognition can never diverge from what
    the ledger scan would actually process.

    Returns:
        ``(header_row_idx, col_indices)`` — the 1-based header row index (or
        None when the anchor label is absent) and ``{key: 0-based column}``.
    """
    anchor = labels[0][1]
    header_row_idx: int | None = None
    col_indices: dict[str, int] = {}
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=scan_depth, values_only=True), 1
    ):
        row_strs = [
            str(c).strip() if c is not None else "" for c in row
        ]
        if anchor in row_strs:
            header_row_idx = row_idx
            occurrences: dict[str, list[int]] = {}
            label_by_key = {key: label for key, label in labels}
            for ci, val in enumerate(row_strs):
                for key, label in labels:
                    if label is not None and val == label:
                        occurrences.setdefault(key, []).append(ci)
            for key, occ in occurrences.items():
                chosen = occ[0]
                if len(occ) > 1:
                    ref_key = (resolve_left_of or {}).get(key)
                    ref_occ = occurrences.get(ref_key) if ref_key else None
                    left = (
                        [c for c in occ if c < ref_occ[0]] if ref_occ else []
                    )
                    if left:
                        chosen = max(left)
                    if warn_duplicates:
                        logger.warning(
                            "Sheet '%s': duplicate '%s' header column — "
                            "using the %s.",
                            sheet_name, label_by_key[key],
                            "occurrence nearest left of "
                            f"'{label_by_key.get(ref_key, ref_key)}'"
                            if left else "first occurrence",
                        )
                col_indices[key] = chosen
            break
    return header_row_idx, col_indices


def scan_sheet_headers(
    wb,
    target_cols: dict[str, str],
) -> dict[str, dict]:
    """
    Scan monthly tabs for header rows and column indices.

    Returns a dict mapping sheet_name → {header_row, columns}.
    Skips sheets in SKIP_SHEET_NAMES and sheets without the
    source date column. Header-row location and duplicate-column
    resolution live in :func:`find_header_row` (single owner).
    """
    sheet_maps: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        if is_skip_sheet(sheet_name):
            continue
        ws = wb[sheet_name]
        header_row_idx, col_indices_local = find_header_row(
            ws,
            (
                ("source", target_cols["source_date"]),
                ("currency", target_cols["currency"]),
                ("out_rate", target_cols["out_rate"]),
            ),
            # Duplicate 'Date' headers resolve to the column nearest left
            # of 'EX Rate' (the export-entry date the legacy formulas use),
            # not the first occurrence (the invoice date).
            resolve_left_of={"source": "out_rate"},
            sheet_name=sheet_name,
        )

        if header_row_idx is None or "source" not in col_indices_local:
            logger.info(
                "Sheet '%s' missing source date column — skipped.",
                sheet_name,
            )
            continue

        sheet_maps[sheet_name] = {
            "header_row": header_row_idx,
            "columns": col_indices_local,
        }

    return sheet_maps


def inject_xlookup_formulas(
    wb,
    sheet_maps: dict[str, dict],
    parse_date_fn: Callable,
    emit_fn: Callable[[str], None] | None = None,
    dry_run: bool = False,
    buffer_rows: int = PREFORMAT_BUFFER_ROWS,
    rate_type: str = "buying_transfer",
    exrate_col_map: dict[str, str] | None = None,
) -> None:
    """
    Inject XLOOKUP formulas into monthly tabs.

    Writes a SINGLE IFS formula per row to the "EX Rate" column
    that dynamically checks the Cur column for the currency.

    This means:
      - Formula can be dragged down without breaking
      - Currency is checked inside the formula via IFS()
      - THB → 1, USD → ExRate col (varies by rate_type),
        EUR → ExRate col (varies by rate_type)
      - Additional currencies (GBP/JPY/CNY) via exrate_col_map

    CRITICAL: Date Normalization
    Monthly tab dates may be stored as TEXT STRINGS (e.g.,
    "10/03/2025") which lookups cannot match against the
    DATE SERIAL NUMBERS in ExRate. We normalize by writing
    the parsed date object back to the cell.

    Args:
        wb: openpyxl Workbook.
        sheet_maps: Dict from scan_sheet_headers().
        parse_date_fn: Callable to parse cell values to date objects.
        emit_fn: Optional status callback.
        dry_run: If True, don't write; just report what would change.
        buffer_rows: Number of rows below data to pre-format.
        rate_type: API field name for the selected rate type
            ("buying_transfer", "selling", "buying_sight", "mid_rate").
            Determines which ExRate columns are referenced.
        exrate_col_map: Optional dict mapping currency code → ExRate
            column letter for additional currencies beyond USD/EUR.
    """
    # ── Map rate_type → ExRate column letters for USD and EUR ─────
    # Resolved from the single layout source (core/exrate_sheet.py):
    # selling → C/E, buying_transfer → B/D. The ledger rate_type is
    # restricted to buying_transfer/selling and normalized upstream in
    # engine, so no other value reaches here.
    fixed_letters = exrate_fixed_letters(rate_type)
    usd_col = fixed_letters["USD"]
    eur_col = fixed_letters["EUR"]

    def _guarded_lookup(date_ref: str, rate_col: str) -> str:
        """One IFS branch value: exact-match XLOOKUP, guarded twice.

        The "" 4th arg covers NOT-FOUND only. Weekend/holiday rows EXIST
        in the ExRate sheet with blank rate cells, so an unguarded lookup
        returns a reference to an empty cell which Excel renders as
        numeric 0 — accountants would see EX Rate = 0.0000 and dependent
        THB amounts compute 0. The IF guard renders found-but-empty
        results as blank "" instead. Exact-match semantics (match_mode 0)
        are preserved: NO rollback, NO carry-forward.

        WHOLE-COLUMN references ($A:$A / $B:$B): a row-pinned $A$2:$A$N
        range went stale whenever the master grew without re-injection
        (user-dragged formula copies, misrouted/skipped sheets, the
        standalone updater growing the master inside a ledger workbook).
        Row-1 text headers can never exact-match a date serial, and the
        lookup/return columns stay aligned by construction, so the
        unbounded range is safe and self-healing.
        """
        lookup = (
            f"_xlfn.XLOOKUP({date_ref},"
            f"ExRate!$A:$A,"
            f"ExRate!${rate_col}:${rate_col},\"\",0)"
        )
        return f"IFERROR(IF({lookup}=\"\",\"\",{lookup}),\"\")"

    for sheet_name, mapping in sheet_maps.items():
        ws = wb[sheet_name]
        cols = mapping["columns"]
        src_idx = cols["source"] + 1
        cur_idx = cols.get("currency")
        out_rate_idx = cols.get("out_rate")
        if out_rate_idx is None or cur_idx is None:
            continue
        out_col = out_rate_idx + 1  # 1-indexed
        cur_col = cur_idx + 1       # 1-indexed

        # Column letters for cell references in formulas
        date_letter = get_column_letter(src_idx)
        cur_letter = get_column_letter(cur_col)

        skipped = 0
        written = 0
        overwritten = 0
        # Rows whose Date or EX Rate cell is a non-anchor MergedCell —
        # read-only, so no formula can be injected and the date is not
        # normalized. Collected so the skip is SURFACED (warning + emit)
        # instead of silently leaving the row untouched.
        merged_rows: list[int] = []

        # ── Last DATA row (bounds the injection loop AND the preformat
        #    base) ────────────────────────────────────────────────────
        # A row counts as data when ANY of its source-date / Cur /
        # EX Rate cells holds a non-blank value. Bounding on ws.max_row
        # instead grew every tab without bound: the preformat pass below
        # styles buffer rows (which inflates max_row on save), so the
        # NEXT run's injection loop reached into the previous run's
        # empty buffer and wrote formulas there — +buffer_rows of sheet
        # growth per run under the daily scheduler.
        last_data_row = mapping["header_row"]
        for row_idx in range(
            mapping["header_row"] + 1, (ws.max_row or 0) + 1
        ):
            if not all(
                _is_blank_value(ws.cell(row=row_idx, column=col).value)
                for col in (src_idx, cur_col, out_col)
            ):
                last_data_row = row_idx

        for row_idx in range(
            mapping["header_row"] + 1, last_data_row + 1
        ):
            src_cell = ws.cell(row=row_idx, column=src_idx)
            out_cell = ws.cell(row=row_idx, column=out_col)
            cur_cell = ws.cell(row=row_idx, column=cur_col)

            # Merged cells are read-only at non-anchor positions — record
            # the row so the operator hears about it (warning below).
            if isinstance(src_cell, MergedCell) or isinstance(
                out_cell, MergedCell
            ):
                merged_rows.append(row_idx)
                continue

            # ── Currency Normalization (in-place strip) ────
            # Excel's IFS comparison ({cur_ref}="USD") is whitespace-
            # sensitive while the Python warning/audit mirror strips
            # (core/exrate_updater.py) — a " USD " row rendered BLANK
            # in Excel yet was recorded as filled with no warning.
            # Strip the Cur cell in place so the formula and the
            # mirror resolve identically. Case is left untouched:
            # Excel's "=" comparison is case-insensitive, which
            # matches the mirror's .upper().
            if not isinstance(cur_cell, MergedCell) and isinstance(
                cur_cell.value, str
            ):
                stripped_cur = cur_cell.value.strip()
                if stripped_cur != cur_cell.value:
                    cur_cell.value = stripped_cur if stripped_cur else None

            # ── Date Normalization ─────────────────────────
            inv_date = parse_date_fn(src_cell.value)
            if inv_date:
                existing_fmt = src_cell.number_format or "General"
                src_cell.value = inv_date
                if existing_fmt in (
                    "General", "@", "0", "general",
                ):
                    src_cell.number_format = "dd mmm yyyy"
                else:
                    src_cell.number_format = existing_fmt

            # ── Skip rows with neither date nor currency ───
            # Spacer rows inside the data extent (and any buffer rows a
            # pre-fix run already polluted) stay truly blank instead of
            # accreting an IFS formula that can only ever render "".
            cur_value = (
                None if isinstance(cur_cell, MergedCell) else cur_cell.value
            )
            if _is_blank_value(src_cell.value) and _is_blank_value(cur_value):
                continue

            # ── Build the expected XLOOKUP formula ─────────
            date_ref = f"{date_letter}{row_idx}"
            cur_ref = f"{cur_letter}{row_idx}"

            # Core IFS branches: THB, USD, EUR
            ifs_branches = (
                f"{cur_ref}=\"THB\",1,"
                f"{cur_ref}=\"USD\","
                f"{_guarded_lookup(date_ref, usd_col)},"
                f"{cur_ref}=\"EUR\","
                f"{_guarded_lookup(date_ref, eur_col)}"
            )

            # Additional currency branches from exrate_col_map
            if exrate_col_map:
                for ccy, col_letter in exrate_col_map.items():
                    if ccy in ("USD", "EUR", "THB"):
                        continue  # already handled above
                    # Validate before interpolating into the formula string;
                    # a bad code/column would otherwise corrupt the whole row.
                    if not _CCY_RE.match(ccy) or not _COL_RE.match(col_letter):
                        logger.warning(
                            "Skipping invalid exrate_col_map entry: "
                            "ccy=%r col_letter=%r",
                            ccy, col_letter,
                        )
                        continue
                    ifs_branches += (
                        f",{cur_ref}=\"{ccy}\","
                        f"{_guarded_lookup(date_ref, col_letter)}"
                    )

            formula = (
                f"=IF(OR({cur_ref}=\"\",{date_ref}=\"\"),\"\","
                f"_xlfn.IFS("
                f"{ifs_branches},"
                f"TRUE,\"\"))"
            )

            # ── Skip-if-identical: exact formula match ─────
            existing_val = out_cell.value
            if (
                isinstance(existing_val, str)
                and existing_val == formula
            ):
                skipped += 1
                continue

            # Track if we're replacing an old formula
            if existing_val is not None and isinstance(existing_val, str) and existing_val.startswith("="):
                overwritten += 1

            zero_touch_write(ws, row_idx, out_col, formula)
            written += 1

        if skipped or overwritten or written:
            logger.info(
                "Sheet '%s': %d identical (skipped), "
                "%d old formulas replaced, %d new written",
                sheet_name, skipped, overwritten,
                written - overwritten,
            )
            if dry_run and emit_fn:
                emit_fn(
                    f"[SIM] {sheet_name}: Would inject {written} formulas "
                    f"(replaced {overwritten}) and normalize {written} dates"
                )
            elif emit_fn:
                emit_fn(
                    f"{sheet_name}: {skipped} skipped, "
                    f"{overwritten} replaced, "
                    f"{written - overwritten} new"
                )

        if merged_rows:
            # Cap the listed rows at 10 to bound message size (4GB target).
            listed = ", ".join(map(str, merged_rows[:10]))
            suffix = ", …" if len(merged_rows) > 10 else ""
            msg = (
                f"{sheet_name}: {len(merged_rows)} row(s) have merged "
                f"Date/EX Rate cells — EX Rate left untouched "
                f"(rows {listed}{suffix})"
            )
            logger.warning("%s", msg)
            if emit_fn:
                emit_fn(f"[SIM] {msg}" if dry_run else msg)

        # ── Pre-format Date column for manual entry ───────────
        # Buffer rows ONLY (last_data_row+1 .. +buffer_rows): data-row
        # date formats are owned by the normalization branch above,
        # which preserves user formats. Basing this range on ws.max_row
        # both clobbered every data row to DD/MM/YYYY and re-extended
        # the styled region by buffer_rows on every run (the styled
        # cells inflate max_row), growing the sheet without bound.
        for r in range(last_data_row + 1, last_data_row + buffer_rows + 1):
            cell = ws.cell(row=r, column=src_idx)
            if not isinstance(cell, MergedCell):
                cell.number_format = "DD/MM/YYYY"


def write_custom_exrate_data(
    ws,
    rate_data: dict[str, dict[str, dict[date, float]]],
    col_specs: list[tuple[str, str]],
    headers: list[str],
    all_dates: list[date],
    holidays_set: set[date],
    holidays_names: dict[date, str],
) -> None:
    """
    Write multi-currency ExRate data with styling to a worksheet.

    Used by the custom ExRate path in update_exrate_standalone
    for non-standard currency/rate-type combinations.

    Args:
        ws: Target worksheet.
        rate_data: Nested dict: rate_data[ccy][api_field][date] = value.
        col_specs: List of (currency, api_field) per data column.
        headers: Column header labels.
        all_dates: Sorted list of dates to write.
        holidays_set: Set of holiday dates.
        holidays_names: Map of date → holiday name.
    """
    # ── Styles ────────────────────────────────────────────────────
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1A365D", end_color="1A365D", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Calibri", size=10)
    date_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    holiday_fill = PatternFill(
        start_color="FFF3CD", end_color="FFF3CD", fill_type="solid",
    )
    weekend_fill = PatternFill(
        start_color="E8E8E8", end_color="E8E8E8", fill_type="solid",
    )

    # ── Clear existing content ────────────────────────────────────
    # delete_rows drops the whole used range in one shot instead of
    # walking max_row × max_col cells (which blows up memory on an
    # inflated used-range and violates the featherweight limit).
    existing_rows = ws.max_row or 0
    if existing_rows:
        ws.delete_rows(1, existing_rows)

    # ── Write headers ─────────────────────────────────────────────
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Column widths ─────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    for i in range(len(col_specs)):
        col_letter = get_column_letter(i + 2)
        ws.column_dimensions[col_letter].width = 18
    last_col_letter = get_column_letter(len(headers))
    ws.column_dimensions[last_col_letter].width = 40

    # ── Data rows ─────────────────────────────────────────────────
    for row_offset, d in enumerate(all_dates):
        row_idx = row_offset + 2
        is_weekend = d.weekday() >= 5
        is_holiday = d in holidays_set

        # Date
        cell_date = ws.cell(row=row_idx, column=1, value=d)
        cell_date.number_format = "DD/MM/YYYY"
        cell_date.font = data_font
        cell_date.alignment = date_align
        cell_date.border = thin_border

        # Rate columns
        for col_offset, (ccy, api_field) in enumerate(col_specs):
            val = rate_data.get(ccy, {}).get(api_field, {}).get(d)
            cell = ws.cell(row=row_idx, column=col_offset + 2, value=val)
            cell.number_format = "0.0000"
            cell.font = data_font
            cell.border = thin_border

        # Holiday/Weekend label
        if is_weekend and is_holiday:
            label = f"Weekend; {holidays_names.get(d, 'Holiday')}"
        elif is_weekend:
            label = "Weekend"
        elif is_holiday:
            label = holidays_names.get(d, "Holiday")
        else:
            label = ""

        cell_label = ws.cell(
            row=row_idx, column=len(headers), value=label,
        )
        cell_label.font = data_font
        cell_label.border = thin_border

        # Row fill
        if is_holiday:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=ci).fill = holiday_fill
        elif is_weekend:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=ci).fill = weekend_fill
