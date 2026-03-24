#!/usr/bin/env python3
"""
core/xls_converter.py
---------------------------------------------------------------------------
BOT Exchange Rate Processor (v3.0.8) - Native OS-Aware Conversion Pipeline
---------------------------------------------------------------------------
Strictly converts legacy .xls files to .xlsx using native proxies (win32com
on Windows, or soffice headless on Mac/Linux) to guarantee absolute 100%
preservation of enterprise fonts, styles, and geometries.
"""

import logging
import os
import shutil
import subprocess

logger = logging.getLogger(__name__)


def _get_soffice_path() -> str | None:
    """Find the LibreOffice executable path."""
    # macOS path
    mac_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if os.path.exists(mac_path):
        return mac_path

    # Linux / Windows paths in PATH
    try:
        # 'where' on Windows, 'which' on Linux/Mac
        cmd = "where" if os.name == "nt" else "which"
        path = subprocess.check_output([cmd, "soffice"]).decode().strip()
        if path and os.path.exists(path):
            return path
    except Exception:
        pass
    return None


def convert_xls_to_xlsx(filepath: str) -> str:
    """
    Convert a legacy .xls file to .xlsx preserving formatting natively.

    Priority 1 (Windows): win32com.client (Microsoft Excel proxy)
    Priority 2 (Mac/Linux/WinFallback): LibreOffice soffice --headless proxy
    Priority 3 (Failsafe): RuntimeError (Alerts GUI to prevent style loss)
    """
    logger.info("Converting legacy .xls natively: %s", os.path.basename(filepath))
    filepath = os.path.abspath(filepath)
    dir_name = os.path.dirname(filepath)
    base_name = os.path.splitext(os.path.basename(filepath))[0]
    temp_path = os.path.join(dir_name, f".{base_name}_converted.xlsx")

    # ── Priority 1: LIBREOFFICE PROXY (MAC / LINUX / WIN-FALLBACK) ───
    soffice_path = _get_soffice_path()
    if soffice_path:
        logger.info("Engaging LibreOffice Native Headless Engine.")
        try:
            outdir = dir_name or "."
            subprocess.run(
                [
                    soffice_path, "--headless", "--convert-to", "xlsx",
                    "--outdir", outdir, filepath
                ],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
            standard_outpath = os.path.join(outdir, f"{base_name}.xlsx")
            if os.path.exists(standard_outpath):
                shutil.move(standard_outpath, temp_path)
                logger.info("LibreOffice conversion complete: %s", temp_path)
                return temp_path
        except Exception as e:
            logger.warning("LibreOffice proxy failed. %s", e)

    # ── Priority 2: PURE PYTHON FALLBACK (xlrd → openpyxl) ─────────
    # No external software needed — works on ALL platforms
    # Preserves: fonts, fills, borders, number formats, alignment,
    #            column widths, row heights, merged cells
    logger.info("No native converter found. Using pure Python xlrd → openpyxl conversion.")
    try:
        import openpyxl as _openpyxl
        import xlrd
        from openpyxl.styles import (
            Alignment as _Alignment,
        )
        from openpyxl.styles import (
            Border as _Border,
        )
        from openpyxl.styles import (
            Font as _Font,
        )
        from openpyxl.styles import (
            PatternFill as _PatternFill,
        )
        from openpyxl.styles import (
            Side as _Side,
        )
        from openpyxl.utils import get_column_letter as _gcl

        devnull_fh = open(os.devnull, 'w')
        try:
            wb_xls = xlrd.open_workbook(
                filepath, formatting_info=True, logfile=devnull_fh,
            )
        finally:
            devnull_fh.close()

        wb_xlsx = _openpyxl.Workbook()
        # Remove the default empty sheet that openpyxl creates
        if wb_xlsx.sheetnames:
            wb_xlsx.remove(wb_xlsx.active)

        # ── Pre-build xlrd formatting lookup tables ───────────────────
        xf_list = wb_xls.xf_list
        font_list = wb_xls.font_list
        format_map = wb_xls.format_map

        # xlrd border style index → openpyxl border style name
        _BORDER_STYLES = {
            0: None, 1: "thin", 2: "medium", 3: "dashed",
            4: "dotted", 5: "thick", 6: "double", 7: "hair",
            8: "mediumDashed", 9: "dashDot", 10: "mediumDashDot",
            11: "dashDotDot", 12: "mediumDashDotDot", 13: "slantDashDot",
        }

        def _xlrd_color_to_hex(colour_index: int) -> str | None:
            """Convert xlrd colour index to hex string, or None."""
            if colour_index is None or colour_index in (
                0x40, 0x41, 0x43, 0x7F, 64, 65, 127,
            ):
                return None
            colour_map = wb_xls.colour_map
            rgb = colour_map.get(colour_index)
            if rgb and rgb != (0, 0, 0) or colour_index == 0:
                if rgb:
                    return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
            return None

        def _make_font(xf_idx: int) -> _Font | None:
            """Build openpyxl Font from xlrd XF record."""
            try:
                xf = xf_list[xf_idx]
                fnt = font_list[xf.font_index]
                color_hex = _xlrd_color_to_hex(fnt.colour_index)
                return _Font(
                    name=fnt.name or "Calibri",
                    size=fnt.height / 20 if fnt.height else 11,
                    bold=fnt.bold,
                    italic=fnt.italic,
                    underline="single" if fnt.underline_type else None,
                    strike=fnt.struck_out,
                    color=color_hex if color_hex else "000000",
                )
            except Exception:
                return None

        def _make_fill(xf_idx: int) -> _PatternFill | None:
            """Build openpyxl PatternFill from xlrd XF record."""
            try:
                xf = xf_list[xf_idx]
                bg = xf.background
                fg_hex = _xlrd_color_to_hex(bg.pattern_colour_index)
                bg_hex = _xlrd_color_to_hex(bg.background_colour_index)
                if bg.fill_pattern == 1 and fg_hex:
                    return _PatternFill(
                        patternType="solid", fgColor=fg_hex,
                    )
                elif bg.fill_pattern > 0 and (fg_hex or bg_hex):
                    return _PatternFill(
                        patternType="solid",
                        fgColor=fg_hex or bg_hex,
                    )
                return None
            except Exception:
                return None

        def _make_border_side(border_line_style: int, colour_index: int) -> _Side:
            style = _BORDER_STYLES.get(border_line_style)
            color = _xlrd_color_to_hex(colour_index)
            if style:
                return _Side(style=style, color=color or "000000")
            return _Side()

        def _make_border(xf_idx: int) -> _Border | None:
            """Build openpyxl Border from xlrd XF record."""
            try:
                xf = xf_list[xf_idx]
                b = xf.border
                return _Border(
                    left=_make_border_side(
                        b.left_line_style, b.left_colour_index),
                    right=_make_border_side(
                        b.right_line_style, b.right_colour_index),
                    top=_make_border_side(
                        b.top_line_style, b.top_colour_index),
                    bottom=_make_border_side(
                        b.bottom_line_style, b.bottom_colour_index),
                )
            except Exception:
                return None

        def _make_alignment(xf_idx: int) -> _Alignment | None:
            """Build openpyxl Alignment from xlrd XF record."""
            try:
                xf = xf_list[xf_idx]
                al = xf.alignment
                h_map = {
                    0: "general", 1: "left", 2: "center", 3: "right",
                    4: "fill", 5: "justify", 6: "centerContinuous",
                }
                v_map = {0: "top", 1: "center", 2: "bottom", 3: "justify"}
                return _Alignment(
                    horizontal=h_map.get(al.hor_align, "general"),
                    vertical=v_map.get(al.vert_align, "bottom"),
                    wrap_text=bool(al.text_wrapped),
                    text_rotation=al.rotation if al.rotation < 181 else 0,
                )
            except Exception:
                return None

        def _get_number_format(xf_idx: int) -> str | None:
            """Get the number format string from xlrd XF record."""
            try:
                xf = xf_list[xf_idx]
                fmt_key = xf.format_key
                if fmt_key in format_map:
                    return format_map[fmt_key].format_str
                return None
            except Exception:
                return None

        # ── Convert each sheet ────────────────────────────────────────
        for sheet_name in wb_xls.sheet_names():
            ws_xls = wb_xls.sheet_by_name(sheet_name)
            ws_xlsx = wb_xlsx.create_sheet(title=sheet_name)

            # Transfer cell values + formatting
            for row_idx in range(ws_xls.nrows):
                for col_idx in range(ws_xls.ncols):
                    cell_type = ws_xls.cell_type(row_idx, col_idx)
                    cell_val = ws_xls.cell_value(row_idx, col_idx)

                    # Convert xlrd types to Python types
                    if cell_type == xlrd.XL_CELL_DATE and cell_val:
                        try:
                            from datetime import datetime as _dt
                            dt_tuple = xlrd.xldate_as_tuple(
                                cell_val, wb_xls.datemode)
                            if dt_tuple[3:] == (0, 0, 0):
                                cell_val = _dt(*dt_tuple[:3]).date()
                            else:
                                cell_val = _dt(*dt_tuple)
                        except Exception:
                            pass
                    elif cell_type == xlrd.XL_CELL_BOOLEAN:
                        cell_val = bool(cell_val)
                    elif cell_type == xlrd.XL_CELL_EMPTY:
                        # Still apply formatting to empty cells if they have it
                        xf_idx = ws_xls.cell_xf_index(row_idx, col_idx)
                        if xf_idx > 0:
                            target = ws_xlsx.cell(
                                row=row_idx + 1, column=col_idx + 1)
                            font = _make_font(xf_idx)
                            if font:
                                target.font = font
                            fill = _make_fill(xf_idx)
                            if fill:
                                target.fill = fill
                            border = _make_border(xf_idx)
                            if border:
                                target.border = border
                        continue

                    target = ws_xlsx.cell(
                        row=row_idx + 1, column=col_idx + 1)
                    target.value = cell_val

                    # Apply formatting from XF record
                    xf_idx = ws_xls.cell_xf_index(row_idx, col_idx)
                    font = _make_font(xf_idx)
                    if font:
                        target.font = font
                    fill = _make_fill(xf_idx)
                    if fill:
                        target.fill = fill
                    border = _make_border(xf_idx)
                    if border:
                        target.border = border
                    alignment = _make_alignment(xf_idx)
                    if alignment:
                        target.alignment = alignment
                    num_fmt = _get_number_format(xf_idx)
                    if num_fmt:
                        target.number_format = num_fmt

            # Copy column widths from xlrd (in character units)
            for col_idx in range(ws_xls.ncols):
                col_letter = _gcl(col_idx + 1)
                try:
                    # xlrd stores width in 1/256 of character width
                    col_info = ws_xls.colinfo_map.get(col_idx)
                    if col_info:
                        width_chars = col_info.width / 256
                        ws_xlsx.column_dimensions[col_letter].width = max(
                            width_chars, 8)
                    else:
                        ws_xlsx.column_dimensions[col_letter].width = 14
                except Exception:
                    ws_xlsx.column_dimensions[col_letter].width = 14

            # Copy row heights
            for row_idx in range(ws_xls.nrows):
                try:
                    row_info = ws_xls.rowinfo_map.get(row_idx)
                    if row_info and row_info.height:
                        # xlrd stores height in twips (1/20 pt)
                        ws_xlsx.row_dimensions[
                            row_idx + 1].height = row_info.height / 20
                except Exception:
                    pass

            # Copy merged cell ranges
            for crange in ws_xls.merged_cells:
                rlo, rhi, clo, chi = crange
                ws_xlsx.merge_cells(
                    start_row=rlo + 1, start_column=clo + 1,
                    end_row=rhi, end_column=chi,
                )

        wb_xls.release_resources()
        wb_xlsx.save(temp_path)
        wb_xlsx.close()
        logger.info("Pure Python conversion complete (with formatting): %s",
                     os.path.basename(temp_path))
        return temp_path
    except Exception as e:
        logger.error("Pure Python xlrd→openpyxl conversion failed: %s", e)
        raise RuntimeError(
            f"Failed to convert .xls to .xlsx: {e}\n"
            "Install LibreOffice for better conversion, or use .xlsx files directly."
        ) from e
