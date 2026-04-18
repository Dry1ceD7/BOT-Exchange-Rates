#!/usr/bin/env python3
"""
core/enterprise.py
---------------------------------------------------------------------------
Enterprise utility helpers:
 - Validation summaries
 - Holiday overlay loading
 - Multi-source FX fallback
 - Webhook notifications
 - Job history persistence
 - Secret masking helpers
---------------------------------------------------------------------------
"""

from __future__ import annotations

import csv
import json
import logging
import os
from dataclasses import dataclass
from datetime import date, datetime
from typing import Dict, List, Optional

import httpx
import openpyxl

from core.paths import get_project_root

logger = logging.getLogger(__name__)

_ALLOWED_CURRENCIES = {"USD", "EUR", "THB"}
_SKIP_SHEETS = {"ExRate", "Exrate USD", "Exrate EUR"}


def mask_secret(value: Optional[str], visible: int = 4) -> str:
    """Mask a secret string for UI/log output."""
    if not value:
        return ""
    if len(value) <= visible:
        return "*" * len(value)
    return f"{'*' * (len(value) - visible)}{value[-visible:]}"


def _parse_excel_date(value) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y%m%d"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    return None


def _find_header_indexes(ws) -> Optional[Dict[str, int]]:
    for row_idx in range(1, 11):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, (ws.max_column or 1) + 1)]
        row_str = [str(v).strip() if v is not None else "" for v in row_vals]
        if "Date" in row_str and "Cur" in row_str and "EX Rate" in row_str:
            return {
                "row": row_idx,
                "date": row_str.index("Date") + 1,
                "cur": row_str.index("Cur") + 1,
                "out": row_str.index("EX Rate") + 1,
            }
    return None


def validate_ledger_file(filepath: str) -> Dict[str, int]:
    """Return validation counters for a single ledger file."""
    result = {
        "missing_dates": 0,
        "unmatched_currencies": 0,
        "skipped_rows": 0,
        "scanned_rows": 0,
    }
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False, read_only=True)
    except (OSError, ValueError):
        return result

    try:
        for sheet_name in wb.sheetnames:
            if sheet_name in _SKIP_SHEETS:
                continue
            ws = wb[sheet_name]
            hdr = _find_header_indexes(ws)
            if not hdr:
                continue
            for row_idx in range(hdr["row"] + 1, (ws.max_row or 1) + 1):
                date_val = ws.cell(row=row_idx, column=hdr["date"]).value
                cur_val = ws.cell(row=row_idx, column=hdr["cur"]).value
                out_val = ws.cell(row=row_idx, column=hdr["out"]).value
                cur = str(cur_val).strip().upper() if cur_val is not None else ""
                if not cur:
                    continue
                result["scanned_rows"] += 1
                if _parse_excel_date(date_val) is None:
                    result["missing_dates"] += 1
                if cur not in _ALLOWED_CURRENCIES:
                    result["unmatched_currencies"] += 1
                if isinstance(out_val, str) and out_val.strip().startswith("="):
                    result["skipped_rows"] += 1
    finally:
        try:
            wb.close()
        except OSError:
            pass
    return result


def summarize_validation(filepaths: List[str]) -> Dict[str, object]:
    """Aggregate validation summary across file paths."""
    per_file = []
    totals = {
        "missing_dates": 0,
        "unmatched_currencies": 0,
        "skipped_rows": 0,
        "scanned_rows": 0,
    }
    for fp in filepaths:
        info = validate_ledger_file(fp)
        per_file.append({"file": os.path.basename(fp), **info})
        for k in totals:
            totals[k] += int(info.get(k, 0))
    return {
        "file_count": len(filepaths),
        "totals": totals,
        "per_file": per_file,
    }


def load_holiday_overlays(path: str) -> List[tuple[str, str]]:
    """Load holiday overlays from CSV, JSON, or TXT."""
    if not path:
        return []
    if not os.path.exists(path):
        logger.warning("Holiday overlay file not found: %s", path)
        return []
    ext = os.path.splitext(path)[1].lower()
    items: List[tuple[str, str]] = []
    try:
        if ext == ".csv":
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    d = str(row.get("date", "")).strip()
                    n = str(row.get("name", "Company Holiday")).strip()
                    if d:
                        items.append((d, n))
        elif ext == ".json":
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                for row in data:
                    if isinstance(row, dict):
                        d = str(row.get("date", "")).strip()
                        n = str(row.get("name", "Company Holiday")).strip()
                        if d:
                            items.append((d, n))
        else:
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    raw = line.strip()
                    if not raw or raw.startswith("#"):
                        continue
                    if "," in raw:
                        d, n = raw.split(",", 1)
                        items.append((d.strip(), n.strip() or "Company Holiday"))
                    else:
                        items.append((raw, "Company Holiday"))
    except (OSError, ValueError, TypeError, json.JSONDecodeError) as e:
        logger.warning("Holiday overlay load failed: %s", e)
        return []
    return items


@dataclass
class FallbackRateRecord:
    period: str
    currency: str
    buying_transfer: Optional[float] = None
    buying_sight: Optional[float] = None
    selling: Optional[float] = None
    mid_rate: Optional[float] = None


async def fetch_fallback_rates(
    start_date: date,
    end_date: date,
    currency: str,
    client: httpx.AsyncClient,
) -> List[FallbackRateRecord]:
    """Fetch fallback FX rates from Frankfurter and map to BOT-like fields."""
    url = f"https://api.frankfurter.app/{start_date.isoformat()}..{end_date.isoformat()}"
    params = {"from": currency.upper(), "to": "THB"}
    resp = await client.get(url, params=params, timeout=20.0)
    resp.raise_for_status()
    payload = resp.json()
    out: List[FallbackRateRecord] = []
    rates = payload.get("rates", {}) if isinstance(payload, dict) else {}
    if not isinstance(rates, dict):
        return out
    for d_str, day_rates in rates.items():
        if not isinstance(day_rates, dict):
            continue
        val = day_rates.get("THB")
        if val is None:
            continue
        try:
            rate = float(val)
        except (TypeError, ValueError):
            continue
        out.append(
            FallbackRateRecord(
                period=d_str,
                currency=currency.upper(),
                buying_transfer=rate,
                buying_sight=rate,
                selling=rate,
                mid_rate=rate,
            )
        )
    return out


def send_webhook_notification(webhook_url: str, payload: dict) -> bool:
    """Send a simple JSON webhook notification."""
    if not webhook_url:
        return False
    try:
        resp = httpx.post(webhook_url, json=payload, timeout=5.0)
        if 200 <= resp.status_code < 300:
            return True
        logger.warning("Webhook notification failed: HTTP %s", resp.status_code)
        return False
    except (httpx.RequestError, ValueError) as e:
        logger.warning("Webhook notification error: %s", e)
        return False


def _default_history_path() -> str:
    return os.path.join(get_project_root(), "data", "logs", "job_history.jsonl")


def record_job_history(entry: dict, history_path: Optional[str] = None) -> str:
    """Append a job entry to JSONL history."""
    path = history_path or _default_history_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    row = dict(entry)
    row.setdefault("timestamp", datetime.now().isoformat(timespec="seconds"))
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\n")
    return path


def load_job_history_stats(limit: int = 30, history_path: Optional[str] = None) -> dict:
    """Read recent job history and return compact stats."""
    path = history_path or _default_history_path()
    if not os.path.exists(path):
        return {"runs": 0, "success_runs": 0, "failed_runs": 0, "avg_runtime_sec": 0.0}
    rows = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rows.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    except OSError:
        return {"runs": 0, "success_runs": 0, "failed_runs": 0, "avg_runtime_sec": 0.0}

    if limit > 0:
        rows = rows[-limit:]
    runs = len(rows)
    success_runs = sum(1 for r in rows if int(r.get("failed", 0)) == 0)
    failed_runs = runs - success_runs
    runtimes = [float(r.get("duration_sec", 0.0)) for r in rows if r.get("duration_sec") is not None]
    avg_runtime = sum(runtimes) / len(runtimes) if runtimes else 0.0
    return {
        "runs": runs,
        "success_runs": success_runs,
        "failed_runs": failed_runs,
        "avg_runtime_sec": round(avg_runtime, 2),
    }
