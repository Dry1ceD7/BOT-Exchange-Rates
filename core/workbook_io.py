#!/usr/bin/env python3
"""
core/workbook_io.py
---------------------------------------------------------------------------
BOT Exchange Rate Processor — Workbook I/O guardrails
---------------------------------------------------------------------------
Small, shared helpers for the in-place openpyxl save pipeline.

ensure_disk_space is the single source of truth for the pre-save free-space
guard used by both the ledger write pipeline and the standalone updater. It is
deliberately implemented with a module-level ``import shutil`` + a late
``shutil.disk_usage(...)`` lookup so that tests which patch the shared shutil
module object (e.g. ``monkeypatch.setattr(engine_mod.shutil, 'disk_usage', ...)``)
see the patch — module objects are singletons.
"""

import contextlib
import gc
import logging
import shutil
from collections.abc import Callable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from core.constants import MAX_FILE_SIZE_MB, is_skip_sheet

logger = logging.getLogger(__name__)


class WorkbookVerifyError(OSError):
    """Post-write verification failed; the original file was left untouched.

    Subclasses OSError so every existing save-failure handler (the engine
    batch loop, the GUI handlers, the ExRate dialog worker) surfaces it
    through the same error path as a disk/permission failure.
    """


def ensure_disk_space(target_dir: Path, min_mb: int) -> None:
    """Raise OSError if free disk space at target_dir is below min_mb.

    Args:
        target_dir: Directory the workbook will be saved into.
        min_mb: Minimum required free space in megabytes.

    Raises:
        OSError: If free space is below the configured minimum. The message
            is byte-identical to the legacy inline guard so callers and tests
            matching on "Insufficient disk space" keep working.
    """
    free_mb = shutil.disk_usage(target_dir).free // (1024 * 1024)
    if free_mb < min_mb:
        raise OSError(
            f"Insufficient disk space ({free_mb}MB free, "
            f"need {min_mb}MB). File NOT saved."
        )


def atomic_write_text(path: Path, payload: str) -> None:
    """Atomically (over)write a small text file (the round-7 temp + replace
    idiom — text twin of :func:`atomic_save`).

    Writes ``payload`` to a sibling temp file in the SAME directory then
    ``Path.replace`` swaps it in, so the replace stays on one filesystem and
    a crash mid-write leaves the previous good file untouched. On any
    failure the partial temp file is removed and never left behind.

    NOTE: ``core.config_manager.SettingsManager._save_locked`` deliberately
    does NOT use this helper: several long-lived SettingsManager instances
    save concurrently, so the unique temp name from
    ``tempfile.NamedTemporaryFile`` is load-bearing there — the fixed
    ``.tmp~`` sibling used here would let two savers race on one temp path.
    """
    tmp_path = path.with_name(f"{path.name}.tmp~")
    try:
        tmp_path.write_text(payload, encoding="utf-8")
        tmp_path.replace(path)
    except BaseException:
        with contextlib.suppress(OSError):
            tmp_path.unlink()
        raise


def is_standalone_exrate_workbook(
    filepath: str,
    *,
    date_header: str = "Date",
    currency_header: str = "Cur",
) -> bool:
    """Read-only probe: is ``filepath`` a standalone ExRate workbook?

    A standalone ExRate workbook has an ``ExRate`` sheet and NO sheet the
    ledger pipeline would process. "Would process" is decided by the SAME
    primitive the ledger scan uses (``core.excel_io.find_header_row``, same
    10-row depth, same anchor semantics: a mapped source-date column) — by
    construction the probe can never disagree with ``scan_sheet_headers``.
    The old 5-row 'Date'+'Cur'-in-one-row heuristic could: a ledger whose
    headers sit in rows 6-10 (or whose Cur label is absent from the header
    row) was misrouted onto the standalone path on every run AFTER the
    first — the run that creates the ExRate sheet — freezing its ledger
    formulas forever while the master kept growing.

    Single owner of the probe shared by
    ``LedgerEngine._detect_standalone_exrate`` (which passes its
    ``target_cols`` labels) and main.py's headless labeller (which keeps the
    literal ``'Date'``/``'Cur'`` defaults — the engine's default target_cols
    are those same literals, so the two callers' match semantics were
    already identical).

    Featherweight: read-only load, try/finally close + gc. ANY probe failure
    (including OSError on a locked/corrupt file) returns False — the probe
    must never mislabel a ledger or crash the caller.
    """
    if not filepath.lower().endswith(".xlsx"):
        return False
    # Featherweight pre-check: stat the size BEFORE opening. The engine's
    # memory guardrail rejects oversized files later anyway, so probing one
    # here would only burn memory parsing a workbook that can never be
    # processed. Both probe callers (the engine's standalone detection and
    # main.py's headless labeller) share this guard.
    try:
        if Path(filepath).stat().st_size > MAX_FILE_SIZE_MB * 1024 * 1024:
            logger.debug(
                "Standalone detection probe skipped (file exceeds %dMB): %s",
                MAX_FILE_SIZE_MB, filepath,
            )
            return False
    except OSError:
        return False
    wb = None
    try:
        # Local import: workbook_io is imported by excel_io's save path, so
        # a top-level import here would be circular.
        from core.excel_io import find_header_row

        wb = openpyxl.load_workbook(filepath, read_only=True)
        if "ExRate" not in wb.sheetnames:
            return False
        for sheet_name in wb.sheetnames:
            if is_skip_sheet(sheet_name):
                continue
            ws = wb[sheet_name]
            header_row_idx, cols = find_header_row(
                ws,
                (
                    ("source", date_header),
                    ("currency", currency_header),
                ),
                warn_duplicates=False,
            )
            if header_row_idx is not None and "source" in cols:
                # The ledger scan would process this sheet → normal
                # ledger, not standalone. (Same condition as
                # scan_sheet_headers' skip check.)
                return False
        return True
    except Exception as exc:  # noqa: BLE001 — probe must never propagate
        logger.debug("Standalone detection probe failed: %s", exc)
        return False
    finally:
        if wb is not None:
            with contextlib.suppress(OSError):
                wb.close()
        del wb
        gc.collect()


def atomic_save(
    wb,
    filepath: str,
    verify: Callable[[openpyxl.Workbook], None] | None = None,
) -> None:
    """Save an openpyxl workbook atomically over an existing file.

    ``wb.save(filepath)`` opens the target ZipFile in 'w' mode, truncating the
    original immediately — a crash mid-save destroys the ledger. We instead
    save to a sibling temp file in the SAME directory (so the replace stays on
    one filesystem and is atomic) and only swap it in once the write fully
    succeeds. On any failure the original is left untouched and the temp file
    is removed.

    Args:
        wb: An openpyxl Workbook with a ``save(path)`` method.
        filepath: Destination path to overwrite atomically.
        verify: Optional callback receiving the saved TEMP file reopened with
            ``load_workbook(read_only=True, data_only=False)``. Raise from it
            to abort: the temp file is unlinked, the original is never
            replaced, and the failure is re-raised as
            :class:`WorkbookVerifyError`. This is the Layer-1 exactness hard
            gate — what reached disk is proven to match the in-memory
            expectation BEFORE it replaces the user's file.

    Raises:
        WorkbookVerifyError: When ``verify`` raised; the original file on
            disk is byte-for-byte untouched.
    """
    target = Path(filepath)
    tmp_path = target.with_name(f"{target.name}.tmp~")
    try:
        wb.save(str(tmp_path))
        if verify is not None:
            _run_verifier(tmp_path, verify)
        # Path.replace is atomic on the same filesystem (same dir guarantees it).
        tmp_path.replace(target)
    except BaseException:
        # Clean up the partial temp file; never leave it behind. The original
        # file is still intact because the replace never ran.
        with contextlib.suppress(OSError):
            tmp_path.unlink()
        raise


def _run_verifier(tmp_path: Path, verify: Callable) -> None:
    """Reopen the saved temp file read-only and run ``verify`` against it.

    ANY exception from the reopen or the verifier is wrapped in
    :class:`WorkbookVerifyError` with a clear message; ``atomic_save``'s
    outer handler then unlinks the temp so the original is never replaced.
    The reopened handle is closed + gc'd on every exit (house style).

    The temp is opened as a BINARY HANDLE (not a path) because openpyxl's
    path-based loader rejects the ``.tmp~`` extension — same pattern as
    ``core.rate_audit.StandaloneRateAuditor._read_exrate_dates``.
    """
    fh = None
    reopened = None
    try:
        try:
            fh = tmp_path.open("rb")
            reopened = openpyxl.load_workbook(
                fh, read_only=True, data_only=False
            )
            verify(reopened)
        except Exception as exc:
            raise WorkbookVerifyError(
                "Post-write verification failed — original file left "
                f"untouched: {exc}"
            ) from exc
    finally:
        if reopened is not None:
            with contextlib.suppress(OSError):
                reopened.close()
        if fh is not None:
            with contextlib.suppress(OSError):
                fh.close()
        del reopened
        gc.collect()


def build_cell_verifier(
    expected: dict[str, dict[int, dict[int, object]]],
) -> Callable[[openpyxl.Workbook], None]:
    """Build an ``atomic_save`` verify callback from expected cell values.

    Args:
        expected: ``{sheet_name: {row: {col: value}}}`` (1-based row/col)
            collected from the in-memory workbook at the write sites. ``None``
            asserts a blank cell; a str asserts an exact match (this is how
            injected formula strings are verified — openpyxl round-trips them
            verbatim, leading ``=`` included); int/float/Decimal assert
            ``Decimal(str(actual)) == Decimal(str(expected))`` so a 4dp
            Decimal written in memory must reparse to exactly that value.

    Returns:
        A callable raising ValueError on the first mismatch. Featherweight:
        ONE forward read-only pass per sheet (iter_rows) — never random
        access on the reopened temp file, which would re-parse the XML.
    """

    def _verify(wb) -> None:
        for sheet_name, rows in expected.items():
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"sheet '{sheet_name}' missing from saved file"
                )
            ws = wb[sheet_name]
            unseen = set(rows)
            for row_idx, row_vals in enumerate(
                ws.iter_rows(values_only=True), 1
            ):
                cols = rows.get(row_idx)
                if cols is None:
                    continue
                unseen.discard(row_idx)
                for col_idx, exp in cols.items():
                    actual = (
                        row_vals[col_idx - 1]
                        if col_idx <= len(row_vals) else None
                    )
                    if not _cell_matches(exp, actual):
                        raise ValueError(
                            f"{sheet_name}!{get_column_letter(col_idx)}"
                            f"{row_idx}: expected {exp!r}, saved file "
                            f"holds {actual!r}"
                        )
            # Rows beyond the saved sheet's extent: blank expectations pass,
            # anything else means written data never reached the file.
            for row_idx in sorted(unseen):
                for col_idx, exp in rows[row_idx].items():
                    if exp is not None:
                        raise ValueError(
                            f"{sheet_name}!{get_column_letter(col_idx)}"
                            f"{row_idx}: expected {exp!r}, row missing "
                            "from saved file"
                        )

    return _verify


def _cell_matches(expected: object, actual: object) -> bool:
    """True when a reopened cell value equals the in-memory expectation.

    Blank → None; str → exact equality; numbers → Decimal-string round-trip
    (the Mathematical Truth rule: the saved number must reparse to exactly
    the intended Decimal); date/datetime → normalized datetime equality
    (openpyxl reads date cells back as ``datetime``).
    """
    if expected is None:
        return actual is None
    if isinstance(expected, str):
        return expected == actual
    if isinstance(expected, bool):  # bool is an int subclass — check first
        return expected == actual
    if isinstance(expected, (int, float, Decimal)):
        if isinstance(actual, bool) or not isinstance(
            actual, (int, float, Decimal)
        ):
            return False
        try:
            return Decimal(str(actual)) == Decimal(str(expected))
        except InvalidOperation:
            return False
    if isinstance(expected, (datetime, date)):
        exp_dt = (
            expected if isinstance(expected, datetime)
            else datetime(expected.year, expected.month, expected.day)
        )
        act_dt = (
            datetime(actual.year, actual.month, actual.day)
            if isinstance(actual, date) and not isinstance(actual, datetime)
            else actual
        )
        return act_dt == exp_dt
    return expected == actual
