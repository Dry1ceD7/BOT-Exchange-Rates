#!/usr/bin/env python3
"""
core/exrate_sheet.py
---------------------------------------------------------------------------
BOT Exchange Rate Processor - Master ExRate Sheet Builder
---------------------------------------------------------------------------
Separated from engine.py for SFFB compliance (<200 lines).
Builds and updates the unified "ExRate" master tab in Excel workbooks.
"""

import logging
import re
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.constants import bot_today
from core.constants import parse_date as _shared_parse_date
from core.logic import safe_to_decimal

logger = logging.getLogger(__name__)

# ── ExRate standard layout — SINGLE SOURCE OF TRUTH ──────────────────────
# Fixed columns: A=Date, B=USD Buying TT, C=USD Selling, D=EUR Buying TT,
# E=EUR Selling. Extra currencies are appended from EXRATE_EXTRA_START_COL;
# the Holidays/Weekend label always sits in the column AFTER the last rate
# column (see exrate_holidays_col). Consumed by core/excel_io.py (index +
# formula columns), core/exrate_updater.py (audit/warning resolution) and
# core/rate_audit.py (layout validation) — change it HERE only.
EXRATE_SHEET_NAME = "ExRate"
EXRATE_DATE_HEADER = "Date"
EXRATE_HOLIDAYS_HEADER = "Holidays/Weekend"

# Each entry: (1-based column, header label, currency, BOT api rate_type).
EXRATE_RATE_COLUMNS: tuple[tuple[int, str, str, str], ...] = (
    (2, "USD Buying TT Rate", "USD", "buying_transfer"),
    (3, "USD Selling Rate", "USD", "selling"),
    (4, "EUR Buying TT Rate", "EUR", "buying_transfer"),
    (5, "EUR Selling Rate", "EUR", "selling"),
)

# First column for appended extra currencies (immediately after E).
EXRATE_EXTRA_START_COL = 6


def exrate_holidays_col(extra_count: int) -> int:
    """1-based Holidays/Weekend column for a sheet with N extra currencies."""
    return EXRATE_EXTRA_START_COL + extra_count


def exrate_index_key(currency: str, rate_type: str) -> str:
    """In-memory lookup key for a fixed rate column ("usd_buying", ...).

    These are the keys build_exrate_index (core/excel_io.py) and the
    updater's fixed-column resolution use: the BOT api field
    "buying_transfer" shortens to "buying".
    """
    suffix = "buying" if rate_type == "buying_transfer" else "selling"
    return f"{currency.lower()}_{suffix}"


def _fixed_columns_for(rate_type: str) -> tuple[tuple[int, str, str, str], ...]:
    """The EXRATE_RATE_COLUMNS entries for one ledger rate type.

    The ledger rate_type is normalized upstream to buying_transfer/selling;
    anything that is not "selling" resolves to the Buying TT columns
    (preserving the historical else-branch behavior).
    """
    rt = "selling" if rate_type == "selling" else "buying_transfer"
    return tuple(e for e in EXRATE_RATE_COLUMNS if e[3] == rt)


def exrate_fixed_index_keys(rate_type: str) -> dict[str, str]:
    """``{currency: index_key}`` for the fixed USD/EUR columns of one rate type."""
    return {
        ccy: exrate_index_key(ccy, rt)
        for _col, _label, ccy, rt in _fixed_columns_for(rate_type)
    }


def exrate_fixed_letters(rate_type: str) -> dict[str, str]:
    """``{currency: column_letter}`` for the fixed USD/EUR columns of one rate type."""
    return {
        ccy: get_column_letter(col)
        for col, _label, ccy, _rt in _fixed_columns_for(rate_type)
    }


def update_master_exrate_sheet(
    wb: openpyxl.Workbook,
    usd_buying_rates: dict[date, Decimal],
    usd_selling_rates: dict[date, Decimal],
    eur_buying_rates: dict[date, Decimal],
    eur_selling_rates: dict[date, Decimal],
    holidays_list: list[date],
    holidays_names: dict[date, str],
    start_date: date,
    end_date: date | None = None,
    extra_currency_rates: dict[str, dict[date, Decimal]] | None = None,
) -> dict[str, str]:
    """
    Creates or updates a unified "ExRate" master tab.

    Columns: Date | USD Buying TT Rate | USD Selling Rate |
             EUR Buying TT Rate | EUR Selling Rate |
             [<CCY> Rate ...extra...] | Holidays/Weekend

    Holiday/Weekend Overlap Rule (semicolon separator):
      - Weekend only → "Weekend"
      - Holiday on weekday → "[Holiday Name]"
      - Holiday on weekend → "Weekend; [Holiday Name]"

    Args:
        end_date: Last calendar date to populate. Defaults to the BOT business
            date (bot_today, Asia/Bangkok) when None so the standard ledger
            path keeps its prior behavior; the GUI manual-range path passes the
            user's explicit end date.
        extra_currency_rates: Optional ``{ccy: {date: Decimal}}`` for non-USD/
            EUR currencies. Each gets ONE appended column (the user's selected
            rate type) inserted between EUR Selling and Holidays, with the same
            blank weekend/holiday handling (no rate written on those rows).
            Iterated in dict order, so callers pass an order-stable mapping
            (e.g. built from sorted codes).

    Returns:
        ``{ccy: column_letter}`` for every appended extra currency so the
        caller can wire those columns into the ledger XLOOKUP formula
        (``inject_xlookup_formulas(exrate_col_map=...)``). USD/EUR are NOT in
        the map — they occupy the fixed B-E columns the formula already knows.
    """
    SHEET_NAME = EXRATE_SHEET_NAME
    HEADER_ROW = 1
    DATA_START_ROW = 2
    # Copy — the union below must never mutate the caller's mapping.
    extra_rates = dict(extra_currency_rates or {})
    # Union previously appended extras (read from the live sheet's header
    # row) into the layout: a run that does not re-select a currency must
    # NEVER wipe its column — the standalone standard path passes no extras
    # at all, and before this union it repurposed column F as the Holidays
    # label and cleared everything beyond it, silently destroying GBP/CNY/...
    # values that ledger XLOOKUPs still reference. PRIOR sheet order comes
    # first (already-injected formulas point at those absolute columns),
    # newly selected codes append after.
    prior_codes: list[str] = []
    if SHEET_NAME in wb.sheetnames:
        prior_codes = [
            ccy for ccy, _col in _read_prior_extra_codes(wb[SHEET_NAME])
        ]
    extra_codes = prior_codes + [
        ccy for ccy in extra_rates if ccy not in prior_codes
    ]
    for ccy in extra_codes:
        # No fresh data for a carried-over currency: empty mapping → the
        # existing-sheet fallback in _merge_rate_data preserves its values.
        extra_rates.setdefault(ccy, {})
    HEADERS = [EXRATE_DATE_HEADER]
    HEADERS += [label for _col, label, _ccy, _rt in EXRATE_RATE_COLUMNS]
    HEADERS.extend(f"{ccy} Rate" for ccy in extra_codes)
    HEADERS.append(EXRATE_HOLIDAYS_HEADER)

    # Column index (1-based) of each appended extra currency + the trailing
    # Holidays column (layout constants above: extras start after E).
    extra_col_index = {
        ccy: EXRATE_EXTRA_START_COL + offset
        for offset, ccy in enumerate(extra_codes)
    }
    holidays_col = exrate_holidays_col(len(extra_codes))
    exrate_col_map = {
        ccy: get_column_letter(idx) for ccy, idx in extra_col_index.items()
    }

    # ── Get or create the sheet ──────────────────────────────────────
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)

    # Always write/refresh headers with enterprise styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1A365D", end_color="1A365D", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    for ccy in extra_codes:
        ws.column_dimensions[exrate_col_map[ccy]].width = 16
    ws.column_dimensions[get_column_letter(holidays_col)].width = 40

    # ── Read existing data from the sheet ────────────────────────────
    existing_data = _read_existing_data(ws, DATA_START_ROW)

    # ── Build ALL calendar dates ─────────────────────────────────────
    holidays_set = set(holidays_list)
    if end_date is None:
        end_date = bot_today()
    all_dates = _build_date_range(start_date, end_date, existing_data)

    # ── Build the merged dataset ────────────────────────────────────
    merged = _merge_rate_data(
        all_dates, existing_data, holidays_set, holidays_names,
        usd_buying_rates, usd_selling_rates,
        eur_buying_rates, eur_selling_rates,
        extra_rates,
    )

    # ── Write data ───────────────────────────────────────────────────
    # Clear any legacy/helper columns beyond our current layout
    total_cols = len(HEADERS)
    if ws.max_column and ws.max_column > total_cols:
        for row_idx in range(1, (ws.max_row or 1) + 1):
            for col in range(total_cols + 1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).value = None

    if ws.max_row and ws.max_row >= DATA_START_ROW:
        ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)

    _write_merged_data(
        ws, merged, holidays_set, thin_border, DATA_START_ROW,
        extra_col_index, holidays_col,
    )
    return exrate_col_map


def _read_prior_extra_codes(ws) -> list[tuple[str, int]]:
    """Appended ``<CCY> Rate`` columns on the live sheet, in sheet order.

    Scans header row 1 from ``EXRATE_EXTRA_START_COL`` rightward until the
    Holidays/Weekend column (or anything that is not a ``<CCY> Rate``
    header). These columns must survive a run that does not re-select
    them: :func:`_read_existing_data` round-trips their values and
    :func:`update_master_exrate_sheet` unions them into the layout.
    """
    out: list[tuple[str, int]] = []
    col = EXRATE_EXTRA_START_COL
    max_col = ws.max_column or 0
    while col <= max_col:
        val = ws.cell(row=1, column=col).value
        if not isinstance(val, str):
            break
        text = val.strip()
        if text == EXRATE_HOLIDAYS_HEADER or not text:
            break
        m = re.fullmatch(r"([A-Z]{3}) Rate", text)
        if not m:
            break
        out.append((m.group(1), col))
        col += 1
    return out


def _read_existing_data(ws, data_start_row: int) -> dict[date, dict]:
    """Reads existing rate data from the ExRate sheet.

    Reads the fixed USD/EUR columns (B-E), every appended ``<CCY> Rate``
    column under its ``extra:<CCY>`` key, and the Holidays label from its
    TRUE position (column 6 only when no extras exist). Before the extras
    were read back, _merge_rate_data's existing-sheet fallback could never
    fire for them — any re-run silently wiped previously written
    GBP/CNY/... values.
    """
    prior_extras = _read_prior_extra_codes(ws)
    holidays_col = exrate_holidays_col(len(prior_extras))
    existing: dict[date, dict] = {}
    if ws.max_row and ws.max_row >= data_start_row:
        for row_idx in range(data_start_row, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            row_date = _parse_cell_date(cell_val)
            if row_date:
                entry = {
                    "usd_buy": ws.cell(row=row_idx, column=2).value,
                    "usd_sell": ws.cell(row=row_idx, column=3).value,
                    "eur_buy": ws.cell(row=row_idx, column=4).value,
                    "eur_sell": ws.cell(row=row_idx, column=5).value,
                    "holidays_weekend": ws.cell(
                        row=row_idx, column=holidays_col,
                    ).value,
                }
                for ccy, col_idx in prior_extras:
                    entry[f"extra:{ccy}"] = ws.cell(
                        row=row_idx, column=col_idx,
                    ).value
                existing[row_date] = entry
    return existing


def _parse_cell_date(cell_val) -> date | None:
    """Parse a date from a cell value (shared parser, full format superset)."""
    return _shared_parse_date(cell_val)


def _build_date_range(
    start: date, end: date, existing: dict[date, dict]
) -> set:
    """Build the full set of calendar dates to populate."""
    all_dates = set()
    current = start
    while current <= end:
        all_dates.add(current)
        current += timedelta(days=1)
    all_dates |= set(existing.keys())
    return {d for d in all_dates if d >= start}


def _merge_rate_data(
    all_dates, existing_data, holidays_set, holidays_names,
    usd_buying_rates, usd_selling_rates,
    eur_buying_rates, eur_selling_rates,
    extra_currency_rates: dict[str, dict[date, Decimal]] | None = None,
) -> dict[date, dict]:
    """Merge API rates with existing sheet data (API priority).

    Weekend/holiday rows carry no BOT rate of their own, so their rate cells
    are left BLANK — only the Date and the Holidays/Weekend label are written.
    BOT publishes no rate on those days, so no value is fabricated for them:
    no carry-forward of the prior trading-day rate. The ledger XLOOKUP is
    exact-match and intentionally yields "" for a weekend/holiday-dated
    transaction. This matches the v3.2.8 behavior.

    Existing-value fallback (trading days only): a value read back from the
    sheet is re-validated through safe_to_decimal (string-built Decimal,
    quantized 4dp) before it is re-written — never echoed as a raw openpyxl
    float. Non-numeric sheet residue quantizes to None (cell goes blank).
    Each sheet-sourced date+column is logged at DEBUG.

    Non-trading (weekend/holiday) rows NEVER take the sheet fallback. A
    fresh API/cache rate for that exact date is honored (genuine
    BOT-published weekend rate, API priority as always), but a sheet-only
    value is dropped and logged at INFO — whether it matches the
    carry-forward signature (value equals the nearest prior trading-day's
    value in the same column, as fabricated by v3.4.0/v3.5.0 builds of
    2026-06-04..06-09) or not: per the frozen invariant, weekend/holiday
    rows are Date + label only, so any sheet-sourced rate there has no BOT
    backing. The signature only classifies the log line.

    ``extra_currency_rates`` ({ccy: {date: Decimal}}) is treated identically
    under per-currency keys ``f"extra:{ccy}"`` — a GBP/JPY weekend row also
    stays blank.
    """
    extra_currency_rates = extra_currency_rates or {}
    extra_keys = {ccy: f"extra:{ccy}" for ccy in extra_currency_rates}
    rate_keys = ("usd_buy", "usd_sell", "eur_buy", "eur_sell")
    all_keys = list(rate_keys) + list(extra_keys.values())

    merged: dict[date, dict] = {}
    # Nearest prior trading-day value per column (4dp Decimal) — the value
    # the v3.4.0/v3.5.0 carry-forward would have copied into a weekend or
    # holiday row. Used only to classify cleanup log lines.
    prior_trading: dict[str, Decimal] = {}
    for d in sorted(all_dates):
        existing = existing_data.get(d, {})
        is_weekend = d.weekday() >= 5
        is_holiday = d in holidays_set
        is_trading = not is_weekend and not is_holiday

        # Keep rate values as Decimal end-to-end — NEVER cast to float.
        # float() corrupts 4dp precision (34.5650 -> 34.564999...).
        # openpyxl writes Decimal cells natively.
        row_rates = {
            "usd_buy": usd_buying_rates.get(d),
            "usd_sell": usd_selling_rates.get(d),
            "eur_buy": eur_buying_rates.get(d),
            "eur_sell": eur_selling_rates.get(d),
        }
        for ccy, key in extra_keys.items():
            row_rates[key] = extra_currency_rates[ccy].get(d)

        holiday_label = ""
        if is_weekend and is_holiday:
            holiday_label = f"Weekend; {holidays_names.get(d, 'Holiday')}"
        elif is_weekend:
            holiday_label = "Weekend"
        elif is_holiday:
            holiday_label = holidays_names.get(d, "Holiday")

        entry: dict = {}
        for key in all_keys:
            # API value wins; otherwise fall back to whatever was already on
            # the sheet. NO weekend/holiday carry-forward — those rows keep a
            # blank rate cell (only Date + label survive), matching v3.2.8.
            if row_rates[key] is not None:
                # Genuine BOT value for this exact date — kept even on a
                # weekend/holiday (API priority, defensive).
                entry[key] = row_rates[key]
                continue
            sheet_val = existing.get(key)
            if sheet_val is None or sheet_val == "":
                entry[key] = None
                continue
            if is_trading:
                # Sheet fallback re-validated: string-built Decimal, 4dp.
                # A raw openpyxl float is never echoed back to the cell.
                entry[key] = safe_to_decimal(sheet_val)
                if entry[key] is not None:
                    logger.debug(
                        "ExRate merge: %s %s=%s sourced from existing sheet "
                        "(no BOT value for this date)",
                        d.isoformat(), key, entry[key],
                    )
            else:
                # Weekend/holiday cleanup: sheet-only value has no BOT
                # backing — drop it (cell stays blank, Date + label only).
                entry[key] = None
                dropped = safe_to_decimal(sheet_val)
                if dropped is not None and dropped == prior_trading.get(key):
                    logger.info(
                        "ExRate merge cleanup: dropped carry-forward rate "
                        "%s on %s (%s) — equals nearest prior trading-day "
                        "value (v3.4.0/v3.5.0 fabrication signature)",
                        dropped, d.isoformat(), key,
                    )
                else:
                    logger.info(
                        "ExRate merge cleanup: dropped sheet-only rate %s "
                        "on non-trading day %s (%s) — no BOT rate published",
                        dropped if dropped is not None else sheet_val,
                        d.isoformat(), key,
                    )

        if is_trading:
            for key in all_keys:
                if entry[key] is not None:
                    prior_trading[key] = safe_to_decimal(entry[key])

        entry["holidays_weekend"] = holiday_label
        merged[d] = entry
    return merged


def _write_merged_data(
    ws, merged, holidays_set, thin_border, start_row,
    extra_col_index: dict[str, int] | None = None,
    holidays_col: int = 6,
):
    """Write the merged rate data to the worksheet.

    Args:
        extra_col_index: ``{ccy: 1-based column}`` for appended extra
            currencies (their merged-entry key is ``f"extra:{ccy}"``).
        holidays_col: 1-based column of the trailing Holidays/Weekend label.
    """
    extra_col_index = extra_col_index or {}
    data_font = Font(name="Calibri", size=10)
    date_align = Alignment(horizontal="center")
    num_align = Alignment(horizontal="right")
    holiday_fill = PatternFill(
        start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
    )
    weekend_fill = PatternFill(
        start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"
    )

    current_row = start_row
    for d in sorted(merged.keys()):
        entry = merged[d]
        is_weekend = d.weekday() >= 5
        is_holiday = d in holidays_set

        cell_date = ws.cell(row=current_row, column=1, value=d)
        cell_date.number_format = "DD/MM/YYYY"
        cell_date.font = data_font
        cell_date.alignment = date_align
        cell_date.border = thin_border

        rate_cols = [
            (2, "usd_buy"), (3, "usd_sell"),
            (4, "eur_buy"), (5, "eur_sell"),
        ]
        rate_cols.extend(
            (col, f"extra:{ccy}") for ccy, col in extra_col_index.items()
        )
        for col, key in rate_cols:
            cell = ws.cell(row=current_row, column=col, value=entry[key])
            if entry[key] is not None:
                cell.number_format = "0.0000"
            cell.font = data_font
            cell.alignment = num_align
            cell.border = thin_border

        cell_hw = ws.cell(
            row=current_row, column=holidays_col,
            value=entry["holidays_weekend"],
        )
        cell_hw.font = data_font
        cell_hw.border = thin_border

        if is_holiday:
            for col in range(1, holidays_col + 1):
                ws.cell(row=current_row, column=col).fill = holiday_fill
        elif is_weekend:
            for col in range(1, holidays_col + 1):
                ws.cell(row=current_row, column=col).fill = weekend_fill

        current_row += 1
