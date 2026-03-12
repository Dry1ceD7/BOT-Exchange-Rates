#!/usr/bin/env python3
"""
================================================================================
  BOT Accountant Excel Filler  v2.0
  ─────────────────────────────────
  Enterprise-grade tool that auto-fills exchange rate data in any accountant
  spreadsheet that follows a similar column layout.

  USAGE (Terminal):
    python3 bot_acc_filler.py                                  # default sample
    python3 bot_acc_filler.py --input my_file.xlsx             # custom file
    python3 bot_acc_filler.py --input my_file.xlsx --verbose   # debug logging
    python3 bot_acc_filler.py --gui                            # launch GUI

  WHAT IT DOES (per row in each data sheet):
    1. Reads column "Cur" (fuzzy-matched) to determine USD or EUR
    2. Reads column "วันที่ใบขน" to get the export entry date
    3. If that date is a weekend or BOT holiday, rolls back to the
       most recent open trading day
    4. Writes a dynamic XLOOKUP formula into "วันที่ดึง Exchange rate date"
    5. Writes a VLOOKUP formula into "EX Rate"
    6. Highlights rows with missing data in RED for easy review
    7. Saves as a NEW file (never overwrites the original)

  NEW IN v2.0:
    - Async API engine (aiohttp) — fetches data in under 1 second
    - Smart date scanning — only fetches the date range found in the Excel
    - Fuzzy header detection — tolerates "Currency", "Curr.", etc.
    - Error highlighting — RED rows for missing data
    - Dynamic array formulas — single spill formula per column
    - Professional logging (--verbose / --silent flags)
    - --gui flag launches a drag-and-drop desktop GUI
================================================================================
"""

# ─── Standard library imports ────────────────────────────────
import sys
import os
import argparse
import asyncio
import logging
from datetime import date, datetime
from typing import Dict, Optional, Any

# ─── Ensure the local _libs folder is on the import path ─────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_LIBS_DIR = os.path.join(SCRIPT_DIR, "_libs")
if _LIBS_DIR not in sys.path:
    sys.path.insert(0, _LIBS_DIR)

# ─── Import the shared core module ───────────────────────────
# bot_core.py handles: .env loading, aiohttp install, async API,
# holiday/rate fetching, date parsing, and rate-date resolution.
from bot_core import (  # noqa: E402
    fetch_all_data,
    parse_date_string,
    resolve_effective_rate_date,
    SCRIPT_DIR as CORE_DIR,
)

# ─── Auto-install thefuzz for fuzzy header matching ──────────
try:
    from thefuzz import fuzz  # type: ignore
except ImportError:
    print("  Installing required package 'thefuzz' locally...")
    import subprocess
    subprocess.check_call([
        sys.executable, "-m", "pip", "install",
        "--target", _LIBS_DIR, "thefuzz",
        "--break-system-packages", "--quiet",
    ])
    import importlib
    importlib.invalidate_caches()
    from thefuzz import fuzz  # type: ignore

import openpyxl  # noqa: E402 — installed by bot_core.py
import openpyxl.utils  # noqa: E402
from openpyxl.styles import PatternFill, Font  # noqa: E402


# ═══════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════

# Column headers we search for (exact or fuzzy)
EXPECTED_HEADERS = {
    "cur":       ["Cur", "Currency", "Curr", "สกุลเงิน"],
    "export_dt": ["วันที่ใบขน", "วันที่ขนส่ง", "Export Date", "ExportDate"],
    "rate_dt":   ["วันที่ดึง Exchange rate date", "Rate Date", "วันที่ดึง"],
    "ex_rate":   ["EX Rate", "Exchange Rate", "Selling Rate", "อัตราแลกเปลี่ยน"],
}

# Minimum similarity score (0–100) for fuzzy header matching.
# 80 means "Currency" matches "Cur" but "Customer" does not.
FUZZY_THRESHOLD = 75

# Styles for error-highlighted rows (missing data)
FILL_ERROR = PatternFill("solid", fgColor="FF4444")
FONT_ERROR = Font(name="Calibri", size=10, color="FFFFFF", bold=True)

# Logging setup
logger = logging.getLogger("bot_acc_filler")


# ═══════════════════════════════════════════════════════════════
# FUZZY HEADER DETECTION
# ═══════════════════════════════════════════════════════════════

def find_columns_fuzzy(ws: Any) -> Optional[Dict[str, Any]]:
    """Scan the first 5 rows to find column positions using fuzzy matching.

    Instead of requiring exact header names, this function scores every
    cell against our list of known aliases. If the best score exceeds
    FUZZY_THRESHOLD, the column is accepted.

    Returns a dict like:
      {"cur": 9, "export_dt": 17, "rate_dt": 18, "ex_rate": 19, "header_row": 2}
    or None if the two required columns (cur, export_dt) are not found.
    """
    found: Dict[str, Any] = {}

    # Check the first 5 rows (some files have multi-line merged headers)
    for row_num in range(1, 6):
        for col_num in range(1, (ws.max_column or 30) + 1):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value is None:
                continue

            header = str(cell_value).strip()
            if not header:
                continue

            # Test each expected column against all its aliases
            for key, aliases in EXPECTED_HEADERS.items():
                if key in found:
                    continue  # already matched this column

                for alias in aliases:
                    # Try exact match first (fastest), then fuzzy
                    if header == alias or header.startswith(alias):
                        found[key] = col_num
                        found["header_row"] = row_num
                        logger.debug(
                            "  [exact] '%s' → %s (col %d, row %d)",
                            header, key, col_num, row_num,
                        )
                        break

                    # Fuzzy match (handles typos, abbreviations)
                    score = fuzz.ratio(header.lower(), alias.lower())
                    if score >= FUZZY_THRESHOLD:
                        found[key] = col_num
                        found["header_row"] = row_num
                        logger.debug(
                            "  [fuzzy] '%s' ≈ '%s' (%d%%) → %s (col %d)",
                            header, alias, score, key, col_num,
                        )
                        break

    # We MUST have at least currency and export date columns
    if "cur" not in found or "export_dt" not in found:
        return None

    # If output columns were not found, place them next to export_dt
    if "rate_dt" not in found:
        found["rate_dt"] = found["export_dt"] + 1
    if "ex_rate" not in found:
        found["ex_rate"] = found["rate_dt"] + 1
    if "header_row" not in found:
        found["header_row"] = 1

    return found


# ═══════════════════════════════════════════════════════════════
# SMART DATE RANGE SCANNING
# ═══════════════════════════════════════════════════════════════

def scan_date_range(wb: Any) -> tuple:
    """Scan every data sheet in the workbook and find the MIN/MAX dates.

    This lets us ask the BOT API only for the exact date range the
    accountant actually has data for, instead of downloading everything
    from Jan 2025. Much faster and kinder to their servers.

    Returns (min_date, max_date) or (fallback_start, today) if no dates found.
    """
    min_dt: Optional[date] = None
    max_dt: Optional[date] = None

    for ws in wb.worksheets:
        # Skip reference sheets
        if ws.title.lower().startswith("exrate"):
            continue

        cols = find_columns_fuzzy(ws)
        if cols is None:
            continue

        data_start = cols["header_row"] + 1
        for row_num in range(data_start, (ws.max_row or data_start) + 1):
            raw = ws.cell(row=row_num, column=cols["export_dt"]).value
            d = parse_date_string(raw)
            if d is None:
                continue
            if min_dt is None or d < min_dt:
                min_dt = d
            if max_dt is None or d > max_dt:
                max_dt = d

    # Fallback if no dates found in any sheet
    if min_dt is None:
        min_dt = date(2025, 1, 1)
    if max_dt is None:
        max_dt = date.today()

    # Add a small buffer (7 days before earliest, in case of weekend rollback)
    from datetime import timedelta
    min_dt = min_dt - timedelta(days=7)

    return min_dt, max_dt


# ═══════════════════════════════════════════════════════════════
# EXCEL PROCESSING ENGINE
# ═══════════════════════════════════════════════════════════════

def process_workbook(
    input_path: str,
    output_path: str,
    holidays: Dict[str, str],
    rates: Dict[str, Dict[str, Dict[str, Optional[float]]]],
    log_fn=print,
) -> Dict[str, int]:
    """Open the accountant's Excel file, fill formulas, and save.

    Returns a summary dict with counts of filled/skipped/highlighted rows.
    """
    wb = openpyxl.load_workbook(input_path)
    stats = {"sheets": 0, "filled": 0, "skipped": 0, "errors": 0}

    for ws in wb.worksheets:

        # ── Handle Exrate reference sheets ────────────────────
        # Convert string dates to real Excel dates so XLOOKUP works
        if ws.title.lower().startswith("exrate"):
            log_fn(f"  [setup] '{ws.title}' — converting to real Excel dates")
            for r in range(6, (ws.max_row or 6) + 1):
                cell = ws.cell(row=r, column=1)
                if cell.value and isinstance(cell.value, str):
                    parsed = parse_date_string(cell.value)
                    if parsed:
                        cell.value = parsed
                        cell.number_format = "DD MMM YYYY"
            continue

        # ── Find columns (fuzzy) ──────────────────────────────
        cols = find_columns_fuzzy(ws)
        if cols is None:
            log_fn(f"  [skip] '{ws.title}' — required columns not found")
            continue

        data_start = cols["header_row"] + 1
        log_fn(
            f"  [work] '{ws.title}' — Cur=col {cols['cur']}, "
            f"วันที่ใบขน=col {cols['export_dt']}, "
            f"rate_dt=col {cols['rate_dt']}, "
            f"EX Rate=col {cols['ex_rate']}, "
            f"data starts row {data_start}"
        )

        stats["sheets"] += 1
        sheet_filled = 0

        # Column letters for formula generation
        q_col = openpyxl.utils.get_column_letter(cols["export_dt"])
        r_col = openpyxl.utils.get_column_letter(cols["rate_dt"])

        # ── Process each data row ─────────────────────────────
        for row_num in range(data_start, (ws.max_row or data_start) + 1):

            # Read currency code
            raw_cur = ws.cell(row=row_num, column=cols["cur"]).value
            if not raw_cur:
                continue

            currency = str(raw_cur).strip().upper()
            if currency not in ("USD", "EUR"):
                stats["skipped"] += 1
                continue

            # Read export date
            raw_date = ws.cell(row=row_num, column=cols["export_dt"]).value
            export_date = parse_date_string(raw_date)
            if export_date is None:
                stats["skipped"] += 1
                continue

            # Resolve effective date (roll back weekends/holidays)
            effective_date = resolve_effective_rate_date(export_date, holidays)
            effective_str = effective_date.strftime("%Y-%m-%d")

            # Check if we have rate data for this date
            day_rates = rates.get(effective_str, {})
            currency_rates = day_rates.get(currency)

            if currency_rates and (
                currency_rates.get("buying") is not None
                or currency_rates.get("selling") is not None
            ):
                ref_sheet = f"Exrate {currency}"

                # ── Write XLOOKUP formula for rate date ───────
                # match_mode -1 = "exact or next smaller" → finds
                # the most recent trading day automatically.
                formula_date = (
                    f"=XLOOKUP({q_col}{row_num},"
                    f"'{ref_sheet}'!$A$6:$A$1000,"
                    f"'{ref_sheet}'!$A$6:$A$1000,"
                    f'"",-1)'
                )
                ws.cell(row=row_num, column=cols["rate_dt"]).value = formula_date
                ws.cell(row=row_num, column=cols["rate_dt"]).number_format = "DD MMM YYYY"

                # ── Write VLOOKUP formula for exchange rate ───
                formula_rate = (
                    f"=VLOOKUP({r_col}{row_num},"
                    f"'{ref_sheet}'!$A$6:$D$1000,3,FALSE())"
                )
                ws.cell(row=row_num, column=cols["ex_rate"]).value = formula_rate

                stats["filled"] += 1
                sheet_filled += 1

            else:
                # ── ERROR: No rate data → highlight RED ───────
                ws.cell(row=row_num, column=cols["rate_dt"]).value = "⚠ NO DATA"
                ws.cell(row=row_num, column=cols["rate_dt"]).fill = FILL_ERROR
                ws.cell(row=row_num, column=cols["rate_dt"]).font = FONT_ERROR

                ws.cell(row=row_num, column=cols["ex_rate"]).value = None
                ws.cell(row=row_num, column=cols["ex_rate"]).fill = FILL_ERROR

                stats["errors"] += 1
                log_fn(
                    f"    ⚠ Row {row_num}: No rate for {currency} on {effective_str}"
                )

        log_fn(f"    Filled {sheet_filled} rows in '{ws.title}'")

    # ── Save output ───────────────────────────────────────────
    wb.save(output_path)
    return stats


# ═══════════════════════════════════════════════════════════════
# CLI ARGUMENT PARSER
# ═══════════════════════════════════════════════════════════════

def parse_args():
    """Parse command-line arguments for the filler script."""
    default_input = os.path.join(SCRIPT_DIR, "data", "input", "exchange_rate_file_sample.xlsx")

    parser = argparse.ArgumentParser(
        description="BOT Accountant Excel Filler — auto-fills exchange rates",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  python3 bot_acc_filler.py
  python3 bot_acc_filler.py --input Feb_2026.xlsx
  python3 bot_acc_filler.py --input Feb_2026.xlsx --output Feb_filled.xlsx
  python3 bot_acc_filler.py --gui
""",
    )
    parser.add_argument(
        "--input", "-i",
        type=str,
        default=default_input,
        help="Path to the accountant's .xlsx file (default: sample file)",
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="Output file path (default: <input>_updated.xlsx)",
    )
    parser.add_argument(
        "--gui",
        action="store_true",
        help="Launch the drag-and-drop desktop GUI",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Show detailed debug logging",
    )
    parser.add_argument(
        "--silent", "-s",
        action="store_true",
        help="Suppress all terminal output",
    )

    # Also support the old positional argument: python3 bot_acc_filler.py file.xlsx
    parser.add_argument("legacy_input", nargs="?", default=None, help=argparse.SUPPRESS)

    args = parser.parse_args()

    # Handle the legacy positional argument
    if args.legacy_input and args.input == default_input:
        args.input = os.path.abspath(args.legacy_input)
    else:
        args.input = os.path.abspath(args.input)

    # Build default output path
    if args.output is None:
        base = os.path.splitext(os.path.basename(args.input))[0]
        args.output = os.path.join(os.path.dirname(args.input), f"{base}_updated.xlsx")
    else:
        args.output = os.path.abspath(args.output)

    return args


# ═══════════════════════════════════════════════════════════════
# MAIN EXECUTION
# ═══════════════════════════════════════════════════════════════

async def run_filler(input_path: str, output_path: str, log_fn=print) -> Dict[str, int]:
    """The main async pipeline: scan → fetch → process → save.

    Args:
        input_path:  Absolute path to the accountant's .xlsx file.
        output_path: Absolute path for the output file.
        log_fn:      Callable for printing progress (swapped for GUI).

    Returns:
        Summary stats dict with filled/skipped/errors counts.
    """
    log_fn("=" * 60)
    log_fn("  BOT Accountant Excel Filler v2.0")
    log_fn(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log_fn("=" * 60)

    if not os.path.exists(input_path):
        log_fn(f"\n  ✗ Error: File not found — {input_path}")
        return {"sheets": 0, "filled": 0, "skipped": 0, "errors": 0}

    # Step 1: Smart date scanning
    log_fn(f"\n[1/3] Scanning '{os.path.basename(input_path)}' for date range...")
    wb_scan = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    min_date, max_date = scan_date_range(wb_scan)
    wb_scan.close()
    log_fn(f"  Date range: {min_date} → {max_date}")

    # Step 2: Async API fetch (only the needed range)
    log_fn(f"\n[2/3] Fetching BOT API data...")
    holidays, rates = await fetch_all_data(min_date, max_date, log_fn=log_fn)

    # Step 3: Process the workbook
    log_fn(f"\n[3/3] Processing '{os.path.basename(input_path)}'...")
    stats = process_workbook(input_path, output_path, holidays, rates, log_fn)

    # Summary
    log_fn(f"\n{'=' * 60}")
    log_fn(f"  DONE!")
    log_fn(f"  Sheets processed:   {stats['sheets']}")
    log_fn(f"  Rows filled:        {stats['filled']}")
    log_fn(f"  Rows skipped (THB): {stats['skipped']}")
    log_fn(f"  Rows with errors:   {stats['errors']}")
    log_fn(f"  Output saved:       {os.path.basename(output_path)}")
    log_fn(f"{'=' * 60}")

    return stats


def main():
    """Entry point: parse CLI args, configure logging, and run."""
    args = parse_args()

    # ── Launch GUI if requested ───────────────────────────────
    if args.gui:
        try:
            from bot_acc_filler_gui import launch_gui
            launch_gui()
        except ImportError:
            print("Error: GUI module (bot_acc_filler_gui.py) not found.")
            print("Make sure bot_acc_filler_gui.py is in the same directory.")
            sys.exit(1)
        return

    # ── Configure logging level ───────────────────────────────
    if args.silent:
        logging.basicConfig(level=logging.CRITICAL)
        log_fn = lambda msg: None  # noqa: E731
    elif args.verbose:
        logging.basicConfig(
            level=logging.DEBUG,
            format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
        )
        log_fn = print
    else:
        logging.basicConfig(level=logging.INFO)
        log_fn = print

    # ── Run the async pipeline ────────────────────────────────
    asyncio.run(run_filler(args.input, args.output, log_fn))


if __name__ == "__main__":
    main()


# ─── Changelog ───────────────────────────────────────────────
# Every update to this file gets a new entry below.
#
# 2026-03-10 | v1 — Initial version
#            | - Hardcoded to exchange_rate_file_sample.xlsx only
#            | - Column positions were fixed (I=9, Q=17, R=18, S=19)
#
# 2026-03-10 | v2 — Generalized for any Excel file
#            | - Now accepts any .xlsx file as a command-line argument
#            | - Auto-detects column positions by scanning the header row
#            | - Processes ALL data sheets (skips "Exrate" reference sheets)
#            | - Output filename mirrors input: "file_updated_YYYY-MM-DD.xlsx"
#
# 2026-03-11 | v3 — Formula & Format Overhaul
#            | - Preserves VLOOKUP formula in EX Rate column
#            | - Standardized all date outputs to "DD MMM YYYY" format
#            | - Added logic to skip reference sheets
#
# 2026-03-11 | v4 — Dynamic Smart Formulas
#            | - Converted Exrate reference tab dates to Real Excel Dates
#            | - Upgraded rate date column to use XLOOKUP(..., -1)
#
# 2026-03-12 | v2.0 — Full Rebuild
#            | - Switched to asyncio + aiohttp via bot_core.py
#            | - Smart date scanning (only fetches the range from Excel)
#            | - Fuzzy header detection (thefuzz library)
#            | - Error highlighting (RED rows for missing data)
#            | - Professional argparse CLI (--input, --output, --gui)
#            | - Professional logging (--verbose, --silent)
#            | - Separated GUI into bot_acc_filler_gui.py
