#!/usr/bin/env python3
"""
tests/test_ledger_processing.py
---------------------------------------------------------------------------
Unit tests for core/ledger_processing.py — near-pure ledger helpers.
Focus: prescan_target_dates duplicate-header determinism (first-wins).
---------------------------------------------------------------------------
"""

from datetime import date
from decimal import Decimal

import openpyxl
import pytest

from core.anomaly_guard import AnomalyGuard
from core.ledger_processing import (
    classify_currencies,
    prescan_target_dates,
    prescan_target_dates_and_currencies,
    run_anomaly_check,
)

TARGET_COLS = {"source_date": "Date", "currency": "Cur", "out_rate": "EX Rate"}


def _write_workbook(tmp_path, rows, header):
    """Build a one-tab workbook with the given header + data rows."""
    filepath = tmp_path / "ledger.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jan"
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(str(filepath))
    wb.close()
    return str(filepath)


class TestPrescanTargetDates:

    def test_scans_source_date_column(self, tmp_path):
        path = _write_workbook(
            tmp_path,
            rows=[
                [date(2025, 1, 7), "USD", None],
                [date(2025, 1, 8), "EUR", None],
            ],
            header=["Date", "Cur", "EX Rate"],
        )
        dates = prescan_target_dates(path, TARGET_COLS)
        assert dates == {date(2025, 1, 7), date(2025, 1, 8)}

    def test_duplicate_date_header_uses_first_column(self, tmp_path, caplog):
        """A duplicate 'Date' RIGHT of 'EX Rate' never wins.

        Column A holds the real dates; the duplicate column D (right of
        EX Rate) holds DIFFERENT dates. Resolution considers only
        occurrences LEFT of 'EX Rate', so column A's dates are returned
        and the collision is logged.
        """
        path = _write_workbook(
            tmp_path,
            rows=[
                # A=real date, B=Cur, C=EX Rate, D=duplicate "Date" column.
                [date(2025, 1, 7), "USD", None, date(2030, 12, 31)],
                [date(2025, 1, 8), "EUR", None, date(2031, 12, 31)],
            ],
            header=["Date", "Cur", "EX Rate", "Date"],
        )
        with caplog.at_level("WARNING"):
            dates = prescan_target_dates(path, TARGET_COLS)
        # Only the first "Date" column (A) is scanned — never the duplicate (D).
        assert dates == {date(2025, 1, 7), date(2025, 1, 8)}
        assert date(2030, 12, 31) not in dates
        assert date(2031, 12, 31) not in dates
        # The collision is logged.
        assert any(
            "duplicate" in r.message.lower() and "Date" in r.message
            for r in caplog.records
        )

    def test_duplicate_date_header_uses_export_entry_column(self, tmp_path):
        """Two 'Date' columns LEFT of 'EX Rate' → the nearer one wins.

        Mirrors the real production ledgers: invoice Date in column A,
        export-entry Date immediately left of EX Rate. The written
        formulas resolve rates by the export-entry date, so the fetch
        window must come from THAT column — first-occurrence resolution
        would fetch a window for the invoice dates instead.
        """
        path = _write_workbook(
            tmp_path,
            rows=[
                # A=invoice date, B=Cur, C=export-entry date, D=EX Rate.
                [date(2025, 1, 7), "USD", date(2025, 2, 14), None],
                [date(2025, 1, 8), "EUR", date(2025, 2, 17), None],
            ],
            header=["Date", "Cur", "Date", "EX Rate"],
        )
        dates = prescan_target_dates(path, TARGET_COLS)
        assert dates == {date(2025, 2, 14), date(2025, 2, 17)}


class TestPrescanCurrencies:
    """Currency collection powering the multi-currency ledger path."""

    def test_collects_distinct_currency_codes(self, tmp_path):
        path = _write_workbook(
            tmp_path,
            rows=[
                [date(2025, 1, 7), "USD", None],
                [date(2025, 1, 8), "gbp", None],   # lower-case → normalized
                [date(2025, 1, 9), " EUR ", None],  # whitespace → trimmed
                [date(2025, 1, 10), "GBP", None],   # dup → collapsed
            ],
            header=["Date", "Cur", "EX Rate"],
        )
        dates, currencies = prescan_target_dates_and_currencies(
            path, TARGET_COLS,
        )
        assert dates == {
            date(2025, 1, 7), date(2025, 1, 8),
            date(2025, 1, 9), date(2025, 1, 10),
        }
        assert currencies == {"USD", "GBP", "EUR"}

    def test_dates_only_wrapper_matches(self, tmp_path):
        """prescan_target_dates returns the same dates as the combined scan."""
        path = _write_workbook(
            tmp_path,
            rows=[[date(2025, 1, 7), "USD", None]],
            header=["Date", "Cur", "EX Rate"],
        )
        dates_only = prescan_target_dates(path, TARGET_COLS)
        dates_combined, _ = prescan_target_dates_and_currencies(
            path, TARGET_COLS,
        )
        assert dates_only == dates_combined == {date(2025, 1, 7)}

    def test_skip_sheet_currencies_ignored(self, tmp_path):
        """Currencies on a SKIP sheet (ExRate) must not be collected."""
        filepath = tmp_path / "with_exrate.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jan"
        ws.append(["Date", "Cur", "EX Rate"])
        ws.append([date(2025, 6, 2), "GBP", None])
        ws_ex = wb.create_sheet("ExRate")
        ws_ex.append(["Date", "Cur"])
        ws_ex.append([date(1999, 1, 1), "XXX"])
        wb.save(str(filepath))
        wb.close()
        _dates, currencies = prescan_target_dates_and_currencies(
            str(filepath), TARGET_COLS,
        )
        assert currencies == {"GBP"}
        assert "XXX" not in currencies


class TestClassifyCurrencies:
    """classify_currencies splits scanned codes into extra vs unsupported."""

    def test_usd_eur_thb_are_dropped(self):
        extra, unsupported = classify_currencies({"USD", "EUR", "THB"})
        # All handled by the core IFS branches — nothing to fetch / warn about.
        assert extra == []
        assert unsupported == []

    def test_supported_extra_currencies_sorted(self):
        extra, unsupported = classify_currencies({"SGD", "GBP", "USD"})
        # Sorted for deterministic ExRate column ordering.
        assert extra == ["GBP", "SGD"]
        assert unsupported == []

    def test_jpy_routed_to_unsupported_path(self):
        # JPY is quoted by BOT per 100 yen, so it is deliberately excluded
        # from LEDGER_SUPPORTED_CURRENCIES and must take the unsupported path.
        extra, unsupported = classify_currencies({"JPY", "GBP"})
        assert extra == ["GBP"]
        assert unsupported == ["JPY"]

    def test_unsupported_currency_flagged(self):
        extra, unsupported = classify_currencies({"GBP", "XYZ", "ABC"})
        assert extra == ["GBP"]
        assert unsupported == ["ABC", "XYZ"]


class TestRunAnomalyCheckExtraCurrencies:
    """F42: extra (non-USD/EUR) series go through the REAL AnomalyGuard;
    F25: flagged (currency, date) pairs are reported via anomalous_out.
    Alert-only throughout — the function can only emit and report."""

    # Mon → Tue: 1-day gap, ~19% jump (>5% threshold).
    PREV_DAY = date(2025, 1, 6)
    JUMP_DAY = date(2025, 1, 7)

    def test_real_guard_flags_extra_currency_jump(self):
        emitted = []
        flagged: set[tuple[str, date]] = set()

        count = run_anomaly_check(
            AnomalyGuard(threshold_pct=5.0),
            lambda msg, etype: emitted.append((msg, etype)),
            {}, {}, {}, {},
            extra_currency_rates={
                "GBP": {
                    self.PREV_DAY: Decimal("42.0000"),
                    self.JUMP_DAY: Decimal("50.0000"),
                },
            },
            extra_rate_type="buying_transfer",
            anomalous_out=flagged,
        )

        assert count == 1
        assert flagged == {("GBP", self.JUMP_DAY)}
        assert len(emitted) == 1
        msg, etype = emitted[0]
        assert etype == "warning"
        assert "ANOMALY" in msg
        assert "GBP" in msg

    def test_extra_currency_within_threshold_not_flagged(self):
        emitted = []
        flagged: set[tuple[str, date]] = set()

        count = run_anomaly_check(
            AnomalyGuard(threshold_pct=5.0),
            lambda msg, etype: emitted.append((msg, etype)),
            {}, {}, {}, {},
            extra_currency_rates={
                "GBP": {
                    self.PREV_DAY: Decimal("42.0000"),
                    self.JUMP_DAY: Decimal("42.5000"),
                },
            },
            anomalous_out=flagged,
        )

        assert count == 0
        assert flagged == set()
        assert emitted == []

    def test_defaults_keep_legacy_four_series_contract(self):
        """No extra args → the bundle stays exactly the four fixed series
        (backward compatibility for existing callers)."""
        seen = {}

        class _Guard:
            def check_rates_bulk(self, bundle):
                seen["bundle"] = bundle
                return []

        count = run_anomaly_check(
            _Guard(), lambda msg, etype: None, {}, {}, {}, {},
        )
        assert count == 0
        assert set(seen["bundle"]) == {
            "USD_buying_transfer", "USD_selling",
            "EUR_buying_transfer", "EUR_selling",
        }


class TestPrescanMemoization:
    """Round 11: the Smart-Date pass and process_ledger share ONE read-only
    workbook open per unchanged file via an opt-in memo keyed on
    (abspath, st_mtime_ns, st_size) + header labels."""

    @pytest.fixture(autouse=True)
    def _clean_memo(self):
        import core.ledger_processing as lp
        with lp._PRESCAN_CACHE_LOCK:
            lp._PRESCAN_CACHE.clear()
        yield
        with lp._PRESCAN_CACHE_LOCK:
            lp._PRESCAN_CACHE.clear()

    def _count_opens(self, monkeypatch):
        opens = {"n": 0}
        real_load = openpyxl.load_workbook

        def _counting(*a, **k):
            opens["n"] += 1
            return real_load(*a, **k)

        monkeypatch.setattr(openpyxl, "load_workbook", _counting)
        return opens

    def test_second_scan_of_unchanged_file_uses_memo(
        self, tmp_path, monkeypatch,
    ):
        path = _write_workbook(
            tmp_path,
            [(date(2025, 1, 7), "USD", None), (date(2025, 1, 8), "EUR", None)],
            ["Date", "Cur", "EX Rate"],
        )
        opens = self._count_opens(monkeypatch)

        first = prescan_target_dates_and_currencies(
            path, TARGET_COLS, use_cache=True,
        )
        second = prescan_target_dates_and_currencies(
            path, TARGET_COLS, use_cache=True,
        )
        assert opens["n"] == 1  # one physical open, second served from memo
        assert first == second
        assert first[0] == {date(2025, 1, 7), date(2025, 1, 8)}
        assert first[1] == {"USD", "EUR"}

    def test_memo_returns_defensive_copies(self, tmp_path, monkeypatch):
        path = _write_workbook(
            tmp_path, [(date(2025, 1, 7), "USD", None)],
            ["Date", "Cur", "EX Rate"],
        )
        a = prescan_target_dates_and_currencies(
            path, TARGET_COLS, use_cache=True,
        )
        a[0].add(date(1999, 1, 1))  # mutate the returned set
        b = prescan_target_dates_and_currencies(
            path, TARGET_COLS, use_cache=True,
        )
        assert date(1999, 1, 1) not in b[0]  # memo not poisoned

    def test_modified_file_is_rescanned(self, tmp_path, monkeypatch):
        path = _write_workbook(
            tmp_path, [(date(2025, 1, 7), "USD", None)],
            ["Date", "Cur", "EX Rate"],
        )
        opens = self._count_opens(monkeypatch)
        prescan_target_dates_and_currencies(path, TARGET_COLS, use_cache=True)

        # Modify the file (new row → new mtime_ns + size → new identity).
        wb = openpyxl.load_workbook(path)
        wb["Jan"].append([date(2025, 1, 9), "GBP", None])
        wb.save(path)
        wb.close()
        opens["n"] = 0  # ignore the edit's own open

        dates, ccys = prescan_target_dates_and_currencies(
            path, TARGET_COLS, use_cache=True,
        )
        assert opens["n"] == 1  # memo missed → real rescan
        assert date(2025, 1, 9) in dates
        assert "GBP" in ccys

    def test_no_memo_without_opt_in(self, tmp_path, monkeypatch):
        path = _write_workbook(
            tmp_path, [(date(2025, 1, 7), "USD", None)],
            ["Date", "Cur", "EX Rate"],
        )
        opens = self._count_opens(monkeypatch)
        prescan_target_dates_and_currencies(path, TARGET_COLS)
        prescan_target_dates_and_currencies(path, TARGET_COLS)
        assert opens["n"] == 2  # default path is untouched (opt-in only)

    def test_smart_date_prescan_feeds_the_ledger_scan(
        self, tmp_path, monkeypatch,
    ):
        """The headline dedup: prescan_oldest_date scans the file once and
        the subsequent process_ledger-style scan reuses that exact result —
        eliminating the provably duplicate read-only open per file."""
        from core.prescan import prescan_oldest_date

        path = _write_workbook(
            tmp_path,
            [(date(2025, 1, 7), "USD", None), (date(2025, 1, 3), "EUR", None)],
            ["Date", "Cur", "EX Rate"],
        )
        opens = self._count_opens(monkeypatch)

        oldest, detected = prescan_oldest_date([path])
        assert (oldest, detected) == (date(2025, 1, 3), True)

        # process_ledger's scan (same labels, use_cache=True) hits the memo.
        dates, ccys = prescan_target_dates_and_currencies(
            path, TARGET_COLS, use_cache=True,
        )
        assert opens["n"] == 1
        assert dates == {date(2025, 1, 7), date(2025, 1, 3)}
        assert ccys == {"USD", "EUR"}
