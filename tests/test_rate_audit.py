#!/usr/bin/env python3
"""Unit tests for core/rate_audit.py — the pure ExRate-vs-BOT comparison.

Builds a tiny in-memory ExRate sheet (no disk, no network) and asserts the
scanner's financial invariants, above all: weekend/holiday rows are NEVER
touched (blank by design), and only trading-day cells that differ from BOT are
corrected.
"""
from datetime import date
from decimal import Decimal

import openpyxl

from core.rate_audit import (
    EXRATE_RATE_COLUMNS,
    LAYOUT_ERROR_MSG,
    apply_corrections,
    rate_key,
    scan_exrate_corrections,
    validate_exrate_layout,
)

HEADERS = [
    "Date", "USD Buying TT Rate", "USD Selling Rate",
    "EUR Buying TT Rate", "EUR Selling Rate", "Holidays/Weekend",
]


def _sheet(rows, headers=None):
    """Build an ExRate worksheet. rows = list of (date, ub, us, eb, es, label)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExRate"
    ws.append(headers if headers is not None else HEADERS)
    for r in rows:
        ws.append(list(r))
    return wb, ws


def _bot(**by_key):
    """Build a bot_rates map from key=date->Decimal dicts (key like USD_buy)."""
    alias = {
        "USD_buy": rate_key("USD", "buying_transfer"),
        "USD_sell": rate_key("USD", "selling"),
        "EUR_buy": rate_key("EUR", "buying_transfer"),
        "EUR_sell": rate_key("EUR", "selling"),
    }
    return {alias[k]: v for k, v in by_key.items()}


D = Decimal


class TestScanCorrections:
    def test_correct_value_yields_no_change(self):
        d = date(2026, 5, 27)  # Wednesday — trading day
        wb, ws = _sheet([(d, D("32.4507"), D("32.7790"),
                          D("37.0"), D("37.5"), "")])
        bot = _bot(
            USD_buy={d: D("32.4507")}, USD_sell={d: D("32.7790")},
            EUR_buy={d: D("37.0000")}, EUR_sell={d: D("37.5000")},
        )
        report = scan_exrate_corrections(ws, bot, set())
        assert report.change_count == 0
        assert report.scanned_rows == 1
        assert report.compared_cells == 4
        wb.close()

    def test_wrong_value_is_flagged_with_bot_value(self):
        d = date(2026, 5, 27)
        wb, ws = _sheet([(d, D("32.0000"), D("32.7790"),
                          D("37.0"), D("37.5"), "")])
        bot = _bot(
            USD_buy={d: D("32.4507")}, USD_sell={d: D("32.7790")},
            EUR_buy={d: D("37.0000")}, EUR_sell={d: D("37.5000")},
        )
        report = scan_exrate_corrections(ws, bot, set())
        assert report.change_count == 1
        ch = report.changes[0]
        assert ch.currency == "USD" and ch.rate_type == "buying_transfer"
        assert ch.old_value == D("32.0000")
        assert ch.new_value == D("32.4507")
        assert ch.cell == "B2"
        wb.close()

    def test_blank_trading_day_cell_is_filled(self):
        d = date(2026, 5, 27)  # trading day
        wb, ws = _sheet([(d, None, D("32.7790"), D("37.0"), D("37.5"), "")])
        bot = _bot(USD_buy={d: D("32.4507")})
        report = scan_exrate_corrections(ws, bot, set())
        # USD buy is filled from BOT; the others have no BOT value (skipped).
        usd_buy = [c for c in report.changes if c.col == 2]
        assert len(usd_buy) == 1
        assert usd_buy[0].old_value is None
        assert usd_buy[0].new_value == D("32.4507")
        assert "missing" in usd_buy[0].reason
        wb.close()

    def test_weekend_row_is_never_touched(self):
        sat = date(2026, 5, 23)  # Saturday
        # Cell holds a WRONG value and BOT happens to have data — must NOT fix.
        wb, ws = _sheet([(sat, D("99.9999"), None, None, None, "Weekend")])
        bot = _bot(USD_buy={sat: D("32.4507")})
        report = scan_exrate_corrections(ws, bot, set())
        assert report.change_count == 0
        assert report.compared_cells == 0  # weekend skipped before comparison
        wb.close()

    def test_holiday_row_is_never_touched(self):
        mon = date(2026, 5, 4)  # treat as a BOT holiday
        wb, ws = _sheet([(mon, D("99.9999"), None, None, None, "Coronation Day")])
        bot = _bot(USD_buy={mon: D("32.1000")})
        report = scan_exrate_corrections(ws, bot, {mon})
        assert report.change_count == 0
        wb.close()

    def test_value_with_no_bot_data_is_unverifiable_not_changed(self):
        d = date(2026, 5, 27)
        wb, ws = _sheet([(d, D("32.4507"), None, None, None, "")])
        report = scan_exrate_corrections(ws, _bot(), set())  # empty BOT map
        assert report.change_count == 0
        assert report.unverifiable == 1  # the USD buy cell could not be verified
        wb.close()

    def test_float_noise_does_not_false_positive(self):
        # A legacy float cell that rounds to BOT's 4dp value must NOT be flagged.
        d = date(2026, 5, 27)
        wb, ws = _sheet([(d, 32.50009999, None, None, None, "")])
        bot = _bot(USD_buy={d: D("32.5001")})
        report = scan_exrate_corrections(ws, bot, set())
        assert report.change_count == 0
        wb.close()

    def test_bot_value_over_4dp_does_not_false_flag_correct_cell(self):
        # A >4dp BOT value must be quantized to 4dp before comparing, so an
        # already-correct 4dp cell is NOT spuriously flagged by extra digits.
        d = date(2026, 5, 27)
        wb, ws = _sheet([(d, D("32.4507"), None, None, None, "")])
        bot = _bot(USD_buy={d: D("32.45071")})  # 5dp → 32.4507 at 4dp
        report = scan_exrate_corrections(ws, bot, set())
        assert report.change_count == 0
        wb.close()

    def test_correction_value_is_always_4dp(self):
        # When a correction IS written, the new value is quantized to exactly
        # 4dp — a >4dp BOT/cache value must never reach the rate cell.
        d = date(2026, 5, 27)
        wb, ws = _sheet([(d, D("99.0000"), None, None, None, "")])
        bot = _bot(USD_buy={d: D("32.45079")})  # 5dp → 32.4508 at 4dp
        report = scan_exrate_corrections(ws, bot, set())
        assert report.change_count == 1
        nv = report.changes[0].new_value
        assert nv == D("32.4508")
        assert nv.as_tuple().exponent == -4  # exactly 4dp, never 5
        wb.close()

    def test_unparseable_nonempty_cell_is_not_overwritten(self):
        # A formula/garbage string in a rate cell is unverifiable — NOT a blank
        # to fill. The auditor must leave it untouched (no silent overwrite).
        d = date(2026, 5, 27)
        wb, ws = _sheet([(d, "=B1*2", None, None, None, "")])
        bot = _bot(USD_buy={d: D("32.4507")})
        report = scan_exrate_corrections(ws, bot, set())
        assert report.change_count == 0
        assert report.unverifiable >= 1
        wb.close()

    def test_multiple_currencies_and_types(self):
        d = date(2026, 5, 27)
        wb, ws = _sheet([(d, D("1.0"), D("2.0"), D("3.0"), D("4.0"), "")])
        bot = _bot(
            USD_buy={d: D("32.4507")}, USD_sell={d: D("32.7790")},
            EUR_buy={d: D("37.0000")}, EUR_sell={d: D("37.5000")},
        )
        report = scan_exrate_corrections(ws, bot, set())
        assert report.change_count == 4
        by_col = {c.col: c.new_value for c in report.changes}
        assert by_col == {
            2: D("32.4507"), 3: D("32.7790"),
            4: D("37.0000"), 5: D("37.5000"),
        }
        wb.close()


class TestLayoutGuard:
    """F9 — a custom-layout 'ExRate' sheet (e.g. GBP columns) must be refused;
    auditing it would overwrite foreign-currency cells with USD/EUR values."""

    def test_custom_layout_sheet_yields_zero_corrections(self):
        d = date(2026, 5, 27)  # trading day
        gbp_headers = [
            "Date", "GBP Buying TT", "GBP Selling", "Holidays/Weekend",
        ]
        wb, ws = _sheet(
            [(d, D("42.1234"), D("42.5678"), "")], headers=gbp_headers,
        )
        bot = _bot(USD_buy={d: D("32.4507")}, USD_sell={d: D("32.7790")})
        report = scan_exrate_corrections(ws, bot, set())
        assert report.layout_error == LAYOUT_ERROR_MSG
        assert report.change_count == 0
        assert report.scanned_rows == 0
        assert report.compared_cells == 0
        # The GBP cells survive untouched even if apply is (wrongly) called.
        apply_corrections(ws, report)
        assert ws.cell(row=2, column=2).value == D("42.1234")
        assert ws.cell(row=2, column=3).value == D("42.5678")
        wb.close()

    def test_extra_currency_column_beyond_e_is_valid(self):
        # Standard A-E layout + an appended "GBP Rate" column F: still the
        # standard sheet — the audit runs on B-E and never touches F.
        d = date(2026, 5, 27)
        headers = HEADERS[:5] + ["GBP Rate", "Holidays/Weekend"]
        wb, ws = _sheet(
            [(d, D("32.0000"), D("32.7790"), D("37.0"), D("37.5"),
              D("42.1234"), "")],
            headers=headers,
        )
        bot = _bot(
            USD_buy={d: D("32.4507")}, USD_sell={d: D("32.7790")},
            EUR_buy={d: D("37.0000")}, EUR_sell={d: D("37.5000")},
        )
        report = scan_exrate_corrections(ws, bot, set())
        assert report.layout_error is None
        assert report.change_count == 1
        assert report.changes[0].cell == "B2"
        assert all(c.col <= 5 for c in report.changes)  # F is out of scope
        wb.close()

    def test_header_check_tolerates_whitespace_and_case(self):
        d = date(2026, 5, 27)
        headers = [
            "  date ", " USD BUYING TT RATE", "usd selling rate ",
            " EUR Buying TT Rate", "eur SELLING rate", "Holidays/Weekend",
        ]
        wb, ws = _sheet(
            [(d, D("32.0000"), None, None, None, "")], headers=headers,
        )
        bot = _bot(USD_buy={d: D("32.4507")})
        report = scan_exrate_corrections(ws, bot, set())
        assert report.layout_error is None
        assert report.change_count == 1
        assert report.changes[0].new_value == D("32.4507")
        wb.close()

    def test_missing_or_blank_header_row_is_rejected(self):
        # No header row at all (e.g. data starts at row 1) is non-standard.
        d = date(2026, 5, 27)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws.append([d, D("32.0000"), None, None, None, ""])
        assert validate_exrate_layout(ws) == LAYOUT_ERROR_MSG
        report = scan_exrate_corrections(
            ws, _bot(USD_buy={d: D("32.4507")}), set(),
        )
        assert report.layout_error == LAYOUT_ERROR_MSG
        assert report.change_count == 0
        wb.close()


class TestApplyCorrections:
    def test_apply_writes_new_values_and_format(self):
        d = date(2026, 5, 27)
        wb, ws = _sheet([(d, D("32.0000"), None, None, None, "")])
        bot = _bot(USD_buy={d: D("32.4507")})
        report = scan_exrate_corrections(ws, bot, set())
        assert report.change_count == 1
        assert report.applied is False

        apply_corrections(ws, report)
        assert report.applied is True
        cell = ws.cell(row=2, column=2)
        assert cell.value == D("32.4507")
        assert isinstance(cell.value, Decimal)
        assert cell.number_format == "0.0000"
        wb.close()

    def test_apply_is_noop_when_no_changes(self):
        d = date(2026, 5, 27)
        wb, ws = _sheet([(d, D("32.4507"), None, None, None, "")])
        bot = _bot(USD_buy={d: D("32.4507")})
        report = scan_exrate_corrections(ws, bot, set())
        apply_corrections(ws, report)
        assert ws.cell(row=2, column=2).value == D("32.4507")
        wb.close()


def test_columns_constant_matches_exrate_sheet_layout():
    # Guard against drift from core/exrate_sheet.py's fixed B-E layout.
    cols = [c[0] for c in EXRATE_RATE_COLUMNS]
    assert cols == [2, 3, 4, 5]
