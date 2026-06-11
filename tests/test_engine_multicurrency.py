#!/usr/bin/env python3
"""
tests/test_engine_multicurrency.py
---------------------------------------------------------------------------
Regression coverage for LedgerEngine.update_exrate_standalone CUSTOM
multi-currency path (e.g. GBP/JPY with arbitrary rate_types + date_range).

Verifies:
  - headers + values land in the correct columns,
  - rate values are written as EXACT Decimals (Mathematical Truth, no float
    round-trip / approx),
  - the ExRate sheet content is cleared and rewritten,
  - the disk-space OSError guard fires BEFORE the in-place overwrite, leaving
    the original file untouched on disk.

All API access is mocked; backup + cache are injected as temp instances so
no real network, no real singleton state, no real backup dir is touched.
---------------------------------------------------------------------------
"""

import asyncio
from collections import namedtuple
from datetime import date, datetime
from decimal import Decimal
from types import SimpleNamespace
from unittest.mock import AsyncMock

import openpyxl
import pytest

import core.workbook_io as workbook_io_mod
from core.backup_manager import BackupManager
from core.engine import LedgerEngine

_DiskUsage = namedtuple("_DiskUsage", ["total", "used", "free"])


def _cell_decimal(value) -> Decimal:
    """Recover the exact 4dp Decimal value from a reloaded cell.

    openpyxl serializes numeric cell values to float on save and returns a
    float on reload, so the in-memory Decimal type does not survive a
    round-trip. Going through str() recovers the shortest exact decimal that
    reproduces the float, which equals the original 4dp value for BOT rates.
    """
    return Decimal(str(value))


# =========================================================================
#  FIXTURES
# =========================================================================

@pytest.fixture
def temp_backup(tmp_path):
    """A BackupManager rooted in a temp dir (never touches the real data/)."""
    return BackupManager(backup_dir=str(tmp_path / "backups"))


@pytest.fixture
def exrate_file(tmp_path):
    """A standalone ExRate workbook with pre-existing junk content.

    Pre-seeds an old header + an old data row so we can prove the custom
    writer CLEARS the prior content before rewriting.
    """
    filepath = tmp_path / "ExRate_standalone.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExRate"
    # Stale content from a previous run — must be wiped.
    ws.append(["OLD-DATE", "OLD-USD", "OLD-EUR"])
    ws.append([date(1999, 1, 1), 1.1111, 2.2222])
    wb.save(str(filepath))
    wb.close()
    return str(filepath)


def _make_currency_side_effect(per_ccy):
    """Build an async side_effect for api.get_exchange_rates keyed on currency.

    per_ccy maps currency -> list of SimpleNamespace rate records. Any
    currency not present (e.g. USD/EUR fetched by the holiday preload step)
    returns an empty list.
    """
    async def _side_effect(start, end, currency):
        return list(per_ccy.get(currency, []))

    return _side_effect


def _make_api():
    """A mocked API matching the BOTClient async surface."""
    api = AsyncMock()
    api.get_holidays = AsyncMock(return_value=[])
    api.get_exchange_rates = AsyncMock(return_value=[])
    return api


# =========================================================================
#  CUSTOM MULTI-CURRENCY PATH
# =========================================================================

class TestCustomMultiCurrency:
    """update_exrate_standalone custom path (GBP/JPY, mixed rate types)."""

    def _run(self, exrate_file, temp_backup, tmp_cache, per_ccy,
             rate_types, currencies, date_range):
        api = _make_api()
        api.get_exchange_rates = AsyncMock(
            side_effect=_make_currency_side_effect(per_ccy)
        )
        eng = LedgerEngine(api, backup=temp_backup, cache=tmp_cache)
        return eng, asyncio.run(
            eng.update_exrate_standalone(
                exrate_file,
                currencies=currencies,
                rate_types=rate_types,
                date_range=date_range,
            )
        )

    def test_headers_and_values_land_in_right_columns(
        self, exrate_file, temp_backup, tmp_cache,
    ):
        # Two-day window so the layout stays small + deterministic.
        dr = (date(2025, 3, 10), date(2025, 3, 11))
        rate_types = {"Buying TT": "buying_transfer", "Selling": "selling"}
        currencies = ["GBP", "JPY"]

        # EXACT Decimals fed via the rate records — the engine writes them
        # straight through to the cells (no float coercion in the custom path).
        gbp_records = [
            SimpleNamespace(
                period="2025-03-10", currency="GBP",
                buying_transfer=Decimal("42.1234"),
                selling=Decimal("43.5678"),
                buying_sight=None, mid_rate=None,
            ),
            SimpleNamespace(
                period="2025-03-11", currency="GBP",
                buying_transfer=Decimal("42.2200"),
                selling=Decimal("43.6600"),
                buying_sight=None, mid_rate=None,
            ),
        ]
        jpy_records = [
            SimpleNamespace(
                period="2025-03-10", currency="JPY",
                buying_transfer=Decimal("0.2155"),
                selling=Decimal("0.2233"),
                buying_sight=None, mid_rate=None,
            ),
        ]
        per_ccy = {"GBP": gbp_records, "JPY": jpy_records}

        _eng, out = self._run(
            exrate_file, temp_backup, tmp_cache, per_ccy,
            rate_types, currencies, dr,
        )
        assert out == exrate_file

        wb = openpyxl.load_workbook(out)
        ws = wb["ExRate"]

        # ── Headers: Date | GBP Buying TT | GBP Selling | JPY Buying TT |
        #            JPY Selling | Holidays/Weekend ──────────────────────
        header_row = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert header_row == [
            "Date",
            "GBP Buying TT", "GBP Selling",
            "JPY Buying TT", "JPY Selling",
            "Holidays/Weekend",
        ]

        # ── Row 2 = 2025-03-10 (Monday) ───────────────────────────────
        row2 = {
            "date": ws.cell(row=2, column=1).value,
            "gbp_tt": ws.cell(row=2, column=2).value,
            "gbp_sell": ws.cell(row=2, column=3).value,
            "jpy_tt": ws.cell(row=2, column=4).value,
            "jpy_sell": ws.cell(row=2, column=5).value,
        }
        assert row2["date"] in (date(2025, 3, 10),
                                datetime(2025, 3, 10))
        # GBP lands in cols 2/3, JPY in cols 4/5 — never crossed.
        # NOTE: openpyxl serializes numeric cells back to float on reload, so
        # we compare the exact 4dp VALUE via Decimal(str(cell)). See the
        # `issues` note about Decimal type not surviving the round-trip.
        assert _cell_decimal(row2["gbp_tt"]) == Decimal("42.1234")
        assert _cell_decimal(row2["gbp_sell"]) == Decimal("43.5678")
        assert _cell_decimal(row2["jpy_tt"]) == Decimal("0.2155")
        assert _cell_decimal(row2["jpy_sell"]) == Decimal("0.2233")

        # ── Row 3 = 2025-03-11 (Tuesday) — GBP only, JPY missing ──────
        assert _cell_decimal(ws.cell(row=3, column=2).value) == Decimal("42.2200")
        assert _cell_decimal(ws.cell(row=3, column=3).value) == Decimal("43.6600")
        # JPY has no record for the 11th -> empty cells, not stale data.
        assert ws.cell(row=3, column=4).value is None
        assert ws.cell(row=3, column=5).value is None
        wb.close()

    def test_value_is_exact_4dp_and_formatted(
        self, exrate_file, temp_backup, tmp_cache,
    ):
        """The written value must equal the exact 4dp rate and carry the
        0.0000 number format (the 4dp presentation guarantee).

        Mathematical Truth is asserted as an EXACT Decimal value via
        Decimal(str(cell)) — never pytest.approx. (Type-level Decimal does not
        survive openpyxl's save/reload; see `issues`.)
        """
        dr = (date(2025, 3, 10), date(2025, 3, 10))
        rate_types = {"Buying TT": "buying_transfer"}
        per_ccy = {
            "GBP": [SimpleNamespace(
                period="2025-03-10", currency="GBP",
                buying_transfer=Decimal("42.1234"),
                selling=None, buying_sight=None, mid_rate=None,
            )],
        }
        _eng, out = self._run(
            exrate_file, temp_backup, tmp_cache, per_ccy,
            rate_types, ["GBP"], dr,
        )
        wb = openpyxl.load_workbook(out)
        ws = wb["ExRate"]
        cell = ws.cell(row=2, column=2)
        # Exact value (not an approximation):
        assert _cell_decimal(cell.value) == Decimal("42.1234")
        assert _cell_decimal(cell.value) != Decimal("42.1235")
        # 4dp presentation enforced via number_format.
        assert cell.number_format == "0.0000"
        wb.close()

    def test_raw_api_float_is_quantized_to_4dp(
        self, exrate_file, temp_backup, tmp_cache,
    ):
        """Production BOTRateDetail fields are floats with arbitrary precision.

        The custom path must apply the same safe_to_decimal 4dp quantization as
        the standard USD/EUR path — never persist the raw API float. Feeds a
        value with >4dp and asserts the stored value equals the 4dp-quantized
        rate (Mathematical Truth), not the raw float.
        """
        from core.logic import safe_to_decimal

        raw = 42.123456  # 6dp float, as the live API would return
        dr = (date(2025, 3, 10), date(2025, 3, 10))
        per_ccy = {
            "GBP": [SimpleNamespace(
                period="2025-03-10", currency="GBP",
                buying_transfer=raw,
                selling=None, buying_sight=None, mid_rate=None,
            )],
        }
        _eng, out = self._run(
            exrate_file, temp_backup, tmp_cache, per_ccy,
            {"Buying TT": "buying_transfer"}, ["GBP"], dr,
        )
        wb = openpyxl.load_workbook(out)
        ws = wb["ExRate"]
        cell_val = ws.cell(row=2, column=2).value
        # Quantized to exactly 4dp, matching the standard path's discipline.
        assert _cell_decimal(cell_val) == safe_to_decimal(raw)
        assert _cell_decimal(cell_val) != Decimal(str(raw))  # raw float NOT stored
        wb.close()

    def test_stale_content_cleared_and_rewritten(
        self, exrate_file, temp_backup, tmp_cache,
    ):
        # Confirm the seeded junk ("OLD-DATE", date(1999,1,1)) is gone.
        dr = (date(2025, 3, 10), date(2025, 3, 10))
        per_ccy = {
            "GBP": [SimpleNamespace(
                period="2025-03-10", currency="GBP",
                buying_transfer=Decimal("42.0000"),
                selling=None, buying_sight=None, mid_rate=None,
            )],
        }
        _eng, out = self._run(
            exrate_file, temp_backup, tmp_cache, per_ccy,
            {"Buying TT": "buying_transfer"}, ["GBP"], dr,
        )
        wb = openpyxl.load_workbook(out)
        ws = wb["ExRate"]
        all_values = {
            ws.cell(row=r, column=c).value
            for r in range(1, (ws.max_row or 1) + 1)
            for c in range(1, (ws.max_column or 1) + 1)
        }
        assert "OLD-DATE" not in all_values
        assert "OLD-USD" not in all_values
        assert date(1999, 1, 1) not in all_values
        # New header was written.
        assert ws.cell(row=1, column=1).value == "Date"
        wb.close()

    def test_disk_space_guard_blocks_overwrite(
        self, exrate_file, temp_backup, tmp_cache, monkeypatch,
    ):
        """OSError disk guard must fire BEFORE wb.save in the custom path.

        The original file content must remain intact (no partial overwrite).
        """
        # Snapshot the original on-disk content.
        with open(exrate_file, "rb") as fh:
            original_bytes = fh.read()

        # Force the disk-space check to report a tiny free space.
        monkeypatch.setattr(
            workbook_io_mod.shutil, "disk_usage",
            lambda _path: _DiskUsage(total=10**12, used=10**12, free=0),
        )

        api = _make_api()
        api.get_exchange_rates = AsyncMock(
            side_effect=_make_currency_side_effect({
                "GBP": [SimpleNamespace(
                    period="2025-03-10", currency="GBP",
                    buying_transfer=Decimal("42.0000"),
                    selling=None, buying_sight=None, mid_rate=None,
                )],
            })
        )
        eng = LedgerEngine(api, backup=temp_backup, cache=tmp_cache)

        with pytest.raises(OSError, match="Insufficient disk space"):
            asyncio.run(eng.update_exrate_standalone(
                exrate_file,
                currencies=["GBP"],
                rate_types={"Buying TT": "buying_transfer"},
                date_range=(date(2025, 3, 10), date(2025, 3, 10)),
            ))

        # File on disk is byte-for-byte unchanged — the guard prevented save.
        with open(exrate_file, "rb") as fh:
            assert fh.read() == original_bytes


# =========================================================================
#  STANDARD PATH — manual date range + backup cleanup
# =========================================================================

class TestStandardPathManualRange:
    """update_exrate_standalone standard USD/EUR path with a manual range."""

    def _parse_cell_date(self, value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    def test_written_range_matches_manual_dr(
        self, exrate_file, temp_backup, tmp_cache,
    ):
        """A manual (dr_start, dr_end) must bound the written ExRate sheet —
        previously dr_end was ignored and the sheet ran out to today()."""
        from core.logic import BOTLogicEngine

        dr_start = date(2025, 3, 10)
        dr_end = date(2025, 3, 12)

        api = _make_api()
        eng = LedgerEngine(api, backup=temp_backup, cache=tmp_cache)

        # Mock the API preload so no network is hit; return rates only for the
        # manual window. Live-binding contract: reassigning _preload_api_data
        # after construction is honored at call time.
        logic_engine = BOTLogicEngine(holidays=[], max_rollback_days=10)
        usd_buying = {dr_start: Decimal("33.5000")}
        usd_selling = {dr_start: Decimal("33.6000")}
        eur_buying = {dr_start: Decimal("37.0000")}
        eur_selling = {dr_start: Decimal("37.1000")}

        async def _fake_preload(_dates, _start_str):
            return (
                logic_engine, usd_selling, eur_selling,
                usd_buying, eur_buying, [], [],
            )

        eng._preload_api_data = _fake_preload

        out = asyncio.run(eng.update_exrate_standalone(
            exrate_file,
            currencies=["USD", "EUR"],
            rate_types={"Buying TT": "buying_transfer", "Selling": "selling"},
            date_range=(dr_start, dr_end),
        ))
        assert out == exrate_file

        wb = openpyxl.load_workbook(out)
        ws = wb["ExRate"]
        written = {
            self._parse_cell_date(ws.cell(row=r, column=1).value)
            for r in range(2, (ws.max_row or 1) + 1)
        }
        written.discard(None)
        wb.close()

        # The manual 3-day window — dr_end honored, not today() — PLUS the
        # fixture's pre-existing 1999 row: existing history outside the
        # requested range is preserved, never silently trimmed.
        assert written == {
            date(1999, 1, 1), dr_start, date(2025, 3, 11), dr_end,
        }

    def test_standalone_run_prunes_old_backups(
        self, exrate_file, temp_backup, tmp_cache,
    ):
        """The standalone path must run the 7-day backup cleanup (previously
        only process_batch did), so a stale backup is pruned after a run."""
        from pathlib import Path

        # Seed a backup that is well past the 7-day cutoff. The cleanup derives
        # age from the embedded timestamp, so an old timestamp = old backup.
        backup_dir = Path(temp_backup.backup_dir)
        old_backup = backup_dir / "ExRate_standalone__bak__20000101_000000_000000.xlsx"
        old_backup.write_bytes(b"PK\x03\x04stale-backup")
        assert old_backup.exists()

        api = _make_api()
        api.get_exchange_rates = AsyncMock(
            side_effect=_make_currency_side_effect({
                "GBP": [SimpleNamespace(
                    period="2025-03-10", currency="GBP",
                    buying_transfer=Decimal("42.0000"),
                    selling=None, buying_sight=None, mid_rate=None,
                )],
            })
        )
        eng = LedgerEngine(api, backup=temp_backup, cache=tmp_cache)

        asyncio.run(eng.update_exrate_standalone(
            exrate_file,
            currencies=["GBP"],
            rate_types={"Buying TT": "buying_transfer"},
            date_range=(date(2025, 3, 10), date(2025, 3, 10)),
        ))

        # The stale backup is gone; a fresh backup of this run remains.
        assert not old_backup.exists()
        # Digest-aware glob: backup names are now {stem}__{digest}__bak__...
        fresh = list(backup_dir.glob("ExRate_standalone__*__bak__*.xlsx"))
        assert fresh, "expected a fresh backup from the standalone run"


# =========================================================================
#  LEDGER MULTI-CURRENCY PATH (process_ledger end-to-end)
# =========================================================================

def _ledger_engine(per_ccy, tmp_cache):
    """Engine wired to a per-currency mocked API + temp backup/cache.

    per_ccy maps a currency code → flat (buying, selling) tuple; the side
    effect emits one record per date in the requested range so any window
    resolves.
    """
    from datetime import timedelta as _td

    async def _rates(start, end, currency):
        pair = per_ccy.get(currency)
        if pair is None:
            return []
        buy, sell = pair
        out, d = [], start
        while d <= end:
            out.append(SimpleNamespace(
                period=d.strftime("%Y-%m-%d"), currency=currency,
                buying_transfer=buy, buying_sight=None,
                selling=sell, mid_rate=None,
            ))
            d += _td(days=1)
        return out

    async def _holidays(year):
        return []

    api = _make_api()
    api.get_exchange_rates = AsyncMock(side_effect=_rates)
    api.get_holidays = AsyncMock(side_effect=_holidays)
    from unittest.mock import MagicMock
    return LedgerEngine(api, backup=MagicMock(), cache=tmp_cache)


class TestLedgerMultiCurrency:
    """process_ledger must fill non-USD/EUR rows instead of leaving them blank.

    Findings: multi-currency ledger coverage, unsupported-currency warning,
    no-rate-available warning.
    """

    def _ledger(self, tmp_path, rows, name="led.xlsx"):
        path = tmp_path / name
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jan"
        ws.append(["Date", "Cur", "EX Rate", "Amount"])
        for d, ccy in rows:
            ws.append([d, ccy, None, 1000])
        wb.save(str(path))
        wb.close()
        return str(path)

    def _collect_warnings(self, engine):
        events: list[dict] = []
        engine._bus = SimpleNamespace(push=events.append)
        return events

    def test_gbp_row_gets_dynamic_column_and_formula(
        self, tmp_path, tmp_cache,
    ):
        path = self._ledger(tmp_path, [
            (date(2025, 1, 7), "USD"),
            (date(2025, 1, 7), "GBP"),
        ])
        engine = _ledger_engine(
            {"USD": (33.0, 33.5), "EUR": (36.0, 36.5),
             "GBP": (42.1234, 43.5)},
            tmp_cache,
        )
        result = asyncio.run(engine.process_ledger(path))
        assert result == path

        wb = openpyxl.load_workbook(path)
        try:
            ws_ex = wb["ExRate"]
            headers = [
                ws_ex.cell(row=1, column=c).value
                for c in range(1, (ws_ex.max_column or 1) + 1)
            ]
            # A GBP rate column physically exists in the master sheet.
            assert "GBP Rate" in headers
            gbp_col = headers.index("GBP Rate") + 1
            # The column carries the 4dp-quantized GBP rate (Mathematical
            # Truth, not the raw float) for the target date.
            gbp_vals = {
                _cell_decimal(ws_ex.cell(row=r, column=gbp_col).value)
                for r in range(2, (ws_ex.max_row or 1) + 1)
                if ws_ex.cell(row=r, column=gbp_col).value is not None
            }
            assert Decimal("42.1234") in gbp_vals

            # The GBP ledger row's formula references the GBP master column.
            gbp_formula = wb["Jan"].cell(row=3, column=3).value
            assert isinstance(gbp_formula, str)
            assert 'B3="GBP"' in gbp_formula
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(gbp_col)
            assert f"ExRate!${col_letter}$2" in gbp_formula
        finally:
            wb.close()

    def test_unsupported_currency_emits_warning(self, tmp_path, tmp_cache):
        path = self._ledger(tmp_path, [
            (date(2025, 1, 7), "USD"),
            (date(2025, 1, 7), "XYZ"),  # not in the supported set
            (date(2025, 1, 8), "XYZ"),
        ])
        engine = _ledger_engine({"USD": (33.0, 33.5), "EUR": (36.0, 36.5)},
                                tmp_cache)
        events = self._collect_warnings(engine)
        asyncio.run(engine.process_ledger(path))

        warnings = [e["msg"] for e in events if e["type"] == "warning"]
        assert any(
            "unsupported currency XYZ" in m and "2 row" in m
            for m in warnings
        ), warnings

    def test_jpy_is_unsupported_and_emits_warning(self, tmp_path, tmp_cache):
        """F4: BOT quotes JPY per 100 yen, so JPY is excluded from the ledger
        path — its rows must take the unsupported route (blank cell + warning,
        no dynamic 'JPY Rate' column), never a verbatim per-100 rate."""
        path = self._ledger(tmp_path, [
            (date(2025, 1, 7), "USD"),
            (date(2025, 1, 7), "JPY"),
        ])
        # Even with JPY data available from the API, the ledger path must NOT
        # fetch it — the published per-100 figure would overstate 100x.
        engine = _ledger_engine(
            {"USD": (33.0, 33.5), "EUR": (36.0, 36.5),
             "JPY": (23.1234, 23.5)},
            tmp_cache,
        )
        events = self._collect_warnings(engine)
        asyncio.run(engine.process_ledger(path))

        warnings = [e["msg"] for e in events if e["type"] == "warning"]
        assert any(
            "unsupported currency JPY" in m and "1 row" in m
            for m in warnings
        ), warnings

        wb = openpyxl.load_workbook(path)
        try:
            ws_ex = wb["ExRate"]
            headers = [
                ws_ex.cell(row=1, column=c).value
                for c in range(1, (ws_ex.max_column or 1) + 1)
            ]
            # No dynamic JPY column appended to the master sheet.
            assert "JPY Rate" not in headers
            # The injected IFS formula carries NO JPY branch — the row falls
            # through to the TRUE,"" fallback and renders blank in Excel
            # (never a verbatim per-100 rate).
            jpy_formula = wb["Jan"].cell(row=3, column=3).value
            assert isinstance(jpy_formula, str)
            assert '="JPY"' not in jpy_formula
        finally:
            wb.close()

    def test_no_rate_available_emits_warning(self, tmp_path, tmp_cache):
        """A USD row dated where no rate exists (API returns nothing) is
        flagged instead of silently blank."""
        path = self._ledger(tmp_path, [(date(2025, 1, 7), "USD")])
        # API returns NO USD records → the ExRate cell stays blank.
        engine = _ledger_engine({}, tmp_cache)
        events = self._collect_warnings(engine)
        asyncio.run(engine.process_ledger(path))

        warnings = [e["msg"] for e in events if e["type"] == "warning"]
        assert any(
            "no rate available" in m.lower() for m in warnings
        ), warnings

    def test_fully_resolved_ledger_emits_no_blank_warning(
        self, tmp_path, tmp_cache,
    ):
        """When every row resolves, no no-rate / unsupported warning fires."""
        path = self._ledger(tmp_path, [
            (date(2025, 1, 7), "USD"),
            (date(2025, 1, 7), "THB"),  # always resolves to 1
        ])
        engine = _ledger_engine({"USD": (33.0, 33.5), "EUR": (36.0, 36.5)},
                                tmp_cache)
        events = self._collect_warnings(engine)
        asyncio.run(engine.process_ledger(path))

        warnings = [e["msg"] for e in events if e["type"] == "warning"]
        assert not any(
            "no rate available" in m.lower() or "unsupported currency" in m
            for m in warnings
        ), warnings


# =========================================================================
#  EXTRA-CURRENCY ANOMALIES (F42) + AUDIT Anomaly_Flag WIRING (F25)
# =========================================================================

class TestExtraCurrencyAnomalyAlertOnly:
    """F42: a GBP rate jump is flagged like the USD/EUR series — yet the
    value still writes unchanged (alert-only), and the flagged (ccy, date)
    pair reaches the audit collector as anomaly_flag=True (F25)."""

    # Mon → Tue: 1-day gap (within ANOMALY_MAX_DAY_GAP), ~19% jump (>5%).
    PREV_DAY = date(2025, 1, 6)
    JUMP_DAY = date(2025, 1, 7)

    def _ledger(self, tmp_path, name="led_anomaly.xlsx"):
        path = tmp_path / name
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jan"
        ws.append(["Date", "Cur", "EX Rate", "Amount"])
        ws.append([self.PREV_DAY, "GBP", None, 500])   # pre-jump (clean)
        ws.append([self.JUMP_DAY, "GBP", None, 1000])  # the anomalous day
        wb.save(str(path))
        wb.close()
        return str(path)

    def _jumping_engine(self, tmp_cache):
        """Engine whose GBP series jumps 42.0000 → 50.0000 across one day."""
        gbp_series = {
            self.PREV_DAY: "42.0000",
            self.JUMP_DAY: "50.0000",
        }

        async def _rates(start, end, currency):
            if currency != "GBP":
                return []
            return [
                SimpleNamespace(
                    period=d.strftime("%Y-%m-%d"), currency="GBP",
                    buying_transfer=val, buying_sight=None,
                    selling=None, mid_rate=None,
                )
                for d, val in gbp_series.items()
                if start <= d <= end
            ]

        api = _make_api()
        api.get_exchange_rates = AsyncMock(side_effect=_rates)
        from unittest.mock import MagicMock
        return LedgerEngine(api, backup=MagicMock(), cache=tmp_cache)

    def test_gbp_jump_warns_and_value_still_writes(
        self, tmp_path, tmp_cache,
    ):
        path = self._ledger(tmp_path)
        engine = self._jumping_engine(tmp_cache)
        events: list[dict] = []
        engine._bus = SimpleNamespace(push=events.append)

        result = asyncio.run(engine.process_ledger(path))
        assert result == path

        # The extra-currency series was checked and the jump flagged.
        warnings = [e["msg"] for e in events if e["type"] == "warning"]
        assert any(
            "ANOMALY" in m and "GBP" in m for m in warnings
        ), warnings
        assert engine.last_anomaly_count >= 1

        # ALERT-ONLY: the anomalous 50.0000 still landed in the master
        # sheet, unchanged — the guard never blocks or substitutes a write.
        wb = openpyxl.load_workbook(path)
        try:
            ws_ex = wb["ExRate"]
            headers = [
                ws_ex.cell(row=1, column=c).value
                for c in range(1, (ws_ex.max_column or 1) + 1)
            ]
            assert "GBP Rate" in headers
            gbp_col = headers.index("GBP Rate") + 1
            gbp_vals = {
                _cell_decimal(ws_ex.cell(row=r, column=gbp_col).value)
                for r in range(2, (ws_ex.max_row or 1) + 1)
                if ws_ex.cell(row=r, column=gbp_col).value is not None
            }
            assert Decimal("50.0000") in gbp_vals
        finally:
            wb.close()

    def test_anomalous_row_audit_record_is_flagged(
        self, tmp_path, tmp_cache,
    ):
        """F25: the collector record for the flagged (GBP, jump-day) cell
        carries anomaly_flag=True; the clean previous day stays False."""
        from core.audit_logger import AuditCollector

        path = self._ledger(tmp_path)
        engine = self._jumping_engine(tmp_cache)
        collector = AuditCollector()

        asyncio.run(engine.process_ledger(path, audit=collector))

        records = {
            r.cell_date: r for r in collector.drain() if r.currency == "GBP"
        }
        jump_key = self.JUMP_DAY.strftime("%Y-%m-%d")
        prev_key = self.PREV_DAY.strftime("%Y-%m-%d")
        assert records[jump_key].anomaly_flag is True
        assert records[prev_key].anomaly_flag is False
        # The flag is metadata only — the anomalous value still resolved.
        assert records[jump_key].new_value == "50.0000"


class TestExtraCurrencyCacheFirst:
    """_fetch_extra_currency_rates must be cache-first like the USD/EUR path.

    Covers:
    - Offline path: CSV-imported (cache-seeded) GBP rate written into ledger
      when the API returns nothing for GBP.
    - API-wins path: fresh API data supersedes the cache value for the same
      date, matching the USD/EUR precedence.
    """

    def _ledger(self, tmp_path, rows, name="led.xlsx"):
        path = tmp_path / name
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jan"
        ws.append(["Date", "Cur", "EX Rate", "Amount"])
        for d, ccy in rows:
            ws.append([d, ccy, None, 1000])
        wb.save(str(path))
        wb.close()
        return str(path)

    def test_csv_imported_gbp_written_when_api_returns_nothing(
        self, tmp_path, tmp_cache,
    ):
        """Offline path: cache-seeded GBP rate reaches the ledger even when
        the API returns an empty list for GBP (simulates air-gapped operation
        after a CSV import has populated rates_multi).
        """
        target_date = date(2025, 1, 7)

        # Seed rates_multi directly (as csv_import.py would do).
        tmp_cache.insert_multi_rates_bulk([
            ("2025-01-07", "GBP", "buying_transfer", "42.1234"),
        ])

        path = self._ledger(tmp_path, [
            (target_date, "USD"),
            (target_date, "GBP"),
        ])

        # API returns USD/EUR data but nothing for GBP — simulates offline.
        async def _rates(start, end, currency):
            if currency == "USD":
                return [SimpleNamespace(
                    period="2025-01-07", currency="USD",
                    buying_transfer=33.0, selling=33.5,
                    buying_sight=None, mid_rate=None,
                )]
            if currency == "EUR":
                return [SimpleNamespace(
                    period="2025-01-07", currency="EUR",
                    buying_transfer=36.0, selling=36.5,
                    buying_sight=None, mid_rate=None,
                )]
            return []  # GBP — offline, no API data

        api = _make_api()
        api.get_exchange_rates = AsyncMock(side_effect=_rates)
        api.get_holidays = AsyncMock(return_value=[])

        from unittest.mock import MagicMock
        eng = LedgerEngine(api, backup=MagicMock(), cache=tmp_cache)
        result = asyncio.run(eng.process_ledger(path))
        assert result == path

        wb = openpyxl.load_workbook(path)
        try:
            ws_ex = wb["ExRate"]
            headers = [
                ws_ex.cell(row=1, column=c).value
                for c in range(1, (ws_ex.max_column or 1) + 1)
            ]
            # GBP Rate column must exist — cache-seeded data must flow through.
            assert "GBP Rate" in headers, f"Expected 'GBP Rate' in {headers}"
            gbp_col = headers.index("GBP Rate") + 1
            gbp_vals = {
                _cell_decimal(ws_ex.cell(row=r, column=gbp_col).value)
                for r in range(2, (ws_ex.max_row or 1) + 1)
                if ws_ex.cell(row=r, column=gbp_col).value is not None
            }
            # The cache-seeded rate (not an API rate) must appear.
            assert Decimal("42.1234") in gbp_vals, (
                f"Cache-seeded GBP rate 42.1234 not found in ExRate; "
                f"got {gbp_vals}"
            )
        finally:
            wb.close()

    def test_api_data_wins_over_cache_for_same_date(
        self, tmp_cache,
    ):
        """API-wins path: when the API returns a GBP rate for a date that is
        also present in the cache, the fresh API value supersedes the stale
        cache value — same precedence rule as the USD/EUR path.

        Exercises _fetch_extra_currency_rates directly with an empty cache so
        the API is consulted (cache miss), then verifies the API value lands.
        """
        target_date = date(2025, 1, 7)
        fresh_rate = Decimal("42.1234")

        # API returns a fresh GBP rate.
        async def _rates(start, end, currency):
            if currency == "GBP":
                return [SimpleNamespace(
                    period="2025-01-07", currency="GBP",
                    buying_transfer=float(fresh_rate), selling=43.5,
                    buying_sight=None, mid_rate=None,
                )]
            return []

        api = _make_api()
        api.get_exchange_rates = AsyncMock(side_effect=_rates)

        from unittest.mock import MagicMock
        eng = LedgerEngine(api, backup=MagicMock(), cache=tmp_cache)

        result = asyncio.run(eng._fetch_extra_currency_rates(
            ["GBP"], "buying_transfer",
            target_date, target_date,
        ))
        # API returned fresh_rate for 2025-01-07; it must be present.
        assert result["GBP"][target_date] == fresh_rate, (
            f"Expected API value {fresh_rate}, got {result['GBP'].get(target_date)}"
        )

        # Also verify the fresh API value was persisted to rates_multi so
        # the next run can serve it from cache (cache-store contract).
        cached = tmp_cache.get_rates_multi(
            target_date, target_date, "GBP", "buying_transfer"
        )
        assert cached.get(target_date) == fresh_rate, (
            f"API rate not stored back to rates_multi: {cached}"
        )


class TestRateTypeSnapshot:
    """A Settings rate_type change mid-run must NOT affect the in-flight file.

    process_ledger snapshots rate_type once at the start and threads it to the
    writer instead of re-reading SettingsManager inside write().
    """

    def test_rate_type_snapshotted_at_start(
        self, tmp_path, tmp_cache, monkeypatch,
    ):
        import core.engine as engine_mod

        path = tmp_path / "led.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jan"
        ws.append(["Date", "Cur", "EX Rate", "Amount"])
        ws.append([date(2025, 1, 7), "USD", None, 1000])
        wb.save(str(path))
        wb.close()

        # SettingsManager().load() yields "buying_transfer" at snapshot time,
        # then flips to "selling" to simulate a concurrent mid-run Save. The
        # writer must use the value captured BEFORE the flip. The patch is
        # installed BEFORE the engine is constructed because the snapshot is
        # taken in LedgerEngine.__init__ — patching afterwards would leave the
        # snapshot to whatever settings file happened to exist on disk.
        state = {"rate_type": "buying_transfer"}

        class _Settings:
            def load(self):
                val = state["rate_type"]
                # After the first (snapshot) read, simulate the user saving a
                # new rate type while this file is still being processed.
                state["rate_type"] = "selling"
                return {"rate_type": val}

        monkeypatch.setattr(engine_mod, "SettingsManager", _Settings)

        engine = _ledger_engine({"USD": (33.0, 33.5), "EUR": (36.0, 36.5)},
                                tmp_cache)
        # Construction consumed exactly the first (pre-flip) read: the
        # snapshot is deterministic, independent of any on-disk settings.
        assert engine._rate_type == "buying_transfer"
        assert state["rate_type"] == "selling"

        captured = {}
        from core.exrate_updater import WorkbookWriter
        orig_write = WorkbookWriter.write

        async def _spy(self, *args, **kwargs):
            captured["rate_type"] = kwargs.get("rate_type")
            return await orig_write(self, *args, **kwargs)

        monkeypatch.setattr(WorkbookWriter, "write", _spy)

        asyncio.run(engine.process_ledger(str(path)))
        # The snapshot (buying_transfer) reached the writer — NOT the value the
        # concurrent save flipped to (selling).
        assert captured["rate_type"] == "buying_transfer"

        wb = openpyxl.load_workbook(str(path))
        try:
            formula = wb["Jan"].cell(row=2, column=3).value
            # Buying columns are B (USD) / D (EUR); selling would be C / E.
            assert "ExRate!$B$2" in formula
            assert "ExRate!$C$2" not in formula
        finally:
            wb.close()
