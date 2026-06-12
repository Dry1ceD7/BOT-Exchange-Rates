#!/usr/bin/env python3
"""
tests/golden/build_fixtures.py
---------------------------------------------------------------------------
Golden-master fixture builder and scenario runner.
---------------------------------------------------------------------------
Deterministically constructs the input workbooks, runs the REAL write paths
(LedgerEngine.process_batch and update_exrate_standalone) with the BOT API
mocked at the same seam the engine test suite uses (an api object whose
async get_exchange_rates / get_holidays return a FIXED hardcoded dataset —
never the network), and serializes the complete user-visible output:

  - every ExRate master-sheet cell (dates, 4dp rate values as exact strings,
    weekend/holiday rows carrying Date + label only with blank rate cells),
  - every injected ledger IFS/XLOOKUP formula string verbatim,
  - the audit CSV header and row shape (timestamps excluded).

The committed tests/golden/expected_*.json files are the frozen golden
masters; tests/test_golden_master.py re-runs both scenarios and compares.

REGENERATION POLICY
    python tests/golden/build_fixtures.py --regen
rewrites the expected_*.json files from CURRENT behavior. Regenerating is a
deliberate act: it must happen ONLY in a commit that explicitly declares and
justifies the user-visible behavior change the new snapshots encode. Running
the script with no flag (or --check) compares current behavior against the
committed snapshots without writing anything.

Determinism notes:
  - bot_today() is pinned to FIXED_TODAY in every module that binds it on the
    write path (core.engine, core.exrate_sheet, core.exrate_updater).
  - SettingsManager is replaced so rate_type/anomaly threshold cannot drift
    with the developer's local settings file.
  - The engine-owned AuditLogger is redirected into the scenario work dir and
    audit-log pruning is disabled, so data/ in the repo is never touched.
"""

import argparse
import asyncio
import contextlib
import csv
import functools
import json
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace
from unittest import mock
from unittest.mock import MagicMock

# Allow direct script execution (python tests/golden/build_fixtures.py) by
# putting the repo root on sys.path before the core imports.
_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import openpyxl  # noqa: E402

GOLDEN_DIR = Path(__file__).resolve().parent

EXPECTED_FILES = {
    "ledger_exrate": GOLDEN_DIR / "expected_ledger_exrate.json",
    "ledger_formulas": GOLDEN_DIR / "expected_ledger_formulas.json",
    "ledger_audit": GOLDEN_DIR / "expected_ledger_audit.json",
    "standalone_exrate": GOLDEN_DIR / "expected_standalone_exrate.json",
    "realistic_exrate": GOLDEN_DIR / "expected_realistic_exrate.json",
    "realistic_formulas": GOLDEN_DIR / "expected_realistic_formulas.json",
    "realistic_audit": GOLDEN_DIR / "expected_realistic_audit.json",
}

# ── Fixed business calendar ──────────────────────────────────────────────
# Friday 2025-01-10 is the pinned BOT business date ("today"), so the ledger
# ExRate sheet spans computed_start (Mon 2024-12-30) .. 2025-01-10: 12 rows
# including one Saturday, one Sunday, and two holidays.
FIXED_TODAY = date(2025, 1, 10)

HOLIDAYS_BY_YEAR = {
    2024: [("2024-12-31", "New Year's Eve")],
    2025: [("2025-01-01", "New Year's Day")],
}

# ── Fixed rate dataset ───────────────────────────────────────────────────
# One record per BOT trading day in the preload window (2024-12-20 ..
# 2025-01-10). Weekends and the two holidays above publish NO rate, exactly
# like the live BOT API. All values are exact 4dp strings; day-over-day
# changes stay far below the 5 percent anomaly threshold.
# Columns: date, USD buy, USD sell, EUR buy, EUR sell, GBP buy, GBP sell.
RATE_ROWS = [
    ("2024-12-20", "34.0512", "34.3209", "35.4023", "36.1217", "42.8101", "43.5604"),
    ("2024-12-23", "34.1020", "34.3718", "35.4521", "36.1722", "42.8650", "43.6158"),
    ("2024-12-24", "34.1535", "34.4231", "35.5034", "36.2240", "42.9203", "43.6711"),
    ("2024-12-25", "34.2048", "34.4744", "35.5546", "36.2757", "42.9755", "43.7263"),
    ("2024-12-26", "34.1560", "34.4256", "35.5051", "36.2261", "42.9258", "43.6766"),
    ("2024-12-27", "34.2071", "34.4767", "35.5560", "36.2770", "42.9810", "43.7318"),
    ("2024-12-30", "34.2580", "34.5276", "35.6068", "36.3278", "43.0361", "43.7869"),
    ("2025-01-02", "34.3088", "34.5784", "35.6575", "36.3785", "43.0911", "43.8419"),
    ("2025-01-03", "34.2599", "34.5295", "35.6080", "36.3290", "43.0414", "43.7922"),
    ("2025-01-06", "34.3105", "34.5801", "35.6584", "36.3794", "43.0962", "43.8470"),
    ("2025-01-07", "34.3610", "34.6306", "35.7087", "36.4297", "43.1509", "43.9017"),
    ("2025-01-08", "34.3114", "34.5810", "35.6590", "36.3800", "43.1010", "43.8518"),
    ("2025-01-09", "34.3617", "34.6313", "35.7092", "36.4302", "43.1556", "43.9064"),
    ("2025-01-10", "34.4119", "34.6815", "35.7593", "36.4803", "43.2101", "43.9609"),
]

# Column slice of RATE_ROWS per currency: (buying_transfer, selling).
_CCY_COLUMNS = {"USD": (1, 2), "EUR": (3, 4), "GBP": (5, 6)}

# ── Ledger input rows (day-first date strings, written verbatim) ─────────
# 12 rows spanning the first business week of January 2025 plus a Saturday
# (04/01), a Sunday (05/01), and the New Year's Day holiday (01/01).
LEDGER_ROWS = [
    ("02/01/2025", "USD"),
    ("02/01/2025", "EUR"),
    ("03/01/2025", "GBP"),
    ("04/01/2025", "USD"),  # Saturday — no BOT rate
    ("05/01/2025", "EUR"),  # Sunday — no BOT rate
    ("06/01/2025", "USD"),
    ("06/01/2025", "GBP"),
    ("07/01/2025", "EUR"),
    ("08/01/2025", "USD"),
    ("01/01/2025", "USD"),  # New Year's Day holiday — no BOT rate
    ("09/01/2025", "GBP"),
    ("10/01/2025", "EUR"),
]

STANDARD_EXRATE_HEADERS = [
    "Date", "USD Buying TT Rate", "USD Selling Rate",
    "EUR Buying TT Rate", "EUR Selling Rate", "Holidays/Weekend",
]

# ── Realistic production-ledger fixture ──────────────────────────────────
# Mirrors the REAL 16-sheet ledger workbooks: 12 month tabs in the
# production column layout (NO | Date(invoice) | Thai detail | Cur |
# Date(export-entry) | EX Rate | Amount) with the header at row 3 below a
# Crystal-Reports-style preamble, plus a PI sheet, the pre-existing
# 'Exrate USD'/'Exrate EUR' historical tabs (SKIP_SHEET_NAMES members that
# must round-trip untouched), and a pre-existing ExRate master carrying one
# prior trading-day row the run must preserve. Only Jan holds data — the
# production shape at the start of a year.
REALISTIC_MONTHS = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]
REALISTIC_HEADER_ROW = 3
REALISTIC_HEADERS = [
    "NO", "Date", "รายละเอียด", "Cur", "Date", "EX Rate", "Amount",
]
REALISTIC_PREAMBLE = "บริษัท ตัวอย่างการค้า จำกัด — Crystal Reports Export"

# (invoice_date, thai_detail, cur, export_entry_date, amount). The
# export-entry Date (column E, nearest left of EX Rate) is the
# rate-resolving one; invoice dates deliberately differ so a regression
# that re-binds the source column to the first 'Date' changes the output.
REALISTIC_JAN_ROWS = [
    ("30/12/2024", "ค่าสินค้านำเข้า",  "USD", "02/01/2025", 18500),
    ("02/01/2025", "ค่าขนส่งระหว่างประเทศ", "EUR", "03/01/2025", 7200),
    ("03/01/2025", "ค่าสินค้านำเข้า",  "GBP", "06/01/2025", 5100),
    ("03/01/2025", "ค่าธรรมเนียมธนาคาร", "USD", "04/01/2025", 950),   # Saturday
    ("31/12/2024", "ค่าสินค้านำเข้า",  "EUR", "01/01/2025", 4400),   # holiday
    ("06/01/2025", "ค่าสินค้าในประเทศ", "THB", "07/01/2025", 30000),
    ("06/01/2025", "ค่าอะไหล่นำเข้า",  "JPY", "08/01/2025", 2100),   # per-100 excluded
    ("08/01/2025", "ค่าสินค้านำเข้า",  "USD", "10/01/2025", 12750),
]

# One historical trading-day row already on the pre-existing ExRate master
# (before the fetch window) — the run must PRESERVE it, never trim history.
REALISTIC_PRIOR_MASTER_ROW = (
    date(2024, 12, 16), "33.9876", "34.2587", "35.3399", "36.0601",
)

# Historical per-currency tabs found in older production books. They are
# SKIP_SHEET_NAMES members: never scanned, never written.
REALISTIC_HISTORICAL_TABS = {
    "Exrate USD": [(date(2024, 12, 27), 34.2071)],
    "Exrate EUR": [(date(2024, 12, 27), 35.5560)],
}


# ========================================================================
#  INPUT WORKBOOK BUILDERS
# ========================================================================

def build_ledger_workbook(filepath) -> str:
    """Build the deterministic one-month ledger workbook (tab 'Jan')."""
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jan"
    ws.append(["Date", "Cur", "EX Rate", "Amount"])
    for raw_date, ccy in LEDGER_ROWS:
        ws.append([raw_date, ccy, None, 1000])
    wb.save(str(filepath))
    wb.close()
    return str(filepath)


def build_realistic_ledger_workbook(filepath) -> str:
    """Build the deterministic 16-sheet production-shape ledger workbook.

    12 month tabs (production layout, header at row 3 under a Crystal-style
    preamble; data in Jan only), a PI sheet, the 'Exrate USD'/'Exrate EUR'
    historical tabs, and a pre-existing ExRate master with one prior
    trading-day row.
    """
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for month in REALISTIC_MONTHS:
        ws = wb.create_sheet(month)
        ws.append([REALISTIC_PREAMBLE])   # row 1 — report title
        ws.append([])                     # row 2 — blank preamble row
        ws.append(REALISTIC_HEADERS)      # row 3 — production header
        if month == "Jan":
            for no, (inv, detail, ccy, entry, amount) in enumerate(
                REALISTIC_JAN_ROWS, 1,
            ):
                # Production books carry a FORMULA in the NO column (a
                # running row number) on most rows — it must round-trip
                # untouched by injection. Keep one plain int (row 1) so
                # both shapes are pinned.
                no_val = (
                    no if no == 1
                    else f"=ROW()-{REALISTIC_HEADER_ROW}"
                )
                ws.append([no_val, inv, detail, ccy, entry, None, amount])

    ws_pi = wb.create_sheet("PI")
    ws_pi.append(["PI No", "Customer", "Value"])
    ws_pi.append(["PI-2025-001", "ACME Industrial GmbH", 12000])
    ws_pi.append(["PI-2025-002", "Siam Trading Co., Ltd.", 8400])

    for tab, rows in REALISTIC_HISTORICAL_TABS.items():
        ws_h = wb.create_sheet(tab)
        ws_h.append(["Date", "Rate"])
        for d, rate in rows:
            ws_h.append([d, rate])

    ws_ex = wb.create_sheet("ExRate")
    ws_ex.append(STANDARD_EXRATE_HEADERS)
    prior_date, usd_b, usd_s, eur_b, eur_s = REALISTIC_PRIOR_MASTER_ROW
    ws_ex.append([prior_date, float(usd_b), float(usd_s),
                  float(eur_b), float(eur_s), ""])

    wb.save(str(filepath))
    wb.close()
    return str(filepath)


def build_standard_exrate_workbook(filepath) -> str:
    """Build the deterministic standalone USD/EUR ExRate workbook (header only)."""
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExRate"
    ws.append(STANDARD_EXRATE_HEADERS)
    wb.save(str(filepath))
    wb.close()
    return str(filepath)


# ========================================================================
#  MOCKED API (same seam as tests/test_engine*.py)
# ========================================================================

def make_mock_api():
    """A mocked BOTClient surface emitting the fixed dataset, never the network."""
    by_date = {
        datetime.strptime(row[0], "%Y-%m-%d").date(): row for row in RATE_ROWS
    }

    async def _rates(start, end, currency):
        cols = _CCY_COLUMNS.get(currency)
        if cols is None:
            return []
        out = []
        for d in sorted(by_date):
            if start <= d <= end:
                row = by_date[d]
                out.append(SimpleNamespace(
                    period=d.strftime("%Y-%m-%d"),
                    currency=currency,
                    buying_transfer=float(row[cols[0]]),
                    buying_sight=None,
                    selling=float(row[cols[1]]),
                    mid_rate=None,
                ))
        return out

    async def _holidays(year):
        return [
            SimpleNamespace(date=d_str, description=name)
            for d_str, name in HOLIDAYS_BY_YEAR.get(year, [])
        ]

    api = MagicMock()
    api.get_exchange_rates = _rates
    api.get_holidays = _holidays
    return api


# ========================================================================
#  DETERMINISM PATCHES
# ========================================================================

@contextlib.contextmanager
def deterministic_patches(log_dir):
    """Pin every environment-dependent seam on the write path.

    - bot_today() pinned to FIXED_TODAY in each module-level binding,
    - SettingsManager replaced (rate_type=buying_transfer, threshold=5.0),
    - the engine-owned AuditLogger redirected into ``log_dir``,
    - audit-log pruning disabled so the real data/logs is never touched.
    """
    from core.audit_logger import AuditLogger

    class _FixedSettings:
        def load(self):
            return {
                "rate_type": "buying_transfer",
                "anomaly_threshold_pct": 5.0,
            }

    def _fixed_today():
        return FIXED_TODAY

    with contextlib.ExitStack() as stack:
        for target in (
            "core.engine.bot_today",
            "core.exrate_sheet.bot_today",
            "core.exrate_updater.bot_today",
        ):
            stack.enter_context(mock.patch(target, _fixed_today))
        stack.enter_context(
            mock.patch("core.engine.SettingsManager", _FixedSettings)
        )
        stack.enter_context(mock.patch(
            "core.engine.AuditLogger",
            functools.partial(AuditLogger, log_dir=str(log_dir)),
        ))
        stack.enter_context(mock.patch(
            "core.engine.cleanup_old_audit_logs", lambda *a, **k: 0,
        ))
        yield


# ========================================================================
#  OUTPUT SNAPSHOT SERIALIZERS
# ========================================================================

def _cell_repr(value):
    """JSON-stable representation of a reloaded cell value.

    None stays None; dates/datetimes become ISO strings; numbers go through
    str() (the shortest exact decimal repr of the stored float), so a 4dp
    Decimal written by the engine round-trips to its exact digit string.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def snapshot_exrate_sheet(filepath) -> dict:
    """Serialize EVERY cell of the ExRate sheet plus its number formats."""
    wb = openpyxl.load_workbook(str(filepath))
    try:
        ws = wb["ExRate"]
        n_cols = ws.max_column or 1
        headers = [ws.cell(row=1, column=c).value for c in range(1, n_cols + 1)]
        rows = []
        date_formats: set[str] = set()
        rate_formats: set[str] = set()
        for r in range(2, (ws.max_row or 1) + 1):
            row_vals = []
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                row_vals.append(_cell_repr(cell.value))
                if c == 1:
                    date_formats.add(cell.number_format)
                elif c < n_cols and cell.value is not None:
                    rate_formats.add(cell.number_format)
            rows.append(row_vals)
        return {
            "headers": headers,
            "rows": rows,
            "date_number_formats": sorted(date_formats),
            "rate_number_formats": sorted(rate_formats),
        }
    finally:
        wb.close()


def snapshot_ledger_sheet(filepath, sheet: str = "Jan") -> dict:
    """Serialize every data row of a monthly tab: normalized date, currency,
    and the injected EX Rate formula string verbatim."""
    wb = openpyxl.load_workbook(str(filepath))
    try:
        ws = wb[sheet]
        rows = []
        for r in range(2, (ws.max_row or 1) + 1):
            rows.append({
                "row": r,
                "date": _cell_repr(ws.cell(row=r, column=1).value),
                "date_number_format": ws.cell(row=r, column=1).number_format,
                "cur": _cell_repr(ws.cell(row=r, column=2).value),
                "formula": ws.cell(row=r, column=3).value,
            })
        return {"sheet": sheet, "rows": rows}
    finally:
        wb.close()


def snapshot_realistic_ledger(filepath) -> dict:
    """Serialize the production-shape workbook's user-visible output.

    Per month tab: the header row, ``max_row`` (locks the bounded
    last-data-row + PREFORMAT_BUFFER_ROWS extent — the unbounded-growth
    regression guard), and every row holding any value in the production
    columns (NO/invoice/detail/Cur/entry/EX Rate). The PI and
    'Exrate USD'/'Exrate EUR' tabs are dumped whole: they must round-trip
    byte-identical through the run (skip-sheet contract).
    """
    wb = openpyxl.load_workbook(str(filepath))
    try:
        tabs = {}
        for sheet in REALISTIC_MONTHS:
            ws = wb[sheet]
            rows = []
            for r in range(REALISTIC_HEADER_ROW + 1, (ws.max_row or 1) + 1):
                entry_cell = ws.cell(row=r, column=5)
                row_repr = {
                    "row": r,
                    "no": _cell_repr(ws.cell(row=r, column=1).value),
                    "invoice_date": _cell_repr(ws.cell(row=r, column=2).value),
                    "detail": _cell_repr(ws.cell(row=r, column=3).value),
                    "cur": _cell_repr(ws.cell(row=r, column=4).value),
                    "entry_date": _cell_repr(entry_cell.value),
                    "entry_date_number_format": entry_cell.number_format,
                    "formula": ws.cell(row=r, column=6).value,
                    "amount": _cell_repr(ws.cell(row=r, column=7).value),
                }
                if any(
                    v is not None for k, v in row_repr.items()
                    if k not in ("row", "entry_date_number_format")
                ):
                    rows.append(row_repr)
            tabs[sheet] = {
                "header_row": REALISTIC_HEADER_ROW,
                "max_row": ws.max_row,
                "rows": rows,
            }
        passthrough = {
            sheet: [
                [_cell_repr(v) for v in row]
                for row in wb[sheet].iter_rows(values_only=True)
            ]
            for sheet in ("PI", *REALISTIC_HISTORICAL_TABS)
        }
        return {"tabs": tabs, "passthrough": passthrough}
    finally:
        wb.close()


def snapshot_audit_csv(filepath) -> dict:
    """Serialize the audit CSV shape: full header, every row WITHOUT the
    volatile Timestamp column (column 0). Blank separator rows stay []."""
    with open(filepath, encoding="utf-8-sig", newline="") as fh:
        raw = list(csv.reader(fh))
    return {
        "headers": raw[0],
        "rows": [row[1:] if row else [] for row in raw[1:]],
    }


# ========================================================================
#  SCENARIO RUNNERS (the real write paths)
# ========================================================================

def run_ledger_scenario(workdir) -> dict:
    """Run process_batch on the golden ledger; return all three snapshots."""
    from core.database import CacheDB
    from core.engine import BatchManifest, LedgerEngine

    workdir = Path(workdir)
    workdir.mkdir(parents=True, exist_ok=True)
    ledger_path = build_ledger_workbook(workdir / "golden_ledger.xlsx")
    cache = CacheDB(db_path=str(workdir / "golden_cache.db"))
    try:
        with deterministic_patches(workdir / "logs"):
            engine = LedgerEngine(
                make_mock_api(), backup=MagicMock(), cache=cache,
            )
            manifest = BatchManifest(workdir / "batch_state.json")
            success, failed, errors = asyncio.run(
                engine.process_batch([ledger_path], manifest=manifest)
            )
        if (success, failed) != (1, 0):
            raise RuntimeError(f"golden ledger scenario failed: {errors}")
        return {
            "exrate": snapshot_exrate_sheet(ledger_path),
            "ledger": snapshot_ledger_sheet(ledger_path),
            "audit": snapshot_audit_csv(engine.last_audit_path),
        }
    finally:
        cache.close()


def run_realistic_scenario(workdir) -> dict:
    """Run process_batch on the 16-sheet production-shape golden ledger."""
    from core.database import CacheDB
    from core.engine import BatchManifest, LedgerEngine

    workdir = Path(workdir)
    workdir.mkdir(parents=True, exist_ok=True)
    ledger_path = build_realistic_ledger_workbook(
        workdir / "golden_realistic.xlsx"
    )
    cache = CacheDB(db_path=str(workdir / "golden_cache.db"))
    try:
        with deterministic_patches(workdir / "logs"):
            engine = LedgerEngine(
                make_mock_api(), backup=MagicMock(), cache=cache,
            )
            manifest = BatchManifest(workdir / "batch_state.json")
            success, failed, errors = asyncio.run(
                engine.process_batch([ledger_path], manifest=manifest)
            )
        if (success, failed) != (1, 0):
            raise RuntimeError(f"golden realistic scenario failed: {errors}")
        return {
            "exrate": snapshot_exrate_sheet(ledger_path),
            "ledger": snapshot_realistic_ledger(ledger_path),
            "audit": snapshot_audit_csv(engine.last_audit_path),
        }
    finally:
        cache.close()


def run_standalone_scenario(workdir) -> dict:
    """Run update_exrate_standalone (standard USD/EUR path, manual range)."""
    from core.database import CacheDB
    from core.engine import LedgerEngine

    workdir = Path(workdir)
    workdir.mkdir(parents=True, exist_ok=True)
    exrate_path = build_standard_exrate_workbook(workdir / "golden_exrate.xlsx")
    cache = CacheDB(db_path=str(workdir / "golden_cache.db"))
    try:
        with deterministic_patches(workdir / "logs"):
            engine = LedgerEngine(
                make_mock_api(), backup=MagicMock(), cache=cache,
            )
            out = asyncio.run(engine.update_exrate_standalone(
                exrate_path,
                currencies=["USD", "EUR"],
                rate_types={
                    "Buying TT": "buying_transfer",
                    "Selling": "selling",
                },
                date_range=(date(2024, 12, 30), FIXED_TODAY),
            ))
        return {"exrate": snapshot_exrate_sheet(out)}
    finally:
        cache.close()


# ========================================================================
#  EXPECTED-FILE HELPERS + CLI
# ========================================================================

def load_expected(key: str) -> dict:
    """Load one committed expected_*.json golden master."""
    return json.loads(EXPECTED_FILES[key].read_text(encoding="utf-8"))


def _current_payloads() -> dict[str, dict]:
    """Run both scenarios in a throwaway dir and return all four payloads."""
    with tempfile.TemporaryDirectory(prefix="bot_golden_") as td:
        td = Path(td)
        ledger = run_ledger_scenario(td / "ledger")
        standalone = run_standalone_scenario(td / "standalone")
        realistic = run_realistic_scenario(td / "realistic")
    return {
        "ledger_exrate": ledger["exrate"],
        "ledger_formulas": ledger["ledger"],
        "ledger_audit": ledger["audit"],
        "standalone_exrate": standalone["exrate"],
        "realistic_exrate": realistic["exrate"],
        "realistic_formulas": realistic["ledger"],
        "realistic_audit": realistic["audit"],
    }


def regenerate() -> None:
    """Rewrite every expected_*.json from CURRENT behavior.

    Only do this in a commit that explicitly declares and justifies the
    behavior change (see the module docstring's regeneration policy).
    """
    for key, payload in _current_payloads().items():
        EXPECTED_FILES[key].write_text(
            json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        print(f"wrote {EXPECTED_FILES[key]}")


def check() -> int:
    """Compare current behavior against the committed snapshots."""
    drift = 0
    payloads = _current_payloads()
    for key, payload in payloads.items():
        path = EXPECTED_FILES[key]
        if not path.is_file():
            print(f"MISSING {path}")
            drift += 1
            continue
        if payload != load_expected(key):
            print(f"DRIFT   {path}")
            drift += 1
        else:
            print(f"OK      {path}")
    return 1 if drift else 0


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[2])
    parser.add_argument(
        "--regen", action="store_true",
        help="rewrite expected_*.json from current behavior "
             "(requires a commit declaring the behavior change)",
    )
    args = parser.parse_args(argv)
    if args.regen:
        regenerate()
        return 0
    return check()


if __name__ == "__main__":
    raise SystemExit(main())
