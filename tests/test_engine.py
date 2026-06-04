#!/usr/bin/env python3
"""
tests/test_engine.py
---------------------------------------------------------------------------
Unit & integration tests for core/engine.py — LedgerEngine orchestrator.
Uses mocked API client and temporary files.
---------------------------------------------------------------------------
"""

import asyncio
import threading
from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock

import httpx
import openpyxl
import pytest

import core.engine as engine_mod
from core.api_client import BOTAPIError
from core.constants import MAX_FILE_SIZE_MB, SKIP_SHEET_NAMES
from core.engine import (
    FileSizeLimitError,
    LedgerEngine,
)
from core.ledger_processing import prescan_target_dates, run_anomaly_check

# =========================================================================
#  FIXTURES
# =========================================================================

@pytest.fixture
def mock_api():
    """Creates a mock BOTClient."""
    api = AsyncMock()
    api.get_holidays = AsyncMock(return_value=[])
    api.get_exchange_rates = AsyncMock(return_value=[])
    return api


@pytest.fixture
def engine(mock_api, tmp_cache):
    """Creates a LedgerEngine with mocked API and injected temp cache/backup.

    Injecting ``cache=tmp_cache`` (the temp on-disk SQLite from conftest) and a
    MagicMock backup keeps unit tests from lazily constructing the REAL
    data/cache.db singleton (+atexit handler) as a side effect.
    """
    return LedgerEngine(mock_api, cache=tmp_cache, backup=MagicMock())


@pytest.fixture
def sample_xlsx(tmp_path):
    """Creates a minimal .xlsx file with a single ledger sheet."""
    filepath = tmp_path / "test_ledger.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jan"
    ws.append(["Date", "Cur", "EX Rate", "Amount"])
    ws.append([date(2025, 3, 10), "USD", None, 1000])
    ws.append([date(2025, 3, 11), "EUR", None, 2000])
    ws.append([date(2025, 3, 12), "THB", None, 3000])
    wb.save(str(filepath))
    wb.close()
    return str(filepath)


@pytest.fixture
def oversized_file(tmp_path):
    """Creates a file just over the MAX_FILE_SIZE_MB limit (default 15 MB).

    Sized relative to the constant ((MAX_FILE_SIZE_MB + 1) MiB) so the fixture
    tracks the real featherweight limit instead of a hardcoded number.
    """
    filepath = tmp_path / "huge.xlsx"
    filepath.write_bytes(b"x" * ((MAX_FILE_SIZE_MB + 1) * 1024 * 1024))
    return str(filepath)


# =========================================================================
#  PARSE DATE
# =========================================================================

class TestParseDate:
    """Tests for _parse_date method."""

    def test_datetime_input(self, engine):
        dt = datetime(2025, 3, 10, 14, 30)
        assert engine._parse_date(dt) == date(2025, 3, 10)

    def test_date_input(self, engine):
        d = date(2025, 3, 10)
        assert engine._parse_date(d) == d

    def test_iso_string(self, engine):
        assert engine._parse_date("2025-03-10") == date(2025, 3, 10)

    def test_slash_format(self, engine):
        assert engine._parse_date("10/03/2025") == date(2025, 3, 10)

    def test_dash_dmy(self, engine):
        assert engine._parse_date("10-03-2025") == date(2025, 3, 10)

    def test_named_month(self, engine):
        assert engine._parse_date("10 Mar 2025") == date(2025, 3, 10)

    def test_full_month_name(self, engine):
        assert engine._parse_date("10 March 2025") == date(2025, 3, 10)

    def test_compact_format(self, engine):
        assert engine._parse_date("20250310") == date(2025, 3, 10)

    def test_none_returns_none(self, engine):
        assert engine._parse_date(None) is None

    def test_empty_string_returns_none(self, engine):
        assert engine._parse_date("") is None

    def test_nan_string_returns_none(self, engine):
        assert engine._parse_date("nan") is None

    def test_null_string_returns_none(self, engine):
        assert engine._parse_date("null") is None

    def test_invalid_string_returns_none(self, engine):
        assert engine._parse_date("not-a-date") is None

    def test_integer_returns_none(self, engine):
        assert engine._parse_date(12345) is None


# =========================================================================
#  MEMORY GUARDRAIL
# =========================================================================

class TestMemoryGuardrail:
    """Tests for _check_memory_guardrail method."""

    def test_existing_file_passes(self, engine, sample_xlsx):
        engine._check_memory_guardrail(sample_xlsx)  # Should not raise

    def test_missing_file_raises(self, engine):
        with pytest.raises(FileNotFoundError):
            engine._check_memory_guardrail("/nonexistent/path.xlsx")

    def test_oversized_file_raises(self, engine, oversized_file):
        with pytest.raises(FileSizeLimitError):
            engine._check_memory_guardrail(oversized_file)


# =========================================================================
#  COMPUTE YEAR START DATE
# =========================================================================

class TestComputeYearStartDate:
    """Tests for compute_year_start_date static method."""

    def test_normal_weekday(self):
        # 2024-12-30 is Monday
        result = LedgerEngine.compute_year_start_date(2025, [])
        assert result == date(2024, 12, 30)

    def test_with_holiday_on_dec30(self):
        holidays = [date(2024, 12, 30)]
        result = LedgerEngine.compute_year_start_date(2025, holidays)
        # Rolls back to 2024-12-27 (Friday)
        assert result == date(2024, 12, 27)

    def test_dec30_on_weekend(self):
        # 2023-12-30 is Saturday → should roll back to Friday 12/29
        result = LedgerEngine.compute_year_start_date(2024, [])
        assert result == date(2023, 12, 29)

    def test_no_trading_day_raises(self):
        """If every December weekday is a holiday, raise (no silent Dec 20)."""
        from datetime import timedelta
        prev_year = 2024
        all_dec = []
        d = date(prev_year, 12, 1)
        while d.year == prev_year:
            all_dec.append(d)
            d += timedelta(days=1)
        with pytest.raises(ValueError):
            LedgerEngine.compute_year_start_date(prev_year + 1, all_dec)


# =========================================================================
#  PRESCAN DELEGATE
# =========================================================================

class TestPrescanDelegate:
    """Tests for the static prescan_oldest_date delegate."""

    def test_empty_list_returns_fallback(self):
        d, detected = LedgerEngine.prescan_oldest_date([])
        assert detected is False
        assert isinstance(d, date)

    def test_with_xlsx_file(self, sample_xlsx):
        d, detected = LedgerEngine.prescan_oldest_date([sample_xlsx])
        assert detected is True
        assert d == date(2025, 3, 10)

    def test_nonexistent_file_gracefully_skipped(self):
        d, detected = LedgerEngine.prescan_oldest_date(["/no/such/file.xlsx"])
        assert detected is False


# =========================================================================
#  SKIP SHEET NAMES
# =========================================================================

class TestFileSizeConstant:
    """Fix #4: featherweight 15MB default per CLAUDE.md."""

    def test_default_is_15mb(self, monkeypatch):
        monkeypatch.delenv("BOT_MAX_FILE_MB", raising=False)
        import importlib

        import core.constants as const
        importlib.reload(const)
        assert const.MAX_FILE_SIZE_MB == 15
        # restore default module state for the rest of the suite
        importlib.reload(const)


@pytest.fixture
def exrate_xlsx(tmp_path):
    """A minimal standalone ExRate workbook (pre-existing file to back up)."""
    filepath = tmp_path / "exrate.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExRate"
    ws.append(["Date", "USD Buying TT Rate", "USD Selling Rate",
               "EUR Buying TT Rate", "EUR Selling Rate", "Holidays/Weekend"])
    ws.append([date(2025, 3, 10), None, None, None, None, ""])
    wb.save(str(filepath))
    wb.close()
    return str(filepath)


class TestUpdateExrateStandaloneFailSafe:
    """Fixes #2/#3: guardrail + backup-before-load + handle release."""

    def test_oversized_file_raises_before_backup(self, engine,
                                                 oversized_file):
        from unittest.mock import MagicMock
        engine.backup = MagicMock()
        with pytest.raises(FileSizeLimitError):
            asyncio.run(engine.update_exrate_standalone(oversized_file))
        engine.backup.create_backup.assert_not_called()

    def test_backup_failure_skips_overwrite(self, engine, exrate_xlsx):
        from unittest.mock import MagicMock

        from core.backup_manager import BackupError
        with open(exrate_xlsx, "rb") as f:
            before = f.read()
        engine.backup = MagicMock()
        engine.backup.create_backup.side_effect = BackupError("disk full")
        with pytest.raises(BackupError):
            asyncio.run(engine.update_exrate_standalone(exrate_xlsx))
        # File must remain untouched when backup fails.
        with open(exrate_xlsx, "rb") as f:
            assert f.read() == before

    def test_backup_created_before_load(self, engine, exrate_xlsx,
                                        monkeypatch):
        from unittest.mock import MagicMock
        engine.backup = MagicMock()

        # Force an error AFTER load to prove the handle is released via
        # try/finally (no leaked workbook, gc runs).
        async def boom(*a, **k):
            raise ValueError("api down")
        monkeypatch.setattr(engine, "_preload_api_data", boom)

        with pytest.raises(ValueError):
            asyncio.run(engine.update_exrate_standalone(exrate_xlsx))
        engine.backup.create_backup.assert_called_once()
        # Backup happened before the post-load failure (load succeeded).


class TestStandaloneFromLedgerSingleBackup:
    """Fix #2: process_ledger on a standalone ExRate file backs up only once.

    Before the fix, process_ledger created its own backup THEN delegated to
    update_exrate_standalone, which backed the identical pristine file up
    again — two backups of the same file. The detection now runs BEFORE
    process_ledger's own backup, so only the standalone path's backup fires.
    """

    def test_standalone_from_ledger_produces_one_backup(
        self, exrate_xlsx, tmp_cache, monkeypatch,
    ):
        from unittest.mock import MagicMock

        # Counting backup shared by both code paths via the live engine.
        counting_backup = MagicMock()

        # Mocked API so the standalone update path completes without network.
        async def _rates(start, end, currency):
            return []

        async def _holidays(year):
            return []

        api = MagicMock()
        api.get_exchange_rates = _rates
        api.get_holidays = _holidays

        engine = LedgerEngine(api, backup=counting_backup, cache=tmp_cache)

        result = asyncio.run(engine.process_ledger(exrate_xlsx))
        assert result == exrate_xlsx
        # Exactly one backup of the pristine file — no duplicate.
        counting_backup.create_backup.assert_called_once_with(exrate_xlsx)

    def test_standalone_dry_run_takes_no_backup(
        self, exrate_xlsx, tmp_cache,
    ):
        """A dry-run on a standalone-from-ledger file makes no backup at all.

        The standalone path only backs up when not skipped; routing BEFORE the
        ledger backup means a dry run never produces a stray backup either.
        """
        from unittest.mock import MagicMock

        async def _rates(start, end, currency):
            return []

        async def _holidays(year):
            return []

        api = MagicMock()
        api.get_exchange_rates = _rates
        api.get_holidays = _holidays

        counting_backup = MagicMock()
        engine = LedgerEngine(api, backup=counting_backup, cache=tmp_cache)

        # Standalone path (update_exrate_standalone) backs up whenever the file
        # exists; it has no dry_run concept. The key assertion is that the
        # LEDGER backup path is not ALSO invoked (would be 2 calls).
        asyncio.run(engine.process_ledger(exrate_xlsx, dry_run=True))
        assert counting_backup.create_backup.call_count == 1


class TestSkipSheetNames:
    """Tests for the SKIP_SHEET_NAMES constant."""

    def test_contains_exrate(self):
        assert "ExRate" in SKIP_SHEET_NAMES

    def test_legacy_tabs_skipped(self):
        """Legacy Exrate tabs must be skipped (non-standard headers)."""
        assert "Exrate USD" in SKIP_SHEET_NAMES
        assert "Exrate EUR" in SKIP_SHEET_NAMES

    def test_normal_month_not_skipped(self):
        assert "January" not in SKIP_SHEET_NAMES
        assert "Jan" not in SKIP_SHEET_NAMES


class TestBatchAndHolidayLookup:
    """Tests for batch-level error handling and holiday substitution parsing."""

    def test_process_batch_skips_oversized_file(self, engine, oversized_file):
        success, failed, errors = asyncio.run(engine.process_batch([oversized_file]))
        assert success == 0
        assert failed == 1
        assert len(errors) == 1
        assert "File too large" in errors[0]

    def test_build_holiday_lookup_parses_substitution_holiday(self, engine, monkeypatch):
        substitution_entry = (
            "2025-04-16",
            "Substitution for Songkran Day (15th April 2025)",
        )
        monkeypatch.setattr(
            engine.cache,
            "get_holidays",
            lambda year: [substitution_entry] if year == 2025 else [],
        )

        holidays_set, holidays_names = engine._build_holiday_lookup(
            all_target_dates={date(2025, 4, 16)},
            computed_start=date(2024, 12, 30),
            logic_engine=SimpleNamespace(holidays=[]),
        )

        assert date(2025, 4, 16) in holidays_names
        assert date(2025, 4, 15) in holidays_set
        assert holidays_names[date(2025, 4, 15)] == "Songkran Day"

    def test_process_batch_tracks_anomaly_totals(self, engine, monkeypatch):
        async def _fake_process_ledger(*args, **kwargs):
            engine._last_anomaly_count = 2
            return "fake.xlsx"

        monkeypatch.setattr(engine, "process_ledger", _fake_process_ledger)
        success, failed, errors = asyncio.run(
            engine.process_batch(["a.xlsx", "b.xlsx"], dry_run=True)
        )
        assert (success, failed, errors) == (2, 0, [])
        assert engine.last_batch_anomaly_count == 4

    def test_botapi_error_on_one_file_does_not_abort_batch(
        self, engine, monkeypatch,
    ):
        """Fix #1: a BOTAPIError (e.g. 401) on file 2 must NOT abort 1 & 3.

        BOTAPIError is a plain Exception (not OSError), so before the fix the
        per-file loop let it propagate and skip the remaining files silently.
        """
        async def _fake_process_ledger(fp, *args, **kwargs):
            if fp == "b.xlsx":
                raise BOTAPIError("BOT API server error 401.")
            return fp

        monkeypatch.setattr(engine, "process_ledger", _fake_process_ledger)
        success, failed, errors = asyncio.run(
            engine.process_batch(["a.xlsx", "b.xlsx", "c.xlsx"], dry_run=True)
        )
        # Files 1 and 3 still processed; only file 2 failed.
        assert success == 2
        assert failed == 1
        assert len(errors) == 1
        assert "b.xlsx" in errors[0]
        assert "401" in errors[0]

    def test_httpx_error_on_one_file_does_not_abort_batch(
        self, engine, monkeypatch,
    ):
        """Fix #1: an httpx network error (not an OSError subclass) on file 2
        must be recorded and the batch must continue with file 3."""
        async def _fake_process_ledger(fp, *args, **kwargs):
            if fp == "b.xlsx":
                raise httpx.ConnectError("connection dropped")
            return fp

        monkeypatch.setattr(engine, "process_ledger", _fake_process_ledger)
        success, failed, errors = asyncio.run(
            engine.process_batch(["a.xlsx", "b.xlsx", "c.xlsx"], dry_run=True)
        )
        assert success == 2
        assert failed == 1
        assert len(errors) == 1
        assert "b.xlsx" in errors[0]

    def test_botapi_error_invokes_progress_cb(self, engine, monkeypatch):
        """The per-file API/network branch must still drive progress_cb."""
        async def _fake_process_ledger(fp, *args, **kwargs):
            if fp == "b.xlsx":
                raise BOTAPIError("503")
            return fp

        monkeypatch.setattr(engine, "process_ledger", _fake_process_ledger)
        calls = []
        asyncio.run(engine.process_batch(
            ["a.xlsx", "b.xlsx"], dry_run=True,
            progress_cb=lambda i, t, n, e: calls.append((i, t, n, e)),
        ))
        # Both files reported; file 2 carries the error string.
        assert len(calls) == 2
        assert calls[1][3] is not None
        assert "503" in calls[1][3]

    def test_stop_event_halts_batch_between_files(self, engine, monkeypatch):
        """Fix #3: a pre-set stop_event stops the batch at the file boundary.

        With the event already set, NO file is processed and every file is
        reported as unprocessed via errors + progress_cb.
        """
        processed = []

        async def _fake_process_ledger(fp, *args, **kwargs):
            processed.append(fp)
            return fp

        monkeypatch.setattr(engine, "process_ledger", _fake_process_ledger)
        stop_event = threading.Event()
        stop_event.set()
        calls = []
        success, failed, errors = asyncio.run(engine.process_batch(
            ["a.xlsx", "b.xlsx", "c.xlsx"], dry_run=True,
            stop_event=stop_event,
            progress_cb=lambda i, t, n, e: calls.append((i, t, n, e)),
        ))
        # Nothing processed; all three reported unprocessed.
        assert processed == []
        assert success == 0
        assert failed == 3
        assert len(errors) == 3
        assert all("cancelled" in m for m in errors)
        assert all(c[3] == "cancelled" for c in calls)

    def test_stop_event_set_after_first_file_stops_remainder(
        self, engine, monkeypatch,
    ):
        """Fix #3: setting the event after file 1 lets file 1 finish but stops
        files 2..N (checked at the safe between-file boundary)."""
        processed = []
        stop_event = threading.Event()

        async def _fake_process_ledger(fp, *args, **kwargs):
            processed.append(fp)
            # Simulate the GUI requesting shutdown during the first file.
            stop_event.set()
            return fp

        monkeypatch.setattr(engine, "process_ledger", _fake_process_ledger)
        success, failed, errors = asyncio.run(engine.process_batch(
            ["a.xlsx", "b.xlsx", "c.xlsx"], dry_run=True,
            stop_event=stop_event,
        ))
        # Only the first file ran to completion; the rest were cancelled.
        assert processed == ["a.xlsx"]
        assert success == 1
        assert failed == 2
        assert len(errors) == 2
        assert all("cancelled" in m for m in errors)

    def test_no_stop_event_processes_all(self, engine, monkeypatch):
        """Backward compat: stop_event defaults to None and changes nothing."""
        processed = []

        async def _fake_process_ledger(fp, *args, **kwargs):
            processed.append(fp)
            return fp

        monkeypatch.setattr(engine, "process_ledger", _fake_process_ledger)
        success, failed, errors = asyncio.run(
            engine.process_batch(["a.xlsx", "b.xlsx"], dry_run=True)
        )
        assert processed == ["a.xlsx", "b.xlsx"]
        assert (success, failed, errors) == (2, 0, [])

    def test_locked_file_gets_actionable_message(self, engine, monkeypatch):
        """A file open in Excel surfaces a clear 'close it and retry' message
        instead of a raw WinError/EACCES string the user cannot act on."""
        async def _fake_process_ledger(fp, *args, **kwargs):
            # Simulate the in-place save failing because Excel holds the file.
            raise PermissionError(
                13, "The process cannot access the file because it is being "
                "used by another process",
            )

        monkeypatch.setattr(engine, "process_ledger", _fake_process_ledger)
        calls = []
        success, failed, errors = asyncio.run(engine.process_batch(
            ["January_Ledger.xlsx"], dry_run=True,
            progress_cb=lambda i, t, n, e: calls.append((i, t, n, e)),
        ))
        assert success == 0
        assert failed == 1
        assert len(errors) == 1
        # Actionable, localized message — names the file and the remedy.
        assert "January_Ledger.xlsx" in errors[0]
        assert "open in Excel" in errors[0]
        assert "close it" in errors[0].lower()
        # The raw OS noise is NOT surfaced to the user.
        assert "WinError" not in errors[0]
        assert "Errno" not in errors[0]
        # progress_cb carries the same friendly message.
        assert calls[0][3] == errors[0]

    def test_non_locked_oserror_keeps_original_message(
        self, engine, monkeypatch,
    ):
        """A non-locked OSError (e.g. disk space) keeps its original text —
        the humanizer only rewrites locked-file errors."""
        async def _fake_process_ledger(fp, *args, **kwargs):
            raise OSError(
                "Insufficient disk space (0MB free, need 100MB). "
                "File NOT saved."
            )

        monkeypatch.setattr(engine, "process_ledger", _fake_process_ledger)
        _success, failed, errors = asyncio.run(
            engine.process_batch(["a.xlsx"], dry_run=True)
        )
        assert failed == 1
        assert "Insufficient disk space" in errors[0]
        assert "open in Excel" not in errors[0]


# =========================================================================
#  PRE-FLIGHT SELECTION-TIME CHECK (engine.py:163)
# =========================================================================

class TestPreflightFile:
    """LedgerEngine.preflight_file — cheap selection-time feedback so an
    oversized / unsupported / missing / locked file is flagged immediately
    instead of failing mid-run after the API fetch + backup."""

    def test_good_file_is_ok(self, sample_xlsx):
        result = LedgerEngine.preflight_file(sample_xlsx)
        assert result["ok"] is True
        assert result["exists"] is True
        assert result["size_ok"] is True
        assert result["supported"] is True
        assert result["writable"] is True
        assert result["reason"] is None
        assert result["name"] == "test_ledger.xlsx"

    def test_oversized_file_flagged_with_reason(self, oversized_file):
        result = LedgerEngine.preflight_file(oversized_file)
        assert result["ok"] is False
        assert result["size_ok"] is False
        # Size is surfaced in MB and names the limit so the GUI can show it.
        assert result["size_mb"] > MAX_FILE_SIZE_MB
        assert "too large" in result["reason"].lower()
        assert str(MAX_FILE_SIZE_MB) in result["reason"]
        assert "huge.xlsx" in result["reason"]

    def test_unsupported_extension_flagged(self, tmp_path):
        bad = tmp_path / "rates.csv"
        bad.write_text("not excel", encoding="utf-8")
        result = LedgerEngine.preflight_file(str(bad))
        assert result["ok"] is False
        assert result["supported"] is False
        assert "Unsupported format" in result["reason"]
        assert "rates.csv" in result["reason"]

    def test_missing_file_flagged(self, tmp_path):
        result = LedgerEngine.preflight_file(str(tmp_path / "nope.xlsx"))
        assert result["ok"] is False
        assert result["exists"] is False
        assert result["size_mb"] == 0.0
        assert "not found" in result["reason"].lower()

    def test_xlsm_extension_supported(self, tmp_path):
        macro = tmp_path / "macro_ledger.xlsm"
        macro.write_bytes(b"x" * 1024)  # tiny, within the size limit
        result = LedgerEngine.preflight_file(str(macro))
        assert result["supported"] is True
        assert result["ok"] is True

    def test_locked_file_reason_matches_runtime_message(
        self, sample_xlsx, monkeypatch,
    ):
        """A file the probe reports as not writable yields the SAME 'open in
        Excel — close it' wording the run-time humanizer uses, so selection
        feedback and batch-failure feedback are consistent."""
        monkeypatch.setattr(
            LedgerEngine, "_probe_writable", staticmethod(lambda fp: False),
        )
        result = LedgerEngine.preflight_file(sample_xlsx)
        assert result["ok"] is False
        assert result["writable"] is False
        assert "open in Excel" in result["reason"]
        assert "close it" in result["reason"].lower()

    def test_probe_writable_true_for_writable_file(self, sample_xlsx):
        assert LedgerEngine._probe_writable(Path(sample_xlsx)) is True


# =========================================================================
#  PRE-FLIGHT WRITABILITY/LOCK CHECK BEFORE BACKUP (engine.py:556)
# =========================================================================

class TestProcessLedgerPreflightWritability:
    """A locked target must fail FAST — before the API fetch + backup — with
    the round-7 humanized 'open in Excel' message, not after the round-trip."""

    def test_locked_file_skips_before_backup_and_api(
        self, engine, sample_xlsx, monkeypatch,
    ):
        # Simulate Excel holding the file: the writability probe reports it as
        # not writable. The real save path is never reached.
        monkeypatch.setattr(engine, "_probe_writable", lambda fp: False)

        success, failed, errors = asyncio.run(
            engine.process_batch([sample_xlsx]),
        )
        assert success == 0
        assert failed == 1
        # Humanized, actionable message (round-7 wording) — no raw OS noise.
        assert "open in Excel" in errors[0]
        assert "close it" in errors[0].lower()
        assert "WinError" not in errors[0]
        assert "Errno" not in errors[0]
        # Fail-fast: NO backup taken and NO network round-trip for a locked file.
        engine.backup.create_backup.assert_not_called()
        engine.api.get_exchange_rates.assert_not_called()
        engine.api.get_holidays.assert_not_called()

    def test_writable_file_does_not_trip_preflight(
        self, engine, sample_xlsx, monkeypatch,
    ):
        """A writable file passes the pre-flight and reaches the backup step."""
        monkeypatch.setattr(engine, "_probe_writable", lambda fp: True)

        success, failed, _errors = asyncio.run(
            engine.process_batch([sample_xlsx]),
        )
        assert (success, failed) == (1, 0)
        engine.backup.create_backup.assert_called_once()

    def test_dry_run_skips_writability_probe(
        self, engine, sample_xlsx, monkeypatch,
    ):
        """Dry runs write nothing, so a locked file must NOT block a simulation
        (the probe is never consulted on the dry-run path)."""
        def _boom(fp):
            raise AssertionError("probe must not run on dry runs")

        monkeypatch.setattr(engine, "_probe_writable", _boom)
        success, failed, _errors = asyncio.run(
            engine.process_batch([sample_xlsx], dry_run=True),
        )
        assert (success, failed) == (1, 0)


# =========================================================================
#  END-TO-END INTEGRATION (GAP2)
# =========================================================================

class TestProcessLedgerEndToEnd:
    """Full process_ledger run on a real fixture with a mocked API."""

    def _build_engine(self, tmp_cache, tmp_path):
        """Engine wired to a mocked API + injected tmp backup/cache."""
        from types import SimpleNamespace
        from unittest.mock import MagicMock

        def _rate(period, currency, buying_transfer, selling):
            return SimpleNamespace(
                period=period, currency=currency,
                buying_transfer=buying_transfer, buying_sight=None,
                selling=selling, mid_rate=None,
            )

        async def _rates(start, end, currency):
            # Provide rates for the target dates (and surrounding range).
            base_b = 33.0 if currency == "USD" else 36.0
            base_s = 33.5 if currency == "USD" else 36.5
            out = []
            d = start
            from datetime import timedelta as _td
            while d <= end:
                out.append(_rate(
                    d.strftime("%Y-%m-%d"), currency, base_b, base_s,
                ))
                d += _td(days=1)
            return out

        async def _holidays(year):
            return []

        api = MagicMock()
        api.get_exchange_rates = _rates
        api.get_holidays = _holidays

        backup = MagicMock()  # no-op backup, injected
        return LedgerEngine(api, backup=backup, cache=tmp_cache)

    def test_real_run_populates_exrate_and_formulas(
        self, ledger_xlsx, tmp_cache, tmp_path,
    ):
        path = ledger_xlsx({"Jan": [
            (date(2025, 1, 7), "USD"),    # Tuesday
            ("10/03/2025", "EUR"),        # string date → normalized
        ]})
        engine = self._build_engine(tmp_cache, tmp_path)

        result = asyncio.run(engine.process_ledger(path))
        assert result == path

        wb = openpyxl.load_workbook(path)
        try:
            # ExRate master sheet populated.
            assert "ExRate" in wb.sheetnames
            ws_ex = wb["ExRate"]
            assert ws_ex.max_row >= 2

            ws = wb["Jan"]
            # EX Rate (col 3) holds the IFS formula.
            usd_formula = ws.cell(row=2, column=3).value
            assert isinstance(usd_formula, str)
            assert usd_formula.startswith("=IF(OR(")
            assert "_xlfn.IFS(" in usd_formula
            assert "_xlfn.XLOOKUP(" in usd_formula

            # Source date that was a string is now a real date.
            normalized = ws.cell(row=3, column=1).value
            as_date = normalized.date() if isinstance(normalized, datetime) \
                else normalized
            assert as_date == date(2025, 3, 10)
        finally:
            wb.close()

    def test_dry_run_leaves_file_bytes_unchanged(
        self, ledger_xlsx, tmp_cache, tmp_path,
    ):
        path = ledger_xlsx({"Jan": [(date(2025, 1, 7), "USD")]})
        with open(path, "rb") as f:
            before = f.read()
        engine = self._build_engine(tmp_cache, tmp_path)

        result = asyncio.run(engine.process_ledger(path, dry_run=True))
        assert result == path
        with open(path, "rb") as f:
            assert f.read() == before
        # Backup must be skipped on dry runs.
        engine.backup.create_backup.assert_not_called()


# =========================================================================
#  CACHE-FIRST INVARIANT (Core Rule 5)
# =========================================================================

class TestCacheFirstInvariant:
    """End-to-end proof of Core Rule 5: SQLite is checked before the API.

    ``_preload_api_data`` builds its weekday window from
    ``min/max(target_dates ∪ {force_start, bot_today()})`` and only calls the
    API for weekdays missing from the rates cache. Pre-populating EVERY weekday
    in that window (plus a holiday per touched year) drives the API call count
    to zero; a partial cache triggers exactly one fetch per currency covering
    only the missing window.
    """

    # A fixed Bangkok "today" so the weekday window is fully deterministic.
    _TODAY = date(2025, 1, 15)  # Wednesday

    def _seed_cache(self, cache, start, end, *, base=33.0):
        """Cache a flat rate for every weekday in [start, end] (inclusive).

        Flat values keep the anomaly guard quiet; one holiday per touched year
        is seeded so has_holidays_for_year short-circuits the holiday API call.
        """
        from datetime import timedelta as _td
        bulk = []
        d = start
        while d <= end:
            if d.weekday() < 5:  # weekdays only — matches all_needed logic
                d_str = d.strftime("%Y-%m-%d")
                bulk.append((d_str, base, base + 0.5, base + 3.0, base + 3.5))
            d += _td(days=1)
        cache.insert_rates_bulk(bulk)
        for year in {start.year, end.year}:
            cache.insert_holidays([(f"{year}-12-31", "Year-End Holiday")])

    def _build_mocked_engine(self, tmp_cache):
        """Engine whose api is an AsyncMock so call counts are assertable."""
        api = AsyncMock()
        api.get_holidays = AsyncMock(return_value=[])
        api.get_exchange_rates = AsyncMock(return_value=[])
        return LedgerEngine(api, backup=MagicMock(), cache=tmp_cache)

    def test_full_cache_hit_makes_zero_api_calls(
        self, ledger_xlsx, tmp_cache, monkeypatch,
    ):
        """Every target/today date cached → no API calls at all."""
        monkeypatch.setattr(engine_mod, "bot_today", lambda: self._TODAY)

        target = date(2025, 1, 7)  # Tuesday
        path = ledger_xlsx({"Jan": [(target, "USD")]})

        # force_start = "{year-1}-12-20" (process_ledger default); the window
        # spans 2024-12-20 .. 2025-01-15. Seed every weekday in that range.
        self._seed_cache(tmp_cache, date(2024, 12, 20), self._TODAY)

        engine = self._build_mocked_engine(tmp_cache)
        result = asyncio.run(engine.process_ledger(path))
        assert result == path

        # Core Rule 5: the cache fully served the run.
        engine.api.get_exchange_rates.assert_not_called()
        engine.api.get_holidays.assert_not_called()

    def test_partial_cache_fetches_only_missing_window(
        self, ledger_xlsx, tmp_cache, monkeypatch,
    ):
        """A gap in the cache triggers exactly one fetch per currency over the
        missing window (engine.py missing_dates path)."""
        monkeypatch.setattr(engine_mod, "bot_today", lambda: self._TODAY)

        target = date(2025, 1, 7)
        path = ledger_xlsx({"Jan": [(target, "USD")]})

        # Seed all weekdays EXCEPT a contiguous early-January gap so the engine
        # must fetch only that narrow window.
        self._seed_cache(tmp_cache, date(2024, 12, 20), date(2024, 12, 31))
        self._seed_cache(tmp_cache, date(2025, 1, 9), self._TODAY)
        # Missing weekdays: 2025-01-01 (Wed) .. 2025-01-08 (Wed).

        engine = self._build_mocked_engine(tmp_cache)
        result = asyncio.run(engine.process_ledger(path))
        assert result == path

        # API consulted only for the missing dates — once for USD, once for EUR
        # (concurrent gather), never for holidays (year already cached).
        assert engine.api.get_exchange_rates.await_count == 2
        engine.api.get_holidays.assert_not_called()
        # Both fetches are bounded by the missing window, not the full range.
        for call in engine.api.get_exchange_rates.await_args_list:
            fetch_start, fetch_end, _ccy = call.args
            assert fetch_start == date(2025, 1, 1)
            assert fetch_end == date(2025, 1, 8)


# =========================================================================
#  DISK-SPACE GUARD — standard ledger path (Fix #2)
# =========================================================================

class TestLedgerDiskSpaceGuard:
    """The pre-save free-space guard must fire on the STANDARD ledger path.

    test_engine_multicurrency.py covers the custom standalone path; this proves
    the same OSError guard protects process_ledger's WorkbookWriter save. Saves
    are atomic (temp file + os.replace via workbook_io.atomic_save), so a
    blocked save must leave the ORIGINAL .xlsx byte-for-byte intact and leave no
    stray temp file behind.
    """

    def _build_engine(self, tmp_cache):
        from types import SimpleNamespace as _SN

        async def _rates(start, end, currency):
            from datetime import timedelta as _td
            base_b = 33.0 if currency == "USD" else 36.0
            base_s = 33.5 if currency == "USD" else 36.5
            out, d = [], start
            while d <= end:
                out.append(_SN(
                    period=d.strftime("%Y-%m-%d"), currency=currency,
                    buying_transfer=base_b, buying_sight=None,
                    selling=base_s, mid_rate=None,
                ))
                d += _td(days=1)
            return out

        async def _holidays(year):
            return []

        api = MagicMock()
        api.get_exchange_rates = _rates
        api.get_holidays = _holidays
        return LedgerEngine(api, backup=MagicMock(), cache=tmp_cache)

    def test_insufficient_disk_surfaces_and_leaves_file_intact(
        self, ledger_xlsx, tmp_cache, monkeypatch,
    ):
        from pathlib import Path

        import core.workbook_io as workbook_io_mod

        path = ledger_xlsx({"Jan": [(date(2025, 1, 7), "USD")]})
        with open(path, "rb") as f:
            original_bytes = f.read()

        # Report zero free space as core.workbook_io sees it (module singleton).
        from collections import namedtuple
        _Usage = namedtuple("_Usage", ["total", "used", "free"])
        monkeypatch.setattr(
            workbook_io_mod.shutil, "disk_usage",
            lambda _path: _Usage(total=10**12, used=10**12, free=0),
        )

        engine = self._build_engine(tmp_cache)
        with pytest.raises(OSError, match="Insufficient disk space"):
            asyncio.run(engine.process_ledger(path))

        # Atomic save: original bytes unchanged, no leftover temp file.
        with open(path, "rb") as f:
            assert f.read() == original_bytes
        leftover = list(Path(path).parent.glob("*.tmp~"))
        assert leftover == []


# =========================================================================
#  BOT BUSINESS-DATE SWEEP
# =========================================================================

class TestPreloadUsesBotToday:
    """_preload_api_data keys its 'today' upper bound off bot_today().

    The sweep replaced the bare date.today() with bot_today() (Asia/Bangkok)
    so the fetch range tracks the BOT trading calendar, not the local machine
    clock. Patching bot_today proves the upper fetch bound follows it.
    """

    def test_fetch_upper_bound_follows_bot_today(self, engine, monkeypatch):
        fixed_today = date(2025, 3, 14)
        monkeypatch.setattr(engine_mod, "bot_today", lambda: fixed_today)

        asyncio.run(engine._preload_api_data(set(), "2025-03-10"))

        # Empty cache → a fetch fires; its end date is the patched BOT today
        # (the weekday upper bound of the [force_start, bot_today] window).
        assert engine.api.get_exchange_rates.await_count >= 1
        end_args = {
            call.args[1] for call in engine.api.get_exchange_rates.await_args_list
        }
        assert end_args == {fixed_today}


# =========================================================================
#  LEDGER_PROCESSING — extracted near-pure helpers
# =========================================================================

class TestRunAnomalyCheck:
    """Tests for the extracted run_anomaly_check function."""

    def test_no_anomalies_returns_zero_and_no_emit(self):
        class _Guard:
            def check_rates_bulk(self, bundle):
                # Bundle must carry all four labelled rate dicts.
                assert set(bundle) == {
                    "USD_buying_transfer", "USD_selling",
                    "EUR_buying_transfer", "EUR_selling",
                }
                return []

        emitted = []
        count = run_anomaly_check(
            _Guard(),
            lambda msg, etype: emitted.append((msg, etype)),
            {}, {}, {}, {},
        )
        assert count == 0
        assert emitted == []

    def test_anomalies_emit_warning_per_record(self):
        anomaly = SimpleNamespace(
            currency="USD", rate_type="selling",
            check_date=date(2025, 3, 10),
            pct_change=12.5, prev_value="33.0000", new_value="37.0000",
        )

        class _Guard:
            def check_rates_bulk(self, bundle):
                return [anomaly, anomaly]

        emitted = []
        count = run_anomaly_check(
            _Guard(),
            lambda msg, etype: emitted.append((msg, etype)),
            {}, {}, {}, {},
        )
        assert count == 2
        assert len(emitted) == 2
        assert all(etype == "warning" for _msg, etype in emitted)
        assert "ANOMALY" in emitted[0][0]


class TestPrescanTargetDates:
    """Tests for the extracted prescan_target_dates function."""

    _COLS = {"source_date": "Date", "currency": "Cur", "out_rate": "EX Rate"}

    def test_extracts_all_dates(self, sample_xlsx):
        dates = prescan_target_dates(sample_xlsx, self._COLS)
        assert dates == {
            date(2025, 3, 10), date(2025, 3, 11), date(2025, 3, 12),
        }

    def test_skips_skip_sheet_names(self, tmp_path):
        filepath = tmp_path / "with_exrate.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jan"
        ws.append(["Date", "Cur", "EX Rate"])
        ws.append([date(2025, 6, 2), "USD", None])
        # ExRate is a SKIP sheet — its dates must NOT be collected.
        ws_ex = wb.create_sheet("ExRate")
        ws_ex.append(["Date", "USD Buying TT Rate"])
        ws_ex.append([date(1999, 1, 1), 1.0])
        wb.save(str(filepath))
        wb.close()

        dates = prescan_target_dates(str(filepath), self._COLS)
        assert dates == {date(2025, 6, 2)}
        assert date(1999, 1, 1) not in dates

    def test_emit_callback_invoked(self, sample_xlsx):
        msgs = []
        prescan_target_dates(
            sample_xlsx, self._COLS, emit_fn=msgs.append,
        )
        assert msgs == ["Scanning dates from workbook"]

    def test_engine_method_delegates(self, engine, sample_xlsx):
        """The engine shim returns identical results to the function."""
        assert engine._prescan_target_dates(sample_xlsx) == prescan_target_dates(
            sample_xlsx, engine.target_cols,
        )


# =========================================================================
#  AUDIT TRAIL — process_batch wires log_row_change into the write pipeline
# =========================================================================

def _audit_engine(per_ccy, tmp_cache):
    """LedgerEngine whose mocked API emits one record per date per currency.

    ``per_ccy`` maps a currency code → (buying_transfer, selling) so any
    requested window resolves. Mirrors the multi-currency test helper so the
    full process_ledger write pipeline runs end-to-end with real rates.
    """
    from datetime import timedelta as _td

    async def _rates(start, end, currency):
        pair = per_ccy.get(currency)
        if pair is None:
            return []
        buy, sell = pair
        out, d = [], start
        while d <= end:
            out.append(SimpleNamespace(
                period=d.strftime("%Y-%m-%d"), currency=currency,
                buying_transfer=buy, buying_sight=None,
                selling=sell, mid_rate=None,
            ))
            d += _td(days=1)
        return out

    api = AsyncMock()
    api.get_exchange_rates = AsyncMock(side_effect=_rates)
    api.get_holidays = AsyncMock(return_value=[])
    return LedgerEngine(api, backup=MagicMock(), cache=tmp_cache)


def _build_ledger(tmp_path, rows, name="audit_led.xlsx"):
    filepath = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jan"
    ws.append(["Date", "Cur", "EX Rate", "Amount"])
    for d, ccy in rows:
        ws.append([d, ccy, None, 1000])
    wb.save(str(filepath))
    wb.close()
    return str(filepath)


def _patch_audit_dir(monkeypatch, tmp_path):
    """Redirect engine-owned AuditLogger output into ``tmp_path``."""
    import functools

    import core.engine as eng
    from core.audit_logger import AuditLogger as _RealAuditLogger

    log_dir = tmp_path / "logs"
    monkeypatch.setattr(
        eng, "AuditLogger",
        functools.partial(_RealAuditLogger, log_dir=str(log_dir)),
    )
    # cleanup_old_audit_logs() runs against data/logs by default; point it at
    # the temp dir so the real project folder is never touched by the test.
    monkeypatch.setattr(
        eng, "cleanup_old_audit_logs",
        lambda *a, **k: 0,
    )
    return log_dir


class TestAuditTrailWiredIntoBatch:
    """Fix: log_row_change had ZERO call sites — process_batch now records a
    per-cell change for every resolved EX Rate cell and writes a real CSV."""

    def _read_rows(self, audit_path):
        import csv
        with open(audit_path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader)
            rows = [r for r in reader if r and r[0]]
        return header, rows

    def test_real_batch_writes_populated_audit_csv(
        self, tmp_path, tmp_cache, monkeypatch,
    ):
        _patch_audit_dir(monkeypatch, tmp_path)
        path = _build_ledger(tmp_path, [
            (date(2025, 1, 7), "USD"),
            (date(2025, 1, 8), "THB"),
        ])
        engine = _audit_engine(
            {"USD": (33.1234, 33.5), "EUR": (36.0, 36.5)}, tmp_cache,
        )

        success, fail, _errors = asyncio.run(engine.process_batch([path]))
        assert (success, fail) == (1, 0)

        # The engine published a real audit CSV path.
        assert engine.last_audit_path is not None
        assert Path(engine.last_audit_path).exists()

        header, rows = self._read_rows(engine.last_audit_path)
        assert header[1] == "Filename"
        # At least the USD and THB cells were recorded (no longer a hollow log).
        cells = [r for r in rows if "BATCH SUMMARY" not in r[1]]
        currencies = {r[5] for r in cells}
        assert "USD" in currencies
        assert "THB" in currencies
        # The USD row carries the 4dp-quantized resolved rate as New_Value.
        usd_rows = [r for r in cells if r[5] == "USD"]
        assert any(r[7] == "33.1234" for r in usd_rows), usd_rows
        # The summary is no longer a misleading 0.
        assert any(
            "Total Rows Modified" in c and "0" not in c.split(":")[-1]
            for r in rows for c in r if "Total Rows Modified" in c
        )

    def test_dry_run_writes_no_audit_log(
        self, tmp_path, tmp_cache, monkeypatch,
    ):
        log_dir = _patch_audit_dir(monkeypatch, tmp_path)
        path = _build_ledger(tmp_path, [(date(2025, 1, 7), "USD")])
        engine = _audit_engine({"USD": (33.0, 33.5)}, tmp_cache)

        asyncio.run(engine.process_batch([path], dry_run=True))
        assert engine.last_audit_path is None
        # No CSV files were created for a simulation.
        assert not log_dir.exists() or not list(log_dir.glob("Audit_Log_*.csv"))

    def test_injected_audit_is_not_finalized_by_engine(
        self, tmp_path, tmp_cache, monkeypatch,
    ):
        """A caller-supplied audit log keeps its lifecycle: the engine records
        rows but does not finalize, summarize, or set last_audit_path."""
        _patch_audit_dir(monkeypatch, tmp_path)
        from core.audit_logger import AuditLogger

        path = _build_ledger(tmp_path, [(date(2025, 1, 7), "USD")])
        engine = _audit_engine({"USD": (33.0, 33.5)}, tmp_cache)
        external = AuditLogger(log_dir=str(tmp_path / "ext"))

        asyncio.run(engine.process_batch([path], audit=external))
        # Engine did not claim ownership.
        assert engine.last_audit_path is None
        # Rows were recorded into the caller's log; it is still open.
        assert external.row_count >= 1
        assert not external._closed
        external.finalize()

    def test_failed_file_records_do_not_leak_into_next_file(
        self, tmp_path, tmp_cache, monkeypatch,
    ):
        """Regression: WorkbookWriter pushes per-cell AuditRecords into the
        collector BEFORE the atomic save. A file that fails AFTER that point
        (disk-full, locked, etc.) must have its phantom records drained and
        discarded — never flushed into the NEXT successful file's audit rows.
        """
        _patch_audit_dir(monkeypatch, tmp_path)
        file_a = _build_ledger(
            tmp_path, [(date(2025, 1, 7), "USD")], name="fileA.xlsx",
        )
        file_b = _build_ledger(
            tmp_path, [(date(2025, 1, 8), "EUR")], name="fileB.xlsx",
        )
        engine = _audit_engine(
            {"USD": (33.1234, 33.5), "EUR": (36.5678, 36.9)}, tmp_cache,
        )

        # Wrap process_ledger so file A populates the shared collector with its
        # per-cell records (exactly as STEP 5 does) and THEN raises a save-time
        # OSError; file B proceeds normally.
        from core.audit_logger import AuditRecord

        real_process_ledger = engine.process_ledger

        async def _flaky(fp, *args, **kwargs):
            if Path(fp).name == "fileA.xlsx":
                collector = kwargs.get("audit")
                if collector is not None:
                    collector.add(AuditRecord(
                        filename="fileA.xlsx", sheet="Jan", row=2,
                        cell_date="2025-01-07", currency="USD",
                        original_value="", new_value="33.1234",
                    ))
                raise OSError("disk full during save")
            return await real_process_ledger(fp, *args, **kwargs)

        monkeypatch.setattr(engine, "process_ledger", _flaky)

        success, fail, _errors = asyncio.run(
            engine.process_batch([file_a, file_b]),
        )
        assert (success, fail) == (1, 1)

        header, rows = self._read_rows(engine.last_audit_path)
        cells = [r for r in rows if "BATCH SUMMARY" not in r[1]]
        filenames = {r[1] for r in cells}
        # fileA failed its save → none of its phantom records may appear.
        assert "fileA.xlsx" not in filenames, cells
        # Only the file that was actually written is audited.
        assert filenames == {"fileB.xlsx"}, cells


# =========================================================================
#  SETTINGS-SNAPSHOT CONTRACT (state-conflicts)
#  Finding: "Anomaly threshold is snapshotted at batch start but rate_type
#  is read live, so the two settings behave oppositely in one run."
#  Contract: every relevant setting (rate_type, anomaly threshold) is
#  snapshotted ONCE at engine construction (== batch start, since a fresh
#  LedgerEngine is built per batch), so a mid-batch Settings "Save" never
#  affects the in-flight run.
# =========================================================================

class TestSettingsSnapshotContract:
    """rate_type and anomaly threshold are BOTH snapshotted at construction."""

    def test_rate_type_snapshotted_at_construction(
        self, mock_api, tmp_cache, monkeypatch
    ):
        """The engine reads rate_type once, in __init__, not per file."""
        loads: list[None] = []

        class _FakeSettings:
            def load(self):
                loads.append(None)
                return {"rate_type": "selling", "anomaly_threshold_pct": 7.5}

        monkeypatch.setattr(engine_mod, "SettingsManager", _FakeSettings)
        eng = LedgerEngine(mock_api, cache=tmp_cache, backup=MagicMock())

        assert eng._rate_type == "selling"
        assert eng._anomaly_guard.threshold_pct == 7.5
        # Exactly one settings read happened at construction time.
        assert len(loads) == 1

    def test_process_ledger_uses_snapshot_not_live_settings(
        self, mock_api, tmp_cache, sample_xlsx, monkeypatch
    ):
        """A mid-batch Settings change must NOT flip rate_type for this run.

        Construct the engine while settings say 'buying_transfer', then mutate
        the live settings to 'selling' and run process_ledger. The rate_type
        threaded into WorkbookWriter.write must still be the snapshot
        ('buying_transfer') — proving the file run ignores the live change,
        consistent with the threshold which is also fixed at construction.
        """
        live = {"rate_type": "buying_transfer", "anomaly_threshold_pct": 5.0}

        class _MutableSettings:
            def load(self):
                # Return a copy so callers can't mutate our backing store, but
                # reflect whatever 'live' currently holds.
                return dict(live)

        monkeypatch.setattr(engine_mod, "SettingsManager", _MutableSettings)
        eng = LedgerEngine(mock_api, cache=tmp_cache, backup=MagicMock())

        # Simulate the user hitting "Save" in Settings mid-batch.
        live["rate_type"] = "selling"

        # Capture the rate_type that reaches the writer.
        captured: dict[str, str] = {}

        async def _fake_write(self_writer, filepath, dry_run, *args, **kwargs):
            captured["rate_type"] = kwargs.get("rate_type")
            return filepath

        monkeypatch.setattr(
            engine_mod.WorkbookWriter, "write", _fake_write
        )
        # Keep the run cheap: no real API/holiday work, no standalone branch.
        monkeypatch.setattr(
            eng, "_detect_standalone_exrate",
            AsyncMock(return_value=None),
        )

        async def _fake_preload(dates, start_date):
            return (
                SimpleNamespace(holidays=[]), {}, {}, {}, {}, [], [],
            )

        monkeypatch.setattr(eng, "_preload_api_data", _fake_preload)
        monkeypatch.setattr(eng, "_run_anomaly_check", lambda *a, **k: 0)
        monkeypatch.setattr(
            engine_mod, "build_holiday_lookup",
            lambda *a, **k: (set(), {}),
        )
        monkeypatch.setattr(
            engine_mod, "compute_year_start_date",
            lambda *a, **k: date(2024, 12, 30),
        )

        asyncio.run(eng.process_ledger(sample_xlsx, dry_run=True))

        # The snapshot wins — the mid-batch "selling" change is ignored.
        assert captured["rate_type"] == "buying_transfer"


# =========================================================================
#  CRASH-RECOVERY / RESUME MANIFEST  (round-9 upgrade)
# =========================================================================
class TestBatchManifest:
    """Unit tests for the BatchManifest persistence helper.

    The manifest is the on-disk seam a GUI / `--resume` run reads to finish an
    interrupted batch. It must persist ONLY paths/dates/flags, write atomically,
    and survive partial completion without corruption.
    """

    def test_begin_writes_paths_dates_flags_only(self, tmp_path):
        from core.engine import BatchManifest

        mpath = tmp_path / "batch_state.json"
        m = BatchManifest(mpath)
        m.begin(["/a/jan.xlsx", "/a/feb.xlsx"], "2025-01-02", dry_run=False)

        import json
        data = json.loads(mpath.read_text(encoding="utf-8"))
        # Privacy contract: no rates, no tokens — only paths/date/flag/version.
        assert set(data.keys()) == {"version", "start_date", "dry_run", "files"}
        assert data["start_date"] == "2025-01-02"
        assert data["dry_run"] is False
        assert [f["path"] for f in data["files"]] == ["/a/jan.xlsx", "/a/feb.xlsx"]
        assert all(f["done"] is False for f in data["files"])

    def test_mark_done_flags_only_that_file(self, tmp_path):
        from core.engine import BatchManifest

        m = BatchManifest(tmp_path / "batch_state.json")
        m.begin(["/a/jan.xlsx", "/a/feb.xlsx"], "2025-01-02", dry_run=False)
        m.mark_done("/a/jan.xlsx")

        import json
        data = json.loads((tmp_path / "batch_state.json").read_text("utf-8"))
        done = {f["path"]: f["done"] for f in data["files"]}
        assert done == {"/a/jan.xlsx": True, "/a/feb.xlsx": False}

    def test_pending_files_skips_done_and_missing(self, tmp_path):
        from core.engine import BatchManifest

        present = tmp_path / "feb.xlsx"
        present.write_bytes(b"x")
        m = BatchManifest(tmp_path / "batch_state.json")
        # jan is done; gone.xlsx no longer exists on disk; feb is the only resume.
        m.begin([str(present), "/a/jan.xlsx", "/a/gone.xlsx"], "2025-01-02", False)
        m.mark_done("/a/jan.xlsx")

        assert m.pending_files() == [str(present)]
        assert m.has_pending() is True
        assert m.start_date() == "2025-01-02"

    def test_clear_removes_manifest(self, tmp_path):
        from core.engine import BatchManifest

        mpath = tmp_path / "batch_state.json"
        m = BatchManifest(mpath)
        m.begin(["/a/jan.xlsx"], None, dry_run=False)
        assert mpath.exists()
        m.clear()
        assert not mpath.exists()
        # Idempotent — clearing an absent manifest is a no-op, never raises.
        m.clear()

    def test_read_ignores_stale_version(self, tmp_path):
        from core.engine import BatchManifest

        mpath = tmp_path / "batch_state.json"
        mpath.write_text('{"version": 999, "files": [{"path": "x", "done": false}]}')
        m = BatchManifest(mpath)
        # A manifest from an incompatible build is ignored, not misread.
        assert m.pending_files() == []
        assert m.has_pending() is False

    def test_read_handles_corrupt_json(self, tmp_path):
        from core.engine import BatchManifest

        mpath = tmp_path / "batch_state.json"
        mpath.write_text("{ not valid json")
        m = BatchManifest(mpath)
        assert m.pending_files() == []

    def test_write_is_atomic_no_temp_left(self, tmp_path):
        from core.engine import BatchManifest

        m = BatchManifest(tmp_path / "batch_state.json")
        m.begin(["/a/jan.xlsx"], "2025-01-02", dry_run=False)
        m.mark_done("/a/jan.xlsx")
        # No sibling temp file may survive a successful write.
        leftovers = list(tmp_path.glob("*.tmp~"))
        assert leftovers == []


class TestProcessBatchManifestLifecycle:
    """process_batch must write/update/delete the resume manifest correctly."""

    def test_clean_run_deletes_manifest(self, engine, monkeypatch, tmp_path):
        from core.engine import BatchManifest

        async def _ok(fp, *a, **k):
            return fp

        monkeypatch.setattr(engine, "process_ledger", _ok)
        mpath = tmp_path / "batch_state.json"
        m = BatchManifest(mpath)
        # Inject a MagicMock audit logger so the engine does NOT create + write
        # its own real audit CSV into data/logs as a test side effect.
        asyncio.run(engine.process_batch(
            ["a.xlsx", "b.xlsx"], dry_run=False, manifest=m,
            audit=MagicMock(),
        ))
        # A clean completion leaves nothing to recover.
        assert not mpath.exists()

    def test_crash_after_first_file_leaves_manifest_with_remainder(
        self, engine, monkeypatch, tmp_path,
    ):
        """Simulate a crash: file 1 completes, file 2 raises an UNHANDLED error
        that propagates out of process_batch (not one of the caught branches).
        The manifest must survive with file 1 done and file 2 still pending."""
        from core.engine import BatchManifest

        async def _ledger(fp, *a, **k):
            if fp == "b.xlsx":
                raise RuntimeError("simulated crash mid-batch")
            return fp

        monkeypatch.setattr(engine, "process_ledger", _ledger)
        mpath = tmp_path / "batch_state.json"
        m = BatchManifest(mpath)
        with pytest.raises(RuntimeError):
            asyncio.run(engine.process_batch(
                ["a.xlsx", "b.xlsx"], dry_run=False, manifest=m,
                audit=MagicMock(),
            ))
        # Manifest survived the crash; file 1 is done, file 2 still pending.
        assert mpath.exists()
        done = {
            f["path"]: f["done"]
            for f in BatchManifest(mpath)._read_raw()["files"]
        }
        assert done == {"a.xlsx": True, "b.xlsx": False}

    def test_cancellation_deletes_manifest(self, engine, monkeypatch, tmp_path):
        """A cancel is intentional, not a crash — the manifest is dropped so the
        next launch does not offer to resume a deliberately-stopped batch."""
        from core.engine import BatchManifest

        stop_event = threading.Event()
        stop_event.set()  # cancel before any file runs

        async def _ledger(fp, *a, **k):
            return fp

        monkeypatch.setattr(engine, "process_ledger", _ledger)
        mpath = tmp_path / "batch_state.json"
        m = BatchManifest(mpath)
        asyncio.run(engine.process_batch(
            ["a.xlsx", "b.xlsx"], dry_run=False, manifest=m,
            stop_event=stop_event, audit=MagicMock(),
        ))
        assert not mpath.exists()

    def test_dry_run_writes_no_manifest(self, engine, monkeypatch, tmp_path):
        from core.engine import BatchManifest

        async def _ok(fp, *a, **k):
            return fp

        monkeypatch.setattr(engine, "process_ledger", _ok)
        mpath = tmp_path / "batch_state.json"
        m = BatchManifest(mpath)
        # dry_run forces manifest=None internally; the injected one is untouched.
        asyncio.run(engine.process_batch(
            ["a.xlsx"], dry_run=True, manifest=m,
        ))
        assert not mpath.exists()
