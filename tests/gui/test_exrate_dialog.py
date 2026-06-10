#!/usr/bin/env python3
"""
tests/gui/test_exrate_dialog.py
---------------------------------------------------------------------------
Widget-level tests for gui/panels/exrate_dialog.py (show_exrate_dialog).

show_exrate_dialog is a function-based dialog: it creates a CTkToplevel,
builds all widgets as local variables, and wires inner callbacks.  Because
all dialog state is local to the function we use two techniques:

1. Patch ctk.CTkToplevel.__init__ to (a) capture the instance and
   (b) immediately withdraw it so the window never appears.
2. Patch ctk.CTkToplevel.grab_set to a no-op so the test process is not
   blocked on a modal grab that requires a visible window.

All tests require a display; the tk_root fixture skips them on headless CI.
No real filesystem I/O, network, or threads are started.
"""

import contextlib
from datetime import date
from unittest.mock import patch

import pytest

pytestmark = pytest.mark.gui

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _open_dialog(app):
    """Call show_exrate_dialog and return the captured CTkToplevel instance.

    Patches:
    - CTkToplevel.__init__: captures self, then withdraws the window.
    - CTkToplevel.grab_set: no-op (prevents modal block without focus).
    - CTkToplevel.transient: no-op (avoids WM complaints in tests).
    - CTkToplevel.update_idletasks: no-op (prevents geometry recalc errors).
    """
    import customtkinter as ctk

    from gui.panels.exrate_dialog import show_exrate_dialog

    captured = []
    _orig_init = ctk.CTkToplevel.__init__

    def _patched_init(self, *args, **kwargs):
        _orig_init(self, *args, **kwargs)
        self.withdraw()
        captured.append(self)

    with (
        patch.object(ctk.CTkToplevel, "__init__", _patched_init),
        patch.object(ctk.CTkToplevel, "grab_set", lambda self: None),
        patch.object(ctk.CTkToplevel, "transient", lambda self, *a: None),
        patch.object(ctk.CTkToplevel, "update_idletasks", lambda self: None),
    ):
        show_exrate_dialog(app)

    assert captured, "CTkToplevel was not instantiated by show_exrate_dialog"
    return captured[0]


def _find_button(dialog, text_fragment: str):
    """Return the first CTkButton whose text contains *text_fragment*."""
    import customtkinter as ctk

    for widget in dialog.winfo_children():
        if isinstance(widget, ctk.CTkButton) and text_fragment.lower() in str(widget.cget("text")).lower():
            return widget
        # CTk frames can contain nested widgets
        for child in widget.winfo_children():
            if isinstance(child, ctk.CTkButton) and text_fragment.lower() in str(child.cget("text")).lower():
                return child
    return None


def _collect_checkboxes(dialog):
    """Recursively collect all CTkCheckBox instances from the dialog tree."""
    import customtkinter as ctk

    result = []

    def _walk(w):
        for child in w.winfo_children():
            if isinstance(child, ctk.CTkCheckBox):
                result.append(child)
            _walk(child)

    _walk(dialog)
    return result


def _invoke_binding(widget, sequence):
    """Trigger a widget's key binding deterministically.

    event_generate on a withdrawn, unfocused TopLevel is unreliable, so we
    briefly map + focus the window, generate the event with when='now' (so the
    handler runs synchronously), then re-withdraw. This proves the bound
    callback actually fires without depending on real keyboard focus.
    """
    import contextlib as _ctx

    with _ctx.suppress(Exception):
        widget.deiconify()
        widget.update_idletasks()
        widget.focus_force()
        widget.update()
    widget.event_generate(sequence, when="now")
    with _ctx.suppress(Exception):
        if widget.winfo_exists():
            widget.withdraw()


def _collect_comboboxes(dialog):
    """Recursively collect all CTkComboBox instances from the dialog tree."""
    import customtkinter as ctk

    result = []

    def _walk(w):
        for child in w.winfo_children():
            if isinstance(child, ctk.CTkComboBox):
                result.append(child)
            _walk(child)

    _walk(dialog)
    return result


def _collect_switches(dialog):
    """Recursively collect all CTkSwitch instances from the dialog tree."""
    import customtkinter as ctk

    result = []

    def _walk(w):
        for child in w.winfo_children():
            if isinstance(child, ctk.CTkSwitch):
                result.append(child)
            _walk(child)

    _walk(dialog)
    return result


# ---------------------------------------------------------------------------
# Test classes
# ---------------------------------------------------------------------------


class TestExrateDialogCreation:
    """Dialog is created with the correct title and geometry."""

    def test_dialog_is_ctk_toplevel(self, tk_root):
        import customtkinter as ctk

        dialog = _open_dialog(tk_root)
        assert isinstance(dialog, ctk.CTkToplevel)
        dialog.destroy()

    def test_dialog_title(self, tk_root):
        dialog = _open_dialog(tk_root)
        assert dialog.title() == "Create ExRate File"
        dialog.destroy()

    def test_dialog_geometry_440x680_set(self, tk_root):
        """The source calls dialog.geometry('440x680') at construction.
        On a withdrawn window the WM may not apply it before update_idletasks
        runs, so we verify the call was made rather than the resulting string.
        """
        import customtkinter as ctk

        geometry_calls = []
        orig_geometry = ctk.CTkToplevel.geometry

        def _spy_geometry(self, *args, **kwargs):
            if args:
                geometry_calls.append(args[0])
            return orig_geometry(self, *args, **kwargs)

        with patch.object(ctk.CTkToplevel, "geometry", _spy_geometry):
            dialog = _open_dialog(tk_root)

        assert any("440x680" in str(c) for c in geometry_calls), (
            f"dialog.geometry('440x680...') was never called; calls={geometry_calls}"
        )
        dialog.destroy()


class TestExrateDialogCurrencyCheckboxes:
    """Currency checkboxes are built correctly with the right defaults."""

    def test_nine_currency_checkboxes_exist(self, tk_root):
        from gui.panels.exrate_dialog import EXRATE_CURRENCIES

        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)
        # 9 currency + 4 rate type checkboxes = 13 total; filter by label
        ccy_boxes = [
            cb for cb in checkboxes
            if str(cb.cget("text")) in EXRATE_CURRENCIES
        ]
        assert len(ccy_boxes) == len(EXRATE_CURRENCIES) == 9
        dialog.destroy()

    def test_usd_default_on(self, tk_root):
        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)
        usd_boxes = [cb for cb in checkboxes if cb.cget("text") == "USD"]
        assert usd_boxes, "USD checkbox not found"
        assert usd_boxes[0].get() == 1, "USD must be checked by default"
        dialog.destroy()

    def test_eur_default_on(self, tk_root):
        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)
        eur_boxes = [cb for cb in checkboxes if cb.cget("text") == "EUR"]
        assert eur_boxes, "EUR checkbox not found"
        assert eur_boxes[0].get() == 1, "EUR must be checked by default"
        dialog.destroy()

    def test_non_default_currency_off(self, tk_root):
        """GBP, JPY, etc. should be unchecked by default."""
        from gui.panels.exrate_dialog import EXRATE_CURRENCIES

        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)
        defaults_on = {"USD", "EUR"}
        for cb in checkboxes:
            label = str(cb.cget("text"))
            if label in EXRATE_CURRENCIES and label not in defaults_on:
                assert cb.get() == 0, f"{label} should be unchecked by default"
        dialog.destroy()


class TestExrateDialogRateTypeCheckboxes:
    """Rate type checkboxes are built with correct labels and defaults."""

    def test_four_rate_type_checkboxes_exist(self, tk_root):
        from gui.panels.exrate_dialog import EXRATE_RATE_TYPES

        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)
        rate_labels = set(EXRATE_RATE_TYPES.keys())
        rate_boxes = [
            cb for cb in checkboxes
            if str(cb.cget("text")) in rate_labels
        ]
        assert len(rate_boxes) == len(EXRATE_RATE_TYPES) == 4
        dialog.destroy()

    def test_buying_tt_default_on(self, tk_root):
        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)
        boxes = [cb for cb in checkboxes if cb.cget("text") == "Buying TT"]
        assert boxes, "'Buying TT' checkbox not found"
        assert boxes[0].get() == 1, "'Buying TT' must be checked by default"
        dialog.destroy()

    def test_selling_default_on(self, tk_root):
        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)
        boxes = [cb for cb in checkboxes if cb.cget("text") == "Selling"]
        assert boxes, "'Selling' checkbox not found"
        assert boxes[0].get() == 1, "'Selling' must be checked by default"
        dialog.destroy()

    def test_buying_sight_default_off(self, tk_root):
        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)
        boxes = [cb for cb in checkboxes if cb.cget("text") == "Buying Sight"]
        assert boxes, "'Buying Sight' checkbox not found"
        assert boxes[0].get() == 0, "'Buying Sight' must be unchecked by default"
        dialog.destroy()

    def test_mid_rate_default_off(self, tk_root):
        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)
        boxes = [cb for cb in checkboxes if cb.cget("text") == "Mid Rate"]
        assert boxes, "'Mid Rate' checkbox not found"
        assert boxes[0].get() == 0, "'Mid Rate' must be unchecked by default"
        dialog.destroy()


class TestExrateDialogDateModeSwitch:
    """Date mode switch defaults to 'auto'."""

    def test_date_switch_exists(self, tk_root):
        dialog = _open_dialog(tk_root)
        switches = _collect_switches(dialog)
        assert switches, "CTkSwitch for date mode not found"
        dialog.destroy()

    def test_date_switch_default_is_auto(self, tk_root):
        """Switch off-value is 'auto'; the default variable value is 'auto'."""
        dialog = _open_dialog(tk_root)
        switches = _collect_switches(dialog)
        assert switches, "CTkSwitch for date mode not found"
        sw = switches[0]
        # When switch is OFF (default), variable == offvalue == "auto"
        assert sw.get() == "auto", (
            f"Date mode switch should default to 'auto', got {sw.get()!r}"
        )
        dialog.destroy()

    def test_date_switch_text_fragment(self, tk_root):
        dialog = _open_dialog(tk_root)
        switches = _collect_switches(dialog)
        assert switches, "CTkSwitch not found"
        text = str(switches[0].cget("text"))
        assert "manual" in text.lower() or "date" in text.lower(), (
            f"Switch text should describe manual date selection, got {text!r}"
        )
        dialog.destroy()


class TestExrateDialogCreateButton:
    """'Create ExRate File' button is present."""

    def test_create_button_exists(self, tk_root):
        dialog = _open_dialog(tk_root)
        btn = _find_button(dialog, "Create ExRate")
        assert btn is not None, "'Create ExRate File' button not found"
        dialog.destroy()

    def test_create_button_has_command(self, tk_root):
        dialog = _open_dialog(tk_root)
        btn = _find_button(dialog, "Create ExRate")
        assert btn is not None
        # CTkButton stores its command; it should be callable
        cmd = btn.cget("command")
        assert callable(cmd), "Create button command must be callable"
        dialog.destroy()


class TestExrateDialogOnCreateValidation:
    """_on_create() validation: error label shows when selection is empty."""

    def _get_error_label(self, dialog):
        """Return the error CTkLabel (empty text at init, error_text color)."""
        import customtkinter as ctk

        from gui.theme import get_theme

        t = get_theme()
        error_color = t["error_text"]
        for widget in dialog.winfo_children():
            if isinstance(widget, ctk.CTkLabel):
                color = str(widget.cget("text_color"))
                if color == error_color:
                    return widget
        return None

    def test_error_label_starts_empty(self, tk_root):
        dialog = _open_dialog(tk_root)
        lbl = self._get_error_label(dialog)
        assert lbl is not None, "Error label not found in dialog"
        assert lbl.cget("text") == "", "Error label must start empty"
        dialog.destroy()

    def test_no_currency_selected_shows_error(self, tk_root):
        """Unchecking all currencies then pressing Create shows an error."""
        from gui.panels.exrate_dialog import EXRATE_CURRENCIES

        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)
        for cb in checkboxes:
            if str(cb.cget("text")) in EXRATE_CURRENCIES:
                cb.deselect()

        btn = _find_button(dialog, "Create ExRate")
        assert btn is not None
        # Invoke the button command — dialog must NOT be destroyed (validation fail)
        cmd = btn.cget("command")
        cmd()  # should set error text and return without destroying

        lbl = self._get_error_label(dialog)
        assert lbl is not None
        assert "currency" in lbl.cget("text").lower(), (
            f"Expected currency error message, got {lbl.cget('text')!r}"
        )
        dialog.destroy()

    def test_no_rate_type_selected_shows_error(self, tk_root):
        """Unchecking all rate types then pressing Create shows an error."""
        from gui.panels.exrate_dialog import EXRATE_RATE_TYPES

        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)
        rate_labels = set(EXRATE_RATE_TYPES.keys())
        # Ensure at least one currency is selected (defaults: USD/EUR are on)
        # Uncheck all rate types
        for cb in checkboxes:
            if str(cb.cget("text")) in rate_labels:
                cb.deselect()

        btn = _find_button(dialog, "Create ExRate")
        assert btn is not None
        cmd = btn.cget("command")
        cmd()  # should set error text and return

        lbl = self._get_error_label(dialog)
        assert lbl is not None
        assert "rate" in lbl.cget("text").lower(), (
            f"Expected rate type error message, got {lbl.cget('text')!r}"
        )
        dialog.destroy()

    def test_error_label_cleared_on_next_call(self, tk_root):
        """Each _on_create() call clears the previous error first."""
        from gui.panels.exrate_dialog import EXRATE_CURRENCIES, EXRATE_RATE_TYPES

        dialog = _open_dialog(tk_root)
        checkboxes = _collect_checkboxes(dialog)

        # First: uncheck all currencies to trigger error
        for cb in checkboxes:
            if str(cb.cget("text")) in EXRATE_CURRENCIES:
                cb.deselect()

        btn = _find_button(dialog, "Create ExRate")
        assert btn is not None
        cmd = btn.cget("command")
        cmd()  # sets error

        lbl = self._get_error_label(dialog)
        assert lbl is not None
        first_error = lbl.cget("text")
        assert first_error != "", "First call should have set an error"

        # Second call with all rate types unchecked but one currency restored
        for cb in checkboxes:
            if str(cb.cget("text")) == "USD":
                cb.select()
        for cb in checkboxes:
            if str(cb.cget("text")) in set(EXRATE_RATE_TYPES.keys()):
                cb.deselect()

        cmd()  # error cleared then new error set (rate type error)
        second_error = lbl.cget("text")
        # Error label was cleared (even if it now shows a different message)
        assert second_error != first_error, (
            "Error label should be updated on each _on_create() call"
        )
        dialog.destroy()


class TestExrateDialogOnCreateCallsWorker:
    """_on_create() with valid inputs destroys dialog and starts the worker."""

    def test_on_create_valid_destroys_dialog_and_calls_create_file(self, tk_root):
        """With valid selections _on_create destroys the dialog and calls
        _create_exrate_file in a separate thread.  We mock _create_exrate_file
        to prevent real filesystem/network access."""
        dialog = _open_dialog(tk_root)
        # Defaults: USD + EUR checked, Buying TT + Selling checked -> valid

        btn = _find_button(dialog, "Create ExRate")
        assert btn is not None

        called_with = []

        with patch(
            "gui.panels.exrate_dialog._create_exrate_file",
            side_effect=lambda *a, **kw: called_with.append((a, kw)),
        ):
            cmd = btn.cget("command")
            cmd()  # should destroy dialog and call _create_exrate_file

        # Dialog should be destroyed; winfo_exists() returns 0 for destroyed windows
        try:
            exists = dialog.winfo_exists()
        except Exception:
            exists = 0
        assert exists == 0, "Dialog must be destroyed after valid _on_create()"

        assert called_with, "_create_exrate_file must be called on valid input"
        args, kwargs = called_with[0]
        # First positional arg is `app` (tk_root here)
        assert args[0] is tk_root
        # currencies list must contain USD and EUR
        currencies = args[1]
        assert "USD" in currencies
        assert "EUR" in currencies
        # date_range is None in auto mode
        assert kwargs.get("date_range") is None

    def test_on_create_auto_mode_passes_none_date_range(self, tk_root):
        """Auto date mode must pass date_range=None to _create_exrate_file."""
        dialog = _open_dialog(tk_root)
        btn = _find_button(dialog, "Create ExRate")
        assert btn is not None

        captured_kwargs = {}

        with patch(
            "gui.panels.exrate_dialog._create_exrate_file",
            side_effect=lambda *a, **kw: captured_kwargs.update(
                {"args": a, "kwargs": kw}
            ),
        ):
            btn.cget("command")()

        assert captured_kwargs.get("kwargs", {}).get("date_range") is None

    def test_on_create_manual_mode_passes_date_tuple(self, tk_root):
        """Manual date mode must pass a (date, date) tuple to _create_exrate_file."""
        dialog = _open_dialog(tk_root)

        # Toggle the switch to manual
        switches = _collect_switches(dialog)
        assert switches, "Date mode switch not found"
        sw = switches[0]
        sw.select()  # sets variable to onvalue="manual"
        # Invoke the switch command to trigger _toggle_date_mode()
        sw_cmd = sw.cget("command")
        if callable(sw_cmd):
            sw_cmd()

        btn = _find_button(dialog, "Create ExRate")
        assert btn is not None

        captured = {}

        with patch(
            "gui.panels.exrate_dialog._create_exrate_file",
            side_effect=lambda *a, **kw: captured.update(
                {"args": a, "kwargs": kw}
            ),
        ):
            btn.cget("command")()

        dr = captured.get("kwargs", {}).get("date_range")
        assert dr is not None, "Manual mode must pass a date_range tuple"
        assert isinstance(dr, tuple) and len(dr) == 2
        start, end = dr
        assert isinstance(start, date)
        assert isinstance(end, date)


# ---------------------------------------------------------------------------
# Finding 3 — Escape / Enter key handling on the dialog
# ---------------------------------------------------------------------------


class TestExrateDialogKeyBindings:
    """The modal must support Escape-to-cancel and Return-to-confirm (#3)."""

    def test_escape_binding_present(self, tk_root):
        dialog = _open_dialog(tk_root)
        # bind() with no callback returns the bound script(s) as a string.
        binding = dialog.bind("<Escape>")
        assert binding, "Dialog must bind <Escape> to cancel"
        dialog.destroy()

    def test_return_binding_present(self, tk_root):
        dialog = _open_dialog(tk_root)
        binding = dialog.bind("<Return>")
        assert binding, "Dialog must bind <Return> to confirm"
        dialog.destroy()

    def test_return_invokes_create(self, tk_root):
        """The <Return> binding runs the same path as the Create button.

        event_generate is unreliable on a withdrawn, unfocused TopLevel, so we
        invoke the bound callback directly via the funcid Tcl registered for
        the binding — this proves the wiring without needing window focus.
        """
        dialog = _open_dialog(tk_root)
        called = []
        with patch(
            "gui.panels.exrate_dialog._create_exrate_file",
            side_effect=lambda *a, **kw: called.append(a),
        ):
            _invoke_binding(dialog, "<Return>")
        assert called, "<Return> must trigger ExRate creation"

    def test_escape_destroys_dialog(self, tk_root):
        dialog = _open_dialog(tk_root)
        _invoke_binding(dialog, "<Escape>")
        exists = 1
        with contextlib.suppress(Exception):
            exists = dialog.winfo_exists()
        assert exists == 0, "<Escape> must destroy the dialog"


# ---------------------------------------------------------------------------
# Finding 4 — refuse to start while a batch is running
# ---------------------------------------------------------------------------


class TestExrateDialogBatchGuard:
    """ExRate creation must not hijack the shared progress UI mid-batch (#4)."""

    def _get_error_label(self, dialog):
        import customtkinter as ctk

        from gui.theme import get_theme

        error_color = get_theme()["error_text"]
        for widget in dialog.winfo_children():
            if isinstance(widget, ctk.CTkLabel) and (
                str(widget.cget("text_color")) == error_color
            ):
                return widget
        return None

    def test_on_create_blocked_while_batch_running(self, tk_root):
        """_on_create refuses (inline error, no worker) when a batch runs."""
        dialog = _open_dialog(tk_root)
        # Simulate an in-flight batch on the app object.
        tk_root._batch_running = True
        try:
            called = []
            with patch(
                "gui.panels.exrate_dialog._create_exrate_file",
                side_effect=lambda *a, **kw: called.append(a),
            ):
                btn = _find_button(dialog, "Create ExRate")
                btn.cget("command")()
            assert not called, "Must NOT start ExRate creation during a batch"
            lbl = self._get_error_label(dialog)
            assert lbl is not None
            assert "batch" in lbl.cget("text").lower(), (
                f"Expected a batch-busy message, got {lbl.cget('text')!r}"
            )
            # Dialog stays open so the user can retry after the batch.
            assert dialog.winfo_exists() == 1
        finally:
            tk_root._batch_running = False
            with contextlib.suppress(Exception):
                dialog.destroy()

    def test_create_file_early_returns_while_batch_running(self, tk_root):
        """_create_exrate_file itself bails before the save picker (#4)."""
        from gui.panels import exrate_dialog

        tk_root._batch_running = True
        opened = []
        try:
            with patch.object(
                exrate_dialog.filedialog,
                "asksaveasfilename",
                side_effect=lambda *a, **kw: opened.append(True) or "",
            ):
                exrate_dialog._create_exrate_file(
                    tk_root, ["USD"], {"Buying TT": "buying_transfer"},
                )
            assert not opened, "Save picker must not open during a batch"
        finally:
            tk_root._batch_running = False


# ---------------------------------------------------------------------------
# Finding 1 — result summary (row count / date span / currency list)
# ---------------------------------------------------------------------------


class TestExrateSummary:
    """_build_exrate_summary surfaces span + currencies + rate types (#1)."""

    def test_manual_range_day_count_inclusive(self):
        from gui.panels.exrate_dialog import _build_exrate_summary

        dr = (date(2026, 1, 1), date(2026, 1, 10))
        summary = _build_exrate_summary(["USD"], {"Buying TT": "x"}, dr)
        assert "10 days" in summary  # inclusive span
        assert "2026-01-01" in summary
        assert "2026-01-10" in summary

    def test_single_day_is_singular(self):
        from gui.panels.exrate_dialog import _build_exrate_summary

        dr = (date(2026, 3, 5), date(2026, 3, 5))
        summary = _build_exrate_summary(["USD"], {"Selling": "x"}, dr)
        assert "1 day " in summary and "1 days" not in summary

    def test_currencies_and_rate_types_listed(self):
        from gui.panels.exrate_dialog import _build_exrate_summary

        dr = (date(2026, 1, 1), date(2026, 1, 2))
        summary = _build_exrate_summary(
            ["USD", "GBP"], {"Buying TT": "x", "Selling": "y"}, dr
        )
        assert "USD" in summary and "GBP" in summary
        assert "Buying TT" in summary and "Selling" in summary

    def test_auto_mode_uses_current_year(self):
        from gui.panels.exrate_dialog import _build_exrate_summary

        today = date.today()
        summary = _build_exrate_summary(["EUR"], {"Mid Rate": "x"}, None)
        assert f"{today.year}-01-01" in summary
        assert today.strftime("%Y-%m-%d") in summary


# ---------------------------------------------------------------------------
# Worker-path tests (Findings 2 + 5) — no real BOTClient / network.
# ---------------------------------------------------------------------------


class _FakeStatusLabel:
    """Minimal stand-in for app.lbl_status capturing configure() calls."""

    def __init__(self):
        self.kwargs = {}

    def configure(self, **kwargs):
        self.kwargs.update(kwargs)

    def cget(self, key):
        return self.kwargs.get(key, "")


class _FakeApp:
    """Headless app stub: no Tk, runs after() callbacks synchronously."""

    def __init__(self):
        self.lbl_status = _FakeStatusLabel()
        self.event_bus = None
        self._batch_running = False
        self.last_processed_path = None

    def after(self, _delay, func=None, *args):
        if func is not None:
            func(*args)

    def update_idletasks(self):
        pass


def _run_worker_sync(monkeypatch, fake_engine_factory):
    """Run _create_exrate_file's worker inline (no thread) and return the app.

    Stubs out BOTClient / LedgerEngine / httpx.AsyncClient so no tokens or
    network are required, and patches threading.Thread so .start() runs the
    target synchronously inside the test.
    """
    import sys
    import threading as _threading
    import types

    from gui.panels import exrate_dialog

    # Fake httpx.AsyncClient as an async context manager.
    class _FakeAsyncClient:
        def __init__(self, *a, **kw):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, *a):
            return False

    monkeypatch.setattr(exrate_dialog.httpx, "AsyncClient", _FakeAsyncClient)

    # Fake core.api_client.BOTClient and core.engine.LedgerEngine (imported
    # inside the worker, so patch them on their source modules).
    fake_api_mod = types.ModuleType("core.api_client")
    fake_api_mod.CLIENT_TIMEOUT = 1.0
    fake_api_mod.BOTClient = lambda *a, **kw: object()
    fake_engine_mod = types.ModuleType("core.engine")
    fake_engine_mod.LedgerEngine = fake_engine_factory
    monkeypatch.setitem(sys.modules, "core.api_client", fake_api_mod)
    monkeypatch.setitem(sys.modules, "core.engine", fake_engine_mod)

    # Run the worker thread body inline.
    class _InlineThread:
        def __init__(self, target=None, **kw):
            self._target = target

        def start(self):
            self._target()

    monkeypatch.setattr(exrate_dialog.threading, "Thread", _InlineThread)
    monkeypatch.setattr(_threading, "Thread", _InlineThread, raising=False)

    return exrate_dialog


def _write_engine_workbook(filepath, marker="FILLED-EXRATE"):
    """Write a minimal REAL ExRate workbook with a marker cell.

    The success path now read-back-verifies dest (structure: ExRate sheet +
    populated row 2), so success-path fakes must produce a real workbook —
    raw byte markers would fail verification. The C2 marker still proves the
    temp was moved into place (not a blank save).
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExRate"
    ws.append(["Date", "USD Buying TT Rate", "Marker"])
    ws.append([date(2026, 1, 2), 32.4507, marker])
    wb.save(filepath)
    wb.close()


def _read_marker(path):
    """Return the C2 marker of a workbook written by _write_engine_workbook."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True)
    try:
        row2 = next(
            wb["ExRate"].iter_rows(min_row=2, max_row=2, values_only=True),
            None,
        )
        return row2[2] if row2 and len(row2) >= 3 else None
    finally:
        wb.close()


class _RaisingEngine:
    """LedgerEngine stub whose standalone update raises a chosen error."""

    def __init__(self, exc):
        self._exc = exc

    def __call__(self, *a, **kw):
        return self

    async def update_exrate_standalone(self, *a, **kw):
        raise self._exc


class _SuccessEngine:
    """LedgerEngine stub whose standalone update writes a marker + succeeds."""

    def __call__(self, *a, **kw):
        return self

    async def update_exrate_standalone(self, filepath, *a, **kw):
        # The real engine fills the temp file; emulate with a real minimal
        # ExRate workbook (passes the dest read-back verification) carrying
        # a marker that proves the temp was moved into place.
        _write_engine_workbook(filepath)
        return filepath


class TestExrateWorkerSuccess:
    """Findings 1 + 2 happy path: move temp into place, surface a summary."""

    def test_success_moves_temp_and_reports_summary(self, monkeypatch, tmp_path):
        from gui.panels import exrate_dialog

        dest = tmp_path / "ExRate.xlsx"
        _run_worker_sync(monkeypatch, lambda *a, **kw: _SuccessEngine())
        monkeypatch.setattr(
            exrate_dialog.filedialog, "asksaveasfilename",
            lambda *a, **kw: str(dest),
        )

        app = _FakeApp()
        exrate_dialog._create_exrate_file(
            app, ["USD", "EUR"],
            {"Buying TT": "buying_transfer"},
            date_range=(date(2026, 1, 1), date(2026, 1, 10)),
        )

        # Temp was moved into the destination (engine-written content present).
        assert _read_marker(dest) == "FILLED-EXRATE"
        assert not list(tmp_path.glob(".exrate_tmp_*")), "Temp not cleaned up"
        # Success status carries the summary (span + currencies).
        msg = app.lbl_status.cget("text")
        assert "ExRate created" in msg
        assert "10 days" in msg
        assert "USD" in msg and "EUR" in msg
        assert app.last_processed_path == str(dest)


class TestExrateWorkerDataSafety:
    """Finding 2: a failed fetch must NOT destroy the user's chosen file."""

    def test_existing_dest_untouched_on_api_failure(self, monkeypatch, tmp_path):
        import httpx

        from gui.panels import exrate_dialog

        # A pre-existing file the user picked (e.g. a real ledger).
        dest = tmp_path / "ExRate.xlsx"
        dest.write_bytes(b"ORIGINAL-CONTENT")

        engine = _RaisingEngine(httpx.ConnectError("boom"))
        _run_worker_sync(monkeypatch, lambda *a, **kw: engine)

        monkeypatch.setattr(
            exrate_dialog.filedialog, "asksaveasfilename",
            lambda *a, **kw: str(dest),
        )

        app = _FakeApp()
        exrate_dialog._create_exrate_file(
            app, ["USD"], {"Buying TT": "buying_transfer"},
        )

        # The original file must be byte-for-byte intact (never blank-saved).
        assert dest.read_bytes() == b"ORIGINAL-CONTENT", (
            "A failed fetch destroyed the user's existing file"
        )
        # No leftover temp siblings.
        leftovers = list(tmp_path.glob(".exrate_tmp_*"))
        assert not leftovers, f"Temp file not cleaned up: {leftovers}"

    def test_network_error_shows_friendly_message(self, monkeypatch, tmp_path):
        import httpx

        from gui.panels import exrate_dialog

        dest = tmp_path / "ExRate.xlsx"
        engine = _RaisingEngine(httpx.ConnectError("boom"))
        _run_worker_sync(monkeypatch, lambda *a, **kw: engine)
        monkeypatch.setattr(
            exrate_dialog.filedialog, "asksaveasfilename",
            lambda *a, **kw: str(dest),
        )

        app = _FakeApp()
        exrate_dialog._create_exrate_file(
            app, ["USD"], {"Buying TT": "buying_transfer"},
        )

        msg = app.lbl_status.cget("text")
        assert "BOT server" in msg, f"Expected friendly network msg, got {msg!r}"


class _BackupAwareEngine(_SuccessEngine):
    """Success engine exposing a ``backup`` manager like the real LedgerEngine.

    The worker captures ``engine.backup`` inside ``_run`` and uses it to back
    up an existing dest before the temp is moved into place (F10).
    """

    def __init__(self, backup):
        self.backup = backup


class TestExrateWorkerBackupFirst:
    """F10: overwriting an EXISTING dest must back it up before the move."""

    def test_existing_dest_backed_up_before_move(self, monkeypatch, tmp_path):
        from core.backup_manager import BackupManager

        from gui.panels import exrate_dialog

        # A pre-existing workbook the user confirmed overwriting.
        dest = tmp_path / "ExRate.xlsx"
        dest.write_bytes(b"ORIGINAL-CONTENT")
        backup_dir = tmp_path / "backups"

        engine = _BackupAwareEngine(BackupManager(str(backup_dir)))
        _run_worker_sync(monkeypatch, lambda *a, **kw: engine)
        monkeypatch.setattr(
            exrate_dialog.filedialog, "asksaveasfilename",
            lambda *a, **kw: str(dest),
        )

        app = _FakeApp()
        exrate_dialog._create_exrate_file(
            app, ["USD"], {"Buying TT": "buying_transfer"},
        )

        # Dest was replaced by the engine-written temp...
        assert _read_marker(dest) == "FILLED-EXRATE"
        # ...and a real backup of the ORIGINAL dest exists, taken BEFORE the
        # move (its bytes are the pre-overwrite content, not the new sheet).
        backups = list(backup_dir.glob("ExRate__bak__*.xlsx"))
        assert len(backups) == 1, f"Expected 1 dest backup, got {backups}"
        assert backups[0].read_bytes() == b"ORIGINAL-CONTENT", (
            "Backup must capture dest as it was before the overwrite"
        )

    def test_new_dest_creates_no_backup(self, monkeypatch, tmp_path):
        from core.backup_manager import BackupManager

        from gui.panels import exrate_dialog

        # Creation path: dest does not exist yet — nothing to back up.
        dest = tmp_path / "ExRate.xlsx"
        backup_dir = tmp_path / "backups"

        engine = _BackupAwareEngine(BackupManager(str(backup_dir)))
        _run_worker_sync(monkeypatch, lambda *a, **kw: engine)
        monkeypatch.setattr(
            exrate_dialog.filedialog, "asksaveasfilename",
            lambda *a, **kw: str(dest),
        )

        app = _FakeApp()
        exrate_dialog._create_exrate_file(
            app, ["USD"], {"Buying TT": "buying_transfer"},
        )

        assert _read_marker(dest) == "FILLED-EXRATE"
        assert not list(backup_dir.glob("*.xlsx")), (
            "No backup expected when dest did not exist"
        )

    def test_backup_failure_aborts_move_and_keeps_dest(
        self, monkeypatch, tmp_path
    ):
        from core.backup_manager import BackupError

        from gui.panels import exrate_dialog

        dest = tmp_path / "ExRate.xlsx"
        dest.write_bytes(b"ORIGINAL-CONTENT")

        class _FailingBackup:
            def create_backup(self, filepath):
                raise BackupError("disk full")

        engine = _BackupAwareEngine(_FailingBackup())
        _run_worker_sync(monkeypatch, lambda *a, **kw: engine)
        monkeypatch.setattr(
            exrate_dialog.filedialog, "asksaveasfilename",
            lambda *a, **kw: str(dest),
        )

        app = _FakeApp()
        exrate_dialog._create_exrate_file(
            app, ["USD"], {"Buying TT": "buying_transfer"},
        )

        # Fail-safe rule: backup failed, so the overwrite must NOT happen.
        assert dest.read_bytes() == b"ORIGINAL-CONTENT", (
            "Dest was overwritten despite a failed backup"
        )
        # Temp is discarded and the failure is surfaced to the user.
        assert not list(tmp_path.glob(".exrate_tmp_*")), "Temp not cleaned up"
        msg = app.lbl_status.cget("text")
        assert "Failed" in msg, f"Expected failure status, got {msg!r}"


class TestExrateWorkerHumanizedSaveError:
    """Finding 5: a locked/open file yields 'close it in Excel', not raw errno."""

    def test_locked_file_humanized(self, monkeypatch, tmp_path):
        from gui.panels import exrate_dialog

        dest = tmp_path / "ExRate.xlsx"
        # PermissionError is what a Windows/macOS lock surfaces as.
        exc = PermissionError(13, "Permission denied")
        engine = _RaisingEngine(exc)
        _run_worker_sync(monkeypatch, lambda *a, **kw: engine)
        monkeypatch.setattr(
            exrate_dialog.filedialog, "asksaveasfilename",
            lambda *a, **kw: str(dest),
        )

        app = _FakeApp()
        exrate_dialog._create_exrate_file(
            app, ["USD"], {"Buying TT": "buying_transfer"},
        )

        msg = app.lbl_status.cget("text")
        assert "open in Excel" in msg, (
            f"Locked-file error not humanized, got {msg!r}"
        )
        assert "Errno" not in msg, "Raw errno leaked to the user"


# ---------------------------------------------------------------------------
# LOW finding 1 — month/year-aware day comboboxes (no impossible days)
# ---------------------------------------------------------------------------


def _manual_day_combos(tk_root):
    """Open the dialog in manual mode and return (start_day, end_day) combos.

    Combos are collected in widget-tree order: start (year, month, day) then
    end (year, month, day), so indices 2 and 5 are the two day pickers.
    """
    dialog = _open_dialog(tk_root)
    switches = _collect_switches(dialog)
    sw = switches[0]
    sw.select()
    sw_cmd = sw.cget("command")
    if callable(sw_cmd):
        sw_cmd()
    combos = _collect_comboboxes(dialog)
    assert len(combos) == 6, f"Expected 6 date combos, got {len(combos)}"
    return dialog, combos[0], combos[1], combos[2], combos[3], combos[4], combos[5]


class TestExrateDayOptionsAreMonthAware:
    """LOW #1: the day dropdown must reflect the selected month/year."""

    def test_february_non_leap_has_no_day_29_30_31(self, tk_root):
        dialog, s_year, s_month, s_day, *_ = _manual_day_combos(tk_root)
        try:
            s_year.set("2025")  # 2025 is not a leap year
            s_month.set("02")
            # Trigger the month combobox command (calendar.monthrange refresh).
            s_month.cget("command")("02")
            days = list(s_day.cget("values"))
            assert "28" in days
            assert "29" not in days, "Non-leap February must not offer day 29"
            assert "30" not in days and "31" not in days
        finally:
            dialog.destroy()

    def test_february_leap_year_has_day_29(self, tk_root):
        dialog, s_year, s_month, s_day, *_ = _manual_day_combos(tk_root)
        try:
            s_year.set("2024")  # leap year
            s_month.set("02")
            s_month.cget("command")("02")
            days = list(s_day.cget("values"))
            assert "29" in days, "Leap February must offer day 29"
            assert "30" not in days and "31" not in days
        finally:
            dialog.destroy()

    def test_thirty_day_month_has_no_day_31(self, tk_root):
        dialog, s_year, s_month, s_day, *_ = _manual_day_combos(tk_root)
        try:
            s_year.set("2026")
            s_month.set("04")  # April = 30 days
            s_month.cget("command")("04")
            days = list(s_day.cget("values"))
            assert "30" in days
            assert "31" not in days, "30-day month must not offer day 31"
        finally:
            dialog.destroy()

    def test_thirty_one_day_month_offers_day_31(self, tk_root):
        dialog, s_year, s_month, s_day, *_ = _manual_day_combos(tk_root)
        try:
            s_year.set("2026")
            s_month.set("01")  # January = 31 days
            s_month.cget("command")("01")
            assert "31" in list(s_day.cget("values"))
        finally:
            dialog.destroy()

    def test_selected_day_clamped_when_month_shrinks(self, tk_root):
        """A day past the new month's length snaps down (31 → 28 for Feb)."""
        dialog, s_year, s_month, s_day, *_ = _manual_day_combos(tk_root)
        try:
            s_year.set("2025")
            s_month.set("01")
            s_month.cget("command")("01")
            s_day.set("31")  # valid for January
            # Switch to February (non-leap) — 31 is now impossible.
            s_month.set("02")
            s_month.cget("command")("02")
            assert s_day.get() == "28", (
                f"Day should clamp to 28 for non-leap Feb, got {s_day.get()!r}"
            )
        finally:
            dialog.destroy()

    def test_year_change_refreshes_february_leap(self, tk_root):
        """Changing the YEAR (not month) re-evaluates leap-year Feb length."""
        dialog, s_year, s_month, s_day, *_ = _manual_day_combos(tk_root)
        try:
            s_month.set("02")
            s_year.set("2025")  # non-leap
            s_year.cget("command")("2025")
            assert "29" not in list(s_day.cget("values"))
            s_year.set("2024")  # leap
            s_year.cget("command")("2024")
            assert "29" in list(s_day.cget("values"))
        finally:
            dialog.destroy()

    def test_end_day_combo_is_also_month_aware(self, tk_root):
        """The END day picker gets the same treatment as the start picker."""
        dialog, _sy, _sm, _sd, e_year, e_month, e_day = _manual_day_combos(tk_root)
        try:
            e_year.set("2025")
            e_month.set("02")
            e_month.cget("command")("02")
            assert "29" not in list(e_day.cget("values"))
        finally:
            dialog.destroy()


# ---------------------------------------------------------------------------
# LOW finding 2 — busy/cancel affordance for the long standalone fetch
# ---------------------------------------------------------------------------


class _CardApp(_FakeApp):
    """Headless app stub that also exposes a real .card so the transient
    Cancel button (a child of app.card) is actually created."""

    def __init__(self, tk_root):
        super().__init__()
        import customtkinter as ctk

        self.card = ctk.CTkFrame(tk_root)


class _CancellableEngine:
    """LedgerEngine stub that, on its first progress tick, presses the Cancel
    button (proving the real wiring), then expects the second tick to raise."""

    def __init__(self, app):
        self._app = app
        self.ticks = 0

    def __call__(self, *a, **kw):
        return self

    async def update_exrate_standalone(self, filepath, *a, progress_cb=None, **kw):
        import customtkinter as ctk

        # First tick: simulate the user clicking the Cancel button on the card.
        progress_cb("Fetching USD rates...")
        self.ticks += 1
        # Find and invoke the transient Cancel button's command.
        for child in self._app.card.winfo_children():
            if isinstance(child, ctk.CTkButton):
                child.cget("command")()
                break
        # Second tick: the status callback must now raise _ExRateCancelled.
        progress_cb("Fetching EUR rates...")  # should raise
        self.ticks += 1
        # Should never reach here.
        from pathlib import Path as _P

        _P(filepath).write_bytes(b"SHOULD-NOT-HAPPEN")
        return filepath


class TestExrateCancelAffordance:
    """LOW #2: a Cancel button signals the worker to abort the fetch."""

    def test_cancel_button_appears_on_card_during_fetch(self, monkeypatch, tmp_path, tk_root):
        """A transient Cancel button is packed on app.card while fetching."""
        import customtkinter as ctk

        from gui.panels import exrate_dialog

        dest = tmp_path / "ExRate.xlsx"

        seen = {"had_cancel_button": False}

        class _ObservingEngine:
            def __call__(self, *a, **kw):
                return self

            async def update_exrate_standalone(self, filepath, *a, progress_cb=None, **kw):
                # While the fetch runs, the card must carry a Cancel button.
                for child in app.card.winfo_children():
                    if isinstance(child, ctk.CTkButton):
                        seen["had_cancel_button"] = True
                _write_engine_workbook(filepath)
                return filepath

        _run_worker_sync(monkeypatch, lambda *a, **kw: _ObservingEngine())
        monkeypatch.setattr(
            exrate_dialog.filedialog, "asksaveasfilename",
            lambda *a, **kw: str(dest),
        )

        app = _CardApp(tk_root)
        try:
            exrate_dialog._create_exrate_file(
                app, ["USD", "EUR"], {"Buying TT": "buying_transfer"},
            )
            assert seen["had_cancel_button"], (
                "No Cancel button was shown on the card during the fetch"
            )
        finally:
            app.card.destroy()

    def test_cancel_aborts_fetch_and_preserves_dest(self, monkeypatch, tmp_path, tk_root):
        """Clicking Cancel raises _ExRateCancelled and leaves dest untouched."""
        from core.i18n import tr
        from gui.panels import exrate_dialog

        dest = tmp_path / "ExRate.xlsx"
        dest.write_bytes(b"ORIGINAL")  # a pre-existing file must survive

        engine_holder = {}

        def _factory(*a, **kw):
            eng = _CancellableEngine(app)
            engine_holder["engine"] = eng
            return eng

        _run_worker_sync(monkeypatch, _factory)
        monkeypatch.setattr(
            exrate_dialog.filedialog, "asksaveasfilename",
            lambda *a, **kw: str(dest),
        )

        app = _CardApp(tk_root)
        try:
            exrate_dialog._create_exrate_file(
                app, ["USD", "EUR"], {"Buying TT": "buying_transfer"},
            )
            # The engine aborted after the first tick (cancel) — never wrote.
            assert engine_holder["engine"].ticks == 1, (
                "Fetch did not abort on the cancel signal"
            )
            # Original file is byte-for-byte intact (temp discarded, no move).
            assert dest.read_bytes() == b"ORIGINAL"
            assert not list(tmp_path.glob(".exrate_tmp_*")), "Temp not cleaned up"
            # Status reflects cancellation, not an error/traceback.
            msg = app.lbl_status.cget("text")
            assert msg == tr("exrate.cancelled")
        finally:
            app.card.destroy()

    def test_cancel_button_removed_after_completion(self, monkeypatch, tmp_path, tk_root):
        """On success the transient Cancel button is destroyed."""
        import customtkinter as ctk

        from gui.panels import exrate_dialog

        dest = tmp_path / "ExRate.xlsx"
        _run_worker_sync(monkeypatch, lambda *a, **kw: _SuccessEngine())
        monkeypatch.setattr(
            exrate_dialog.filedialog, "asksaveasfilename",
            lambda *a, **kw: str(dest),
        )

        app = _CardApp(tk_root)
        try:
            exrate_dialog._create_exrate_file(
                app, ["USD"], {"Buying TT": "buying_transfer"},
            )
            remaining = [
                c for c in app.card.winfo_children()
                if isinstance(c, ctk.CTkButton)
            ]
            assert not remaining, "Cancel button must be removed after completion"
        finally:
            app.card.destroy()


# ---------------------------------------------------------------------------
# F50/F201 — post-move structural read-back of the destination file
# ---------------------------------------------------------------------------


class TestExrateDestVerification:
    """The moved dest must reopen as a real workbook with a populated
    ExRate sheet; failures surface via the existing worker error path."""

    def test_verify_passes_on_real_exrate_file(self, tmp_path):
        from gui.panels.exrate_dialog import _verify_exrate_dest

        dest = tmp_path / "ok.xlsx"
        _write_engine_workbook(str(dest))
        _verify_exrate_dest(str(dest))  # must not raise

    def test_verify_rejects_missing_exrate_sheet(self, tmp_path):
        import openpyxl

        from gui.panels.exrate_dialog import _verify_exrate_dest

        dest = tmp_path / "no_sheet.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(str(dest))
        wb.close()
        with pytest.raises(ValueError, match="Post-write verification failed"):
            _verify_exrate_dest(str(dest))

    def test_verify_rejects_empty_row_2(self, tmp_path):
        import openpyxl

        from gui.panels.exrate_dialog import _verify_exrate_dest

        dest = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws.append(["Date", "USD Buying TT Rate"])  # header only, no data
        wb.save(str(dest))
        wb.close()
        with pytest.raises(ValueError, match="Post-write verification failed"):
            _verify_exrate_dest(str(dest))

    def test_verify_rejects_unreadable_file(self, tmp_path):
        from gui.panels.exrate_dialog import _verify_exrate_dest

        dest = tmp_path / "garbage.xlsx"
        dest.write_bytes(b"NOT-A-ZIP")
        # zipfile.BadZipFile is NOT an OSError — it must be wrapped in
        # ValueError so the worker's existing except tuple catches it.
        with pytest.raises(ValueError, match="Post-write verification failed"):
            _verify_exrate_dest(str(dest))

    def test_worker_surfaces_verification_failure(self, monkeypatch, tmp_path):
        """An engine run that 'succeeds' but leaves a structurally bad dest
        must be reported as a failure, never as 'ExRate created'."""
        from gui.panels import exrate_dialog

        class _BadFileEngine:
            def __call__(self, *a, **kw):
                return self

            async def update_exrate_standalone(self, filepath, *a, **kw):
                import openpyxl

                wb = openpyxl.Workbook()
                wb.active.title = "Sheet1"  # no ExRate sheet
                wb.save(filepath)
                wb.close()
                return filepath

        dest = tmp_path / "ExRate.xlsx"
        _run_worker_sync(monkeypatch, lambda *a, **kw: _BadFileEngine())
        monkeypatch.setattr(
            exrate_dialog.filedialog, "asksaveasfilename",
            lambda *a, **kw: str(dest),
        )

        app = _FakeApp()
        exrate_dialog._create_exrate_file(
            app, ["USD"], {"Buying TT": "buying_transfer"},
        )

        msg = app.lbl_status.cget("text")
        assert "ExRate created" not in msg, "must not report success"
        assert "Failed" in msg and "verification" in msg.lower(), (
            f"Expected a verification failure status, got {msg!r}"
        )
