#!/usr/bin/env python3
"""
tests/test_excel_io.py
---------------------------------------------------------------------------
Unit tests for core/excel_io.py — the Excel formula engine.
Covers formula injection, date normalization, header scanning,
ExRate indexing, and multi-currency custom writes.
---------------------------------------------------------------------------
"""

from datetime import date, datetime

import openpyxl

from core.constants import parse_date
from core.excel_io import (
    build_exrate_index,
    find_header_row,
    inject_xlookup_formulas,
    scan_sheet_headers,
    write_custom_exrate_data,
)

TARGET_COLS = {"source_date": "Date", "currency": "Cur", "out_rate": "EX Rate"}


# Layout in fixtures: A=Date, B=Cur, C=EX Rate, D=Amount.
def _expected_lookup(date_ref, col, n_marker="{N}"):
    """One guarded IFS branch value: the IF guard renders a found-but-empty
    ExRate cell (weekend/holiday row) as "" instead of numeric 0 (F6)."""
    lookup = (
        f'_xlfn.XLOOKUP({date_ref},'
        f'ExRate!$A$2:$A${n_marker},'
        f'ExRate!${col}$2:${col}${n_marker},"",0)'
    )
    return f'IFERROR(IF({lookup}="","",{lookup}),"")'


def _expected_formula(row, usd_col="B", eur_col="D", extra=None):
    """Reproduce the exact IFS formula inject_xlookup_formulas emits.

    extra: optional list of (ccy, col_letter) appended after the EUR branch.
    """
    date_ref = f"A{row}"
    cur_ref = f"B{row}"
    branches = (
        f'{cur_ref}="THB",1,'
        f'{cur_ref}="USD",'
        f'{_expected_lookup(date_ref, usd_col)},'
        f'{cur_ref}="EUR",'
        f'{_expected_lookup(date_ref, eur_col)}'
    )
    if extra:
        for ccy, col in extra:
            branches += (
                f',{cur_ref}="{ccy}",'
                f'{_expected_lookup(date_ref, col)}'
            )
    return (
        f'=IF(OR({cur_ref}="",{date_ref}=""),"",'
        f'_xlfn.IFS({branches},TRUE,""))'
    )


def _inject(wb, **kwargs):
    """Run scan + inject in one shot with sensible defaults."""
    sheet_maps = scan_sheet_headers(wb, TARGET_COLS)
    inject_xlookup_formulas(
        wb, sheet_maps,
        exrate_last_row=kwargs.pop("N", 30),
        parse_date_fn=parse_date,
        **kwargs,
    )
    return sheet_maps


# =========================================================================
#  inject_xlookup_formulas — exact formula strings
# =========================================================================

class TestInjectFormulaStrings:

    def test_usd_and_eur_buying_formula(self, ledger_xlsx):
        path = ledger_xlsx({"Jan": [
            (date(2025, 1, 7), "USD"),
            (date(2025, 1, 8), "EUR"),
        ]})
        wb = openpyxl.load_workbook(path)
        _inject(wb, N=30)
        ws = wb["Jan"]
        # Row 2 = USD, row 3 = EUR. EX Rate is column C.
        assert ws.cell(row=2, column=3).value == \
            _expected_formula(2).format(N=30)
        assert ws.cell(row=3, column=3).value == \
            _expected_formula(3).format(N=30)
        wb.close()

    def test_thb_branch_emits_literal_one(self, ledger_xlsx):
        path = ledger_xlsx({"Jan": [(date(2025, 1, 7), "THB")]})
        wb = openpyxl.load_workbook(path)
        _inject(wb, N=30)
        formula = wb["Jan"].cell(row=2, column=3).value
        # THB branch is a literal 1, not a lookup.
        assert 'B2="THB",1,' in formula
        wb.close()

    def test_selling_rate_type_refs_c_and_e(self, ledger_xlsx):
        path = ledger_xlsx({"Jan": [(date(2025, 1, 7), "USD")]})
        wb = openpyxl.load_workbook(path)
        _inject(wb, N=30, rate_type="selling")
        assert wb["Jan"].cell(row=2, column=3).value == \
            _expected_formula(2, usd_col="C", eur_col="E").format(N=30)
        wb.close()


# =========================================================================
#  Weekend/holiday blank guard (F6)
# =========================================================================

class TestWeekendBlankGuard:
    """F6: weekend/holiday rows EXIST in the ExRate sheet with blank rate
    cells, so an unguarded XLOOKUP returns a reference to an empty cell
    which Excel renders as numeric 0. Every lookup branch must wrap the
    lookup in an IF guard so a found-but-empty result renders blank ""."""

    def test_every_lookup_branch_carries_blank_guard(self, ledger_xlsx):
        # Saturday 2025-01-04 — the row exists in ExRate with blank rates.
        path = ledger_xlsx({"Jan": [(date(2025, 1, 4), "USD")]})
        wb = openpyxl.load_workbook(path)
        _inject(wb, N=30, exrate_col_map={"GBP": "F"})
        formula = wb["Jan"].cell(row=2, column=3).value
        # Three lookup branches (USD, EUR, GBP), each guarded: the IF
        # compares the lookup result to "" and yields "" when blank, so a
        # weekend-dated row renders blank instead of 0. openpyxl cannot
        # evaluate formulas, so these are string-shape assertions.
        assert formula.count('IFERROR(IF(_xlfn.XLOOKUP(') == 3
        assert formula.count('"",0)="","",_xlfn.XLOOKUP(') == 3
        # No unguarded lookup remains: every XLOOKUP is either the guard
        # condition or the guarded result (exactly 2 per branch).
        assert formula.count('_xlfn.XLOOKUP(') == 6
        wb.close()

    def test_guard_preserves_exact_match_semantics(self, ledger_xlsx):
        path = ledger_xlsx({"Jan": [(date(2025, 1, 4), "USD")]})
        wb = openpyxl.load_workbook(path)
        _inject(wb, N=30)
        formula = wb["Jan"].cell(row=2, column=3).value
        # Every lookup keeps if_not_found="" and match_mode=0 — exact
        # match only, NO rollback, NO carry-forward (approved contract).
        assert formula.count('"",0)') == formula.count('_xlfn.XLOOKUP(')
        assert formula.count('_xlfn.XLOOKUP(') == 4  # USD + EUR, guarded
        wb.close()


# =========================================================================
#  Date normalization
# =========================================================================

class TestDateNormalization:

    def test_string_date_rewritten_to_date_object(self, ledger_xlsx):
        path = ledger_xlsx({"Jan": [("10/03/2025", "USD")]})
        wb = openpyxl.load_workbook(path)
        # Sanity: stored as a text string before injection.
        assert isinstance(wb["Jan"].cell(row=2, column=1).value, str)
        _inject(wb, N=30)
        normalized = wb["Jan"].cell(row=2, column=1).value
        assert isinstance(normalized, (date, datetime))
        as_date = normalized.date() if isinstance(normalized, datetime) \
            else normalized
        assert as_date == date(2025, 3, 10)
        wb.close()


# =========================================================================
#  Extra-currency IFS branch
# =========================================================================

class TestExtraCurrencyBranch:

    def test_gbp_branch_appended_known_skipped(self, ledger_xlsx):
        path = ledger_xlsx({"Jan": [(date(2025, 1, 7), "GBP")]})
        wb = openpyxl.load_workbook(path)
        _inject(wb, N=30, exrate_col_map={
            "GBP": "F", "USD": "B", "EUR": "D", "THB": "Z",
        })
        formula = wb["Jan"].cell(row=2, column=3).value
        # GBP branch appended referencing column F.
        assert 'B2="GBP",' in formula
        assert "ExRate!$F$2:$F$30" in formula
        # USD/EUR/THB are NOT re-appended via the extra branch — they appear
        # exactly once (in the core branches).
        assert formula.count('B2="USD"') == 1
        assert formula.count('B2="EUR"') == 1
        assert formula.count('B2="THB"') == 1
        wb.close()

    def test_invalid_ccy_skipped(self, ledger_xlsx):
        path = ledger_xlsx({"Jan": [(date(2025, 1, 7), "GBP")]})
        wb = openpyxl.load_workbook(path)
        # Bad code (too long / lowercase) and bad column must be skipped.
        _inject(wb, N=30, exrate_col_map={"BADCCY": "F", "JPY": "9X!"})
        formula = wb["Jan"].cell(row=2, column=3).value
        assert "BADCCY" not in formula
        assert "JPY" not in formula
        wb.close()


# =========================================================================
#  MergedCell handling
# =========================================================================

class TestMergedCells:

    def test_merged_source_cell_skipped_no_crash(self, ledger_xlsx):
        path = ledger_xlsx({"Jan": [
            (date(2025, 1, 7), "USD"),
            (date(2025, 1, 8), "USD"),
        ]})
        wb = openpyxl.load_workbook(path)
        ws = wb["Jan"]
        # Merge the EX Rate cell on row 2 → its anchor stays writable but the
        # rest become MergedCell. Merge across C2:D2 so C2 is the anchor; then
        # merge A3 to force a MergedCell in the source column too.
        ws.merge_cells("A3:B3")
        # Should not raise.
        _inject(wb, N=30)
        # Row 2 (unmerged) gets a formula.
        assert isinstance(ws.cell(row=2, column=3).value, str)
        wb.close()


# =========================================================================
#  Idempotency
# =========================================================================

class TestIdempotency:

    def test_rerun_skips_identical(self, ledger_xlsx):
        path = ledger_xlsx({"Jan": [(date(2025, 1, 7), "USD")]})
        wb = openpyxl.load_workbook(path)
        _inject(wb, N=30)
        first = wb["Jan"].cell(row=2, column=3).value

        # Capture emit messages on the second pass to assert "skipped".
        msgs = []
        sheet_maps = scan_sheet_headers(wb, TARGET_COLS)
        inject_xlookup_formulas(
            wb, sheet_maps, exrate_last_row=30,
            parse_date_fn=parse_date,
            emit_fn=msgs.append,
        )
        # Formula unchanged.
        assert wb["Jan"].cell(row=2, column=3).value == first
        # The single row was skipped, not re-written.
        assert any("skipped" in m for m in msgs)
        wb.close()


# =========================================================================
#  build_exrate_index
# =========================================================================

class TestBuildExrateIndex:

    def test_no_exrate_sheet_returns_empty(self, ledger_xlsx):
        path = ledger_xlsx({"Jan": [(date(2025, 1, 7), "USD")]})
        wb = openpyxl.load_workbook(path)
        assert build_exrate_index(wb) == {}
        wb.close()

    def test_date_to_four_col_map(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws.append(["Date", "USD B", "USD S", "EUR B", "EUR S", "Hol"])
        ws.append([date(2025, 1, 7), 33.1, 33.5, 36.1, 36.5, ""])
        idx = build_exrate_index(wb)
        assert set(idx.keys()) == {date(2025, 1, 7)}
        row = idx[date(2025, 1, 7)]
        assert row["usd_buying"] == 33.1
        assert row["usd_selling"] == 33.5
        assert row["eur_buying"] == 36.1
        assert row["eur_selling"] == 36.5
        wb.close()

    def test_datetime_cell_coerced_to_date(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws.append(["Date", "USD B", "USD S", "EUR B", "EUR S", "Hol"])
        ws.append([datetime(2025, 1, 7, 9, 30), 33.1, 33.5, 36.1, 36.5, ""])
        idx = build_exrate_index(wb)
        assert date(2025, 1, 7) in idx
        wb.close()

    def test_extra_currency_columns_indexed(self):
        """With an exrate_col_map, extra-currency columns are indexed under
        'extra:<CCY>' so callers can detect blank multi-currency rows."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        # A=Date, B-E=USD/EUR, F=GBP Rate, G=JPY Rate, H=Holidays.
        ws.append(["Date", "USD B", "USD S", "EUR B", "EUR S",
                   "GBP Rate", "JPY Rate", "Hol"])
        ws.append([date(2025, 1, 7), 33.1, 33.5, 36.1, 36.5,
                   42.12, None, ""])
        idx = build_exrate_index(wb, {"GBP": "F", "JPY": "G"})
        row = idx[date(2025, 1, 7)]
        assert row["extra:GBP"] == 42.12
        # JPY cell is empty → indexed as None (a blank-resolving row).
        assert row["extra:JPY"] is None
        wb.close()

    def test_no_col_map_omits_extra_keys(self):
        """Backward compat: without a map only the fixed USD/EUR keys exist."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws.append(["Date", "USD B", "USD S", "EUR B", "EUR S", "Hol"])
        ws.append([date(2025, 1, 7), 33.1, 33.5, 36.1, 36.5, ""])
        row = build_exrate_index(wb)[date(2025, 1, 7)]
        assert set(row) == {
            "usd_buying", "usd_selling", "eur_buying", "eur_selling",
        }
        wb.close()


# =========================================================================
#  scan_sheet_headers
# =========================================================================

class TestScanSheetHeaders:

    def test_maps_date_cur_exrate(self, ledger_xlsx):
        path = ledger_xlsx({"Jan": [(date(2025, 1, 7), "USD")]})
        wb = openpyxl.load_workbook(path)
        maps = scan_sheet_headers(wb, TARGET_COLS)
        assert "Jan" in maps
        cols = maps["Jan"]["columns"]
        assert cols["source"] == 0
        assert cols["currency"] == 1
        assert cols["out_rate"] == 2
        wb.close()

    def test_skips_skip_sheet_names(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws.append(["Date", "Cur", "EX Rate"])
        ws.append([date(2025, 1, 7), "USD", None])
        maps = scan_sheet_headers(wb, TARGET_COLS)
        assert "ExRate" not in maps
        wb.close()

    def test_skips_sheet_missing_date_col(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jan"
        ws.append(["Cur", "Amount"])  # No Date column.
        ws.append(["USD", 1000])
        maps = scan_sheet_headers(wb, TARGET_COLS)
        assert "Jan" not in maps
        wb.close()

    def test_duplicate_header_uses_first_occurrence(self, caplog):
        """Two 'EX Rate' columns resolve deterministically to the first."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jan"
        # Columns: A=Date, B=Cur, C=EX Rate (first), D=EX Rate (duplicate).
        ws.append(["Date", "Cur", "EX Rate", "EX Rate"])
        ws.append([date(2025, 1, 7), "USD", None, None])
        with caplog.at_level("WARNING"):
            maps = scan_sheet_headers(wb, TARGET_COLS)
        cols = maps["Jan"]["columns"]
        # First "EX Rate" is column index 2 (C), never the duplicate at 3 (D).
        assert cols["out_rate"] == 2
        # The collision is logged so the operator can fix the sheet.
        assert any(
            "duplicate" in r.message.lower() and "EX Rate" in r.message
            for r in caplog.records
        )
        wb.close()


# =========================================================================
#  find_header_row (canonical header-scan primitive)
# =========================================================================

class TestFindHeaderRow:
    """Single owner behind scan_sheet_headers / ledger prescan / prescan."""

    def test_anchor_row_and_zero_based_columns(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Notes"])                      # row 1 — no anchor
        ws.append(["Date", "Cur", "EX Rate"])     # row 2 — header
        row_idx, cols = find_header_row(
            ws, (("source", "Date"), ("currency", "Cur")),
        )
        assert row_idx == 2
        assert cols == {"source": 0, "currency": 1}
        wb.close()

    def test_anchor_absent_returns_none(self):
        wb = openpyxl.Workbook()
        wb.active.append(["Cur", "Amount"])
        row_idx, cols = find_header_row(
            wb.active, (("source", "Date"),),
        )
        assert row_idx is None
        assert cols == {}
        wb.close()

    def test_none_label_is_skipped(self):
        wb = openpyxl.Workbook()
        wb.active.append(["Date", "Cur"])
        row_idx, cols = find_header_row(
            wb.active, (("source", "Date"), ("currency", None)),
        )
        assert row_idx == 1
        assert cols == {"source": 0}
        wb.close()

    def test_warn_duplicates_false_stays_silent(self, caplog):
        """The prescan path resolves first-wins WITHOUT an operator warning."""
        wb = openpyxl.Workbook()
        wb.active.append(["Date", "Date"])
        with caplog.at_level("WARNING"):
            row_idx, cols = find_header_row(
                wb.active, (("source", "Date"),), warn_duplicates=False,
            )
        assert row_idx == 1
        assert cols == {"source": 0}
        assert not any(
            "duplicate" in r.message.lower() for r in caplog.records
        )
        wb.close()

    def test_duplicate_date_resolves_left_of_ex_rate(self):
        """With two 'Date' columns, the source binds NEXT TO 'EX Rate'.

        Mirrors the real production ledgers: invoice Date in column B,
        Cur, then 'Export Entry' + its Date immediately left of EX Rate.
        The legacy in-file formulas resolve rates by the export-entry
        date (XLOOKUP(Q...)), so first-occurrence resolution silently
        switched the rate source to the invoice date — wrong values
        whenever the two dates differ.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            ["NO", "Date", "Inv.No.", "Cur", "Export Entry", "Date",
             "EX Rate"],
        )
        row_idx, cols = find_header_row(
            ws,
            (
                ("source", "Date"),
                ("currency", "Cur"),
                ("out_rate", "EX Rate"),
            ),
            resolve_left_of={"source": "out_rate"},
        )
        assert row_idx == 1
        assert cols["source"] == 5      # the export-entry Date, NOT col 1
        assert cols["currency"] == 3
        assert cols["out_rate"] == 6
        wb.close()

    def test_duplicate_date_without_anchor_falls_back_first(self):
        """resolve_left_of degrades to first-occurrence when the anchor
        column is absent from the sheet."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Cur", "Date"])  # no EX Rate column at all
        row_idx, cols = find_header_row(
            ws,
            (("source", "Date"), ("out_rate", "EX Rate")),
            resolve_left_of={"source": "out_rate"},
        )
        assert cols["source"] == 0
        wb.close()

    def test_duplicate_date_right_of_anchor_is_ignored(self):
        """Only occurrences LEFT of the anchor are eligible."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "EX Rate", "Date"])
        row_idx, cols = find_header_row(
            ws,
            (("source", "Date"), ("out_rate", "EX Rate")),
            resolve_left_of={"source": "out_rate"},
        )
        assert cols["source"] == 0
        wb.close()

    def test_single_date_column_unaffected(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Cur", "EX Rate"])
        row_idx, cols = find_header_row(
            ws,
            (
                ("source", "Date"),
                ("currency", "Cur"),
                ("out_rate", "EX Rate"),
            ),
            resolve_left_of={"source": "out_rate"},
        )
        assert cols["source"] == 0
        wb.close()

    def test_duplicate_resolution_is_logged(self, caplog):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Date", "EX Rate"])
        with caplog.at_level("WARNING"):
            find_header_row(
                ws,
                (("source", "Date"), ("out_rate", "EX Rate")),
                resolve_left_of={"source": "out_rate"},
            )
        assert any(
            "nearest left" in r.message.lower() for r in caplog.records
        )
        wb.close()


class TestInjectShiftedDualDateLayout:
    """The INJECTED formula must reference the export-entry Date column.

    Locks the production layout end-to-end at the write seam: invoice Date
    in column B, Cur in D, export-entry Date in E, EX Rate in F. The
    emitted formula's guard and XLOOKUPs must reference E (the resolved
    source), never B — and the invoice column's cells must be untouched.
    The find_header_row unit tests alone could not catch a regression in
    how scan_sheet_headers' indices flow into the formula builder.
    """

    def test_formula_references_export_entry_column(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jan"
        ws.append(["NO", "Date", "Vendor", "Cur", "Date", "EX Rate"])
        invoice = date(2025, 1, 7)
        entry = date(2025, 2, 14)
        ws.append([1, invoice, "ACME", "USD", entry, None])

        _inject(wb)

        formula = ws.cell(row=2, column=6).value
        assert isinstance(formula, str)
        # Guard + lookups bind to E (export-entry Date) and D (Cur)...
        assert formula.startswith('=IF(OR(D2="",E2=""),"",')
        assert "_xlfn.XLOOKUP(E2," in formula
        # ...and NEVER to the invoice Date in column B.
        assert "B2" not in formula
        # The invoice Date cell itself is untouched by date normalization.
        assert ws.cell(row=2, column=2).value == invoice
        wb.close()

    def test_scan_depth_bounds_the_search(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(5):
            ws.append(["filler"])
        ws.append(["Date", "Cur"])                # row 6 — beyond depth 5
        row_idx, _cols = find_header_row(
            ws, (("source", "Date"),), scan_depth=5,
        )
        assert row_idx is None
        wb.close()


# =========================================================================
#  write_custom_exrate_data
# =========================================================================

class TestWriteCustomExrateData:

    def test_multi_currency_headers_values_and_clear(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        # Prior content that must be wiped by the delete_rows clear.
        ws.append(["STALE", "STALE", "STALE", "STALE", "STALE"])
        ws.append(["old", "old", "old", "old", "old"])

        d1 = date(2025, 1, 7)
        col_specs = [
            ("GBP", "buying_transfer"),
            ("GBP", "selling"),
            ("JPY", "buying_transfer"),
            ("JPY", "selling"),
        ]
        headers = [
            "Date",
            "GBP Buying TT", "GBP Selling",
            "JPY Buying TT", "JPY Selling",
            "Holidays/Weekend",
        ]
        rate_data = {
            "GBP": {
                "buying_transfer": {d1: 42.1},
                "selling": {d1: 42.9},
            },
            "JPY": {
                "buying_transfer": {d1: 0.21},
                "selling": {d1: 0.23},
            },
        }
        write_custom_exrate_data(
            ws, rate_data, col_specs, headers,
            all_dates=[d1], holidays_set=set(), holidays_names={},
        )

        # Headers row 1.
        assert [ws.cell(row=1, column=c).value for c in range(1, 7)] == headers
        # Stale content gone — no "STALE"/"old" anywhere.
        seen = {
            ws.cell(row=r, column=c).value
            for r in range(1, (ws.max_row or 1) + 1)
            for c in range(1, (ws.max_column or 1) + 1)
        }
        assert "STALE" not in seen and "old" not in seen
        # Data row 2: Date + 4 rate cols at correct positions.
        assert ws.cell(row=2, column=1).value == d1
        assert ws.cell(row=2, column=2).value == 42.1   # GBP buying
        assert ws.cell(row=2, column=3).value == 42.9   # GBP selling
        assert ws.cell(row=2, column=4).value == 0.21   # JPY buying
        assert ws.cell(row=2, column=5).value == 0.23   # JPY selling
        wb.close()

    def test_clear_handles_empty_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        # Fresh empty sheet — delete_rows path must not crash.
        write_custom_exrate_data(
            ws, {}, [], ["Date", "Holidays/Weekend"],
            all_dates=[], holidays_set=set(), holidays_names={},
        )
        assert ws.cell(row=1, column=1).value == "Date"
        wb.close()
