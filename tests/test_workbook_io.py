#!/usr/bin/env python3
"""
tests/test_workbook_io.py
---------------------------------------------------------------------------
Unit tests for core/workbook_io.py — the in-place save guardrails.
Covers ensure_disk_space (free-space guard), atomic_save (crash-safe
in-place overwrite via temp file + os.replace), the pre-replace verify
hook (Layer-1 exactness hard gate), and build_cell_verifier.
---------------------------------------------------------------------------
"""

from collections import namedtuple
from decimal import Decimal

import openpyxl
import pytest

import core.workbook_io as workbook_io_mod
from core.workbook_io import (
    WorkbookVerifyError,
    atomic_save,
    build_cell_verifier,
    ensure_disk_space,
)

_DiskUsage = namedtuple("_DiskUsage", ["total", "used", "free"])


# =========================================================================
#  ensure_disk_space
# =========================================================================

class TestEnsureDiskSpace:

    def test_passes_when_space_available(self, tmp_path, monkeypatch):
        monkeypatch.setattr(
            workbook_io_mod.shutil, "disk_usage",
            lambda _p: _DiskUsage(total=10**12, used=0, free=10**12),
        )
        # Plenty of space → no raise.
        ensure_disk_space(tmp_path, 100)

    def test_raises_when_space_low(self, tmp_path, monkeypatch):
        monkeypatch.setattr(
            workbook_io_mod.shutil, "disk_usage",
            lambda _p: _DiskUsage(total=10**12, used=10**12, free=0),
        )
        with pytest.raises(OSError, match="Insufficient disk space"):
            ensure_disk_space(tmp_path, 100)


# =========================================================================
#  atomic_save
# =========================================================================

class TestAtomicSave:

    def _seed_workbook(self, path, marker):
        """Write a small workbook with a sentinel value to a path."""
        wb = openpyxl.Workbook()
        wb.active["A1"] = marker
        wb.save(str(path))
        wb.close()

    def test_success_replaces_content(self, tmp_path):
        target = tmp_path / "ledger.xlsx"
        self._seed_workbook(target, "ORIGINAL")

        # New in-memory workbook with fresh content.
        wb = openpyxl.Workbook()
        wb.active["A1"] = "UPDATED"
        atomic_save(wb, str(target))
        wb.close()

        # Content swapped in.
        reloaded = openpyxl.load_workbook(str(target))
        assert reloaded.active["A1"].value == "UPDATED"
        reloaded.close()
        # No temp file left behind.
        assert not (tmp_path / "ledger.xlsx.tmp~").exists()

    def test_failure_leaves_original_unchanged(self, tmp_path):
        target = tmp_path / "ledger.xlsx"
        self._seed_workbook(target, "ORIGINAL")
        with open(target, "rb") as fh:
            original_bytes = fh.read()

        class _BoomWorkbook:
            """A workbook whose save() raises AFTER writing the temp file,
            simulating a crash mid-save."""
            def save(self, path):
                # Touch the temp file (as a real save would start to) then fail.
                with open(path, "wb") as fh:
                    fh.write(b"PARTIAL")
                raise OSError("simulated mid-save crash")

        with pytest.raises(OSError, match="simulated mid-save crash"):
            atomic_save(_BoomWorkbook(), str(target))

        # Original on disk is byte-for-byte intact — os.replace never ran.
        with open(target, "rb") as fh:
            assert fh.read() == original_bytes
        # The partial temp file was cleaned up.
        assert not (tmp_path / "ledger.xlsx.tmp~").exists()


# =========================================================================
#  atomic_save — verify hook (Layer-1 exactness hard gate)
# =========================================================================

class TestAtomicSaveVerify:

    def _seed_workbook(self, path, marker):
        """Write a small workbook with a sentinel value to a path."""
        wb = openpyxl.Workbook()
        wb.active["A1"] = marker
        wb.save(str(path))
        wb.close()

    def test_verify_error_is_oserror_subclass(self):
        # Existing save-failure handlers catch OSError; the verification
        # failure must flow through the same error path.
        assert issubclass(WorkbookVerifyError, OSError)

    def test_verifier_failure_unlinks_tmp_and_preserves_original(self, tmp_path):
        target = tmp_path / "ledger.xlsx"
        self._seed_workbook(target, "ORIGINAL")
        with open(target, "rb") as fh:
            original_bytes = fh.read()

        wb = openpyxl.Workbook()
        wb.active["A1"] = "UPDATED"

        def _boom(_reopened):
            raise ValueError("B2: expected 32.4507, saved file holds 32.0")

        with pytest.raises(
            WorkbookVerifyError,
            match="Post-write verification failed — original file left untouched",
        ):
            atomic_save(wb, str(target), verify=_boom)
        wb.close()

        # Original on disk is byte-for-byte intact — the replace never ran.
        with open(target, "rb") as fh:
            assert fh.read() == original_bytes
        # The rejected temp file was unlinked.
        assert not (tmp_path / "ledger.xlsx.tmp~").exists()

    def test_honest_save_passes_verification(self, tmp_path):
        target = tmp_path / "ledger.xlsx"
        self._seed_workbook(target, "ORIGINAL")
        with open(target, "rb") as fh:
            original_bytes = fh.read()

        wb = openpyxl.Workbook()
        wb.active["A1"] = "UPDATED"

        seen = {}

        def _check(reopened):
            # The verifier runs against the TEMP file BEFORE the replace —
            # at this moment the target must still hold the original bytes.
            with open(target, "rb") as fh:
                assert fh.read() == original_bytes
            seen["read_only"] = reopened.read_only
            row1 = next(reopened.active.iter_rows(values_only=True))
            seen["a1"] = row1[0]

        atomic_save(wb, str(target), verify=_check)
        wb.close()

        assert seen["a1"] == "UPDATED", "verifier saw the saved temp content"
        assert seen["read_only"] is True
        reloaded = openpyxl.load_workbook(str(target))
        assert reloaded.active["A1"].value == "UPDATED"
        reloaded.close()
        assert not (tmp_path / "ledger.xlsx.tmp~").exists()


# =========================================================================
#  build_cell_verifier
# =========================================================================

class TestBuildCellVerifier:

    FORMULA = '=IF(OR(B2=""),"",_xlfn.IFS(TRUE,""))'

    def _build_wb(self, usd_value):
        """In-memory ExRate workbook: Decimal rate, formula, blank cell."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws["A2"] = "2026-01-02"
        ws["B2"] = usd_value
        ws["C2"] = self.FORMULA
        return wb

    def test_matching_cells_pass(self, tmp_path):
        target = tmp_path / "exrate.xlsx"
        wb = self._build_wb(Decimal("32.4507"))
        expected = {
            "ExRate": {
                2: {2: Decimal("32.4507"), 3: self.FORMULA, 4: None},
            },
        }
        atomic_save(wb, str(target), verify=build_cell_verifier(expected))
        wb.close()
        assert target.exists()

    def test_corruption_simulation_fails_hard(self, tmp_path):
        # The workbook holds 32.0000 but the verifier expects the intended
        # 4dp Decimal 32.4507 — hard failure, nothing persisted.
        target = tmp_path / "exrate.xlsx"
        wb_seed = openpyxl.Workbook()
        wb_seed.active["A1"] = "ORIGINAL"
        wb_seed.save(str(target))
        wb_seed.close()
        with open(target, "rb") as fh:
            original_bytes = fh.read()

        wb = self._build_wb(Decimal("32.0000"))
        expected = {"ExRate": {2: {2: Decimal("32.4507")}}}
        with pytest.raises(WorkbookVerifyError, match="ExRate!B2"):
            atomic_save(wb, str(target), verify=build_cell_verifier(expected))
        wb.close()

        with open(target, "rb") as fh:
            assert fh.read() == original_bytes
        assert not (tmp_path / "exrate.xlsx.tmp~").exists()

    def test_formula_mismatch_fails(self, tmp_path):
        target = tmp_path / "exrate.xlsx"
        wb = self._build_wb(Decimal("32.4507"))
        expected = {"ExRate": {2: {3: '=SUM(A1:A2)'}}}
        with pytest.raises(WorkbookVerifyError, match="ExRate!C2"):
            atomic_save(wb, str(target), verify=build_cell_verifier(expected))
        wb.close()
        assert not target.exists()

    def test_missing_sheet_fails(self, tmp_path):
        target = tmp_path / "exrate.xlsx"
        wb = self._build_wb(Decimal("32.4507"))
        expected = {"Ghost": {2: {2: Decimal("1.0000")}}}
        with pytest.raises(WorkbookVerifyError, match="'Ghost' missing"):
            atomic_save(wb, str(target), verify=build_cell_verifier(expected))
        wb.close()
        assert not target.exists()

    def test_expected_row_beyond_sheet_fails_unless_blank(self, tmp_path):
        target = tmp_path / "exrate.xlsx"

        # A blank expectation on a missing row passes...
        wb = self._build_wb(Decimal("32.4507"))
        ok = {"ExRate": {99: {2: None}}}
        atomic_save(wb, str(target), verify=build_cell_verifier(ok))
        wb.close()

        # ...but a value expectation on a missing row is a hard failure.
        wb2 = self._build_wb(Decimal("32.4507"))
        bad = {"ExRate": {99: {2: Decimal("32.4507")}}}
        with pytest.raises(WorkbookVerifyError, match="row missing"):
            atomic_save(wb2, str(target), verify=build_cell_verifier(bad))
        wb2.close()
