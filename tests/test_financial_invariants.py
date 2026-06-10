#!/usr/bin/env python3
# =========================================================================
#  FROZEN FINANCIAL INVARIANTS — business-approved contract.
#  Do NOT modify assertions to make a code change pass; a change here
#  requires an explicit approval note in the commit message (ADR).
# =========================================================================
"""Spec-level financial invariants exercised across every writer path.

The five invariants locked by this module:

  1. Weekend/holiday rows in any ExRate sheet are Date + label only —
     every rate cell (USD/EUR and extra-currency columns) stays blank.
     No carry-forward or fabricated rates, on any writer path.
  2. Every rate written on a trading day round-trips after save+reload
     as the exact 4dp source Decimal — no float drift beyond 4dp.
  3. safe_to_decimal / format_rate_value quantize to exactly 4 decimal
     places, always via string-constructed Decimal (never from float).
  4. The rate auditor never touches weekend/holiday rows, even when a
     cell on such a row holds a wrong value and BOT has data for it.
  5. Anomaly detection is alert-only: a >threshold jump is reported but
     the anomalous value is still written to the file unchanged.

Writer paths covered: update_master_exrate_sheet (full build and merge
over existing data), write_custom_exrate_data via the
StandaloneExRateUpdater custom path, and core/rate_audit.py scan/apply.
"""
import asyncio
from datetime import date
from decimal import Decimal
from types import SimpleNamespace
from unittest.mock import MagicMock

import openpyxl

from core.anomaly_guard import AnomalyGuard
from core.constants import format_rate_value, parse_date
from core.exrate_sheet import update_master_exrate_sheet
from core.exrate_updater import StandaloneExRateUpdater
from core.ledger_processing import run_anomaly_check
from core.logic import safe_to_decimal
from core.rate_audit import (
    apply_corrections,
    rate_key,
    scan_exrate_corrections,
)

D = Decimal

# Fixed calendar window (March 2026):
THU = date(2026, 3, 5)   # trading day
FRI = date(2026, 3, 6)   # trading day
SAT = date(2026, 3, 7)   # weekend
SUN = date(2026, 3, 8)   # weekend
HOL = date(2026, 3, 9)   # Monday — declared a BOT holiday in fixtures
TUE = date(2026, 3, 10)  # trading day

HOLIDAY_NAME = "Test Holiday"

# ExRate fixed layout: A=Date, B-E=USD/EUR rates, then extras, then label.
USD_EUR_COLS = (2, 3, 4, 5)

AUDIT_HEADERS = [
    "Date", "USD Buying TT Rate", "USD Selling Rate",
    "EUR Buying TT Rate", "EUR Selling Rate", "Holidays/Weekend",
]


# =========================================================================
#  SHARED HELPERS
# =========================================================================

def _rows_by_date(ws, date_col=1, start_row=2):
    """Map each parsed Date-cell value to its 1-based row index."""
    rows = {}
    for r in range(start_row, (ws.max_row or 1) + 1):
        d = parse_date(ws.cell(row=r, column=date_col).value)
        if d is not None:
            rows[d] = r
    return rows


def _build_standard_sheet(wb, *, extra=False):
    """Run the full master-sheet build over THU..TUE with HOL a holiday.

    Trading days THU/FRI/TUE carry rates; SAT/SUN/HOL must stay blank.
    """
    extra_rates = (
        {"GBP": {THU: D("42.1234"), FRI: D("42.5678"), TUE: D("42.9999")}}
        if extra else None
    )
    return update_master_exrate_sheet(
        wb,
        usd_buying_rates={THU: D("34.5650"), FRI: D("34.6000"),
                          TUE: D("34.7000")},
        usd_selling_rates={THU: D("34.7350"), FRI: D("34.8000"),
                           TUE: D("34.9000")},
        eur_buying_rates={THU: D("37.1250"), FRI: D("37.2000"),
                          TUE: D("37.3000")},
        eur_selling_rates={THU: D("37.4450"), FRI: D("37.5000"),
                           TUE: D("37.6000")},
        holidays_list=[HOL],
        holidays_names={HOL: HOLIDAY_NAME},
        start_date=THU,
        end_date=TUE,
        extra_currency_rates=extra_rates,
    )


def _custom_engine(records):
    """Engine stub exposing exactly what the custom updater dereferences.

    HOL is the only holiday; records are returned by the fake API verbatim.
    """
    logic_engine = SimpleNamespace(holidays={HOL})

    async def _preload(_dates, _start):
        return (logic_engine, {}, {}, {}, {}, {}, {})

    async def _get_rates(_start, _end, _ccy):
        return records

    backup = MagicMock()
    backup.create_backup.return_value = "/tmp/exrate.bak.xlsx"
    return SimpleNamespace(
        _check_memory_guardrail=lambda _fp: None,
        _parse_date=parse_date,
        _preload_api_data=_preload,
        _emit=lambda *_a, **_k: None,
        api=SimpleNamespace(get_exchange_rates=_get_rates),
        cache=SimpleNamespace(
            get_holidays=lambda _year: [
                (HOL.strftime("%Y-%m-%d"), HOLIDAY_NAME)
            ],
        ),
        backup=backup,
    )


def _run_custom_updater(tmp_path, records):
    """Drive StandaloneExRateUpdater down the CUSTOM (non-standard) path.

    GBP-only with a single rate type is non-standard, so the run uses
    write_custom_exrate_data. Returns the saved file path.
    """
    fp = tmp_path / "custom_exrate.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "ExRate"
    wb.save(str(fp))
    wb.close()

    updater = StandaloneExRateUpdater(_custom_engine(records))
    asyncio.run(updater.run(
        str(fp),
        progress_cb=None,
        currencies=["GBP"],
        rate_types={"Buying TT": "buying_transfer"},
        date_range=(THU, TUE),
    ))
    return str(fp)


# =========================================================================
#  INVARIANT 1 — weekend/holiday rows are Date + label ONLY (all blank)
# =========================================================================

class TestWeekendHolidayRowsBlank:
    """No writer path may fabricate or carry forward a weekend/holiday rate."""

    def test_full_build_weekend_and_holiday_rate_cells_blank(self, tmp_path):
        wb = openpyxl.Workbook()
        _build_standard_sheet(wb, extra=True)
        fp = tmp_path / "full_build.xlsx"
        wb.save(str(fp))
        wb.close()

        wb2 = openpyxl.load_workbook(str(fp))
        ws = wb2["ExRate"]
        rows = _rows_by_date(ws)
        # Extra GBP column is F (6); label column is G (7).
        for d, label in ((SAT, "Weekend"), (SUN, "Weekend"),
                         (HOL, HOLIDAY_NAME)):
            r = rows[d]
            for col in (*USD_EUR_COLS, 6):
                assert ws.cell(row=r, column=col).value is None, (
                    f"{d} col {col} must be blank — no fabricated rate"
                )
            assert ws.cell(row=r, column=7).value == label
        # Adjacent trading days DID get their own rates (sanity check).
        assert ws.cell(row=rows[FRI], column=2).value is not None
        assert ws.cell(row=rows[TUE], column=2).value is not None
        wb2.close()

    def test_merge_over_existing_data_keeps_weekend_blank(self, tmp_path):
        # First build writes THU..TUE; the second (merge) pass supplies only
        # TUE rates. The weekend/holiday rows from the existing sheet must
        # re-emerge blank — the prior FRI rate is never carried into them.
        wb = openpyxl.Workbook()
        _build_standard_sheet(wb)
        col_map = update_master_exrate_sheet(
            wb,
            usd_buying_rates={TUE: D("35.0000")},
            usd_selling_rates={TUE: D("35.1000")},
            eur_buying_rates={TUE: D("37.9000")},
            eur_selling_rates={TUE: D("38.0000")},
            holidays_list=[HOL],
            holidays_names={HOL: HOLIDAY_NAME},
            start_date=THU,
            end_date=TUE,
        )
        assert col_map == {}
        fp = tmp_path / "merged.xlsx"
        wb.save(str(fp))
        wb.close()

        wb2 = openpyxl.load_workbook(str(fp))
        ws = wb2["ExRate"]
        rows = _rows_by_date(ws)
        for d in (SAT, SUN, HOL):
            for col in USD_EUR_COLS:
                assert ws.cell(row=rows[d], column=col).value is None
        # The merge updated TUE (API wins) and preserved THU from the sheet.
        assert safe_to_decimal(
            ws.cell(row=rows[TUE], column=2).value
        ) == D("35.0000")
        assert safe_to_decimal(
            ws.cell(row=rows[THU], column=2).value
        ) == D("34.5650")
        wb2.close()

    def test_custom_writer_weekend_and_holiday_rows_blank(self, tmp_path):
        records = [
            SimpleNamespace(period="2026-03-05", buying_transfer=42.1234),
            SimpleNamespace(period="2026-03-06", buying_transfer=42.5678),
            SimpleNamespace(period="2026-03-10", buying_transfer=42.9999),
        ]
        fp = _run_custom_updater(tmp_path, records)

        wb = openpyxl.load_workbook(fp)
        ws = wb["ExRate"]
        assert ws.cell(row=1, column=2).value == "GBP Buying TT"
        rows = _rows_by_date(ws)
        for d, label in ((SAT, "Weekend"), (SUN, "Weekend"),
                         (HOL, HOLIDAY_NAME)):
            r = rows[d]
            assert ws.cell(row=r, column=2).value is None, (
                f"{d} GBP cell must be blank — no carry-forward"
            )
            assert ws.cell(row=r, column=3).value == label
        wb.close()


# =========================================================================
#  INVARIANT 2 — trading-day rates round-trip exactly at 4dp
# =========================================================================

class TestDecimalRoundTrip:
    """Saved rate cells reload to the exact 4dp string-built source Decimal."""

    def test_full_build_round_trips_every_rate_cell(self, tmp_path):
        wb = openpyxl.Workbook()
        _build_standard_sheet(wb, extra=True)
        fp = tmp_path / "roundtrip.xlsx"
        wb.save(str(fp))
        wb.close()

        expected = {
            THU: [D("34.5650"), D("34.7350"), D("37.1250"), D("37.4450"),
                  D("42.1234")],
            FRI: [D("34.6000"), D("34.8000"), D("37.2000"), D("37.5000"),
                  D("42.5678")],
            TUE: [D("34.7000"), D("34.9000"), D("37.3000"), D("37.6000"),
                  D("42.9999")],
        }
        wb2 = openpyxl.load_workbook(str(fp))
        ws = wb2["ExRate"]
        rows = _rows_by_date(ws)
        for d, values in expected.items():
            for col, source in zip((2, 3, 4, 5, 6), values):
                cell = ws.cell(row=rows[d], column=col)
                got = Decimal(str(cell.value))
                assert got == source, f"{d} col {col}: {got} != {source}"
                # No drift beyond 4dp: re-quantizing changes nothing.
                assert got == got.quantize(D("0.0000"))
                assert cell.number_format == "0.0000"
        wb2.close()

    def test_custom_writer_round_trips_quantized_api_floats(self, tmp_path):
        # The fake API hands back raw floats; the writer must persist the
        # 4dp string-built Decimal (safe_to_decimal), not the raw float.
        records = [
            SimpleNamespace(period="2026-03-05", buying_transfer=34.565),
            SimpleNamespace(period="2026-03-06", buying_transfer=42.56789),
        ]
        fp = _run_custom_updater(tmp_path, records)

        wb = openpyxl.load_workbook(fp)
        ws = wb["ExRate"]
        rows = _rows_by_date(ws)
        thu = Decimal(str(ws.cell(row=rows[THU], column=2).value))
        fri = Decimal(str(ws.cell(row=rows[FRI], column=2).value))
        assert thu == D("34.5650")
        assert fri == D("42.5679")  # quantized to 4dp before the write
        assert ws.cell(row=rows[THU], column=2).number_format == "0.0000"
        wb.close()

    def test_audit_correction_round_trips_after_save(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws.append(AUDIT_HEADERS)
        ws.append([THU, D("99.0000"), None, None, None, ""])
        bot = {rate_key("USD", "buying_transfer"): {THU: D("32.4507")}}
        report = scan_exrate_corrections(ws, bot, set())
        apply_corrections(ws, report)
        fp = tmp_path / "audited.xlsx"
        wb.save(str(fp))
        wb.close()

        wb2 = openpyxl.load_workbook(str(fp))
        cell = wb2["ExRate"].cell(row=2, column=2)
        assert Decimal(str(cell.value)) == D("32.4507")
        assert cell.number_format == "0.0000"
        wb2.close()


# =========================================================================
#  INVARIANT 3 — 4dp quantization helpers, string-built Decimal only
# =========================================================================

class TestQuantizationHelpers:
    """safe_to_decimal / format_rate_value lock the 4dp Decimal discipline."""

    def test_string_input_exact_4dp(self):
        val = safe_to_decimal("34.5650")
        assert val == D("34.5650")
        assert str(val) == "34.5650"
        assert val.as_tuple().exponent == -4

    def test_float_input_goes_through_str_never_binary(self):
        # Decimal(34.565) directly would be 34.564999999999998... —
        # safe_to_decimal must build from str(value) and quantize to 4dp.
        val = safe_to_decimal(34.565)
        assert val == D("34.5650")
        assert str(val) == "34.5650"
        # The binary-float artifact never leaks into the result.
        assert str(val) != str(Decimal(34.565))

    def test_noisy_legacy_float_quantizes_to_4dp(self):
        assert safe_to_decimal(32.50009999) == D("32.5001")

    def test_short_input_padded_to_exactly_4dp(self):
        val = safe_to_decimal("33.5")
        assert str(val) == "33.5000"
        assert val.as_tuple().exponent == -4

    def test_over_precise_decimal_requantized(self):
        assert safe_to_decimal(D("32.45071")) == D("32.4507")
        assert safe_to_decimal(D("32.45079")) == D("32.4508")
        assert safe_to_decimal(D("32.45079")).as_tuple().exponent == -4

    def test_empty_inputs_return_none(self):
        assert safe_to_decimal(None) is None
        assert safe_to_decimal("") is None

    def test_format_rate_value_decimal_exact(self):
        assert format_rate_value(D("34.5650")) == "34.5650"
        assert format_rate_value(D("33.5")) == "33.5000"

    def test_format_rate_value_float_and_none(self):
        assert format_rate_value(34.56499999999999) == "34.5650"
        assert format_rate_value(None) == ""


# =========================================================================
#  INVARIANT 4 — rate audit never touches weekend/holiday rows
# =========================================================================

class TestAuditSkipsNonTradingRows:
    """A wrong value on a weekend/holiday row is out of the auditor's scope."""

    def test_wrong_saturday_value_gets_zero_corrections(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws.append(AUDIT_HEADERS)
        ws.append([SAT, D("99.9999"), None, None, None, "Weekend"])
        # BOT (anomalously) has a Saturday value — the row is still skipped.
        bot = {rate_key("USD", "buying_transfer"): {SAT: D("32.4507")}}
        report = scan_exrate_corrections(ws, bot, set())
        assert report.change_count == 0
        assert report.compared_cells == 0
        apply_corrections(ws, report)
        # The wrong value survives untouched: out of audit scope by contract.
        assert ws.cell(row=2, column=2).value == D("99.9999")
        wb.close()

    def test_wrong_holiday_value_gets_zero_corrections(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws.append(AUDIT_HEADERS)
        ws.append([HOL, D("99.9999"), None, None, None, HOLIDAY_NAME])
        bot = {rate_key("USD", "buying_transfer"): {HOL: D("32.4507")}}
        report = scan_exrate_corrections(ws, bot, {HOL})
        assert report.change_count == 0
        apply_corrections(ws, report)
        assert ws.cell(row=2, column=2).value == D("99.9999")
        wb.close()

    def test_only_trading_day_corrected_in_mixed_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws.append(AUDIT_HEADERS)
        ws.append([FRI, D("11.1111"), None, None, None, ""])
        ws.append([SAT, D("99.9999"), None, None, None, "Weekend"])
        ws.append([HOL, D("88.8888"), None, None, None, HOLIDAY_NAME])
        bot = {rate_key("USD", "buying_transfer"): {
            FRI: D("34.6000"), SAT: D("34.6000"), HOL: D("34.6000"),
        }}
        report = scan_exrate_corrections(ws, bot, {HOL})
        assert report.change_count == 1
        assert report.changes[0].rate_date == FRI
        apply_corrections(ws, report)
        assert ws.cell(row=2, column=2).value == D("34.6000")
        assert ws.cell(row=3, column=2).value == D("99.9999")  # weekend kept
        assert ws.cell(row=4, column=2).value == D("88.8888")  # holiday kept
        wb.close()


# =========================================================================
#  INVARIANT 5 — anomaly detection is alert-only, never blocks the write
# =========================================================================

class TestAnomalyAlertOnly:
    """A >threshold jump is flagged, yet the value reaches the file as-is."""

    def test_guard_flags_jump_but_sheet_carries_anomalous_value(
        self, tmp_path
    ):
        # THU -> FRI jumps ~9% — well above the 5% threshold.
        usd_buying = {THU: D("33.0000"), FRI: D("36.0000")}
        guard = AnomalyGuard(threshold_pct=5.0)
        anomalies = guard.check_rates_bulk(
            {"USD_buying_transfer": usd_buying}
        )
        assert len(anomalies) == 1
        assert anomalies[0].is_anomaly is True
        assert anomalies[0].new_value == D("36.0000")

        # The anomalous value is still written unchanged to the sheet.
        wb = openpyxl.Workbook()
        update_master_exrate_sheet(
            wb,
            usd_buying_rates=usd_buying,
            usd_selling_rates={},
            eur_buying_rates={},
            eur_selling_rates={},
            holidays_list=[],
            holidays_names={},
            start_date=THU,
            end_date=FRI,
        )
        fp = tmp_path / "anomalous.xlsx"
        wb.save(str(fp))
        wb.close()

        wb2 = openpyxl.load_workbook(str(fp))
        ws = wb2["ExRate"]
        rows = _rows_by_date(ws)
        assert Decimal(str(ws.cell(row=rows[FRI], column=2).value)) == (
            D("36.0000")
        )
        wb2.close()

    def test_run_anomaly_check_emits_warning_and_returns_count(self):
        # The engine-facing wrapper reports via emit_fn and returns the
        # anomaly count — it has no channel to veto or rewrite the rates.
        emitted = []

        def _emit(msg, etype):
            emitted.append((msg, etype))

        count = run_anomaly_check(
            AnomalyGuard(threshold_pct=5.0),
            _emit,
            usd_buying={THU: D("33.0000"), FRI: D("36.0000")},
            usd_selling={},
            eur_buying={},
            eur_selling={},
        )
        assert count == 1
        assert len(emitted) == 1
        assert emitted[0][1] == "warning"
        assert "ANOMALY" in emitted[0][0]

    def test_no_anomaly_within_threshold(self):
        guard = AnomalyGuard(threshold_pct=5.0)
        anomalies = guard.check_rates_bulk(
            {"USD_buying_transfer": {THU: D("33.0000"), FRI: D("33.5000")}}
        )
        assert anomalies == []
