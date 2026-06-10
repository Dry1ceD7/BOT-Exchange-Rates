#!/usr/bin/env python3
"""Integration tests for StandaloneRateAuditor.run — orchestration only.

A fake engine supplies canned BOT rates (no network); a real .xlsx on disk is
read, compared, and (optionally) rewritten. Asserts the end-to-end contract:
wrong trading-day cells are corrected, weekend rows are never touched, the file
is backed up before any write, and a dry run writes nothing.
"""
import asyncio
from datetime import date
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock

import openpyxl
import pytest

import core.rate_audit as rate_audit_mod
from core.constants import parse_date
from core.logic import safe_to_decimal
from core.rate_audit import StandaloneRateAuditor
from core.workbook_io import WorkbookVerifyError

HEADERS = [
    "Date", "USD Buying TT Rate", "USD Selling Rate",
    "EUR Buying TT Rate", "EUR Selling Rate", "Holidays/Weekend",
]
WED = date(2026, 5, 27)   # trading day
SAT = date(2026, 5, 23)   # weekend


def _make_xlsx(tmp_path, usd_buy_value):
    """ExRate sheet: one trading-day row (USD buy = usd_buy_value) + a blank
    weekend row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExRate"
    ws.append(HEADERS)
    ws.append([WED, usd_buy_value, Decimal("32.7790"),
               Decimal("37.0000"), Decimal("37.5000"), ""])
    ws.append([SAT, None, None, None, None, "Weekend"])
    fp = tmp_path / "ledger.xlsx"
    wb.save(str(fp))
    wb.close()
    return str(fp)


def _fake_engine():
    """Engine stub exposing exactly what StandaloneRateAuditor dereferences."""
    usd_buying = {WED: Decimal("32.4507")}
    usd_selling = {WED: Decimal("32.7790")}
    eur_buying = {WED: Decimal("37.0000")}
    eur_selling = {WED: Decimal("37.5000")}
    logic_engine = SimpleNamespace(holidays=set())

    async def _preload(_dates, _start, *, extend_to_today=True):
        # F60: the audit is bounded by the sheet's own dates — it must never
        # ask the preload to extend the fetch window to today.
        assert extend_to_today is False
        return (logic_engine, usd_selling, eur_selling,
                usd_buying, eur_buying, {}, {})

    backup = MagicMock()
    backup.create_backup.return_value = "/tmp/ledger.bak.xlsx"
    return SimpleNamespace(
        _check_memory_guardrail=lambda _fp: None,
        _parse_date=parse_date,
        _preload_api_data=_preload,
        backup=backup,
    )


def test_run_corrects_wrong_cell_and_backs_up(tmp_path):
    eng = _fake_engine()
    fp = _make_xlsx(tmp_path, Decimal("32.0000"))  # USD buy is WRONG
    report = asyncio.run(StandaloneRateAuditor(eng).run(fp, apply=True))

    assert report.change_count == 1
    ch = report.changes[0]
    assert ch.cell == "B2"
    assert ch.new_value == Decimal("32.4507")
    assert report.applied is True
    eng.backup.create_backup.assert_called_once()

    wb = openpyxl.load_workbook(fp)
    ws = wb["ExRate"]
    assert safe_to_decimal(ws.cell(row=2, column=2).value) == Decimal("32.4507")
    # Weekend row stays blank — never filled.
    assert ws.cell(row=3, column=2).value is None
    wb.close()


def test_run_no_changes_when_already_correct(tmp_path):
    eng = _fake_engine()
    fp = _make_xlsx(tmp_path, Decimal("32.4507"))  # already correct
    report = asyncio.run(StandaloneRateAuditor(eng).run(fp, apply=True))
    assert report.change_count == 0
    # Nothing differs → no write and therefore no backup (avoids clutter).
    eng.backup.create_backup.assert_not_called()


def test_dry_run_writes_nothing_and_skips_backup(tmp_path):
    eng = _fake_engine()
    fp = _make_xlsx(tmp_path, Decimal("32.0000"))
    report = asyncio.run(StandaloneRateAuditor(eng).run(fp, apply=False))

    assert report.change_count == 1
    assert report.applied is False
    eng.backup.create_backup.assert_not_called()
    wb = openpyxl.load_workbook(fp)
    ws = wb["ExRate"]
    assert safe_to_decimal(ws.cell(row=2, column=2).value) == Decimal("32.0000")
    wb.close()


def test_run_raises_without_exrate_sheet(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    fp = tmp_path / "no_exrate.xlsx"
    wb.save(str(fp))
    wb.close()
    with pytest.raises(ValueError, match="No ExRate sheet"):
        asyncio.run(StandaloneRateAuditor(_fake_engine()).run(str(fp)))


def test_run_aborts_on_custom_layout_sheet(tmp_path):
    # F9 — an app-built CUSTOM ExRate sheet (GBP columns) must be refused:
    # no corrections, no backup, the file's bytes stay identical.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExRate"
    ws.append(["Date", "GBP Buying TT", "GBP Selling", "Holidays/Weekend"])
    ws.append([WED, Decimal("42.1234"), Decimal("42.5678"), ""])
    fp = tmp_path / "custom.xlsx"
    wb.save(str(fp))
    wb.close()

    before = Path(fp).read_bytes()
    eng = _fake_engine()
    with pytest.raises(ValueError, match="Non-standard ExRate layout"):
        asyncio.run(StandaloneRateAuditor(eng).run(str(fp), apply=True))
    eng.backup.create_backup.assert_not_called()
    assert Path(fp).read_bytes() == before  # nothing was written


def test_run_audits_standard_sheet_with_extra_currency_column(tmp_path):
    # A standard sheet with an appended "GBP Rate" column F is VALID: the
    # audit corrects B-E and leaves the extra-currency column untouched.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExRate"
    ws.append(HEADERS[:5] + ["GBP Rate", "Holidays/Weekend"])
    ws.append([WED, Decimal("32.0000"), Decimal("32.7790"),
               Decimal("37.0000"), Decimal("37.5000"),
               Decimal("42.1234"), ""])
    fp = tmp_path / "extra_ccy.xlsx"
    wb.save(str(fp))
    wb.close()

    eng = _fake_engine()
    report = asyncio.run(StandaloneRateAuditor(eng).run(str(fp), apply=True))
    assert report.layout_error is None
    assert report.change_count == 1
    assert report.changes[0].cell == "B2"
    eng.backup.create_backup.assert_called_once()

    wb2 = openpyxl.load_workbook(str(fp))
    ws2 = wb2["ExRate"]
    assert safe_to_decimal(ws2.cell(row=2, column=2).value) == Decimal("32.4507")
    assert safe_to_decimal(ws2.cell(row=2, column=6).value) == Decimal("42.1234")
    wb2.close()


def test_run_loads_macro_workbook_with_keep_vba(tmp_path, monkeypatch):
    # F7 — the read-write load must pass keep_vba for .xlsm/.xltm so a save
    # cannot silently strip the VBA project. Spy on the call site's kwargs.
    eng = _fake_engine()
    fp = _make_xlsx(tmp_path, Decimal("32.4507"))  # correct → no rewrite
    xlsm = tmp_path / "ledger.xlsm"
    Path(fp).rename(xlsm)

    seen = {}
    real_load = openpyxl.load_workbook
    def _spy(path, **kw):
        if not kw.get("read_only"):
            seen.update(kw)
        return real_load(path, **kw)
    monkeypatch.setattr("core.rate_audit.openpyxl.load_workbook", _spy)

    asyncio.run(StandaloneRateAuditor(eng).run(str(xlsm), apply=True))
    assert seen.get("keep_vba") is True


def test_apply_passes_verifier_to_atomic_save(tmp_path, monkeypatch):
    # F201 — the apply path must hand atomic_save a verifier built from the
    # report's intended 4dp Decimals; the honest save passes it and persists.
    eng = _fake_engine()
    fp = _make_xlsx(tmp_path, Decimal("32.0000"))

    seen = {}
    real_save = rate_audit_mod.atomic_save

    def _spy(wb, filepath, verify=None):
        seen["verify"] = verify
        return real_save(wb, filepath, verify=verify)

    monkeypatch.setattr(rate_audit_mod, "atomic_save", _spy)
    report = asyncio.run(StandaloneRateAuditor(eng).run(fp, apply=True))

    assert report.change_count == 1
    assert callable(seen.get("verify")), "apply path must pass a verifier"
    wb = openpyxl.load_workbook(fp)
    assert safe_to_decimal(
        wb["ExRate"].cell(row=2, column=2).value
    ) == Decimal("32.4507")
    wb.close()


def test_apply_verification_failure_leaves_file_untouched(tmp_path, monkeypatch):
    # F201 corruption simulation — sabotage the in-memory sheet AFTER the
    # report's intended values are fixed, so the saved TEMP disagrees with
    # the verifier's expected 4dp Decimal. Hard failure, no persist: the
    # user's file stays byte-for-byte identical and no temp is left behind.
    eng = _fake_engine()
    fp = _make_xlsx(tmp_path, Decimal("32.0000"))
    before = Path(fp).read_bytes()

    real_apply = rate_audit_mod.apply_corrections

    def _sabotage(ws, report):
        real_apply(ws, report)
        ch = report.changes[0]
        ws.cell(row=ch.row, column=ch.col).value = Decimal("99.9999")
        return report

    monkeypatch.setattr(rate_audit_mod, "apply_corrections", _sabotage)

    with pytest.raises(
        WorkbookVerifyError, match="Post-write verification failed"
    ):
        asyncio.run(StandaloneRateAuditor(eng).run(fp, apply=True))

    assert Path(fp).read_bytes() == before, "original must be untouched"
    assert not list(tmp_path.glob("*.tmp~")), "temp file must be unlinked"


def test_run_loads_plain_xlsx_without_keep_vba(tmp_path, monkeypatch):
    # A plain .xlsx keeps the default loader behavior (keep_vba=False).
    eng = _fake_engine()
    fp = _make_xlsx(tmp_path, Decimal("32.4507"))

    seen = {}
    real_load = openpyxl.load_workbook
    def _spy(path, **kw):
        if not kw.get("read_only"):
            seen.update(kw)
        return real_load(path, **kw)
    monkeypatch.setattr("core.rate_audit.openpyxl.load_workbook", _spy)

    asyncio.run(StandaloneRateAuditor(eng).run(fp, apply=True))
    assert seen.get("keep_vba") is False
