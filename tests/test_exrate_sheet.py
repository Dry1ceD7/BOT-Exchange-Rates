#!/usr/bin/env python3
"""
tests/test_exrate_sheet.py
---------------------------------------------------------------------------
Unit tests for core/exrate_sheet.py — Master ExRate sheet builder.
---------------------------------------------------------------------------
"""

import logging
from datetime import date
from decimal import Decimal

import openpyxl

import core.exrate_sheet as exrate_sheet_mod
from core.exrate_sheet import (
    EXRATE_RATE_COLUMNS,
    _build_date_range,
    _merge_rate_data,
    _parse_cell_date,
    exrate_fixed_index_keys,
    exrate_fixed_letters,
    exrate_holidays_col,
    exrate_index_key,
    update_master_exrate_sheet,
)

# =========================================================================
#  HELPERS
# =========================================================================


class TestLayoutSingleSource:
    """The ExRate layout constants/helpers consumed by excel_io,
    exrate_updater and rate_audit — guard against drift."""

    def test_fixed_rate_columns_are_b_to_e(self):
        assert [c[0] for c in EXRATE_RATE_COLUMNS] == [2, 3, 4, 5]

    def test_fixed_headers_match_standard_layout(self):
        assert [c[1] for c in EXRATE_RATE_COLUMNS] == [
            "USD Buying TT Rate", "USD Selling Rate",
            "EUR Buying TT Rate", "EUR Selling Rate",
        ]

    def test_index_keys_match_build_exrate_index_names(self):
        assert exrate_index_key("USD", "buying_transfer") == "usd_buying"
        assert exrate_index_key("EUR", "selling") == "eur_selling"
        assert exrate_fixed_index_keys("buying_transfer") == {
            "USD": "usd_buying", "EUR": "eur_buying",
        }
        assert exrate_fixed_index_keys("selling") == {
            "USD": "usd_selling", "EUR": "eur_selling",
        }

    def test_fixed_letters_match_ledger_formula_columns(self):
        assert exrate_fixed_letters("buying_transfer") == {
            "USD": "B", "EUR": "D",
        }
        assert exrate_fixed_letters("selling") == {"USD": "C", "EUR": "E"}

    def test_non_selling_rate_type_falls_back_to_buying(self):
        # Historical else-branch behavior: anything not "selling" resolves
        # to the Buying TT columns.
        assert (
            exrate_fixed_index_keys("mid_rate")
            == exrate_fixed_index_keys("buying_transfer")
        )

    def test_holidays_col_resolver(self):
        assert exrate_holidays_col(0) == 6   # standard six-column sheet
        assert exrate_holidays_col(2) == 8   # two appended extra currencies

class TestParseCellDate:
    """Tests for _parse_cell_date helper."""

    def test_date_object(self):
        assert _parse_cell_date(date(2025, 3, 10)) == date(2025, 3, 10)

    def test_iso_string(self):
        assert _parse_cell_date("2025-03-10") == date(2025, 3, 10)

    def test_none_returns_none(self):
        assert _parse_cell_date(None) is None

    def test_invalid_string_returns_none(self):
        assert _parse_cell_date("not-a-date") is None

    def test_integer_returns_none(self):
        assert _parse_cell_date(42) is None

    def test_shared_superset_formats(self):
        """Now uses shared parser — full format superset, not just 2."""
        assert _parse_cell_date("10/03/2025") == date(2025, 3, 10)
        assert _parse_cell_date("10-03-2025") == date(2025, 3, 10)
        assert _parse_cell_date("20250310") == date(2025, 3, 10)
        assert _parse_cell_date("10 Mar 2025") == date(2025, 3, 10)


class TestBuildDateRange:
    """Tests for _build_date_range helper."""

    def test_simple_range(self):
        start = date(2025, 3, 10)
        end = date(2025, 3, 12)
        result = _build_date_range(start, end, {})
        assert start in result
        assert end in result
        assert date(2025, 3, 11) in result
        assert len(result) == 3

    def test_includes_existing_data(self):
        start = date(2025, 3, 10)
        end = date(2025, 3, 10)
        existing = {date(2025, 3, 11): {"usd_buy": 33.0}}
        result = _build_date_range(start, end, existing)
        assert date(2025, 3, 11) in result

    def test_filters_before_start(self):
        start = date(2025, 3, 10)
        end = date(2025, 3, 10)
        existing = {date(2025, 3, 5): {"usd_buy": 33.0}}
        result = _build_date_range(start, end, existing)
        assert date(2025, 3, 5) not in result


class TestMergeRateData:
    """Tests for _merge_rate_data helper."""

    def test_api_rates_override_existing(self):
        all_dates = {date(2025, 3, 10)}
        existing = {date(2025, 3, 10): {"usd_buy": 30.0, "usd_sell": None, "eur_buy": None, "eur_sell": None}}
        usd_b = {date(2025, 3, 10): Decimal("33.5")}
        usd_s = {date(2025, 3, 10): Decimal("33.6")}
        eur_b = {date(2025, 3, 10): Decimal("37.0")}
        eur_s = {date(2025, 3, 10): Decimal("37.1")}

        merged = _merge_rate_data(
            all_dates, existing, set(), {},
            usd_b, usd_s, eur_b, eur_s,
        )
        entry = merged[date(2025, 3, 10)]
        # API overrides existing, value stays exact Decimal (no float cast).
        assert entry["usd_buy"] == Decimal("33.5")
        assert entry["eur_sell"] == Decimal("37.1")

    def test_decimal_precision_preserved_no_float(self):
        """Rate values must stay exact Decimal — never cast to float."""
        d = date(2025, 3, 10)
        usd_b = {d: Decimal("34.5650")}
        merged = _merge_rate_data(
            {d}, {}, set(), {},
            usd_b, {}, {}, {},
        )
        val = merged[d]["usd_buy"]
        assert isinstance(val, Decimal)
        # Exact value preserved (float would yield 34.564999...).
        assert val == Decimal("34.5650")
        assert str(val) == "34.5650"

    def test_decimal_written_to_sheet_exact(self):
        """End-to-end: Decimal survives into the written cell exactly."""
        wb = openpyxl.Workbook()
        d = date(2025, 3, 10)
        update_master_exrate_sheet(
            wb,
            usd_buying_rates={d: Decimal("34.5650")},
            usd_selling_rates={d: Decimal("34.7350")},
            eur_buying_rates={d: Decimal("37.1250")},
            eur_selling_rates={d: Decimal("37.4450")},
            holidays_list=[],
            holidays_names={},
            start_date=d,
        )
        ws = wb["ExRate"]
        cell = ws.cell(row=2, column=2).value
        assert isinstance(cell, Decimal)
        assert cell == Decimal("34.5650")
        wb.close()

    def test_weekend_label(self):
        sat = date(2025, 3, 8)  # Saturday
        all_dates = {sat}
        merged = _merge_rate_data(
            all_dates, {}, set(), {},
            {}, {}, {}, {},
        )
        assert merged[sat]["holidays_weekend"] == "Weekend"

    def test_holiday_label(self):
        d = date(2025, 3, 10)  # Monday
        holidays_set = {d}
        holidays_names = {d: "Test Holiday"}
        merged = _merge_rate_data(
            {d}, {}, holidays_set, holidays_names,
            {}, {}, {}, {},
        )
        assert merged[d]["holidays_weekend"] == "Test Holiday"

    def test_weekend_plus_holiday_label(self):
        sat = date(2025, 3, 8)
        holidays_set = {sat}
        holidays_names = {sat: "Weekend Holiday"}
        merged = _merge_rate_data(
            {sat}, {}, holidays_set, holidays_names,
            {}, {}, {}, {},
        )
        assert "Weekend" in merged[sat]["holidays_weekend"]
        assert "Weekend Holiday" in merged[sat]["holidays_weekend"]


class TestWeekendHolidayRatesBlank:
    """Weekend/holiday rows keep BLANK rate cells — no carry-forward.

    BOT publishes no rate on weekends/holidays, so those rows show only the
    Date + Holidays/Weekend label; no prior trading-day rate is fabricated for
    them. The ledger XLOOKUP is exact-match and intentionally yields "" for a
    weekend/holiday-dated transaction. This matches the v3.2.8 behavior.
    """

    def test_saturday_and_sunday_stay_blank(self):
        fri = date(2025, 3, 7)   # Friday — trading day with a rate
        sat = date(2025, 3, 8)   # Saturday — no rate from BOT
        sun = date(2025, 3, 9)   # Sunday — no rate either
        usd_b = {fri: Decimal("33.5000")}
        merged = _merge_rate_data(
            {fri, sat, sun}, {}, set(), {},
            usd_b, {}, {}, {},
        )
        # Friday keeps its own rate.
        assert merged[fri]["usd_buy"] == Decimal("33.5000")
        # Saturday + Sunday carry NO rate forward — cells stay blank.
        assert merged[sat]["usd_buy"] is None
        assert merged[sun]["usd_buy"] is None

    def test_holiday_stays_blank(self):
        fri = date(2025, 3, 7)   # Friday — trading day with a rate
        mon = date(2025, 3, 10)  # Monday — BOT holiday, no rate
        usd_b = {fri: Decimal("34.0000")}
        merged = _merge_rate_data(
            {fri, date(2025, 3, 8), date(2025, 3, 9), mon},
            {}, {mon}, {mon: "Test Holiday"},
            usd_b, {}, {}, {},
        )
        # Monday holiday has no own rate and none is carried forward.
        assert merged[mon]["usd_buy"] is None

    def test_weekend_with_own_api_rate_is_kept(self):
        # Defensive: API priority still holds. If BOT ever returns a rate for
        # a weekend date, that real value is used — only MISSING rates blank.
        sat = date(2025, 3, 8)   # Saturday
        usd_b = {sat: Decimal("33.9999")}
        merged = _merge_rate_data(
            {sat}, {}, set(), {},
            usd_b, {}, {}, {},
        )
        assert merged[sat]["usd_buy"] == Decimal("33.9999")


# =========================================================================
#  SHEET-SOURCED VALUES — re-validated 4dp fallback + weekend cleanup
# =========================================================================

class TestSheetFallbackRequantized:
    """F47: the trading-day existing-value fallback is re-validated through
    safe_to_decimal (string-built Decimal, 4dp) — a raw openpyxl float is
    never echoed back to the cell — and each sheet-sourced date+column is
    logged."""

    def test_merge_fallback_requantizes_legacy_float(self):
        mon = date(2025, 3, 10)  # Monday — trading day
        existing = {mon: {
            "usd_buy": 34.564999999999998,  # binary-float artifact
            "usd_sell": None, "eur_buy": None, "eur_sell": None,
        }}
        merged = _merge_rate_data(
            {mon}, existing, set(), {},
            {}, {}, {}, {},
        )
        val = merged[mon]["usd_buy"]
        assert isinstance(val, Decimal)
        assert val == Decimal("34.5650")
        assert val.as_tuple().exponent == -4

    def test_existing_6dp_float_rewritten_quantized_4dp(self, caplog):
        """End-to-end: a 6dp float left on the sheet by an older build is
        re-written as the 4dp string-built Decimal, and the sheet-sourced
        date is logged."""
        mon = date(2025, 3, 10)  # Monday — trading day
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("ExRate")
        ws.cell(row=2, column=1, value=mon)
        ws.cell(row=2, column=2, value=32.123456)  # 6dp legacy float
        with caplog.at_level(logging.DEBUG, logger="core.exrate_sheet"):
            update_master_exrate_sheet(
                wb,
                usd_buying_rates={}, usd_selling_rates={},
                eur_buying_rates={}, eur_selling_rates={},
                holidays_list=[], holidays_names={},
                start_date=mon, end_date=mon,
            )
        cell = wb["ExRate"].cell(row=2, column=2)
        assert isinstance(cell.value, Decimal)
        assert cell.value == Decimal("32.1235")
        assert cell.value.as_tuple().exponent == -4
        messages = [r.getMessage() for r in caplog.records]
        assert any(
            "sourced from existing sheet" in m
            and "2025-03-10" in m and "usd_buy" in m
            for m in messages
        )
        wb.close()


class TestWeekendCarryForwardCleanup:
    """F48: v3.4.0/v3.5.0 (2026-06-04..06-09) fabricated weekend/holiday
    carry-forward rates; the merge must drop sheet-sourced values on
    non-trading rows (cell stays blank) and log each drop. Genuine
    BOT-published weekend rates (present in the fresh API/cache data for
    that exact date) are honored."""

    def test_carry_forward_saturday_blanked_on_rebuild(self, caplog):
        """A Saturday cell equal to Friday's value (the carry-forward
        signature) is blanked on rebuild, with an INFO log record."""
        fri = date(2025, 3, 7)   # Friday — trading day
        sat = date(2025, 3, 8)   # Saturday
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("ExRate")
        # Simulate a v3.4.0/v3.5.0 workbook: Friday's rate carried into
        # Saturday (openpyxl reads such cells back as floats).
        ws.cell(row=2, column=1, value=fri)
        ws.cell(row=2, column=2, value=34.565)
        ws.cell(row=3, column=1, value=sat)
        ws.cell(row=3, column=2, value=34.565)
        ws.cell(row=3, column=6, value="Weekend")
        with caplog.at_level(logging.INFO, logger="core.exrate_sheet"):
            update_master_exrate_sheet(
                wb,
                usd_buying_rates={fri: Decimal("34.5650")},
                usd_selling_rates={},
                eur_buying_rates={}, eur_selling_rates={},
                holidays_list=[], holidays_names={},
                start_date=fri, end_date=sat,
            )
        ws = wb["ExRate"]
        # Row 2 = Friday (kept, exact 4dp), row 3 = Saturday (blanked).
        assert ws.cell(row=2, column=2).value == Decimal("34.5650")
        assert ws.cell(row=3, column=2).value is None
        assert ws.cell(row=3, column=6).value == "Weekend"
        messages = [r.getMessage() for r in caplog.records]
        assert any(
            "carry-forward" in m and "2025-03-08" in m and "usd_buy" in m
            for m in messages
        )
        wb.close()

    def test_genuine_api_weekend_rate_survives_stale_sheet_value(
        self, caplog
    ):
        """API priority: a BOT-published weekend rate for that exact date is
        kept — the stale sheet value is simply superseded, not 'cleaned'."""
        sat = date(2025, 3, 8)   # Saturday
        existing = {sat: {
            "usd_buy": 33.5,
            "usd_sell": None, "eur_buy": None, "eur_sell": None,
        }}
        usd_b = {sat: Decimal("33.9999")}
        with caplog.at_level(logging.INFO, logger="core.exrate_sheet"):
            merged = _merge_rate_data(
                {sat}, existing, set(), {},
                usd_b, {}, {}, {},
            )
        assert merged[sat]["usd_buy"] == Decimal("33.9999")
        assert not any(
            "cleanup" in r.getMessage() for r in caplog.records
        )

    def test_sheet_only_weekend_value_without_signature_also_dropped(
        self, caplog
    ):
        """Decision (documented): a Saturday cell DIFFERING from Friday's
        value is still sheet-sourced on a non-trading day with no API
        backing. Per the frozen invariant (weekend/holiday rows are Date +
        label only) it is ALSO dropped — the carry-forward signature only
        classifies the log line. API-backed weekend values stay; sheet-only
        weekend values go."""
        fri = date(2025, 3, 7)   # Friday — trading day
        sat = date(2025, 3, 8)   # Saturday
        existing = {sat: {
            "usd_buy": 77.7777,  # does NOT match Friday's 33.5000
            "usd_sell": None, "eur_buy": None, "eur_sell": None,
        }}
        usd_b = {fri: Decimal("33.5000")}
        with caplog.at_level(logging.INFO, logger="core.exrate_sheet"):
            merged = _merge_rate_data(
                {fri, sat}, existing, set(), {},
                usd_b, {}, {}, {},
            )
        assert merged[fri]["usd_buy"] == Decimal("33.5000")
        assert merged[sat]["usd_buy"] is None
        messages = [r.getMessage() for r in caplog.records]
        assert any(
            "sheet-only" in m and "2025-03-08" in m and "usd_buy" in m
            for m in messages
        )
        assert not any("carry-forward" in m for m in messages)

    def test_holiday_carry_forward_also_dropped(self, caplog):
        """Holiday rows get the same cleanup as weekends."""
        fri = date(2025, 3, 7)   # Friday — trading day
        mon = date(2025, 3, 10)  # Monday — declared holiday
        existing = {mon: {
            "usd_buy": 34.0,     # carried forward from Friday
            "usd_sell": None, "eur_buy": None, "eur_sell": None,
        }}
        usd_b = {fri: Decimal("34.0000")}
        with caplog.at_level(logging.INFO, logger="core.exrate_sheet"):
            merged = _merge_rate_data(
                {fri, date(2025, 3, 8), date(2025, 3, 9), mon},
                existing, {mon}, {mon: "Test Holiday"},
                usd_b, {}, {}, {},
            )
        assert merged[mon]["usd_buy"] is None
        assert any(
            "carry-forward" in r.getMessage()
            and "2025-03-10" in r.getMessage()
            for r in caplog.records
        )


# =========================================================================
#  FULL SHEET UPDATE
# =========================================================================

class TestUpdateMasterExrateSheet:
    """Integration tests for update_master_exrate_sheet."""

    def test_creates_exrate_sheet(self):
        wb = openpyxl.Workbook()
        update_master_exrate_sheet(
            wb,
            usd_buying_rates={date(2025, 3, 10): Decimal("33.5")},
            usd_selling_rates={date(2025, 3, 10): Decimal("33.6")},
            eur_buying_rates={date(2025, 3, 10): Decimal("37.0")},
            eur_selling_rates={date(2025, 3, 10): Decimal("37.1")},
            holidays_list=[],
            holidays_names={},
            start_date=date(2025, 3, 10),
        )
        assert "ExRate" in wb.sheetnames
        ws = wb["ExRate"]
        assert ws.cell(row=1, column=1).value == "Date"
        assert ws.cell(row=1, column=2).value == "USD Buying TT Rate"
        wb.close()

    def test_writes_rate_data(self):
        wb = openpyxl.Workbook()
        d = date(2025, 3, 10)
        update_master_exrate_sheet(
            wb,
            usd_buying_rates={d: Decimal("33.5")},
            usd_selling_rates={d: Decimal("33.6")},
            eur_buying_rates={d: Decimal("37.0")},
            eur_selling_rates={d: Decimal("37.1")},
            holidays_list=[],
            holidays_names={},
            start_date=d,
        )
        ws = wb["ExRate"]
        # Row 2 should have data (row 1 = header). Cells hold exact Decimal.
        assert ws.cell(row=2, column=2).value == Decimal("33.5")
        assert ws.cell(row=2, column=3).value == Decimal("33.6")
        wb.close()

    def test_default_end_date_uses_bot_today(self, monkeypatch):
        """When end_date is None the range runs out to the BOT business date.

        The sweep replaced the bare date.today() default with bot_today()
        (Asia/Bangkok). Patching bot_today proves the default path now keys off
        the BOT calendar rather than the machine's local date.
        """
        fixed_today = date(2025, 3, 12)
        monkeypatch.setattr(
            exrate_sheet_mod, "bot_today", lambda: fixed_today
        )
        wb = openpyxl.Workbook()
        start = date(2025, 3, 10)
        update_master_exrate_sheet(
            wb,
            usd_buying_rates={start: Decimal("33.5")},
            usd_selling_rates={start: Decimal("33.6")},
            eur_buying_rates={start: Decimal("37.0")},
            eur_selling_rates={start: Decimal("37.1")},
            holidays_list=[],
            holidays_names={},
            start_date=start,
            # end_date omitted → defaults to bot_today() (patched).
        )
        ws = wb["ExRate"]
        written = {
            _parse_cell_date(ws.cell(row=r, column=1).value)
            for r in range(2, (ws.max_row or 1) + 1)
        }
        written.discard(None)
        # Range stops at the patched BOT today, inclusive.
        assert written == {start, date(2025, 3, 11), fixed_today}
        wb.close()

    def test_explicit_end_date_bounds_written_range(self):
        """A manual (start, end) writes exactly that range — end not today()."""
        wb = openpyxl.Workbook()
        start = date(2025, 3, 10)
        end = date(2025, 3, 12)
        update_master_exrate_sheet(
            wb,
            usd_buying_rates={start: Decimal("33.5")},
            usd_selling_rates={start: Decimal("33.6")},
            eur_buying_rates={start: Decimal("37.0")},
            eur_selling_rates={start: Decimal("37.1")},
            holidays_list=[],
            holidays_names={},
            start_date=start,
            end_date=end,
        )
        ws = wb["ExRate"]
        written = {
            _parse_cell_date(ws.cell(row=r, column=1).value)
            for r in range(2, (ws.max_row or 1) + 1)
        }
        written.discard(None)
        # Exactly the 3-day manual window — nothing up to today().
        assert written == {start, date(2025, 3, 11), end}
        wb.close()


# =========================================================================
#  EXTRA-CURRENCY COLUMNS (multi-currency ledger master sheet)
# =========================================================================

class TestExtraCurrencyColumns:
    """update_master_exrate_sheet appends a column per extra currency and
    returns the {ccy: column_letter} map for the ledger formula."""

    def test_extra_columns_appended_after_eur_before_holidays(self):
        wb = openpyxl.Workbook()
        d = date(2025, 3, 10)  # Monday
        col_map = update_master_exrate_sheet(
            wb,
            usd_buying_rates={d: Decimal("33.5")},
            usd_selling_rates={d: Decimal("33.6")},
            eur_buying_rates={d: Decimal("37.0")},
            eur_selling_rates={d: Decimal("37.1")},
            holidays_list=[],
            holidays_names={},
            start_date=d,
            end_date=d,
            extra_currency_rates={
                "GBP": {d: Decimal("42.1234")},
                "JPY": {d: Decimal("0.2155")},
            },
        )
        ws = wb["ExRate"]
        headers = [
            ws.cell(row=1, column=c).value
            for c in range(1, (ws.max_column or 1) + 1)
        ]
        # Date, USD x2, EUR x2, then extra cols (dict order), then Holidays.
        assert headers == [
            "Date", "USD Buying TT Rate", "USD Selling Rate",
            "EUR Buying TT Rate", "EUR Selling Rate",
            "GBP Rate", "JPY Rate", "Holidays/Weekend",
        ]
        # Column map points the ledger formula at F (GBP) and G (JPY).
        assert col_map == {"GBP": "F", "JPY": "G"}
        # USD/EUR are NOT in the map (fixed B-E columns).
        assert "USD" not in col_map and "EUR" not in col_map
        wb.close()

    def test_extra_currency_value_exact_decimal(self):
        wb = openpyxl.Workbook()
        d = date(2025, 3, 10)
        update_master_exrate_sheet(
            wb,
            usd_buying_rates={}, usd_selling_rates={},
            eur_buying_rates={}, eur_selling_rates={},
            holidays_list=[], holidays_names={},
            start_date=d, end_date=d,
            extra_currency_rates={"GBP": {d: Decimal("42.1234")}},
        )
        ws = wb["ExRate"]
        cell = ws.cell(row=2, column=6)  # F = GBP
        assert cell.value == Decimal("42.1234")
        assert cell.number_format == "0.0000"
        wb.close()

    def test_extra_currency_weekend_stays_blank(self):
        """A GBP weekend row keeps a blank rate — no carry-forward."""
        fri = date(2025, 3, 7)   # Friday
        sat = date(2025, 3, 8)   # Saturday
        merged = _merge_rate_data(
            {fri, sat}, {}, set(), {},
            {}, {}, {}, {},
            {"GBP": {fri: Decimal("42.5000")}},
        )
        assert merged[fri]["extra:GBP"] == Decimal("42.5000")
        assert merged[sat]["extra:GBP"] is None

    def test_no_extra_currencies_keeps_six_column_layout(self):
        """Backward compat: omitting extra rates keeps the legacy 6 columns and
        returns an empty map."""
        wb = openpyxl.Workbook()
        d = date(2025, 3, 10)
        col_map = update_master_exrate_sheet(
            wb,
            usd_buying_rates={d: Decimal("33.5")},
            usd_selling_rates={d: Decimal("33.6")},
            eur_buying_rates={d: Decimal("37.0")},
            eur_selling_rates={d: Decimal("37.1")},
            holidays_list=[], holidays_names={},
            start_date=d, end_date=d,
        )
        ws = wb["ExRate"]
        assert ws.max_column == 6
        assert ws.cell(row=1, column=6).value == "Holidays/Weekend"
        assert col_map == {}
        wb.close()
