#!/usr/bin/env python3
"""
tests/test_prescan.py
---------------------------------------------------------------------------
Unit tests for core/prescan.py — Smart Date Pre-Scanner.
---------------------------------------------------------------------------
"""

from datetime import date

import openpyxl
import pytest

from core.prescan import DATE_FORMATS, _parse_scan_date, prescan_oldest_date

# =========================================================================
#  HELPERS
# =========================================================================

class TestParseScanDate:
    """Tests for _parse_scan_date helper."""

    FORMATS = DATE_FORMATS

    def test_iso_string(self):
        assert _parse_scan_date("2025-03-10", self.FORMATS) == date(2025, 3, 10)

    def test_date_object(self):
        assert _parse_scan_date(date(2025, 3, 10), self.FORMATS) == date(2025, 3, 10)

    def test_none_returns_none(self):
        assert _parse_scan_date(None, self.FORMATS) is None

    def test_empty_string_returns_none(self):
        assert _parse_scan_date("", self.FORMATS) is None

    def test_nan_returns_none(self):
        assert _parse_scan_date("nan", self.FORMATS) is None

    def test_invalid_returns_none(self):
        assert _parse_scan_date("hello", self.FORMATS) is None

    def test_uses_centralized_formats(self):
        """Fix #5: prescan formats come from the single shared source."""
        from core.constants import DATE_FORMATS as SHARED
        assert tuple(DATE_FORMATS) == tuple(SHARED)
        # Superset coverage preserved.
        assert _parse_scan_date("10/03/2025", self.FORMATS) == date(2025, 3, 10)
        assert _parse_scan_date("20250310", self.FORMATS) == date(2025, 3, 10)


# =========================================================================
#  PRESCAN
# =========================================================================

class TestPrescanOldestDate:
    """Tests for prescan_oldest_date function."""

    @pytest.fixture
    def xlsx_with_dates(self, tmp_path):
        """Creates a .xlsx file with dates in a 'Date' column."""
        filepath = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Amount"])
        ws.append([date(2025, 3, 15), 100])
        ws.append([date(2025, 3, 10), 200])
        ws.append([date(2025, 3, 20), 300])
        wb.save(str(filepath))
        wb.close()
        return str(filepath)

    @pytest.fixture
    def xlsx_no_date_col(self, tmp_path):
        """Creates a .xlsx file without a 'Date' column."""
        filepath = tmp_path / "no_date.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Amount"])
        ws.append(["Alice", 100])
        wb.save(str(filepath))
        wb.close()
        return str(filepath)

    def test_finds_oldest_date(self, xlsx_with_dates):
        oldest, detected = prescan_oldest_date([xlsx_with_dates])
        assert detected is True
        assert oldest == date(2025, 3, 10)

    def test_empty_list_returns_fallback(self):
        oldest, detected = prescan_oldest_date([])
        assert detected is False
        assert isinstance(oldest, date)
        # Fallback should be Dec 28 of previous year
        prev_year = date.today().year - 1
        expected = date(prev_year, 12, 28)
        assert oldest == expected

    def test_nonexistent_file_returns_fallback(self):
        oldest, detected = prescan_oldest_date(["/no/such/file.xlsx"])
        assert detected is False

    def test_file_without_date_column_returns_fallback(self, xlsx_no_date_col):
        oldest, detected = prescan_oldest_date([xlsx_no_date_col])
        assert detected is False

    def test_multiple_files_finds_global_oldest(self, tmp_path):
        """Test scanning multiple files picks the true oldest date."""
        f1 = tmp_path / "file1.xlsx"
        f2 = tmp_path / "file2.xlsx"

        wb1 = openpyxl.Workbook()
        ws1 = wb1.active
        ws1.append(["Date", "Amount"])
        ws1.append([date(2025, 4, 1), 100])
        wb1.save(str(f1))
        wb1.close()

        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.append(["Date", "Amount"])
        ws2.append([date(2025, 2, 15), 200])
        wb2.save(str(f2))
        wb2.close()

        oldest, detected = prescan_oldest_date([str(f1), str(f2)])
        assert detected is True
        assert oldest == date(2025, 2, 15)

    def test_exrate_master_sheet_is_skipped(self, tmp_path):
        """F127: the app's own ExRate master sheet must not skew detection.

        The ExRate sheet carries a "Date" column going back to the year
        start; scanning it would drag the oldest date to dates no ledger row
        actually needs. Only real ledger sheets count.
        """
        fp = tmp_path / "ledger.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExRate"
        ws.append(["Date", "USD Buying TT"])
        ws.append([date(2024, 1, 2), 34.5])  # much older than the ledger

        ledger = wb.create_sheet("Sheet1")
        ledger.append(["Date", "Amount"])
        ledger.append([date(2025, 3, 10), 100])
        wb.save(str(fp))
        wb.close()

        oldest, detected = prescan_oldest_date([str(fp)])
        assert detected is True
        assert oldest == date(2025, 3, 10)

    def test_unreadable_file_is_skipped_not_crashed(self, tmp_path):
        """A locked/permission-denied .xlsx is skipped, not fatal.

        Simulated with a directory at an .xlsx path: ``exists()`` is True but
        ``open('rb')`` raises IsADirectoryError (an OSError), exactly like a
        file held open by Excel on the Windows target. Before the OSError guard
        this aborted the whole headless/scheduled prescan; now the bad file is
        skipped and the readable ledger still yields its date.
        """
        bad = tmp_path / "locked.xlsx"
        bad.mkdir()

        good = tmp_path / "good.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Amount"])
        ws.append([date(2025, 3, 3), 100])
        wb.save(str(good))
        wb.close()

        # Bad file listed first so a regression (un-caught OSError) would abort
        # before the good file is ever scanned.
        oldest, detected = prescan_oldest_date([str(bad), str(good)])
        assert detected is True
        assert oldest == date(2025, 3, 3)

    def test_non_zip_xlsx_is_skipped_not_crashed(self, tmp_path):
        """A non-zip file wearing an .xlsx extension is skipped, not fatal.

        Repro for the masquerading-legacy-file case: a BIFF .xls (or any
        non-zip bytes) renamed to .xlsx makes openpyxl raise
        zipfile.BadZipFile, which is neither OSError nor
        InvalidFileException — before the BadZipFile guard it escaped the
        except tuple and killed the whole headless/scheduled prescan.
        """
        fake = tmp_path / "renamed_legacy.xlsx"
        fake.write_bytes(b"\xd0\xcf\x11\xe0 not a zip, BIFF-style bytes")

        good = tmp_path / "good.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Amount"])
        ws.append([date(2025, 4, 4), 100])
        wb.save(str(good))
        wb.close()

        # Fake file listed first so an un-caught BadZipFile would abort
        # before the good file is ever scanned.
        oldest, detected = prescan_oldest_date([str(fake), str(good)])
        assert detected is True
        assert oldest == date(2025, 4, 4)

    def test_duplicate_date_uses_export_entry_column(self, tmp_path):
        """With two 'Date' columns, the window comes from the one the
        formulas look up — the Date adjacent to 'EX Rate', not the
        invoice Date in column A/B.

        If the prescan kept first-occurrence resolution while the writer
        targets the export-entry column, the fetch window could open too
        late/early for exactly the dates the written XLOOKUPs need.
        """
        fp = tmp_path / "dup.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Cur", "Export Entry", "Date", "EX Rate"])
        # Invoice date deliberately OLDER than the export-entry date: a
        # first-occurrence regression would report 2025-01-01.
        ws.append([date(2025, 1, 1), "USD", "x", date(2025, 3, 3), None])
        wb.save(str(fp))
        wb.close()

        oldest, detected = prescan_oldest_date([str(fp)])
        assert detected is True
        assert oldest == date(2025, 3, 3)


class TestPrescanTruncatedXml:
    """Round 11: a truncated/garbled sheet XML inside a structurally valid
    zip raises xml.etree.ElementTree.ParseError (a SyntaxError subclass) —
    previously NOT in the per-file catch tuple, so one such workbook killed
    the entire headless/scheduled/GUI prescan instead of being skipped."""

    def _good_and_truncated(self, tmp_path):
        import zipfile

        good = tmp_path / "good.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jan"
        ws.append(["Date", "Cur", "EX Rate"])
        for i in range(1, 20):
            ws.append([date(2025, 2, i), "USD", None])
        wb.save(str(good))
        wb.close()

        # Valid zip, sheet1.xml halved → ParseError mid-iteration.
        bad = tmp_path / "bad.xlsx"
        with zipfile.ZipFile(str(good)) as src, \
                zipfile.ZipFile(str(bad), "w") as dst:
            for item in src.infolist():
                data = src.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    data = data[: len(data) // 2]
                dst.writestr(item, data)
        return str(good), str(bad)

    def test_truncated_sheet_xml_skips_file_scans_rest(self, tmp_path):
        good, bad = self._good_and_truncated(tmp_path)
        # Bad file FIRST: the per-file skip must let the good file still scan.
        oldest, detected = prescan_oldest_date([bad, good])
        assert detected is True
        assert oldest == date(2025, 2, 1)

    def test_only_truncated_file_returns_fallback(self, tmp_path):
        _good, bad = self._good_and_truncated(tmp_path)
        oldest, detected = prescan_oldest_date([bad])
        assert detected is False  # fallback date, no crash
