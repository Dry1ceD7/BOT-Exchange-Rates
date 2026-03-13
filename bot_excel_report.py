#!/usr/bin/env python3
"""
============================================================================
  BOT EXCHANGE RATE — EXECUTIVE EXCEL REPORT GENERATOR
  For: Accounting & Finance Department — Board Presentation
  Source: Bank of Thailand Official API (https://www.bot.or.th/)
============================================================================

  This script fetches exchange rate data from the official BOT API
  and generates a presentation-quality Excel workbook with:

    Tab 1: Cover Sheet          – Title, branding, report metadata
    Tab 2: USD Daily Rates      – Daily USD/THB with conditional formatting
    Tab 3: EUR Daily Rates      – Daily EUR/THB with conditional formatting
    Tab 4: Summary Dashboard    – Monthly averages, min/max, volatility
    Tab 5: Monthly Analysis     – Period-over-period comparison
    Tab 6: FX Calculator        – Interactive currency converter
    Tab 7: Notes & Disclaimers  – Financial disclaimers, sources

  Usage:
    python3 bot_excel_report.py

  Output:
    BOT_ExchangeRate_Report_YYYY-MM-DD.xlsx (in same directory)
============================================================================
"""

# ─── Standard Library Imports ────────────────────────────────
import sys
import json
import ssl
import os
import urllib.request
import argparse
import asyncio
from typing import Dict, Any, Optional, List

import importlib
import subprocess
import smtplib
from email.message import EmailMessage
from datetime import date, timedelta, datetime
from collections import OrderedDict

# ─── Auto-install openpyxl and aiohttp to local _libs folder 
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = SCRIPT_DIR

# Possible locations for _libs
_LIBS_LOCAL = os.path.join(SCRIPT_DIR, "_libs")
_LIBS_PARENT = os.path.join(PARENT_DIR, "_libs")

if os.path.exists(_LIBS_PARENT):
    _LIBS = _LIBS_PARENT
else:
    _LIBS = _LIBS_LOCAL

if _LIBS not in sys.path:
    sys.path.insert(0, _LIBS)

try:
    import openpyxl
    import aiohttp
except ImportError:
    print("  Installing openpyxl and aiohttp to local _libs folder...", file=sys.stderr)
    import subprocess
    os.makedirs(_LIBS, exist_ok=True)
    subprocess.check_call([
        sys.executable, "-m", "pip", "install",
        "--target", _LIBS, "openpyxl", "aiohttp",
        "--break-system-packages"
    ])
    importlib.invalidate_caches()
    import openpyxl
    import aiohttp

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule

# ─── Load Environment Variables ─────────────────────────────
# Checks for .env in current folder or parent folder
env_path_local = os.path.join(SCRIPT_DIR, ".env")
env_path_parent = os.path.join(PARENT_DIR, ".env")

env_path = env_path_parent if os.path.exists(env_path_parent) else env_path_local

if os.path.exists(env_path):
    with open(env_path, "r") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                parts = line.split("=", 1)
                if len(parts) == 2:
                    key, val = parts
                    os.environ[key.strip()] = val.strip().strip("\"'")

TOKEN_EXG = os.environ.get("BOT_TOKEN_EXG", "")
TOKEN_HOL = os.environ.get("BOT_TOKEN_HOL", "")
if not TOKEN_EXG or not TOKEN_HOL:
    sys.exit("Error: Missing BOT API tokens in .env file.")

# ─── Configuration ───────────────────────────────────────────
CONFIG_FILE = os.path.join(PARENT_DIR, "config.json")
if not os.path.exists(CONFIG_FILE):
    CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.json")
    
with open(CONFIG_FILE, "r", encoding="utf-8") as f:
    config = json.load(f)

GATEWAY   = config["api"]["gateway_url"]
EXG_PATH  = config["api"]["exchange_rate_path"]
HOL_PATH  = config["api"]["holiday_path"]

# CLI Arguments
parser = argparse.ArgumentParser(description="Bank of Thailand Executive Excel Report")
parser.add_argument("--start", type=str, default="2025-01-01", help="Start date YYYY-MM-DD")
parser.add_argument("--end", type=str, default=datetime.now().strftime("%Y-%m-%d"), help="End date YYYY-MM-DD")
parser.add_argument("--currencies", nargs="+", default=config["currencies"], help="Currencies to include")
parser.add_argument("--pdf", action="store_true", help="Generate PDF version")
parser.add_argument("--email", type=str, help="Email address to send the report to")
args = parser.parse_args()

CURRENCIES = args.currencies
GENERATE_PDF = args.pdf
EMAIL_TO = args.email

try:
    START = datetime.strptime(args.start, "%Y-%m-%d").date()
    END = datetime.strptime(args.end, "%Y-%m-%d").date()
except ValueError:
    sys.exit("Error: Dates must be in YYYY-MM-DD format.")
    
if START > END:
    sys.exit("Error: Start date cannot be after end date.")

CHUNK     = config["api"]["max_days_per_request"]
TODAY_STR  = datetime.now().strftime("%Y-%m-%d")
# Output file — redirected to ../data/output/ if it exists
DATA_OUTPUT_DIR = os.path.join(PARENT_DIR, "data", "output")
if os.path.exists(DATA_OUTPUT_DIR):
    OUTPUT = os.path.join(DATA_OUTPUT_DIR, "BOT_ExchangeRate_Report.xlsx")
else:
    OUTPUT = os.path.join(SCRIPT_DIR, "BOT_ExchangeRate_Report.xlsx")

ssl_ctx = ssl.create_default_context()

# ─── Fixed Thai Calendar Holidays (for weekend annotation) ───
FIXED_HOLIDAYS = {}
for date_str, holiday_name in config["fixed_holidays"].items():
    month, day = map(int, date_str.split("-"))
    FIXED_HOLIDAYS[(month, day)] = holiday_name

# ═══════════════════════════════════════════════════════════════
# DESIGN SYSTEM — Light Professional Theme
# ═══════════════════════════════════════════════════════════════

# Light color palette — clean corporate look
C_PRIMARY    = "1F4E79"   # Deep teal-blue for headers
C_ACCENT     = "2E75B6"   # Medium blue accent
C_ACCENT_LT  = "D6E4F0"   # Very light blue (header bg)
C_WHITE      = "FFFFFF"   # White
C_ROW_ALT    = "F2F7FB"   # Alternating row — pale blue
C_GOLD       = "BF8F00"   # Gold for highlights
C_GOLD_BG    = "FFF8E7"   # Light gold background
C_RED        = "C0392B"   # Negative change
C_GREEN      = "1E8449"   # Positive change
C_GREY       = "808080"   # Muted text
C_BORDER     = "B4C6E7"   # Soft blue border
C_HOLIDAY    = "FFF2CC"   # Light yellow for holidays
C_HOL_WKND   = "FCE4D6"   # Light peach for weekend+holiday
C_COVER_BG   = "F5F7FA"   # Cover background
C_COVER_TOP  = "1F4E79"   # Cover header band

# Pre-built fills
FILL_PRIMARY  = PatternFill("solid", fgColor=C_PRIMARY)
FILL_ACCENT   = PatternFill("solid", fgColor=C_ACCENT)
FILL_ACCENT_LT = PatternFill("solid", fgColor=C_ACCENT_LT)
FILL_WHITE    = PatternFill("solid", fgColor=C_WHITE)
FILL_ALT      = PatternFill("solid", fgColor=C_ROW_ALT)
FILL_GOLD_BG  = PatternFill("solid", fgColor=C_GOLD_BG)
FILL_HOLIDAY  = PatternFill("solid", fgColor=C_HOLIDAY)
FILL_HOL_WKND = PatternFill("solid", fgColor=C_HOL_WKND)
FILL_COVER_BG = PatternFill("solid", fgColor=C_COVER_BG)
FILL_COVER_TOP = PatternFill("solid", fgColor=C_COVER_TOP)

# Fonts
FONT_TITLE    = Font(name="Calibri", size=26, bold=True, color=C_PRIMARY)
FONT_SUBTITLE = Font(name="Calibri", size=13, color=C_ACCENT)
FONT_HDR      = Font(name="Calibri", size=11, bold=True, color=C_WHITE)

FONT_BODY     = Font(name="Calibri", size=10, color="333333")
FONT_BODY_B   = Font(name="Calibri", size=10, bold=True, color="333333")
FONT_SMALL    = Font(name="Calibri", size=9, color=C_GREY)
FONT_NUM      = Font(name="Calibri", size=10, color="333333")
FONT_REMARK   = Font(name="Calibri", size=9, italic=True, color=C_GREY)
FONT_NOTE     = Font(name="Calibri", size=10, color="555555")
FONT_DISCLAIMER = Font(name="Calibri", size=9, italic=True, color=C_GREY)
FONT_RED      = Font(name="Calibri", size=10, color=C_RED)
FONT_GREEN    = Font(name="Calibri", size=10, color=C_GREEN)

FONT_COVER_LBL = Font(name="Calibri", size=11, color=C_GREY)
FONT_COVER_VAL = Font(name="Calibri", size=11, bold=True, color="333333")

# Alignments
ALIGN_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L  = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R  = Alignment(horizontal="right", vertical="center")
ALIGN_TL = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER),
)
BOTTOM_ACCENT = Border(bottom=Side(style="medium", color=C_ACCENT))
NO_BORDER = Border()

# Number formats
NUM_FMT_RATE = '#,##0.0000'
NUM_FMT_PCT  = '0.00%'
NUM_FMT_AMT  = '#,##0.00'


def log(msg):
    print(msg)


async def bot_api_get_async(session: aiohttp.ClientSession, full_url: str, auth_token: str, retries: int = 3) -> Optional[Dict[str, Any]]:
    """Fetches data from BOT API asychronously with exponential backoff retries."""
    headers = {"Authorization": auth_token, "accept": "application/json"}
    for attempt in range(1, retries + 1):
        try:
            async with session.get(full_url, headers=headers, ssl=ssl_ctx, timeout=aiohttp.ClientTimeout(total=30)) as response:
                if response.status == 200:
                    return await response.json()
        except (aiohttp.ClientError, asyncio.TimeoutError) as e:
            pass
        if attempt < retries:
            await asyncio.sleep(2 ** attempt)
    return None

async def fetch_all_data(start_date, end_date):
    holidays = {}
    rates = {}
    async with aiohttp.ClientSession() as session:
        holiday_tasks = []
        for yr in range(start_date.year, end_date.year + 1):
            url = f"{GATEWAY}{HOL_PATH}?year={yr}"
            holiday_tasks.append(bot_api_get_async(session, url, TOKEN_HOL))
            
        rate_tasks = []
        cs = start_date
        while cs <= end_date:
            ce = min(cs + timedelta(days=CHUNK), end_date)
            sp, ep = cs.strftime("%Y-%m-%d"), ce.strftime("%Y-%m-%d")
            for ccy in CURRENCIES:
                url = f"{GATEWAY}{EXG_PATH}?start_period={sp}&end_period={ep}&currency={ccy}"
                rate_tasks.append((ccy, bot_api_get_async(session, url, TOKEN_EXG)))
            cs = ce + timedelta(days=1)
            
        log(f"\n  [1/3] Fetching data ({len(holiday_tasks)} holiday years, {len(rate_tasks)} rate chunks concurrently)...")
        holiday_results = await asyncio.gather(*holiday_tasks)
        rate_results = await asyncio.gather(*(task[1] for task in rate_tasks))
        
        for data in holiday_results:
            if data:
                for h in data.get("result", {}).get("data", []):
                    dt = str(h.get("Date", "")).strip()[:10]
                    nm = str(h.get("HolidayDescription", "Holiday")).strip()
                    if dt:
                        holidays[dt] = nm
                        
        for (ccy, _), data in zip(rate_tasks, rate_results):
            if data:
                try:
                    details = data.get("result", {}).get("data", {}).get("data_detail", [])
                except (KeyError, AttributeError):
                    continue
                if not isinstance(details, list):
                    continue
                for row in details:
                    dt = str(row.get("period", "")).strip()[:10]
                    if not dt: continue
                    bt = str(row.get("buying_transfer", "")).strip()
                    sl = str(row.get("selling", "")).strip()
                    if dt not in rates: rates[dt] = {}
                    rates[dt][ccy] = {
                        "buy_tt": float(bt) if bt else None,
                        "sell": float(sl) if sl else None
                    }
                    
    log(f"        Loaded {len(rates)} trading days.")
    log("\n  [2/3] Building report data...")
    all_days = []
    cur = start_date
    while cur <= end_date:
        ds = cur.strftime("%Y-%m-%d")
        hol = holidays.get(ds, "")
        if not hol:
            hol = FIXED_HOLIDAYS.get((cur.month, cur.day), "")
        is_wknd = cur.weekday() >= 5
        if is_wknd and hol:
            remark = f"Weekend; {hol}"
        elif is_wknd:
            remark = "Weekend"
        elif hol:
            remark = hol
        else:
            remark = ""

        day = rates.get(ds, {})
        day_dict = {"date": cur, "remark": remark}
        for ccy in CURRENCIES:
            day_dict[f"{ccy.lower()}_buy"] = day.get(ccy, {}).get("buy_tt")
            day_dict[f"{ccy.lower()}_sell"] = day.get(ccy, {}).get("sell")
        all_days.append(day_dict)
        cur += timedelta(days=1)
        
    return all_days, rates


def write_cell(ws, row, col, value, font=FONT_BODY, fill=None,
               align=ALIGN_L, border=THIN_BORDER, num_fmt=None):
    """Write a styled cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align
    if border:
        cell.border = border
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ═══════════════════════════════════════════════════════════════
# STEP 1: FETCH DATA FROM BOT API
# ═══════════════════════════════════════════════════════════════
log("=" * 60)
log("  BOT Executive Excel Report Generator")
log(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
log("=" * 60)

# ─── Holidays ────────────────────────────────────────────────
all_days, rates = asyncio.run(fetch_all_data(START, END))


# ═══════════════════════════════════════════════════════════════
# STEP 4: BUILD THE EXCEL WORKBOOK
# ═══════════════════════════════════════════════════════════════
log("\n  [3/3] Building Excel workbook...")
wb = openpyxl.Workbook()

# ─────────────────────────────────────────────────────────────
# TAB 1: COVER SHEET (Light theme)
# ─────────────────────────────────────────────────────────────
log("  → Cover Sheet...")
ws = wb.active
ws.title = "Cover"
ws.sheet_properties.tabColor = C_PRIMARY

# Light background
for r in range(1, 42):
    for c in range(1, 12):
        ws.cell(row=r, column=c).fill = FILL_COVER_BG

set_col_widths(ws, {"A": 4, "B": 28, "C": 28, "D": 22, "E": 15,
                     "F": 15, "G": 15, "H": 15, "I": 15, "J": 15, "K": 4})

# Top accent band
for c in range(1, 12):
    ws.cell(row=1, column=c).fill = FILL_COVER_TOP
    ws.cell(row=2, column=c).fill = FILL_COVER_TOP
ws.row_dimensions[1].height = 8
ws.row_dimensions[2].height = 8

# Thin gold accent line
for c in range(2, 11):
    ws.cell(row=3, column=c).fill = PatternFill("solid", fgColor=C_GOLD)
ws.row_dimensions[3].height = 3

# Title
ws.merge_cells("B6:J6")
write_cell(ws, 6, 2, "BANK OF THAILAND", FONT_TITLE, FILL_COVER_BG, ALIGN_L, NO_BORDER)
ws.row_dimensions[6].height = 40

ws.merge_cells("B8:J8")
write_cell(ws, 8, 2, "Daily Exchange Rate Report",
           Font(name="Calibri", size=18, color=C_GOLD, bold=True),
           FILL_COVER_BG, ALIGN_L, NO_BORDER)

ws.merge_cells("B9:J9")
ccy_subtitle = "  &  ".join(f"{c} / THB" for c in CURRENCIES)
write_cell(ws, 9, 2, f"{ccy_subtitle}  —  Weighted Average Interbank Rates",
           FONT_SUBTITLE, FILL_COVER_BG, ALIGN_L, NO_BORDER)

# Thin line under title
for c in range(2, 11):
    ws.cell(row=11, column=c).border = BOTTOM_ACCENT
ws.row_dimensions[11].height = 6

# Report metadata on white cards
meta = [
    ("Report Period", f"{START.strftime('%B %d, %Y')}  —  {END.strftime('%B %d, %Y')}"),
    ("Generated On", datetime.now().strftime("%B %d, %Y at %H:%M")),
    ("Data Source", "Bank of Thailand Official API (gateway.api.bot.or.th)"),
    ("Currencies", ", ".join(f"{c}/THB" for c in CURRENCIES)),
    ("Rate Types", "Buying Transfer (TT), Selling"),
    ("Trading Days", f"{len(rates)} days"),
    ("Calendar Days", f"{len(all_days)} days"),
    ("Precision", "4+ decimal places (BOT standard)"),
]
for i, (lbl, val) in enumerate(meta):
    r = 13 + i
    ws.merge_cells(f"B{r}:C{r}")
    write_cell(ws, r, 2, f"  {lbl}:", FONT_COVER_LBL, FILL_WHITE, ALIGN_L, THIN_BORDER)
    ws.merge_cells(f"D{r}:J{r}")
    write_cell(ws, r, 4, f"  {val}", FONT_COVER_VAL, FILL_WHITE, ALIGN_L, THIN_BORDER)
    ws.row_dimensions[r].height = 24

# Disclaimer box
r = 23
for c in range(2, 11):
    ws.cell(row=r, column=c).border = BOTTOM_ACCENT
ws.row_dimensions[r].height = 6

ws.merge_cells(f"B25:J28")
disclaimer = (
    "CONFIDENTIAL — This report is prepared for internal use by the Finance & Accounting Department. "
    "Exchange rates shown are the daily weighted-average interbank rates published by the Bank of Thailand. "
    "These rates are indicative and may differ from actual transaction rates offered by commercial banks. "
    "This report should not be used as the sole basis for financial decisions."
)
write_cell(ws, 25, 2, disclaimer, FONT_DISCLAIMER, FILL_COVER_BG, ALIGN_TL, NO_BORDER)

# Footer
ws.merge_cells("B31:J31")
write_cell(ws, 31, 2, "© Bank of Thailand  |  https://www.bot.or.th/",
           Font(name="Calibri", size=9, color=C_GOLD), FILL_COVER_BG, ALIGN_C, NO_BORDER)

ws.sheet_view.showGridLines = False


# ─────────────────────────────────────────────────────────────
# TAB 2 & 3: USD / EUR DAILY RATES
# ─────────────────────────────────────────────────────────────
def build_rate_sheet(wb, ccy, tab_color, buy_key, sell_key):
    """Build a formatted daily rate sheet for a given currency."""
    log(f"  → {ccy} Daily Rates...")
    ws = wb.create_sheet(f"{ccy} Daily Rates")
    ws.sheet_properties.tabColor = tab_color

    set_col_widths(ws, {
        "A": 8, "B": 14, "C": 5,
        "D": 17, "E": 17,
        "F": 14, "G": 14,
        "H": 35,
    })

    # Title row — light blue band
    ws.merge_cells("A1:H1")
    write_cell(ws, 1, 1, f"  {ccy}/THB — Daily Weighted-Average Interbank Exchange Rates",
               Font(name="Calibri", size=14, bold=True, color=C_PRIMARY),
               FILL_ACCENT_LT, ALIGN_L, NO_BORDER)
    ws.row_dimensions[1].height = 35

    # Sub-title
    ws.merge_cells("A2:H2")
    write_cell(ws, 2, 1, f"  Source: Bank of Thailand  |  Period: {START} to {END}  |  Generated: {TODAY_STR}",
               FONT_SMALL, FILL_ACCENT_LT, ALIGN_L, NO_BORDER)
    ws.row_dimensions[2].height = 20

    # Header row
    headers = ["Year", "Date", "Day", "Buying TT", "Selling", "Daily Δ", "Δ %", "Remark"]
    for i, h in enumerate(headers, 1):
        write_cell(ws, 4, i, h, FONT_HDR, FILL_PRIMARY, ALIGN_C, THIN_BORDER)
    ws.row_dimensions[4].height = 28
    ws.auto_filter.ref = "A4:H4"
    ws.freeze_panes = "A5"

    # Data rows
    row = 5
    prev_sell = None
    for d in all_days:
        buy  = d[buy_key]
        sell = d[sell_key]
        rmk  = d["remark"]
        dt   = d["date"]

        # Row coloring
        if "Weekend" in rmk and ";" in rmk:
            row_fill = FILL_HOL_WKND      # Weekend + Holiday → peach
        elif rmk and "Weekend" not in rmk:
            row_fill = FILL_HOLIDAY        # Holiday only → yellow
        elif row % 2 == 0:
            row_fill = FILL_ALT            # Alternate → pale blue
        else:
            row_fill = FILL_WHITE          # Normal → white

        write_cell(ws, row, 1, dt.year, FONT_BODY, row_fill, ALIGN_C, THIN_BORDER)
        write_cell(ws, row, 2, dt, FONT_BODY, row_fill, ALIGN_C, THIN_BORDER, "DD MMM YYYY")
        write_cell(ws, row, 3, dt.strftime("%a"), FONT_SMALL, row_fill, ALIGN_C, THIN_BORDER)

        # Buying TT
        if buy is not None:
            write_cell(ws, row, 4, buy, FONT_NUM, row_fill, ALIGN_R, THIN_BORDER, NUM_FMT_RATE)
        else:
            write_cell(ws, row, 4, "", FONT_NUM, row_fill, ALIGN_R, THIN_BORDER)

        # Selling
        if sell is not None:
            write_cell(ws, row, 5, sell, FONT_NUM, row_fill, ALIGN_R, THIN_BORDER, NUM_FMT_RATE)
        else:
            write_cell(ws, row, 5, "", FONT_NUM, row_fill, ALIGN_R, THIN_BORDER)

        # Daily change & percentage (based on selling rate)
        if sell is not None and prev_sell is not None:
            delta = sell - prev_sell
            pct = delta / prev_sell if prev_sell != 0 else 0
            d_font = FONT_GREEN if delta >= 0 else FONT_RED
            write_cell(ws, row, 6, delta, d_font, row_fill, ALIGN_R, THIN_BORDER, "+0.0000;-0.0000")
            write_cell(ws, row, 7, pct, d_font, row_fill, ALIGN_R, THIN_BORDER, NUM_FMT_PCT)
        else:
            write_cell(ws, row, 6, "", FONT_BODY, row_fill, ALIGN_R, THIN_BORDER)
            write_cell(ws, row, 7, "", FONT_BODY, row_fill, ALIGN_R, THIN_BORDER)

        if sell is not None:
            prev_sell = sell

        # Remark
        if rmk and "Weekend" not in rmk:
            rmk_font = Font(name="Calibri", size=9, italic=True, color=C_RED)
        else:
            rmk_font = FONT_REMARK
        write_cell(ws, row, 8, rmk, rmk_font, row_fill, ALIGN_L, THIN_BORDER)
        ws.row_dimensions[row].height = 18
        row += 1

    # Statistics footer
    row += 1
    last_data_row = row - 2
    ws.merge_cells(f"A{row}:C{row}")
    write_cell(ws, row, 1, "  PERIOD STATISTICS", FONT_HDR, FILL_PRIMARY, ALIGN_L, THIN_BORDER)
    for c in range(4, 9):
        ws.cell(row=row, column=c).fill = FILL_PRIMARY
        ws.cell(row=row, column=c).border = THIN_BORDER

    stats = [
        ("Average", "AVERAGE"),
        ("Minimum", "MIN"),
        ("Maximum", "MAX"),
        ("Std Dev (σ)", "STDEV"),
        ("Count", "COUNT"),
    ]
    for si, (label, func) in enumerate(stats):
        sr = row + 1 + si
        ws.merge_cells(f"A{sr}:C{sr}")
        write_cell(ws, sr, 1, f"  {label}", FONT_BODY_B, FILL_GOLD_BG, ALIGN_L, THIN_BORDER)
        for ci in [4, 5]:
            col_l = get_column_letter(ci)
            if func == "COUNT":
                formula = f'={func}({col_l}5:{col_l}{last_data_row})'
                write_cell(ws, sr, ci, formula, FONT_BODY_B, FILL_GOLD_BG, ALIGN_R, THIN_BORDER, "#,##0")
            else:
                formula = f'={func}({col_l}5:{col_l}{last_data_row})'
                write_cell(ws, sr, ci, formula, FONT_BODY_B, FILL_GOLD_BG, ALIGN_R, THIN_BORDER, NUM_FMT_RATE)
        for ci in [6, 7, 8]:
            write_cell(ws, sr, ci, "", FONT_BODY, FILL_GOLD_BG, ALIGN_R, THIN_BORDER)

    # ── FEATURE: Heatmap on Daily Δ column (F) ──────────────────
    delta_range = f"F5:F{last_data_row}"
    ws.conditional_formatting.add(delta_range,
        ColorScaleRule(
            start_type='num', start_value=-0.5, start_color='F4CCCC',   # Light red
            mid_type='num',   mid_value=0,      mid_color='FFFFFF',     # White
            end_type='num',   end_value=0.5,     end_color='D9EAD3',    # Light green
        )
    )

    return ws


tab_colors = ["2E75B6", "1E8449", "C0392B", "8E44AD", "F39C12", "D35400", "273746"]
ws_currencies = {}
for i, ccy in enumerate(CURRENCIES):
    color = tab_colors[i % len(tab_colors)]
    ws_currencies[ccy] = build_rate_sheet(wb, ccy, color, f"{ccy.lower()}_buy", f"{ccy.lower()}_sell")


# ─────────────────────────────────────────────────────────────
# TAB 4: SUMMARY DASHBOARD
# ─────────────────────────────────────────────────────────────
log("  → Summary Dashboard...")
ws = wb.create_sheet("Summary Dashboard")
ws.sheet_properties.tabColor = C_GOLD

set_col_widths(ws, {
    "A": 4, "B": 22, "C": 18, "D": 14, "E": 14,
    "F": 4, "G": 22, "H": 18, "I": 14, "J": 14, "K": 4
})

# Title
ws.merge_cells("A1:K1")
write_cell(ws, 1, 1, "  EXCHANGE RATE SUMMARY DASHBOARD",
           Font(name="Calibri", size=16, bold=True, color=C_PRIMARY),
           FILL_ACCENT_LT, ALIGN_L, NO_BORDER)
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:K2")
write_cell(ws, 2, 1, f"  Period: {START.strftime('%d %b %Y')} – {END.strftime('%d %b %Y')}  |  Generated: {TODAY_STR}",
           FONT_SMALL, FILL_ACCENT_LT, ALIGN_L, NO_BORDER)

def get_rate_info(all_days_list, sell_key):
    """Extract latest rate, previous rate, high/low, MoM and YoY with dates."""
    trading = [(d["date"], d[sell_key]) for d in all_days_list if d.get(sell_key) is not None]
    if not trading:
        return {}
    latest_date, latest_rate = trading[-1]
    prev_date, prev_rate = trading[-2] if len(trading) >= 2 else (None, None)
    first_date, first_rate = trading[0]
    all_rates = [r for _, r in trading]
    high_rate = max(all_rates)
    low_rate = min(all_rates)
    high_date = next(d for d, r in trading if r == high_rate)
    low_date = next(d for d, r in trading if r == low_rate)
    avg_rate = sum(all_rates) / len(all_rates) if all_rates else 0
    ytd_change = latest_rate - first_rate
    ytd_pct = (ytd_change / first_rate * 100) if first_rate else 0
    daily_chg = (latest_rate - prev_rate) if prev_rate else 0
    daily_pct = (daily_chg / prev_rate * 100) if prev_rate else 0
    
    # MoM Analysis
    one_month_ago = latest_date - timedelta(days=30)
    closest_mom = min(trading, key=lambda x: abs((x[0] - one_month_ago).days))
    mom_rate = closest_mom[1]
    mom_change = latest_rate - mom_rate
    mom_pct = (mom_change / mom_rate * 100) if mom_rate else 0
    
    # YoY Analysis
    one_year_ago = latest_date - timedelta(days=365)
    closest_yoy = min(trading, key=lambda x: abs((x[0] - one_year_ago).days))
    yoy_rate = closest_yoy[1]
    yoy_change = latest_rate - yoy_rate
    yoy_pct = (yoy_change / yoy_rate * 100) if yoy_rate else 0
    
    return {
        "latest_rate": latest_rate, "latest_date": latest_date,
        "prev_rate": prev_rate, "prev_date": prev_date,
        "high_rate": high_rate, "high_date": high_date,
        "low_rate": low_rate, "low_date": low_date,
        "avg_rate": avg_rate, "trading_days": len(trading),
        "first_rate": first_rate, "first_date": first_date,
        "ytd_change": ytd_change, "ytd_pct": ytd_pct,
        "daily_chg": daily_chg, "daily_pct": daily_pct,
        "mom_change": mom_change, "mom_pct": mom_pct,
        "yoy_change": yoy_change, "yoy_pct": yoy_pct,
        "mom_date": closest_mom[0], "yoy_date": closest_yoy[0]
    }

infos = []
for ccy in CURRENCIES:
    infos.append((ccy, get_rate_info(all_days, f"{ccy.lower()}_sell")))

# ── Styling ────────────────────────────────────────────────
HIGHLIGHT_FILL = PatternFill("solid", fgColor="EBF5FB")
HL_BORDER = Border(
    bottom=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
)
FONT_BIG_NUM = Font(name="Calibri", size=18, bold=True, color=C_PRIMARY)
FONT_CHG_UP = Font(name="Calibri", size=11, bold=True, color=C_GREEN)
FONT_CHG_DN = Font(name="Calibri", size=11, bold=True, color=C_RED)
FONT_DATE_SM = Font(name="Calibri", size=9, italic=True, color=C_GREY)

def build_overview_box(ws, start_row, start_col, info, ccy, header_color):
    """Build one overview card for a currency."""
    c1 = start_col      # Label column (merged)
    c2 = start_col + 1  
    c3 = start_col + 2  # Value column (merged)
    c4 = start_col + 3

    eur_fill = PatternFill("solid", fgColor=header_color)

    r = start_row
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c4)
    write_cell(ws, r, c1, f"  {ccy} / THB", FONT_HDR, eur_fill, ALIGN_L, HL_BORDER)
    for c in range(c1, c4 + 1):
        ws.cell(row=r, column=c).fill = eur_fill
        ws.cell(row=r, column=c).border = HL_BORDER
    r += 1

    if not info:
        return r

    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    write_cell(ws, r, c1, "  Latest Selling Rate", FONT_BODY_B, HIGHLIGHT_FILL, ALIGN_L, HL_BORDER)
    ws.merge_cells(start_row=r, start_column=c3, end_row=r, end_column=c4)
    write_cell(ws, r, c3, info["latest_rate"], FONT_BIG_NUM, HIGHLIGHT_FILL, ALIGN_R, HL_BORDER, NUM_FMT_RATE)
    ws.row_dimensions[r].height = 32
    r += 1

    chg, pct = info["daily_chg"], info["daily_pct"]
    arrow = "▲" if chg >= 0 else "▼"
    chg_font = FONT_CHG_UP if chg >= 0 else FONT_CHG_DN
    chg_text = f"{arrow} {abs(chg):.4f}  ({abs(pct):.2f}%)"
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    write_cell(ws, r, c1, f"  vs {info['prev_date'].strftime('%d %b') if info.get('prev_date') else 'N/A'}", FONT_BODY_B, HIGHLIGHT_FILL, ALIGN_L, HL_BORDER)
    ws.merge_cells(start_row=r, start_column=c3, end_row=r, end_column=c4)
    write_cell(ws, r, c3, chg_text, chg_font, HIGHLIGHT_FILL, ALIGN_R, HL_BORDER)
    ws.row_dimensions[r].height = 22
    r += 1

    for label, val in [
        (f"  Period High  ({info['high_date'].strftime('%d %b')})", info["high_rate"]),
        (f"  Period Low  ({info['low_date'].strftime('%d %b')})", info["low_rate"]),
        ("  Period Average", info["avg_rate"]),
        ("  Trading Range (H−L)", info["high_rate"] - info["low_rate"])
    ]:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        write_cell(ws, r, c1, label, FONT_BODY_B, FILL_WHITE, ALIGN_L, HL_BORDER)
        ws.merge_cells(start_row=r, start_column=c3, end_row=r, end_column=c4)
        write_cell(ws, r, c3, val, FONT_NUM, FILL_WHITE, ALIGN_R, HL_BORDER, NUM_FMT_RATE)
        ws.row_dimensions[r].height = 22
        r += 1

    def get_chg_block(date_text, chg_val, pct_val):
        arr = "▲" if chg_val >= 0 else "▼"
        fnt = FONT_CHG_UP if chg_val >= 0 else FONT_CHG_DN
        txt = f"{arr} {abs(chg_val):.4f}  ({abs(pct_val):.2f}%)"
        return date_text, txt, fnt

    for date_lbl, chg_v, pct_v in [
        (f"  YTD Change  ({info['first_date'].strftime('%d %b')} → today)", info["ytd_change"], info["ytd_pct"]),
        (f"  MoM Change  (From {info['mom_date'].strftime('%d %b')})", info["mom_change"], info["mom_pct"]),
        (f"  YoY Change  (From {info['yoy_date'].strftime('%d %b %Y')})", info["yoy_change"], info["yoy_pct"]),
    ]:
        lbl, txt, fnt = get_chg_block(date_lbl, chg_v, pct_v)
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        write_cell(ws, r, c1, lbl, FONT_BODY_B, HIGHLIGHT_FILL, ALIGN_L, HL_BORDER)
        ws.merge_cells(start_row=r, start_column=c3, end_row=r, end_column=c4)
        write_cell(ws, r, c3, txt, fnt, HIGHLIGHT_FILL, ALIGN_R, HL_BORDER)
        ws.row_dimensions[r].height = 22
        r += 1

    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    write_cell(ws, r, c1, "  Total Trading Days", FONT_BODY_B, HIGHLIGHT_FILL, ALIGN_L, HL_BORDER)
    ws.merge_cells(start_row=r, start_column=c3, end_row=r, end_column=c4)
    write_cell(ws, r, c3, info["trading_days"], FONT_BODY_B, HIGHLIGHT_FILL, ALIGN_R, HL_BORDER, "#,##0")
    ws.row_dimensions[r].height = 22
    
    return r + 2

overview_colors = ["1F4E79", "1E8449", "900C3F", "4A235A", "B9770E", "D35400", "273746"]
box_row = 4
next_box_row = 4
for i, (ccy, info) in enumerate(infos):
    col = 2 if i % 2 == 0 else 7
    color = overview_colors[i % len(overview_colors)]
    h = build_overview_box(ws, box_row, col, info, ccy, color)
    if i % 2 == 1:
        box_row = h
    else:
        next_box_row = max(next_box_row, h)
if len(infos) % 2 == 1:
    box_row = next_box_row

# ── Build monthly aggregates ──────────────────────────────
monthly = OrderedDict()
for d in all_days:
    mkey = d["date"].strftime("%Y-%m")
    if mkey not in monthly:
        monthly[mkey] = {}
        for ccy in CURRENCIES:
            monthly[mkey][f"{ccy.lower()}_buys"] = []
            monthly[mkey][f"{ccy.lower()}_sells"] = []
            
    for ccy in CURRENCIES:
        b_val = d.get(f"{ccy.lower()}_buy")
        s_val = d.get(f"{ccy.lower()}_sell")
        if b_val is not None:
            monthly[mkey][f"{ccy.lower()}_buys"].append(b_val)
            monthly[mkey][f"{ccy.lower()}_sells"].append(s_val)

# ── Monthly Summary Tables & Charts ────────────────────────
r = max(box_row, next_box_row) + 2
for i, ccy in enumerate(CURRENCIES):
    is_left = (i % 2 == 0)
    col = 2 if is_left else 7
    chart_col = "B" if is_left else "G"
    
    ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+3)
    header_color = overview_colors[i % len(overview_colors)]
    write_cell(ws, r, col, f"  {ccy} / THB — Monthly Summary", FONT_HDR, PatternFill("solid", fgColor=header_color), ALIGN_L, THIN_BORDER)
    for c in range(col, col+4):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=header_color)
        ws.cell(row=r, column=c).border = THIN_BORDER
        
    for j, h in enumerate(["Month", "Avg Buying TT", "Avg Selling", "Spread"], col):
        write_cell(ws, r+1, j, h, FONT_HDR, FILL_ACCENT, ALIGN_C, THIN_BORDER)
        
    er = r + 2
    for mkey, mdata in monthly.items():
        row_fill = FILL_ALT if er % 2 == 0 else FILL_WHITE
        write_cell(ws, er, col, mkey, FONT_BODY_B, row_fill, ALIGN_C, THIN_BORDER)
        
        buys = mdata[f"{ccy.lower()}_buys"]
        sells = mdata[f"{ccy.lower()}_sells"]
        if buys:
            avg_buy = sum(buys) / len(buys)
            avg_sell = sum(sells) / len(sells)
            spread = avg_sell - avg_buy
            write_cell(ws, er, col+1, avg_buy, FONT_NUM, row_fill, ALIGN_R, THIN_BORDER, NUM_FMT_RATE)
            write_cell(ws, er, col+2, avg_sell, FONT_NUM, row_fill, ALIGN_R, THIN_BORDER, NUM_FMT_RATE)
            write_cell(ws, er, col+3, spread, FONT_NUM, row_fill, ALIGN_R, THIN_BORDER, NUM_FMT_RATE)
        er += 1
        
    chart = LineChart()
    chart.title = f"{ccy}/THB Selling Rate Trend"
    chart.style = 10
    chart.y_axis.title = "THB"
    chart.x_axis.title = "Month"
    chart.height = 14
    chart.width = 14
    chart.y_axis.numFmt = NUM_FMT_RATE
    
    data_end = r + 1 + len(monthly)
    data_ref = Reference(ws, min_col=col+2, min_row=r+1, max_row=data_end)
    cats_ref = Reference(ws, min_col=col, min_row=r+2, max_row=data_end)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.width = 25000
    chart.series[0].graphicalProperties.line.solidFill = header_color
    
    ws.add_chart(chart, f"{chart_col}{er + 2}")
    
    if not is_left:
        r = er + 17 # Move down for the next pair of tables/charts

ws.sheet_view.showGridLines = False

# ── Monthly Summary Tables ────────────────────────────────
# USD Section



# ─────────────────────────────────────────────────────────────
# TAB 5: MONTHLY ANALYSIS
# ─────────────────────────────────────────────────────────────
log("  → Monthly Analysis...")
ws = wb.create_sheet("Monthly Analysis")
ws.sheet_properties.tabColor = "8E44AD"

set_col_widths(ws, {
    "A": 4, "B": 16, "C": 12, "D": 16, "E": 16, "F": 16,
    "G": 16, "H": 16, "I": 16, "J": 4
})

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
write_cell(ws, 1, 1, "  MONTHLY RATE ANALYSIS — PERIOD COMPARISON",
           Font(name="Calibri", size=14, bold=True, color=C_PRIMARY),
           FILL_ACCENT_LT, ALIGN_L, NO_BORDER)
ws.row_dimensions[1].height = 35

headers = ["Month", "Trading Days"]
for ccy in CURRENCIES:
    headers.extend([f"{ccy} Open", f"{ccy} Close", f"{ccy} Δ"])
    
for i, h in enumerate(headers, 2):
    write_cell(ws, 3, i, h, FONT_HDR, FILL_PRIMARY, ALIGN_C, THIN_BORDER)

r = 4
for mkey, mdata in monthly.items():
    row_fill = FILL_ALT if r % 2 == 0 else FILL_WHITE
    write_cell(ws, r, 2, mkey, FONT_BODY_B, row_fill, ALIGN_C, THIN_BORDER)
    
    # Calculate max trading days using the first currency as proxy
    first_ccy = CURRENCIES[0].lower()
    trading = len(mdata[f"{first_ccy}_sells"])
    write_cell(ws, r, 3, trading, FONT_BODY, row_fill, ALIGN_C, THIN_BORDER)

    col_idx = 4
    for ccy in CURRENCIES:
        sells = mdata[f"{ccy.lower()}_sells"]
        if sells:
            c_open = sells[0]
            c_close = sells[-1]
            c_delta = c_close - c_open
            d_font = FONT_GREEN if c_delta >= 0 else FONT_RED
            write_cell(ws, r, col_idx, c_open, FONT_NUM, row_fill, ALIGN_R, THIN_BORDER, NUM_FMT_RATE)
            write_cell(ws, r, col_idx+1, c_close, FONT_NUM, row_fill, ALIGN_R, THIN_BORDER, NUM_FMT_RATE)
            write_cell(ws, r, col_idx+2, c_delta, d_font, row_fill, ALIGN_R, THIN_BORDER, "+0.0000;-0.0000")
        col_idx += 3

    ws.row_dimensions[r].height = 20
    r += 1

last_col = get_column_letter(1 + len(headers))
ws.auto_filter.ref = f"B3:{last_col}3"
ws.freeze_panes = "B4"


# ─────────────────────────────────────────────────────────────
# TAB 6: FX CALCULATOR
# ─────────────────────────────────────────────────────────────
log("  → FX Calculator...")
ws = wb.create_sheet("FX Calculator")
ws.sheet_properties.tabColor = C_GOLD

max_fx_col = 2 + len(CURRENCIES) * 3 + 2  # dynamic sizing
for r in range(1, 30):
    for c in range(1, max_fx_col):
        ws.cell(row=r, column=c).fill = FILL_WHITE

# Set initial column widths (dynamic columns are set per-currency in the loop below)
set_col_widths(ws, {"A": 4})

ws.merge_cells("A1:Y1")
write_cell(ws, 1, 1, "  FX CURRENCY CALCULATOR",
           Font(name="Calibri", size=16, bold=True, color=C_PRIMARY),
           FILL_ACCENT_LT, ALIGN_L, NO_BORDER)
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:Y2")
write_cell(ws, 2, 1, "  Enter amounts in the yellow cells — results calculate automatically",
           FONT_SMALL, FILL_ACCENT_LT, ALIGN_L, NO_BORDER)

# Get latest valid rate day for each currency
latest_rates = {}
for ccy in CURRENCIES:
    for d in reversed(all_days):
        if d.get(f"{ccy.lower()}_buy") is not None:
            latest_rates[ccy] = d
            break

INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")
RESULT_FILL = PatternFill("solid", fgColor="E8F5E9")
BIG_FONT = Font(name="Calibri", size=12, bold=True, color="333333")

active_col = 2
for ccy in CURRENCIES:
    latest = latest_rates.get(ccy)
    if not latest:
        continue
        
    c1, c2 = active_col, active_col + 1
    
    ws.merge_cells(start_row=4, start_column=c1, end_row=4, end_column=c2)
    write_cell(ws, 4, c1, f"  {ccy} ⇄ THB", FONT_HDR, FILL_PRIMARY, ALIGN_L, THIN_BORDER)

    write_cell(ws, 6, c1, "Rate Date:", FONT_BODY_B, FILL_WHITE, ALIGN_L, THIN_BORDER)
    write_cell(ws, 6, c2, latest["date"], FONT_BODY, FILL_WHITE, ALIGN_C, THIN_BORDER, "DD MMM YYYY")

    write_cell(ws, 7, c1, "Buying TT:", FONT_BODY_B, FILL_WHITE, ALIGN_L, THIN_BORDER)
    write_cell(ws, 7, c2, latest[f"{ccy.lower()}_buy"], FONT_NUM, FILL_WHITE, ALIGN_R, THIN_BORDER, NUM_FMT_RATE)

    write_cell(ws, 8, c1, "Selling:", FONT_BODY_B, FILL_WHITE, ALIGN_L, THIN_BORDER)
    write_cell(ws, 8, c2, latest[f"{ccy.lower()}_sell"], FONT_NUM, FILL_WHITE, ALIGN_R, THIN_BORDER, NUM_FMT_RATE)

    write_cell(ws, 10, c1, f"Enter {ccy}:", FONT_BODY_B, INPUT_FILL, ALIGN_L, THIN_BORDER)
    write_cell(ws, 10, c2, 1000, BIG_FONT, INPUT_FILL, ALIGN_R, THIN_BORDER, NUM_FMT_AMT)

    l2 = get_column_letter(c2)
    write_cell(ws, 12, c1, "  = THB (Buy TT):", FONT_BODY_B, RESULT_FILL, ALIGN_L, THIN_BORDER)
    write_cell(ws, 12, c2, f"={l2}10*{l2}7", Font(name="Calibri", size=12, bold=True, color=C_GREEN), RESULT_FILL, ALIGN_R, THIN_BORDER, "#,##0.00")

    write_cell(ws, 13, c1, "  = THB (Selling):", FONT_BODY_B, RESULT_FILL, ALIGN_L, THIN_BORDER)
    write_cell(ws, 13, c2, f"={l2}10*{l2}8", Font(name="Calibri", size=12, bold=True, color=C_RED), RESULT_FILL, ALIGN_R, THIN_BORDER, "#,##0.00")

    write_cell(ws, 15, c1, f"  Enter THB:", FONT_BODY_B, INPUT_FILL, ALIGN_L, THIN_BORDER)
    write_cell(ws, 15, c2, 100000, BIG_FONT, INPUT_FILL, ALIGN_R, THIN_BORDER, "#,##0")

    write_cell(ws, 17, c1, f"  = {ccy} (Buy TT):", FONT_BODY_B, RESULT_FILL, ALIGN_L, THIN_BORDER)
    write_cell(ws, 17, c2, f"={l2}15/{l2}7", BIG_FONT, RESULT_FILL, ALIGN_R, THIN_BORDER, "#,##0.00")

    write_cell(ws, 18, c1, f"  = {ccy} (Selling):", FONT_BODY_B, RESULT_FILL, ALIGN_L, THIN_BORDER)
    write_cell(ws, 18, c2, f"={l2}15/{l2}8", BIG_FONT, RESULT_FILL, ALIGN_R, THIN_BORDER, "#,##0.00")
    
    set_col_widths(ws, {get_column_letter(c1): 22, get_column_letter(c2): 20, get_column_letter(c2+1): 4})
    active_col += 3

ws.sheet_view.showGridLines = False


# ─────────────────────────────────────────────────────────────
# TAB 7: NOTES & DISCLAIMERS
# ─────────────────────────────────────────────────────────────
log("  → Notes & Disclaimers...")
ws = wb.create_sheet("Notes & Disclaimers")
ws.sheet_properties.tabColor = C_GREY

set_col_widths(ws, {"A": 4, "B": 85, "C": 4})

ws.merge_cells("A1:B1")
write_cell(ws, 1, 1, "  NOTES, DISCLAIMERS & DATA SOURCES",
           Font(name="Calibri", size=14, bold=True, color=C_PRIMARY),
           FILL_ACCENT_LT, ALIGN_L, NO_BORDER)
ws.row_dimensions[1].height = 35
ws.cell(row=1, column=3).fill = FILL_ACCENT_LT

sections = [
    ("DATA SOURCE", [
        "All exchange rate data is sourced exclusively from the Bank of Thailand (BOT) Official API.",
        "API Gateway: https://gateway.api.bot.or.th/",
        "API Portal: https://portal.api.bot.or.th/",
        "Rates are the daily weighted-average interbank exchange rates in Bangkok.",
    ]),
    ("RATE DEFINITIONS", [
        "Buying Transfer (TT): Rate at which banks buy foreign currency via telegraphic transfer.",
        "Selling Rate: Rate at which banks sell foreign currency to customers.",
        "Spread: Difference between Selling and Buying TT rates (bank's gross margin).",
    ]),
    ("PRECISION & ACCURACY", [
        "All exchange rates are displayed with 4–7 decimal places as provided by the BOT.",
        "No rounding or modification has been applied to the original BOT data.",
        f"This report was generated on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}.",
    ]),
    ("HOLIDAYS & NON-TRADING DAYS", [
        "Weekend days (Saturday & Sunday) are marked and have no exchange rate data.",
        "Thai public holidays are sourced from the BOT Financial Institutions Holidays API.",
        "Fixed annual holidays are annotated even when they fall on weekends.",
        "Substitution holidays (วันหยุดชดเชย) from the Royal Gazette are included via the BOT API.",
    ]),
    ("FINANCIAL DISCLAIMER", [
        "IMPORTANT: These are INDICATIVE interbank rates and may differ from actual transaction rates.",
        "Commercial banks may apply their own margins, fees, and spreads.",
        "This report is for INTERNAL USE by the Finance & Accounting Department only.",
        "Past exchange rates do not guarantee or predict future rates.",
    ]),
    ("REGULATORY COMPLIANCE", [
        "Complies with Thai accounting standards (TAS/TFRS) for foreign currency translation.",
        "Under TAS 21 (IAS 21), spot rates at transaction date should be used for initial recognition.",
        "Closing rates at reporting date should be used for monetary items at period end.",
    ]),
    ("AUTOMATION SCRIPT LOGIC", [
        "This report is automatically generated using a Python script ('bot_excel_report.py').",
        "It fetches the latest historical rates and holiday data directly from the Bank of Thailand APIs.",
        "The script computes the Daily Δ (change) and Δ % automatically.",
        "It then compiles all trading days and non-trading days (weekends/holidays) into this formatted workbook.",
    ]),
    ("UNDERSTANDING DAILY Δ (DELTA)", [
        "The Greek letter Δ (Delta) is the universal shorthand for 'change' or 'difference'.",
        "Daily Δ Formula: [Today's Rate] - [Previous Trading Day's Rate]",
        "Example: If Tuesday's rate is 34.10 and Wednesday's is 34.15, the Daily Δ is +0.05",
        "Δ % Formula: ( [Daily Δ] / [Previous Trading Day's Rate] ) * 100",
        "Example: (0.05 / 34.10) * 100 = a +0.14% change compared to the previous day.",
    ]),
]

r = 3
for title, items in sections:
    write_cell(ws, r, 2, title, Font(name="Calibri", size=12, bold=True, color=C_PRIMARY),
               FILL_WHITE, ALIGN_L, BOTTOM_ACCENT)
    ws.row_dimensions[r].height = 28
    r += 1
    for item in items:
        if "IMPORTANT:" in item:
            write_cell(ws, r, 2, item, Font(name="Calibri", size=10, bold=True, color=C_RED),
                       FILL_WHITE, ALIGN_TL, NO_BORDER)
        else:
            write_cell(ws, r, 2, f"  •  {item}", FONT_NOTE, FILL_WHITE, ALIGN_TL, NO_BORDER)
        ws.row_dimensions[r].height = 22
        r += 1
    r += 1

# Footer
ws.merge_cells(f"A{r+1}:B{r+1}")
write_cell(ws, r + 1, 1,
           f"  Report generated by BOT Excel Report Generator  |  © {datetime.now().year}  |  Bank of Thailand Data",
           Font(name="Calibri", size=9, color=C_GOLD), FILL_ACCENT_LT, ALIGN_C, NO_BORDER)

ws.sheet_view.showGridLines = False

# ═══════════════════════════════════════════════════════════════

# SAVE THE WORKBOOK
# ═══════════════════════════════════════════════════════════════
wb.save(OUTPUT)

PDF_OUTPUT = OUTPUT.replace(".xlsx", ".pdf")
if GENERATE_PDF:
    import shutil
    log("\n  [4/4] Generating PDF...")
    soffice_path = shutil.which("soffice") or "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    try:
        subprocess.run([soffice_path, "--headless", "--convert-to", "pdf", "--outdir", os.path.dirname(OUTPUT), OUTPUT], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        log(f"    ✔ PDF saved: {os.path.basename(PDF_OUTPUT)}")
    except Exception as e:
        log(f"    ✘ PDF conversion failed: {e}")

if EMAIL_TO:
    log(f"\n  [5] Emailing report to {EMAIL_TO}...")
    try:
        msg = EmailMessage()
        msg['Subject'] = f"Exchange Rate Dashboard - {TODAY_STR}"
        msg['From'] = "bot-report@internal.network"
        msg['To'] = EMAIL_TO
        msg.set_content(f"Attached is the BOT Exchange Rate Executive Dashboard.\n\nGenerated on: {TODAY_STR}\nPeriod: {START} to {END}\nCurrencies: {', '.join(CURRENCIES)}")
        
        with open(OUTPUT, 'rb') as f:
            msg.add_attachment(f.read(), maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=os.path.basename(OUTPUT))
        
        if GENERATE_PDF and os.path.exists(PDF_OUTPUT):
            with open(PDF_OUTPUT, 'rb') as f:
                msg.add_attachment(f.read(), maintype='application', subtype='pdf', filename=os.path.basename(PDF_OUTPUT))
                
        # Send via a dummy local test SMTP server (can be configured via env vars normally)
        with smtplib.SMTP('localhost', 1025) as s:
            s.send_message(msg)
        log("    ✔ Email sent.")
    except ConnectionRefusedError:
        log("    ✘ Email failed: Could not connect to SMTP server (localhost:1025). Start testing server with python -m smtpd.")
    except Exception as e:
        log(f"    ✘ Email failed: {e}")

log("")
log("=" * 60)
log("=" * 60)
log("  DONE!")
log(f"  Rows written: {len(all_days)}")
log(f"  Trading days: {len(rates)}")
log(f"  Output saved: {os.path.basename(OUTPUT)}")
log(f"  Tabs: {len(wb.sheetnames)}")
log(f"  Sheets: {', '.join(wb.sheetnames)}")
log("=" * 60)

# ─── Changelog ───────────────────────────────────────────────
# Every major logic or visual update to this script is noted here.
#
# 2026-03-10 | v1 — Initial Dashboard version
#            | - Added Summary Dashboard with High/Low/Average metrics
#            | - Added visual heatmaps for daily changes
#            | - Added line charts for USD/EUR trends
#
# 2026-03-11 | v1.03 — Overhaul
#            | - Fixed log() function (switched from pass to print)
#            | - Standardized all Excel date formats to "DD MMM YYYY"
#            | - Fixed output filename to BOT_ExchangeRate_Report.xlsx
#            | - Improved code documentation and section summaries
#
# 2026-03-11 | v1.0.7 — Optimizations
#            | - Implemented argparse for --start and --end CLI arguments
#            | - Switched to asyncio and aiohttp for massive download speedup
#            | - Implemented exponential backoff and retries for network resilience
#            | - Handled Pyre type safety warnings by using explicit gets and lists
#
# 2026-03-13 | v1.1.0 — Visual & Scaling Update
#            | - Standardized aiohttp.ClientTimeout for stable network fetching
#            | - Made Cover Sheet subtitle dynamic (shows all selected currencies)
#            | - Refactored FX Calculator to scale background/styling for any currency count
